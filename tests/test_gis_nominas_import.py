"""GIS Nóminas — importación multihoja y extracción."""

from __future__ import annotations

import sqlite3
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from modules.gestion_idse_sua.nominas.import_service import (
    confirm_classifications,
    confirm_period,
    extract_sheet_workers,
    inspect_workbook,
    register_import,
)
from modules.gestion_idse_sua.nominas import repository as repo
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables


def _build_multisheet_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "3 al 9 jun"
    ws["A2"] = "NOMINA"
    ws["B3"] = "NO."
    ws["C3"] = "NOMBRE DE EMPLEADO"
    ws["D3"] = "PLANTA"
    ws["E3"] = "PUESTO"
    ws["F3"] = "CUENTA"
    ws["B4"] = "101"
    ws["C4"] = "JUAN PEREZ LOPEZ"
    ws["D4"] = "FLOTADO"
    ws["E4"] = "Operador"
    ws["F4"] = "123"
    aux = wb.create_sheet("Auxiliar")
    aux["A1"] = "Totales generales"
    hidden = wb.create_sheet("Oculta")
    hidden["A1"] = "NOMBRE DE EMPLEADO"
    hidden.sheet_state = "hidden"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def conn(tmp_path):
    db = tmp_path / "gis_import.db"
    connection = sqlite3.connect(db)
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    connection.commit()
    yield connection
    connection.close()


def test_inspect_multisheet_workbook():
    data = inspect_workbook(_build_multisheet_bytes(), filename="nomina.xlsx")
    assert len(data["sheets"]) == 3
    nomina = data["sheets"][0]
    assert nomina["suggested_classification"] == "nomina"
    assert nomina["estimated_rows"] == 1
    assert data["sheets"][1]["suggested_classification"] == "auxiliar"


def test_register_import_persists_sheets(conn):
    data = _build_multisheet_bytes()
    payload = register_import(conn, file_bytes=data, filename="nomina.xlsx", uploaded_by="tester")
    sheets = conn.execute("SELECT COUNT(*) FROM gis_nomina_sheets WHERE import_id = ?", (payload["import_id"],)).fetchone()[0]
    assert sheets == 3
    stored = conn.execute(
        "SELECT file_content FROM gis_nomina_imports WHERE id = ?", (payload["import_id"],)
    ).fetchone()[0]
    assert bytes(stored) == data


def test_extract_workers_from_fixture(conn):
    fixture = Path("tests/fixtures/nomina_carrier_anon.xlsx").read_bytes()
    reg = register_import(conn, file_bytes=fixture, filename="carrier.xlsx", uploaded_by="tester")
    sheet_id = conn.execute(
        "SELECT id FROM gis_nomina_sheets WHERE import_id = ? AND suggested_classification = 'nomina'",
        (reg["import_id"],),
    ).fetchone()[0]
    conn.execute(
        "UPDATE gis_nomina_sheets SET confirmed_classification = 'nomina' WHERE id = ?",
        (sheet_id,),
    )
    confirm_period(conn, sheet_id, fecha_inicio="01/06/2026", fecha_fin="07/06/2026")
    out = extract_sheet_workers(conn, file_bytes=fixture, sheet_id=sheet_id, headcount_rows=[])
    assert out["workers_count"] == 4


def test_duplicate_period_warns(conn):
    from modules.gestion_idse_sua.nominas.repository import find_conflicting_periods

    conn.execute(
        "INSERT INTO gis_nomina_imports (original_filename, file_hash, uploaded_at, status) VALUES (?,?,?,?)",
        ("a.xlsx", "h1", "2026-01-01", "classified"),
    )
    conn.execute(
        "INSERT INTO gis_nomina_sheets (import_id, sheet_index, sheet_name, is_hidden, confirmed_classification, estimated_rows) VALUES (1,0,'S1',0,'nomina',1)"
    )
    conn.execute(
        "INSERT INTO gis_nomina_sheets (import_id, sheet_index, sheet_name, is_hidden, confirmed_classification, estimated_rows) VALUES (1,1,'S2',0,'nomina',1)"
    )
    confirm_period(conn, 1, fecha_inicio="01/06/2026", fecha_fin="07/06/2026")
    conflicts = find_conflicting_periods(conn, fecha_inicio="01/06/2026", fecha_fin="07/06/2026", exclude_sheet_id=2)
    assert len(conflicts) == 1
    out = confirm_period(conn, 2, fecha_inicio="01/06/2026", fecha_fin="07/06/2026")
    assert out["conflicts"]


def test_import_resume_states_and_archive_restore(conn):
    data = _build_multisheet_bytes()
    registered = register_import(conn, file_bytes=data, filename="nomina.xlsx", uploaded_by="tester")
    import_id = registered["import_id"]
    assert repo.resolve_import_resume(conn, import_id)["state"] == "uploaded"

    sheet_id = int(registered["sheets"][0]["id"])
    conn.execute("UPDATE gis_nomina_imports SET status = 'classified' WHERE id = ?", (import_id,))
    assert repo.resolve_import_resume(conn, import_id)["state"] == "classified"
    confirm_classifications(conn, import_id, {sheet_id: "nomina"})
    assert repo.resolve_import_resume(conn, import_id)["state"] == "period_pending"

    confirm_period(conn, sheet_id, fecha_inicio="01/06/2026", fecha_fin="07/06/2026")
    extract_sheet_workers(conn, file_bytes=data, sheet_id=sheet_id, headcount_rows=[])
    resumed = repo.resolve_import_resume(conn, import_id)
    assert resumed["state"] == "period_confirmed"
    comparative_id = repo.create_comparative(
        conn, period_id=resumed["period_id"], cliente="VITROFLEX", generated_by="tester"
    )
    conn.commit()
    ready = repo.resolve_import_resume(conn, import_id)
    assert ready["state"] == "comparative_ready"
    assert ready["comparative_id"] == comparative_id

    repo.archive_import(conn, import_id, archived_by="tester", reason="QA")
    conn.commit()
    assert repo.resolve_import_resume(conn, import_id)["state"] == "archived"
    archived = repo.get_import(conn, import_id)
    assert archived["archived_by"] == "tester"
    assert archived["archive_reason"] == "QA"
    repo.restore_import(conn, import_id)
    conn.commit()
    assert repo.resolve_import_resume(conn, import_id)["state"] == "comparative_ready"


def test_import_without_extracted_sheets_is_incomplete(conn):
    import_id = repo.create_import(
        conn,
        original_filename="fallida.xlsx",
        file_hash="hash",
        uploaded_by="tester",
        file_content=None,
    )
    conn.commit()
    assert repo.resolve_import_resume(conn, import_id)["state"] == "incomplete"


def test_archive_dependency_summary_keeps_reports_and_movements(conn):
    from modules.gestion_idse_sua.reportes.schema import ensure_gis_monthly_tables

    data = _build_multisheet_bytes()
    registered = register_import(conn, file_bytes=data, filename="nomina.xlsx", uploaded_by="tester")
    import_id = registered["import_id"]
    sheet_id = int(registered["sheets"][0]["id"])
    confirm_classifications(conn, import_id, {sheet_id: "nomina"})
    period = confirm_period(conn, sheet_id, fecha_inicio="01/06/2026", fecha_fin="07/06/2026")
    extract_sheet_workers(conn, file_bytes=data, sheet_id=sheet_id, headcount_rows=[])
    comparative_id = repo.create_comparative(
        conn, period_id=period["period_id"], cliente="VITROFLEX", generated_by="tester"
    )
    result_id = repo.insert_result(
        conn,
        comparative_id,
        {
            "resultado": "Posible baja",
            "semaforo": "rojo",
            "tipo_sugerido": "BAJA",
        },
    )
    repo.mark_result_conversion(conn, result_id, status="converted", movimiento_id="99")
    ensure_gis_monthly_tables(conn)
    report_id = conn.execute(
        """
        INSERT INTO gis_monthly_reports
            (cliente, mes, anio, estado, created_at, updated_at)
        VALUES ('VITROFLEX', 6, 2026, 'generado', '2026-07-01', '2026-07-01')
        """
    ).lastrowid
    conn.execute(
        """
        INSERT INTO gis_monthly_report_weeks
            (report_id, period_id, sort_order, included_at)
        VALUES (?, ?, 1, '2026-07-01')
        """,
        (report_id, period["period_id"]),
    )
    conn.commit()

    dependencies = repo.import_dependencies(conn, import_id)
    assert dependencies["comparatives"] == 1
    assert dependencies["movements"] == 1
    assert dependencies["reports"] == 1
    repo.archive_import(conn, import_id, archived_by="tester")
    conn.commit()
    assert conn.execute("SELECT COUNT(*) FROM gis_nomina_workers").fetchone()[0] == 1
    assert conn.execute("SELECT COUNT(*) FROM gis_monthly_reports").fetchone()[0] == 1
