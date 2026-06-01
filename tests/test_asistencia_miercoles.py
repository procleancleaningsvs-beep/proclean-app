from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from modules.nomina.asistencia_miercoles import (
    build_miercoles_template_file,
    is_miercoles_v3_excel,
    parse_miercoles_v3_excel,
)
from modules.nomina.db import NominaBaseRow


def _build_minimal_miercoles_workbook_bytes() -> bytes:
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Asistencia"
    ws_a["A2"] = "SEMANA:"
    ws_a["B2"] = "NOMINA:"
    ws_a["E2"] = "COORDINADOR:"
    ws_a["A5"] = "NOMBRE DE EMPLEADO"
    ws_a["B5"] = "CLIENTE"
    ws_a["C5"] = "PLANTA"
    ws_a["D5"] = "TURNO"
    ws_a["E5"] = "PUESTO"

    ws_c = wb.create_sheet("Compilado")
    ws_c["A2"] = "SEMANA:"
    ws_c["B2"] = "NOMINA:"
    ws_c["E2"] = "COORDINADOR:"
    headers = [
        "NOMBRE COMPLETO",
        "PUESTO",
        "BANCO",
        "CUENTA",
        "CLIENTES / TURNOS DONDE LABORÓ",
        "V22",
        "S23",
        "D24",
        "L25",
        "M26",
        "M27",
        "J28",
        "DÍAS ÚNICOS LABORADOS",
        "DÍAS ADICIONALES DETECTADOS",
        "EVENTOS ADICIONALES MISMO DÍA",
        "HORAS EXTRA",
        "HORAS EXTRA NORMALES",
        "DÍAS CUBIERTOS NORMALES",
        "VACACIONES LABORADAS",
        "PRIMA VACACIONAL",
        "BONO",
        "DEDUCCIONES",
        "OBSERVACIONES",
    ]
    for col, header in enumerate(headers, start=1):
        ws_c.cell(row=4, column=col, value=header)
    ws_c.cell(row=5, column=1, value="Empleado Demo")
    ws_c.cell(row=5, column=2, value="Operador")
    ws_c.cell(row=5, column=3, value="Banorte")
    ws_c.cell(row=5, column=4, value="1234")
    ws_c.cell(row=5, column=5, value="GM-A")
    ws_c.cell(row=5, column=6, value="A")
    ws_c.cell(row=5, column=7, value="A")
    ws_c.cell(row=5, column=16, value=2)
    ws_c.cell(row=5, column=17, value=1)
    ws_c.cell(row=5, column=18, value=0)
    ws_c.cell(row=5, column=20, value="N/A")
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_detects_miercoles_v3_template_bytes():
    fixture = Path(__file__).resolve().parent.parent / "modules" / "nomina" / "templates_excel" / "Plantilla_Asistencia_Miercoles_v3.xlsx"
    assert fixture.exists()
    assert is_miercoles_v3_excel(fixture.read_bytes()) is True


def test_build_miercoles_template_updates_headers_and_period():
    data = build_miercoles_template_file(
        fecha_inicio=date(2026, 5, 22),
        fecha_fin=date(2026, 5, 28),
        coordinador="QA",
        base_rows=[
            NominaBaseRow(
                nombre_empleado="Trabajador Uno",
                cliente="GM",
                planta="P1",
                puesto="Aux",
                banco="Banorte",
                cuenta="0001",
            )
        ],
    )
    wb = load_workbook(BytesIO(data))
    ws_a = wb["Asistencia"]
    ws_c = wb["Compilado"]
    assert "22 al 28 may 2026" in str(ws_a["A2"].value or "")
    assert "MIERCOLES / AURIGA" in str(ws_a["B2"].value or "")
    assert "QA" in str(ws_a["E2"].value or "")
    assert ws_a["F5"].value == "V22"
    assert ws_a["L5"].value == "J28"
    assert ws_c["F4"].value == "V22"
    assert ws_c["L4"].value == "J28"
    found = False
    for row in range(1, ws_a.max_row + 1):
        if ws_a.cell(row=row, column=1).value == "Trabajador Uno":
            found = True
            break
    assert found


def test_parse_miercoles_v3_reads_compilado_rows(monkeypatch):
    payload = _build_minimal_miercoles_workbook_bytes()
    monkeypatch.setattr(
        "modules.nomina.asistencia_miercoles._try_recalculate_with_libreoffice",
        lambda file_bytes, timeout_sec=180: (file_bytes, None),
    )
    parsed = parse_miercoles_v3_excel(payload, "miercoles.xlsx")
    assert parsed["template_kind"] == "miercoles_v3"
    assert parsed["total_rows"] == 1
    row = parsed["rows"][0]
    assert row["nombre_empleado"] == "Empleado Demo"
    assert row["cliente"] == "GM"
    assert row["dia_1_header"] == "V22"
    assert row["dia_1_value"] == "A"
    assert row["he"] == "2"
