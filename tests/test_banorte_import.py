from __future__ import annotations

from pathlib import Path

import pytest

from modules.nomina.banorte.import_service import (
    import_nomina_banorte_xlsx,
    import_reporte_detallado_xlsx,
)
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.validators import is_banorte_employee_substituted_comment
from modules.nomina.db import ensure_nomina_tables

FIXTURES = Path(__file__).resolve().parent / "fixtures" / "banorte"


@pytest.fixture
def db_path(tmp_path):
    path = tmp_path / "banorte_import.db"
    conn = connect(path)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    return str(path)


def test_special_comment_match_strict():
    assert is_banorte_employee_substituted_comment(
        "El número de empleado ya existía se asignó el número de cuenta como tu número de Empleado"
    )
    assert not is_banorte_employee_substituted_comment("cliente existente acuda a sucursal")


def test_import_altas_exitosos_manuals_and_fallidos(db_path):
    data = (FIXTURES / "synthetic_altas.xlsx").read_bytes()
    result = import_nomina_banorte_xlsx(db_path, data, "synthetic_altas.xlsx", "tester")
    assert result.mutated is True
    assert result.count_excluidos_hoja_fallidos_total == 3
    assert result.count_fallidos_estatus == 2
    assert result.count_fallidos_hoja_sin_estatus == 1
    assert result.count_fallidos_estatus != result.count_excluidos_hoja_fallidos_total

    conn = connect(db_path)
    # FALLIDOS must not create beneficiaries named FALLIDO
    n_bad = conn.execute(
        "SELECT COUNT(*) FROM nomina_banorte_beneficiaries WHERE nombre_original LIKE 'FALLIDO%'"
    ).fetchone()[0]
    assert n_bad == 0
    # special comment effective emp = account
    row = conn.execute(
        "SELECT employee_number_requested, employee_number_effective, account_number, banorte_employee_substituted "
        "FROM nomina_banorte_beneficiaries WHERE nombre_normalizado=?",
        ("CARLA DEMO TRES",),
    ).fetchone()
    assert row is not None
    assert row["banorte_employee_substituted"] == 1
    assert row["employee_number_effective"] == row["account_number"]
    assert row["employee_number_requested"] == "0000000099"
    # lower duplicate wins -> ANA V2 active, prior inactive but still EXITOSO validation
    ana_active = conn.execute(
        "SELECT * FROM nomina_banorte_beneficiaries WHERE account_number='1321000001' AND record_status='ACTIVO'"
    ).fetchone()
    assert ana_active["nombre_original"] == "ANA DEMO UNO V2"
    assert ana_active["validation_status"] == "IMPORTADO_EXITOSO"
    ana_old = conn.execute(
        "SELECT * FROM nomina_banorte_beneficiaries WHERE replaces_id IS NULL AND account_number='1321000001' AND record_status='INACTIVO_REEMPLAZADO'"
    ).fetchone()
    # there should be an inactive previous version
    inactive = conn.execute(
        "SELECT COUNT(*) FROM nomina_banorte_beneficiaries WHERE account_number='1321000001' AND record_status='INACTIVO_REEMPLAZADO'"
    ).fetchone()[0]
    assert inactive >= 1
    # manual complete present
    manual = conn.execute(
        "SELECT * FROM nomina_banorte_beneficiaries WHERE nombre_normalizado='DIANA MANUAL' AND record_status='ACTIVO'"
    ).fetchone()
    assert manual["validation_status"] == "MANUAL_PENDIENTE_VALIDACION"
    # incomplete not imported
    assert (
        conn.execute(
            "SELECT COUNT(*) FROM nomina_banorte_beneficiaries WHERE nombre_normalizado='ELIAS INCOMPLETO'"
        ).fetchone()[0]
        == 0
    )
    conn.close()


def test_reimport_same_sha_without_confirm_no_mutation(db_path):
    data = (FIXTURES / "synthetic_altas.xlsx").read_bytes()
    first = import_nomina_banorte_xlsx(db_path, data, "synthetic_altas.xlsx", "tester")
    assert first.mutated
    conn = connect(db_path)
    before = conn.execute("SELECT COUNT(*) FROM nomina_banorte_beneficiaries").fetchone()[0]
    batches_before = conn.execute("SELECT COUNT(*) FROM nomina_banorte_import_batches").fetchone()[0]
    conn.close()
    second = import_nomina_banorte_xlsx(db_path, data, "synthetic_altas.xlsx", "tester")
    assert second.mutated is False
    assert second.message == "duplicate_sha_confirmation_required"
    conn = connect(db_path)
    assert conn.execute("SELECT COUNT(*) FROM nomina_banorte_beneficiaries").fetchone()[0] == before
    assert conn.execute("SELECT COUNT(*) FROM nomina_banorte_import_batches").fetchone()[0] == batches_before
    conn.close()


def test_reimport_same_sha_with_confirm_idempotent_rows(db_path):
    data = (FIXTURES / "synthetic_altas.xlsx").read_bytes()
    import_nomina_banorte_xlsx(db_path, data, "synthetic_altas.xlsx", "tester")
    conn = connect(db_path)
    before_active = conn.execute(
        "SELECT COUNT(*) FROM nomina_banorte_beneficiaries WHERE record_status='ACTIVO'"
    ).fetchone()[0]
    conn.close()
    again = import_nomina_banorte_xlsx(
        db_path, data, "synthetic_altas.xlsx", "tester", reimport_confirmed=True
    )
    assert again.mutated is True
    assert again.batch_id is not None
    conn = connect(db_path)
    after_active = conn.execute(
        "SELECT COUNT(*) FROM nomina_banorte_beneficiaries WHERE record_status='ACTIVO'"
    ).fetchone()[0]
    # Should not explode active population
    assert after_active == before_active
    no_change = conn.execute(
        "SELECT COUNT(*) FROM nomina_banorte_import_rows WHERE decision='REIMPORT_NO_CHANGE'"
    ).fetchone()[0]
    assert no_change >= 1
    conn.close()


def test_reporte_excludes_fallido_and_validates_manual(db_path):
    altas = (FIXTURES / "synthetic_altas.xlsx").read_bytes()
    import_nomina_banorte_xlsx(db_path, altas, "synthetic_altas.xlsx", "tester")
    reporte = (FIXTURES / "synthetic_reporte.xlsx").read_bytes()
    result = import_reporte_detallado_xlsx(db_path, reporte, "synthetic_reporte.xlsx", "tester")
    assert result.count_fallidos_estatus >= 1
    conn = connect(db_path)
    assert (
        conn.execute(
            "SELECT COUNT(*) FROM nomina_banorte_beneficiaries WHERE nombre_normalizado='FALLIDO REPORTE'"
        ).fetchone()[0]
        == 0
    )
    # manual Diana should be inactivated and replaced by validated
    diana_versions = conn.execute(
        "SELECT validation_status, record_status, replaces_id FROM nomina_banorte_beneficiaries "
        "WHERE nombre_normalizado='DIANA MANUAL' ORDER BY id"
    ).fetchall()
    assert any(r["record_status"] == "INACTIVO_REEMPLAZADO" and r["validation_status"] == "MANUAL_PENDIENTE_VALIDACION" for r in diana_versions)
    assert any(r["record_status"] == "ACTIVO" and r["validation_status"] == "IMPORTADO_EXITOSO" and r["replaces_id"] for r in diana_versions)
    fabian = conn.execute(
        "SELECT employee_number_requested, employee_number_effective, account_number, banorte_employee_substituted "
        "FROM nomina_banorte_beneficiaries WHERE nombre_normalizado='FABIAN ESPECIAL' AND record_status='ACTIVO'"
    ).fetchone()
    assert fabian["banorte_employee_substituted"] == 1
    assert fabian["employee_number_effective"] == fabian["account_number"]
    assert fabian["employee_number_requested"] == "0000000777"
    conn.close()


def test_account_conflict_does_not_auto_activate(db_path):
    data = (FIXTURES / "synthetic_altas.xlsx").read_bytes()
    import_nomina_banorte_xlsx(db_path, data, "synthetic_altas.xlsx", "tester")
    # Force conflict: insert another EXITOSO-like row via second mini workbook would be heavy;
    # use repository SQL through import of a tiny custom workbook.
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "ALTAS"
    headers = [
        "#",
        "Número de empleado",
        "Curp ",
        "Nombre del empleado",
        "Producto",
        "Tipo de tarjeta",
        "Número de tarjeta",
        "Número de cuenta",
        "Cuenta CLABE",
        "Fecha de alta solicitud",
        "Estatus",
        "Comentarios",
    ]
    for i, h in enumerate(headers, start=2):
        ws.cell(3, i, h)
    ws.cell(4, 3, "0000007777")
    ws.cell(4, 4, "ZZZZ900101HDFRRR99")
    ws.cell(4, 5, "PERSONA DISTINTA CONFLICTO")
    ws.cell(4, 9, "1321000002")  # Beto's account
    ws.cell(4, 12, "EXITOSO")
    wb.create_sheet("FALLIDOS")
    buf = io.BytesIO()
    wb.save(buf)
    result = import_nomina_banorte_xlsx(db_path, buf.getvalue(), "conflict.xlsx", "tester")
    assert result.count_conflictos >= 1
    conn = connect(db_path)
    beto = conn.execute(
        "SELECT record_status FROM nomina_banorte_beneficiaries WHERE account_number='1321000002' AND nombre_normalizado='BETO DEMO DOS'"
    ).fetchone()
    assert beto["record_status"] == "ACTIVO"
    conflict = conn.execute(
        "SELECT record_status FROM nomina_banorte_beneficiaries WHERE nombre_normalizado='PERSONA DISTINTA CONFLICTO'"
    ).fetchone()
    assert conflict["record_status"] == "CONFLICTO_CRITICO"
    conn.close()


def test_import_rollback_on_failure(db_path, monkeypatch):
    data = (FIXTURES / "synthetic_altas.xlsx").read_bytes()

    def boom(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "modules.nomina.banorte.import_service._insert_batch",
        boom,
    )
    with pytest.raises(RuntimeError):
        import_nomina_banorte_xlsx(db_path, data, "synthetic_altas.xlsx", "tester")
    conn = connect(db_path)
    assert conn.execute("SELECT COUNT(*) FROM nomina_banorte_beneficiaries").fetchone()[0] == 0
    assert conn.execute("SELECT COUNT(*) FROM nomina_banorte_import_batches").fetchone()[0] == 0
    conn.close()
