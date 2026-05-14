"""Generate small, anonymized fixtures for nomina + CONTPAQ parsers.

This script writes deterministic minimal Excel files used by tests/test_parametros.py.
Real client data must never be committed; this fixture covers only the
structural cases the parser needs to handle.

Run from repo root: ``python tests/fixtures/build_parametros_fixtures.py``.
"""
from pathlib import Path

from openpyxl import Workbook

FIXTURE_DIR = Path(__file__).resolve().parent


def build_nomina_carrier_like() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina"
    ws["A2"] = "NOMINA EJEMPLO"
    # Header row 3 emulating Carrier layout
    headers = [
        ("B", "NO."),
        ("C", "NOMBRE DE EMPLEADO"),
        ("D", "PLANTA"),
        ("E", "PUESTO"),
        ("F", "BANCO"),
        ("G", "CUENTA"),
        ("H", "SALARIO OPERATIVO"),
        ("I", "VALOR X HE"),
    ]
    for letter, label in headers:
        ws[f"{letter}3"] = label
    rows = [
        ("121", "Empleado Demo Uno", "A", "Aux", "Banorte", "1111111111", 2470, 68.75),
        ("138", "Empleado Demo Dos", "A", "Aux", "Bancoppel", "2222222222", 2500, 89.28),
        ("296", "Empleado Sin Salario", "A", "Aux", "Mifel", "3333333333", None, 75.0),
        ("298", "Empleado Sin HE", "A", "Aux", "Banorte", "4444444444", 2500, None),
    ]
    for offset, vals in enumerate(rows, start=4):
        for col_idx, value in enumerate(vals, start=2):
            ws.cell(row=offset, column=col_idx, value=value)
    out = FIXTURE_DIR / "nomina_carrier_anon.xlsx"
    wb.save(out)
    return out


def build_nomina_pepsi_like() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pepsi"
    ws["B2"] = "NOMINA PEPSI EJEMPLO"
    headers = [
        ("B", "NOMBRE DE EMPLEADO"),
        ("C", "LOCALIDAD"),
        ("D", "FRONTERA"),
        ("E", "PUESTO"),
        ("F", "BANCO"),
        ("G", "CUENTA"),
        ("H", "SALARIO OPERATIVO"),
        ("I", "VALOR X HE"),
    ]
    for letter, label in headers:
        ws[f"{letter}3"] = label
    rows = [
        ("Empleada Tijuana Uno", "Tijuana", "VERDADERO", "Aux limpieza", "Banorte", "11", 3500, 110.71),
        ("Empleada Mexicali", "Mexicali", "VERDADERO", "Aux limpieza", "Banorte", "12", 3500, 110.71),
        ("Empleada Monterrey", "Apodaca", "FALSO", "Aux limpieza", "Banorte", "13", 2800, 95.0),
        ("Empleada Sin Loc", None, "VERDADERO", "Aux limpieza", "Banorte", "14", 2800, 95.0),
    ]
    for offset, vals in enumerate(rows, start=4):
        for col_idx, value in enumerate(vals, start=2):
            ws.cell(row=offset, column=col_idx, value=value)
    out = FIXTURE_DIR / "nomina_pepsi_anon.xlsx"
    wb.save(out)
    return out


def build_contpaq_like() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"
    headers = [
        "Código",
        "Fecha de alta",
        "Fecha de baja",
        "Fecha de reingreso",
        "Apellido paterno",
        "Apellido materno",
        "Nombre",
        "Nombre completo",
        "Salario diario",
        "Estatus empleado",
        "Departamento",
        "Zona de salario",
        "Num seguridad social",
        "Puesto",
        "Registro patronal del IMSS",
    ]
    for c, label in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=label)
    rows = [
        ("001", "02/04/2024", "06/03/2025", "", "DEMO", "UNO", "EMPLEADO", "Empleado Demo Uno", 278.8, "A", "LIMPIEZA", "B", "12345678901", "AUXILIAR", "REG-A"),
        ("002", "01/06/2023", "", "", "DEMO", "DOS", "EMPLEADO", "Empleado Demo Dos", 315.04, "A", "LIMPIEZA", "B", "23456789012", "AUXILIAR", "REG-A"),
        ("003", "27/03/2024", "", "", "TIJUANA", "UNO", "EMPLEADA", "Empleada Tijuana Uno", 440.87, "A", "LIMPIEZA", "F", "34567890123", "AUXILIAR", "REG-B"),
        ("004", "05/02/2024", "", "", "FALTA", "NSS", "EMPLEADA", "Empleada Sin NSS", 248.93, "A", "LIMPIEZA", "B", "", "AUXILIAR", "REG-A"),
    ]
    for r_idx, vals in enumerate(rows, start=2):
        for c_idx, val in enumerate(vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    out = FIXTURE_DIR / "contpaq_anon.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    for fn in (build_nomina_carrier_like, build_nomina_pepsi_like, build_contpaq_like):
        path = fn()
        print("wrote", path)
