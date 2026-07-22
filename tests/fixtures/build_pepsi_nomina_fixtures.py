"""Build anonymized Pepsi-layout fixtures from operational evidence (read-only source).

CI uses the committed ``gis_pepsi_nomina_anon.xlsx`` and does not run this script.

Local regeneration only (outside CI):
  set GIS_PEPSI_EVIDENCE_XLSM=<path-to-local-evidence.xlsm>
  py -3 tests/fixtures/build_pepsi_nomina_fixtures.py
"""

from __future__ import annotations

import os
import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

FIXTURE_DIR = Path(__file__).resolve().parent

FAKE_WORKERS = [
    {
        "num": "9001",
        "nombre": "EMPLEADO PRUEBA 001",
        "planta": "PLANTA PRUEBA NORTE",
        "frontera": "FALSO",
        "puesto": "AUX PRUEBA",
        "banco": "BANCO PRUEBA",
        "cuenta": "0000000001",
        "salario": 3500,
        "valor_he": 100.0,
        "codes": ["A", "A", "F", "D", "F", "V", "A"],
    },
    {
        "num": "9002",
        "nombre": "EMPLEADO PRUEBA 002",
        "planta": "PLANTA PRUEBA SUR",
        "frontera": "VERDADERO",
        "puesto": "AUX PRUEBA",
        "banco": "BANCO PRUEBA",
        "cuenta": "0000000002",
        "salario": 3600,
        "valor_he": 110.0,
        "codes": ["A", "D", "A", "I", "A", "D", "A"],
    },
    {
        "num": "9003",
        "nombre": "EMPLEADO PRUEBA 003",
        "planta": "PLANTA PRUEBA ORIENTE",
        "frontera": "FALSO",
        "puesto": "AUX PRUEBA",
        "banco": "BANCO PRUEBA",
        "cuenta": "0000000003",
        "salario": 3400,
        "valor_he": 95.0,
        "codes": ["F", "F", "D", "F", "F", "A", "D"],
    },
]

SHEET_SPECS = {
    "semana normal": "2 al 8 julio",
    "semana vacaciones": "2 al 8 julio",
    "cruce mes": "26 al 1 de julio",
    "conflicto periodo": "26 al 1 de julio",
    "trayectoria w1": "19 al 25 de jun",
    "trayectoria w2": "2 al 8 julio",
}


def _source_path() -> Path:
    raw = (os.environ.get("GIS_PEPSI_EVIDENCE_XLSM") or "").strip()
    if not raw:
        raise FileNotFoundError(
            "Define GIS_PEPSI_EVIDENCE_XLSM con la ruta local del XLSM de evidencia "
            "(solo regeneración local; CI no lo requiere)."
        )
    return Path(raw)


def _copy_top_block(src_ws, dst_ws, *, max_row: int = 6, max_col: int = 19) -> None:
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            value = src_ws.cell(row_idx, col_idx).value
            if isinstance(value, str):
                value = value.replace("PEPSI", "CLIENTE PRUEBA").replace("Pepsi", "Cliente Prueba")
            dst_ws.cell(row_idx, col_idx, value)
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = True
            if dim.width:
                dst_ws.column_dimensions[col_letter].width = dim.width


def _write_worker_row(dst_ws, row_idx: int, worker: dict) -> None:
    dst_ws.cell(row_idx, 2, worker["num"])
    dst_ws.cell(row_idx, 3, worker["nombre"])
    dst_ws.cell(row_idx, 4, worker["planta"])
    dst_ws.cell(row_idx, 5, worker["frontera"])
    dst_ws.cell(row_idx, 6, worker["puesto"])
    dst_ws.cell(row_idx, 7, worker["banco"])
    dst_ws.cell(row_idx, 8, worker["cuenta"])
    dst_ws.cell(row_idx, 9, worker["salario"])
    dst_ws.cell(row_idx, 10, worker["valor_he"])
    for idx, code in enumerate(worker["codes"], start=11):
        dst_ws.cell(row_idx, idx, code)


def _build_weekly_sheet(wb_out: Workbook, src_ws, *, title: str, workers: list[dict]) -> None:
    ws = wb_out.create_sheet(title=title[:31])
    _copy_top_block(src_ws, ws)
    for offset, worker in enumerate(workers):
        _write_worker_row(ws, 7 + offset, worker)


def _build_contpaq_sheet(wb_out: Workbook, src_ws) -> None:
    ws = wb_out.create_sheet(title="CONTPAQi")
    _copy_top_block(src_ws, ws, max_row=3, max_col=9)
    fake_rows = [
        ("001", "01/01/2024", "DEMO", "UNO", "PRUEBA", "Empleado Prueba Uno"),
        ("002", "02/01/2024", "DEMO", "DOS", "PRUEBA", "Empleado Prueba Dos"),
    ]
    for offset, row in enumerate(fake_rows, start=4):
        ws.cell(offset, 2, row[0])
        ws.cell(offset, 3, row[1])
        ws.cell(offset, 4, row[2])
        ws.cell(offset, 5, row[3])
        ws.cell(offset, 6, row[4])
        ws.cell(offset, 7, row[5])


def build_pepsi_nomina_anon_workbook() -> Path:
    source = _source_path()
    if not source.is_file():
        raise FileNotFoundError(f"Evidence workbook not found: {source}")

    tmp = Path(tempfile.mkdtemp()) / "source.xlsm"
    shutil.copy2(source, tmp)
    src_wb = load_workbook(tmp, data_only=False)

    out = Workbook()
    out.remove(out.active)

    _build_weekly_sheet(
        out,
        src_wb["2 al 8 julio"],
        title="Semana Normal Anon",
        workers=[FAKE_WORKERS[0], FAKE_WORKERS[2]],
    )
    _build_weekly_sheet(
        out,
        src_wb["2 al 8 julio"],
        title="Semana Vacaciones Anon",
        workers=[{**FAKE_WORKERS[1], "codes": ["V", "V", "V", "D", "A", "A", "D"]}],
    )
    _build_weekly_sheet(
        out,
        src_wb["26 al 1 de julio"],
        title="Cruce Mes Anon",
        workers=[FAKE_WORKERS[0]],
    )
    conflict = out.create_sheet(title="Conflicto Periodo Anon")
    _copy_top_block(src_wb["26 al 1 de julio"], conflict)
    conflict.cell(4, 2, "Periodo del día 15 al 21 de mayo de 2026")
    _write_worker_row(conflict, 7, FAKE_WORKERS[0])
    _build_weekly_sheet(
        out,
        src_wb["19 al 25 de jun"],
        title="Trayectoria W1 Anon",
        workers=[FAKE_WORKERS[0]],
    )
    _build_weekly_sheet(
        out,
        src_wb["2 al 8 julio"],
        title="Trayectoria W2 Anon",
        workers=[FAKE_WORKERS[0]],
    )
    _build_contpaq_sheet(out, src_wb["CONTPAQi"])

    out_path = FIXTURE_DIR / "gis_pepsi_nomina_anon.xlsx"
    out.save(out_path)
    src_wb.close()
    tmp.unlink(missing_ok=True)
    shutil.rmtree(tmp.parent, ignore_errors=True)
    return out_path


if __name__ == "__main__":
    path = build_pepsi_nomina_anon_workbook()
    print("wrote", path)
