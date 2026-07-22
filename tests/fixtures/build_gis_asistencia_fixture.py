"""Generate GIS nómina fixture with asistencia v4 block (anon)."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from modules.nomina.asistencia_excel import build_asistencia_template_file
from modules.nomina.db import NominaBaseRow

FIXTURE_DIR = Path(__file__).resolve().parent


def build_gis_nomina_asistencia_anon() -> Path:
    data = build_asistencia_template_file(
        fecha_inicio=date(2026, 6, 1),
        fecha_fin=date(2026, 6, 7),
        cliente="PEPSI",
        coordinador="COORD DEMO",
        base_rows=[
            NominaBaseRow(
                nombre_empleado="Empleado Demo Uno",
                cliente="PEPSI",
                planta="A",
                puesto="Aux",
                banco="Banorte",
                cuenta="1111111111",
                numero_empleado="121",
            ),
            NominaBaseRow(
                nombre_empleado="Empleado Demo Dos",
                cliente="PEPSI",
                planta="A",
                puesto="Aux",
                banco="Banorte",
                cuenta="2222222222",
                numero_empleado="138",
            ),
        ],
    )
    wb = load_workbook(BytesIO(data))
    ws = wb["Asistencia"]
    ws.cell(5, 7, "A")
    ws.cell(5, 8, "A")
    ws.cell(5, 9, "F")
    ws.cell(5, 10, "F")
    ws.cell(5, 11, "D")
    ws.cell(5, 12, "F")
    ws.cell(5, 13, "F")
    ws.cell(6, 7, "A")
    ws.cell(6, 8, "F")
    ws.cell(6, 9, "I")
    ws.cell(6, 10, "V")
    ws.cell(6, 11, "A")
    ws.cell(6, 12, "D")
    ws.cell(6, 13, "A")
    out = FIXTURE_DIR / "gis_nomina_asistencia_anon.xlsx"
    wb.save(out)
    wb.close()
    return out


if __name__ == "__main__":
    path = build_gis_nomina_asistencia_anon()
    print("wrote", path)
