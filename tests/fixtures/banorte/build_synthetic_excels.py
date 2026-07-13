"""Build synthetic Banorte Excel fixtures (fictional PII only)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).resolve().parent


def build_altas() -> None:
    wb = Workbook()
    # LAYOUT placeholder
    ws_l = wb.active
    ws_l.title = "LAYOUT"
    ws_l["A1"] = "PAGOS"

    ws = wb.create_sheet("ALTAS")
    ws["B2"] = "NOMINA BANORTE SYNTHETIC"
    headers = [
        "#",
        "Número de empleado",
        "Curp ",
        "Nombre del empleado",
        "Producto",
        "Tipo de tarjeta",
        "Número de tarjeta",
        "Número de cuenta",
        "Cuenta CLABE",
        "Fecha de alta solicitud",
        "Estatus",
        "Comentarios",
        "Rfc",
    ]
    for i, h in enumerate(headers, start=2):
        ws.cell(3, i, h)

    # row4 EXITOSO
    ws.append([])  # will write manually from row 4
    rows = [
        (4, "0000000001", "AAAA900101HDFRRR01", "ANA DEMO UNO", "1321000001", "EXITOSO", None),
        (5, "0000000002", "BBBB900101HDFRRR02", "BETO DEMO DOS", "1321000002", "EXITOSO", None),
        # special comment: requested becomes account
        (
            6,
            "0000000099",
            "CCCC900101HDFRRR03",
            "CARLA DEMO TRES",
            "1321999999",
            "EXITOSO",
            "El número de empleado ya existía se asignó el número de cuenta como tu número de Empleado",
        ),
        # manual complete
        (7, "0000000010", "DDDD900101HDFRRR04", "DIANA MANUAL", "1321000010", None, None),
        # manual incomplete (no account)
        (8, "0000000011", "EEEE900101HDFRRR05", "ELIAS INCOMPLETO", None, None, None),
        # duplicate emp/account lower wins
        (9, "0000000001", "AAAA900101HDFRRR01", "ANA DEMO UNO V2", "1321000001", "EXITOSO", None),
        # name-only conflict later handled in tests separately
    ]
    for r, emp, curp, nombre, acct, est, com in rows:
        ws.cell(r, 2, r - 3)
        ws.cell(r, 3, emp)
        ws.cell(r, 4, curp)
        ws.cell(r, 5, nombre)
        ws.cell(r, 6, "NOMINA BANORTE 2")
        ws.cell(r, 9, acct)
        ws.cell(r, 11, "2026-01-01 10:00:00.0")
        ws.cell(r, 12, est)
        ws.cell(r, 13, com)

    # FALLIDOS sheet: 2 FALLIDO + 1 empty status
    wsf = wb.create_sheet("FALLIDOS")
    wsf["B2"] = "ERRORES"
    for i, h in enumerate(headers, start=2):
        wsf.cell(3, i, h)
    wsf.cell(4, 3, "0000000901")
    wsf.cell(4, 5, "FALLIDO UNO")
    wsf.cell(4, 12, "FALLIDO")
    wsf.cell(5, 3, "0000000902")
    wsf.cell(5, 5, "FALLIDO DOS")
    wsf.cell(5, 12, "FALLIDO")
    wsf.cell(6, 3, "0000000903")
    wsf.cell(6, 5, "FALLIDO VACIO STATUS")
    wsf.cell(6, 12, None)

    wb.save(OUT / "synthetic_altas.xlsx")


def build_reporte() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws["A1"] = "Reporte_Detallado_synthetic"
    headers = [
        "#",
        "Número de empleado",
        "Curp ",
        "Nombre del empleado",
        "Producto",
        "Tipo de tarjeta",
        "Número de tarjeta",
        "Número de cuenta",
        "Cuenta CLABE",
        "Fecha de alta solicitud",
        "Estatus respuesta",
        "Comentarios",
        "Rfc",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(2, i, h)
    # EXITOSO validates previous manual DIANA if linked by emp
    ws.cell(3, 1, "1")
    ws.cell(3, 2, "0000000010")
    ws.cell(3, 3, "DDDD900101HDFRRR04")
    ws.cell(3, 4, "DIANA MANUAL")
    ws.cell(3, 8, "1321000010")
    ws.cell(3, 10, "2026-02-01 10:00:00.0")
    ws.cell(3, 11, "EXITOSO")
    # FALLIDO ignored
    ws.cell(4, 1, "2")
    ws.cell(4, 2, "0000000888")
    ws.cell(4, 4, "FALLIDO REPORTE")
    ws.cell(4, 8, "1321888888")
    ws.cell(4, 11, "FALLIDO")
    # special comment
    ws.cell(5, 1, "3")
    ws.cell(5, 2, "0000000777")
    ws.cell(5, 3, "FFFF900101HDFRRR06")
    ws.cell(5, 4, "FABIAN ESPECIAL")
    ws.cell(5, 8, "1321777777")
    ws.cell(5, 11, "EXITOSO")
    ws.cell(
        5,
        12,
        "El número de empleado ya existía se asignó el número de cuenta como tu número de Empleado",
    )
    wb.save(OUT / "synthetic_reporte.xlsx")


if __name__ == "__main__":
    build_altas()
    build_reporte()
    print("wrote synthetic_altas.xlsx and synthetic_reporte.xlsx")
