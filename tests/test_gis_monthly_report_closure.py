"""Verificación de cierre — Reporte mensual GIS."""

from __future__ import annotations

import json
import sqlite3
import tempfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas import repository as nom_repo
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables
from modules.gestion_idse_sua.reportes.consolidation_service import generate_monthly_report
from modules.gestion_idse_sua.reportes.date_utils import days_in_calendar_month
from modules.gestion_idse_sua.reportes.excel_export import generate_monthly_excel
from modules.gestion_idse_sua.reportes.monthly_status import (
    classify_monthly_status,
    person_has_active_evidence,
)
from modules.gestion_idse_sua.reportes.movement_bridge import convert_events_to_movements
from modules.gestion_idse_sua.reportes.repository import (
    create_report,
    get_report,
    list_report_events,
    list_report_persons,
    list_report_weeks,
    update_event,
)
from modules.gestion_idse_sua.reportes.schema import ensure_gis_monthly_tables
from modules.gestion_idse_sua.template_contract import MENSUAL_SHA256, MENSUAL_SHEETS, mensual_path
from modules.gestion_idse_sua.template_validator import sha256_file, validate_mensual_template


HC = {
    "nombre_completo": "JUAN PEREZ LOPEZ",
    "cliente": "PEPSI",
    "nss": "11111111111",
    "numero_empleado": "101",
    "curp": "PELJ800101HDFRNN09",
    "apellido_paterno": "PEREZ",
    "apellido_materno": "LOPEZ",
    "nombre": "JUAN",
    "sueldo_diario": 500,
    "patron": "12345678901",
}


@pytest.fixture
def conn(tmp_path):
    connection = sqlite3.connect(tmp_path / "closure.db")
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    ensure_gis_monthly_tables(connection)
    yield connection
    connection.close()


def _seed_period(
    conn: sqlite3.Connection,
    *,
    sheet_key: int,
    fecha_inicio: str,
    fecha_fin: str,
    codes: list[str] | None = None,
) -> int:
    conn.execute(
        """
        INSERT OR IGNORE INTO gis_nomina_imports (id, original_filename, file_hash, uploaded_at, status)
        VALUES (?, ?, ?, '2026-01-01', 'extracted')
        """,
        (sheet_key, f"f{sheet_key}.xlsx", f"hash-{sheet_key}"),
    )
    conn.execute(
        """
        INSERT OR IGNORE INTO gis_nomina_sheets (id, import_id, sheet_index, sheet_name, is_hidden, confirmed_classification, estimated_rows)
        VALUES (?, ?, 0, ?, 0, 'nomina', 1)
        """,
        (sheet_key, sheet_key, f"S{sheet_key}"),
    )
    nom_repo.upsert_period(
        conn,
        sheet_key,
        {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "semana_num": sheet_key,
            "source": "manual",
            "cut_warning": None,
        },
        confirmed=True,
    )
    period_id = conn.execute(
        "SELECT id FROM gis_nomina_periods WHERE sheet_id = ?",
        (sheet_key,),
    ).fetchone()[0]
    worker_id = nom_repo.insert_workers(
        conn,
        period_id,
        [
            {
                "row_number": 4,
                "num_empleado": "101",
                "nombre_original": "Juan",
                "nombre_normalizado": "JUAN PEREZ LOPEZ",
                "puesto": "Op",
                "planta_original": "A",
                "planta_normalizada": "A",
                "cuenta": "1",
                "row_json": "{}",
            }
        ],
    )[0]
    nom_repo.update_worker_cliente(conn, worker_id, "PEPSI")
    nom_repo.upsert_match(
        conn,
        worker_id,
        {
            "headcount_key": "hc-101",
            "match_method": "num_empleado",
            "confidence": 1.0,
            "status": "confirmed",
            "nss": "11111111111",
            "curp": "PELJ800101HDFRNN09",
            "hc_nombre": "JUAN PEREZ LOPEZ",
            "hc_json": json.dumps(HC),
        },
    )
    start = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
    seq = codes or ["A"] * 7
    now = "2026-06-01T00:00:00"
    for idx, code in enumerate(seq[:7], start=1):
        nom_repo.insert_attendance(
            conn,
            {
                "worker_id": worker_id,
                "period_id": period_id,
                "column_index": idx,
                "column_number": idx,
                "fecha_iso": (start + timedelta(days=idx - 1)).isoformat(),
                "code_original": code,
                "code_normalized": code,
                "interpretation_status": "ok",
                "created_at": now,
                "updated_at": now,
            },
        )
    conn.commit()
    return int(period_id)


def _four_june_weeks(conn, codes: list[str] | None = None) -> list[int]:
    return [
        _seed_period(conn, sheet_key=1, fecha_inicio="01/06/2026", fecha_fin="07/06/2026", codes=codes),
        _seed_period(conn, sheet_key=8, fecha_inicio="08/06/2026", fecha_fin="14/06/2026", codes=codes),
        _seed_period(conn, sheet_key=15, fecha_inicio="15/06/2026", fecha_fin="21/06/2026", codes=codes),
        _seed_period(conn, sheet_key=22, fecha_inicio="22/06/2026", fecha_fin="28/06/2026", codes=codes),
    ]


def test_snapshot_person_data_immutable_after_source_change(conn):
    period_ids = _four_june_weeks(conn)
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    before = list_report_persons(conn, report_id)[0]
    before_nss = before["nss"]
    before_daily = before["daily_json"]
    before_totals = before["totals_json"]

    conn.execute("UPDATE gis_nomina_matches SET nss = '99999999999' WHERE nss = '11111111111'")
    conn.execute(
        "UPDATE gis_nomina_attendance SET code_normalized = 'F', code_original = 'F' WHERE code_normalized = 'A'"
    )
    conn.commit()

    after = list_report_persons(conn, report_id)[0]
    assert after["nss"] == before_nss
    assert after["daily_json"] == before_daily
    assert after["totals_json"] == before_totals


def test_snapshot_weeks_immutable_after_period_change(conn):
    period_ids = _four_june_weeks(conn)
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    before = list_report_weeks(conn, report_id)[0]["fecha_inicio"]
    conn.execute(
        "UPDATE gis_nomina_periods SET fecha_inicio = '99/99/9999' WHERE id = ?",
        (period_ids[0],),
    )
    conn.commit()
    after = list_report_weeks(conn, report_id)[0]["fecha_inicio"]
    assert after == before


def test_inclusion_vs_todo_el_mes_classification():
    assert person_has_active_evidence([{"code_normalized": "I"}])
    assert not person_has_active_evidence([{"code_normalized": "D"}])
    daily_i = [
        {"fecha_iso": f"2026-06-{d:02d}", "code_normalized": "I", "period_id": pid, "interpretation_status": "ok"}
        for d, pid in [(1, 1), (8, 8), (15, 15), (22, 22)]
    ]
    assert (
        classify_monthly_status(
            daily=daily_i,
            events=[],
            selected_week_count=4,
            weeks_with_presence=4,
            coverage_complete=True,
            selected_period_ids={1, 8, 15, 22},
        )
        == "Todo el mes"
    )
    daily_v = [{"fecha_iso": "2026-06-01", "code_normalized": "V", "period_id": 1, "interpretation_status": "ok"}]
    assert (
        classify_monthly_status(
            daily=daily_v,
            events=[],
            selected_week_count=1,
            weeks_with_presence=1,
            coverage_complete=True,
            selected_period_ids={1},
        )
        != "Todo el mes"
    )
    daily_d = [{"fecha_iso": "2026-06-01", "code_normalized": "D", "period_id": 1, "interpretation_status": "ok"}]
    assert not person_has_active_evidence(daily_d)
    baja = [{"event_type": "posible_baja", "status": "suggested"}]
    assert (
        classify_monthly_status(
            daily=daily_i,
            events=baja,
            selected_week_count=4,
            weeks_with_presence=4,
            coverage_complete=True,
            selected_period_ids={1, 8, 15, 22},
        )
        == "Salida durante el mes"
    )
    partial = [
        {"fecha_iso": "2026-06-01", "code_normalized": "A", "period_id": 1, "interpretation_status": "ok"},
        {"fecha_iso": "2026-06-02", "code_normalized": "D", "period_id": 1, "interpretation_status": "ok"},
    ]
    assert (
        classify_monthly_status(
            daily=partial,
            events=[],
            selected_week_count=4,
            weeks_with_presence=1,
            coverage_complete=True,
            selected_period_ids={1, 8, 15, 22},
        )
        == "Presencia parcial"
    )


def test_multiple_events_stored_and_managed_independently(conn):
    seq = ["A", "A", "F", "F", "F", "F", "A", "A", "F", "F", "F", "F", "A", "A", "F", "F", "F", "F", "A"]
    period_ids = [
        _seed_period(conn, sheet_key=1, fecha_inicio="01/06/2026", fecha_fin="07/06/2026", codes=seq[:7]),
        _seed_period(conn, sheet_key=8, fecha_inicio="08/06/2026", fecha_fin="14/06/2026", codes=seq[7:14]),
        _seed_period(conn, sheet_key=15, fecha_inicio="15/06/2026", fecha_fin="21/06/2026", codes=seq[14:21]),
        _seed_period(conn, sheet_key=22, fecha_inicio="22/06/2026", fecha_fin="28/06/2026", codes=seq[21:] + ["A"]),
    ]
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    events = list_report_events(conn, report_id)
    ops = [e for e in events if e["event_type_suggested"] in {"ALTA", "BAJA"}]
    assert len(ops) >= 4
    ids = {int(e["id"]) for e in ops[:4]}
    assert len(ids) == 4
    e1, e2, e3, e4 = ops[0], ops[1], ops[2], ops[3]
    update_event(conn, e1["id"], estado="descartado")
    update_event(conn, e2["id"], event_type_confirmed="ALTA", fecha_confirmed="2026-06-07", estado="confirmado")
    update_event(conn, e3["id"], event_type_confirmed="BAJA", fecha_confirmed="2026-06-10", estado="confirmado")
    update_event(conn, e4["id"], event_type_confirmed="ALTA", fecha_confirmed="2026-06-14", estado="confirmado")
    conn.commit()
    refreshed = {int(e["id"]): e for e in list_report_events(conn, report_id)}
    assert refreshed[e1["id"]]["estado"] == "descartado"
    assert refreshed[e2["id"]]["fecha_confirmed"] == "2026-06-07"
    assert refreshed[e4["id"]]["event_type_confirmed"] == "ALTA"


@patch("modules.gestion_idse_sua.reportes.movement_bridge.guardar_movimiento")
def test_movement_bridge_traceability_and_idempotency(mock_save, conn, monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    period_ids = _four_june_weeks(conn, codes=["A", "A", "F", "F", "F", "F", "A"])
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    events = [e for e in list_report_events(conn, report_id) if e["event_type_suggested"] in {"ALTA", "BAJA"}]
    assert events
    for event in events[:2]:
        conn.execute(
            """
            UPDATE gis_monthly_report_events
            SET estado='confirmado', event_type_confirmed=?, fecha_confirmed=?
            WHERE id=?
            """,
            (event["event_type_suggested"], "2026-06-07", event["id"]),
        )
    conn.execute(
        "UPDATE gis_monthly_report_persons SET afiliatorios_json=? WHERE report_id=?",
        (json.dumps(HC), report_id),
    )
    conn.commit()
    mock_save.return_value = {"id": "mov-x"}
    ids = [int(events[0]["id"]), int(events[1]["id"])] if len(events) > 1 else [int(events[0]["id"])]
    convert_events_to_movements(conn, event_ids=ids)
    assert mock_save.call_count == len(ids)
    for call in mock_save.call_args_list:
        payload = call.args[0]
        assert payload["origen"] == "gis_reporte_mensual"
        assert payload["tipo_movimiento"] in {"ALTA", "BAJA"}
        assert "MOD" not in str(payload.get("tipo_movimiento", ""))
        alerta = str(payload.get("alerta") or "")
        assert f"reporte_id={report_id}" in alerta
        assert "persona_id=" in alerta
        assert "evento_id=" in alerta
    convert_events_to_movements(conn, event_ids=ids)
    assert mock_save.call_count == len(ids)


@pytest.mark.parametrize("mes,expected_days", [(2, 28), (4, 30), (6, 30), (1, 31)])
def test_excel_month_day_columns(mes, expected_days, conn):
    from calendar import monthrange

    last_day = monthrange(2026, mes)[1]
    starts = [1, 8, 15, 22]
    period_ids = []
    for sheet_key, start in zip((1, 8, 15, 22), starts, strict=True):
        end = min(start + 6, last_day)
        period_ids.append(
            _seed_period(
                conn,
                sheet_key=sheet_key,
                fecha_inicio=f"{start:02d}/{mes:02d}/2026",
                fecha_fin=f"{end:02d}/{mes:02d}/2026",
            )
        )
    report_id = create_report(conn, cliente="PEPSI", mes=mes, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=mes, anio=2026)
    before = sha256_file(mensual_path())
    buf, _ = generate_monthly_excel(conn, report_id)
    assert before == MENSUAL_SHA256
    validate_mensual_template(buf)
    wb = load_workbook(buf)
    assert wb.sheetnames == list(MENSUAL_SHEETS)
    persons = list_report_persons(conn, report_id)
    if not persons:
        pytest.skip("No persons in short-month fixture")
    daily = json.loads(persons[0]["daily_json"])
    assert len(daily) == expected_days
    row = 7
    for day_idx in range(31):
        col_val = wb["Asistencia Mensual"].cell(row, 12 + day_idx).value
        if day_idx < len(daily):
            expected = daily[day_idx].get("code") or ""
            assert (col_val or "") == expected
        else:
            assert col_val in (None, "")
    wb.close()


def test_excel_sua_only_alta_and_no_ref(conn):
    period_ids = _four_june_weeks(conn, codes=["A", "A", "F", "F", "F", "F", "A"])
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    for event in list_report_events(conn, report_id):
        if event["event_type_suggested"] in {"ALTA", "BAJA"}:
            conn.execute(
                "UPDATE gis_monthly_report_events SET estado='confirmado', event_type_confirmed=?, fecha_confirmed='2026-06-07' WHERE id=?",
                (event["event_type_suggested"], event["id"]),
            )
    conn.commit()
    buf, _ = generate_monthly_excel(conn, report_id)
    wb = load_workbook(buf)
    ws = wb["Movimientos Seleccionados"]
    rows = 0
    for r in range(7, 20):
        tipo = ws.cell(r, 2).value
        if not tipo:
            break
        rows += 1
        idse = ws.cell(r, 17).value
        sua = ws.cell(r, 18).value
        assert idse == "Sí"
        if str(tipo).upper() == "BAJA":
            assert sua != "Sí"
        if str(tipo).upper() == "ALTA":
            assert sua == "Sí"
    assert rows >= 1
    for name in MENSUAL_SHEETS:
        for row in wb[name].iter_rows(max_row=30, max_col=40):
            for cell in row:
                if isinstance(cell.value, str):
                    assert "#REF!" not in cell.value
    wb.close()
