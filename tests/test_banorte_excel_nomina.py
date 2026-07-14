"""Banorte Fase 2.1C — Excel nomina import (synthetic fixtures only)."""

from __future__ import annotations

import sqlite3
from io import BytesIO

import pytest
from openpyxl import Workbook

from modules.nomina.banorte.excel_nomina_service import (
    ExcelNominaError,
    inspect_excel,
    prepare_excel_draft,
    preview_excel,
    sha256_bytes,
)
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.schema import ensure_banorte_tables
from modules.nomina.db import ensure_nomina_tables

SECRET = "test-secret-key"


def _build_workbook(
    *,
    hidden_row: int | None = None,
    formula_net: bool = False,
    rows: list[tuple[str, str, float | str | None]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina1"
    ws.cell(row=6, column=1, value="NOMBRE DE EMPLEADO")
    ws.cell(row=6, column=2, value="BANCO")
    ws.cell(row=6, column=3, value="NETO A PAGAR")
    ws.cell(row=6, column=4, value="CUENTA")
    data = rows or [
        ("JUAN PEREZ", "Banorte", 1500.50),
        ("MARIA LOPEZ", "BBVA", 900.00),
        ("PEDRO GOMEZ", "banorte", 200.00),
    ]
    for i, (name, bank, net) in enumerate(data, start=7):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=bank)
        if formula_net and i == 7:
            ws.cell(row=i, column=3, value="=1500.5")
        else:
            ws.cell(row=i, column=3, value=net)
        ws.cell(row=i, column=4, value="1234567890")
    if hidden_row is not None:
        ws.row_dimensions[hidden_row].hidden = True
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_inspect_lists_sheets_and_token(tmp_path):
    raw = _build_workbook()
    out = inspect_excel(raw, "nomina.xlsx", secret_key=SECRET, user="u")
    assert "Nomina1" in out["sheets"]
    assert out["token"]
    assert out["sha256"] == sha256_bytes(raw)


def test_preview_counts_banorte_and_hidden(tmp_path):
    raw = _build_workbook(hidden_row=8)
    ins = inspect_excel(raw, "nomina.xlsx", secret_key=SECRET, user="u")
    prev = preview_excel(
        raw,
        filename="nomina.xlsx",
        sheet="Nomina1",
        token=ins["token"],
        secret_key=SECRET,
        user="u",
    )
    assert prev.banorte_count == 2
    assert prev.excluded_hidden_count == 1
    assert prev.excluded_other_bank_count == 0
    assert prev.total_banorte_cents == 170050


def test_preview_formula_without_cache(tmp_path):
    raw = _build_workbook(formula_net=True)
    ins = inspect_excel(raw, "nomina.xlsx", secret_key=SECRET, user="u")
    prev = preview_excel(
        raw,
        filename="nomina.xlsx",
        sheet="Nomina1",
        token=ins["token"],
        secret_key=SECRET,
        user="u",
    )
    assert prev.blocked_formula_count >= 1


def test_prepare_creates_excel_nomina_draft(tmp_path):
    db = str(tmp_path / "x.db")
    conn = connect(db)
    ensure_nomina_tables(conn)
    ensure_banorte_tables(conn)
    conn.execute(
        """
        INSERT INTO nomina_banorte_beneficiaries (
            nombre_original, nombre_normalizado, employee_number_effective, account_number,
            source_kind, validation_status, record_status,
            imported_at, imported_by, created_at, updated_at
        ) VALUES ('JUAN PEREZ','JUAN PEREZ','1','1111111111','ALTA_MANUAL','IMPORTADO_EXITOSO','ACTIVO','t','u','t','t'),
                 ('PEDRO GOMEZ','PEDRO GOMEZ','2','2222222222','ALTA_MANUAL','IMPORTADO_EXITOSO','ACTIVO','t','u','t','t')
        """
    )
    conn.commit()
    conn.close()
    raw = _build_workbook()
    ins = inspect_excel(raw, "nomina.xlsx", secret_key=SECRET, user="u")
    draft = prepare_excel_draft(
        db,
        "u",
        raw,
        filename="nomina.xlsx",
        sheet="Nomina1",
        token=ins["token"],
        secret_key=SECRET,
    )
    assert draft["origin_kind"] == "EXCEL_NOMINA"
    assert draft["source_sha256"] == ins["sha256"]
    included = [r for r in draft["rows"] if r["included"]]
    assert len(included) == 2


def test_excel_nomina_migration_allows_origin(tmp_path):
    db = str(tmp_path / "m.db")
    conn = connect(db)
    ensure_nomina_tables(conn)
    ensure_banorte_tables(conn)
    sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='nomina_banorte_export_drafts'"
    ).fetchone()[0]
    assert "EXCEL_NOMINA" in sql
    conn.execute(
        """
        INSERT INTO nomina_banorte_export_drafts (
            created_by, updated_by, created_at, updated_at, origin_kind, calculo_id,
            origin_updated_at, origin_hash, status, revision,
            source_filename, source_sha256, source_sheet, source_file_size
        ) VALUES ('u','u','t','t','EXCEL_NOMINA',NULL,'t','hashx','OPEN',1,'f.xlsx','abc','S1',100)
        """
    )
    conn.commit()
    conn.close()


def test_token_sha_mismatch_rejected(tmp_path):
    raw = _build_workbook()
    ins = inspect_excel(raw, "nomina.xlsx", secret_key=SECRET, user="u")
    with pytest.raises(ValueError, match="excel_token_sha_mismatch"):
        preview_excel(
            raw + b"x",
            filename="nomina.xlsx",
            sheet="Nomina1",
            token=ins["token"],
            secret_key=SECRET,
            user="u",
        )


def test_file_too_large():
    with pytest.raises(ExcelNominaError) as ei:
        inspect_excel(b"x" * (25 * 1024 * 1024 + 1), "big.xlsx", secret_key=SECRET, user="u")
    assert ei.value.code == "file_too_large"
