"""GIS Reportes mensuales — schema, selección, consolidación, Excel y movimientos."""

from __future__ import annotations

import json
import sqlite3
from datetime import date
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas import repository as nom_repo
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables
from modules.gestion_idse_sua.reportes.consolidation_service import generate_monthly_report
from modules.gestion_idse_sua.reportes.date_utils import (
    clip_iso_dates_to_month,
    days_in_calendar_month,
    period_intersects_month,
)
from modules.gestion_idse_sua.reportes.excel_export import generate_monthly_excel
from modules.gestion_idse_sua.reportes.monthly_status import person_has_active_evidence
from modules.gestion_idse_sua.reportes.movement_bridge import convert_events_to_movements
from modules.gestion_idse_sua.reportes.period_selection import validate_week_selection
from modules.gestion_idse_sua.reportes.repository import create_report, list_report_events, list_report_persons
from modules.gestion_idse_sua.reportes.schema import GIS_MONTHLY_TABLES, ensure_gis_monthly_tables
from modules.gestion_idse_sua.template_contract import MENSUAL_SHEETS, MENSUAL_SHA256, mensual_path
from modules.gestion_idse_sua.template_validator import sha256_file, validate_mensual_template


HC = [
    {
        "nombre_completo": "JUAN PEREZ LOPEZ",
        "cliente": "PEPSI",
        "nss": "11111111111",
        "numero_empleado": "101",
        "curp": "PELJ800101HDFRNN09",
        "apellido_paterno": "PEREZ",
        "apellido_materno": "LOPEZ",
        "nombre": "JUAN",
        "sueldo_diario": 500,
        "patron": "12345678901",
    }
]


@pytest.fixture
def conn(tmp_path):
    connection = sqlite3.connect(tmp_path / "monthly.db")
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    ensure_gis_monthly_tables(connection)
    yield connection
    connection.close()


def _seed_period(
    conn: sqlite3.Connection,
    *,
    period_id: int,
    fecha_inicio: str,
    fecha_fin: str,
    import_name: str = "f.xlsx",
    file_hash: str | None = None,
) -> int:
    conn.execute(
        """
        INSERT OR IGNORE INTO gis_nomina_imports (id, original_filename, file_hash, uploaded_at, status)
        VALUES (?, ?, ?, '2026-01-01', 'extracted')
        """,
        (period_id, import_name, file_hash or f"hash-{period_id}"),
    )
    conn.execute(
        """
        INSERT OR IGNORE INTO gis_nomina_sheets (id, import_id, sheet_index, sheet_name, is_hidden, confirmed_classification, estimated_rows)
        VALUES (?, ?, 0, ?, 0, 'nomina', 1)
        """,
        (period_id, period_id, f"S{period_id}"),
    )
    nom_repo.upsert_period(
        conn,
        period_id,
        {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "semana_num": period_id,
            "source": "manual",
            "cut_warning": None,
        },
        confirmed=True,
    )
    db_period_id = conn.execute(
        "SELECT id FROM gis_nomina_periods WHERE sheet_id = ?",
        (period_id,),
    ).fetchone()[0]
    worker_id = nom_repo.insert_workers(
        conn,
        db_period_id,
        [
            {
                "row_number": 4,
                "num_empleado": "101",
                "nombre_original": "Juan",
                "nombre_normalizado": "JUAN PEREZ LOPEZ",
                "puesto": "Op",
                "planta_original": "A",
                "planta_normalizada": "A",
                "cuenta": "1",
                "row_json": "{}",
                "cliente_confirmado": "PEPSI",
            }
        ],
    )[0]
    nom_repo.update_worker_cliente(conn, worker_id, "PEPSI")
    nom_repo.upsert_match(
        conn,
        worker_id,
        {
            "headcount_key": "hc-101",
            "match_method": "num_empleado",
            "confidence": 1.0,
            "status": "confirmed",
            "nss": "11111111111",
            "curp": "PELJ800101HDFRNN09",
            "hc_nombre": "JUAN PEREZ LOPEZ",
            "hc_json": json.dumps(HC[0]),
        },
    )
    now = "2026-06-01T00:00:00"
    from datetime import datetime, timedelta

    start = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
    for idx in range(1, 8):
        fecha = (start + timedelta(days=idx - 1)).isoformat()
        code = "A" if idx <= 5 else "D"
        nom_repo.insert_attendance(
            conn,
            {
                "worker_id": worker_id,
                "period_id": db_period_id,
                "column_index": idx,
                "column_number": idx,
                "fecha_iso": fecha,
                "code_original": code,
                "code_normalized": code,
                "interpretation_status": "ok",
                "created_at": now,
                "updated_at": now,
            },
        )
    conn.commit()
    return int(db_period_id)


def test_monthly_tables_idempotent(conn):
    ensure_gis_monthly_tables(conn)
    conn.commit()
    tables = {
        row[0]
        for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'gis_monthly_%'")
    }
    for table in GIS_MONTHLY_TABLES:
        assert table in tables


def test_period_intersects_month_cross_month():
    assert period_intersects_month("26/05/2026", "01/06/2026", mes=6, anio=2026)
    assert period_intersects_month("26/05/2026", "01/06/2026", mes=5, anio=2026)
    assert not period_intersects_month("01/07/2026", "07/07/2026", mes=6, anio=2026)


def test_clip_to_calendar_month():
    records = [
        {"fecha_iso": "2026-05-31", "code_normalized": "A"},
        {"fecha_iso": "2026-06-01", "code_normalized": "A"},
    ]
    june = clip_iso_dates_to_month(records, mes=6, anio=2026)
    assert len(june) == 1
    assert june[0]["fecha_iso"] == "2026-06-01"


def test_person_inclusion_rules():
    assert person_has_active_evidence([{"code_normalized": "A"}])
    assert person_has_active_evidence([{"code_normalized": "I"}])
    assert not person_has_active_evidence([{"code_normalized": "D"}])
    assert not person_has_active_evidence([{"code_normalized": "F"}])


def test_validate_week_selection_limits(conn):
    for pid in range(1, 8):
        _seed_period(conn, period_id=pid, fecha_inicio=f"{pid:02d}/06/2026", fecha_fin=f"{pid+6:02d}/06/2026")
    too_few = validate_week_selection(conn, [1, 2, 3], cliente="PEPSI", mes=6, anio=2026)
    assert not too_few["ok"]
    too_many = validate_week_selection(conn, list(range(1, 8)), cliente="PEPSI", mes=6, anio=2026)
    assert not too_many["ok"]
    ok = validate_week_selection(conn, [1, 2, 3, 4], cliente="PEPSI", mes=6, anio=2026)
    assert ok["ok"]


def test_generate_monthly_report_consolidates_person(conn):
    period_ids = []
    for start in (1, 8, 15, 22):
        period_ids.append(
            _seed_period(
                conn,
                period_id=start,
                fecha_inicio=f"{start:02d}/06/2026",
                fecha_fin=f"{start+6:02d}/06/2026",
            )
        )
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026, created_by="test")
    out = generate_monthly_report(
        conn,
        report_id=report_id,
        period_ids=period_ids,
        cliente="PEPSI",
        mes=6,
        anio=2026,
    )
    assert out["persons"] == 1
    persons = list_report_persons(conn, report_id)
    assert persons[0]["estado_mensual"] in {"Todo el mes", "Presencia parcial", "Ingreso durante el mes"}


def test_february_has_28_days():
    assert len(days_in_calendar_month(mes=2, anio=2026)) == 28


def test_monthly_excel_export(conn):
    period_ids = [_seed_period(conn, period_id=i, fecha_inicio=f"{i:02d}/06/2026", fecha_fin=f"{i+6:02d}/06/2026") for i in (1, 8, 15, 22)]
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    before = sha256_file(mensual_path())
    buf, _ = generate_monthly_excel(conn, report_id, username="test")
    assert before == MENSUAL_SHA256
    validate_mensual_template(buf)
    wb = load_workbook(buf)
    assert wb.sheetnames == list(MENSUAL_SHEETS)
    assert wb["Resumen"]["B6"].value == "PEPSI"
    wb.close()


@patch("modules.gestion_idse_sua.reportes.movement_bridge.guardar_movimiento")
def test_monthly_movement_conversion_idempotent(mock_save, conn, monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    period_ids = [_seed_period(conn, period_id=i, fecha_inicio=f"{i:02d}/06/2026", fecha_fin=f"{i+6:02d}/06/2026") for i in (1, 8, 15, 22)]
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    events = list_report_events(conn, report_id)
    if not events:
        pytest.skip("No trajectory events in fixture")
    event_id = events[0]["id"]
    conn.execute(
        "UPDATE gis_monthly_report_events SET estado='confirmado', event_type_confirmed='ALTA', fecha_confirmed='2026-06-01' WHERE id=?",
        (event_id,),
    )
    conn.execute(
        """
        UPDATE gis_monthly_report_persons
        SET afiliatorios_json=?
        WHERE report_id=?
        """,
        (json.dumps(HC[0]), report_id),
    )
    conn.commit()
    mock_save.return_value = {"id": "mov-m1"}
    out = convert_events_to_movements(conn, event_ids=[event_id])
    assert out["converted_ids"] == ["mov-m1"]
    again = convert_events_to_movements(conn, event_ids=[event_id])
    assert again["converted_ids"] == ["mov-m1"]
    assert mock_save.call_count == 1
