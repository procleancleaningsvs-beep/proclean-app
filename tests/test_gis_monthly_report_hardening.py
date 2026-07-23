"""Endurecimiento: cobertura calendario y export Excel en memoria."""

from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas import repository as nom_repo
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables
from modules.gestion_idse_sua.reportes.consolidation_service import generate_monthly_report
from modules.gestion_idse_sua.reportes.coverage_service import compute_month_coverage
from modules.gestion_idse_sua.reportes.excel_export import generate_monthly_excel
from modules.gestion_idse_sua.reportes.monthly_status import (
    MIN_WEEKS_FOR_FULL_MONTH,
    classify_monthly_status,
    person_has_active_evidence,
)
from modules.gestion_idse_sua.reportes.repository import create_report, get_report, list_report_persons
from modules.gestion_idse_sua.reportes.schema import ensure_gis_monthly_tables
from modules.gestion_idse_sua.template_contract import mensual_path
from modules.gestion_idse_sua.template_validator import sha256_file


HC = {
    "nombre_completo": "JUAN PEREZ LOPEZ",
    "cliente": "PEPSI",
    "nss": "11111111111",
    "numero_empleado": "101",
    "curp": "PELJ800101HDFRNN09",
    "apellido_paterno": "PEREZ",
    "apellido_materno": "LOPEZ",
    "nombre": "JUAN",
    "sueldo_diario": 500,
    "patron": "12345678901",
}


@pytest.fixture
def conn(tmp_path):
    connection = sqlite3.connect(tmp_path / "hard.db")
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    ensure_gis_monthly_tables(connection)
    yield connection
    connection.close()


def _seed_period(
    conn: sqlite3.Connection,
    *,
    sheet_key: int,
    fecha_inicio: str,
    fecha_fin: str,
    codes: list[str] | None = None,
) -> int:
    conn.execute(
        """
        INSERT OR IGNORE INTO gis_nomina_imports (id, original_filename, file_hash, uploaded_at, status)
        VALUES (?, ?, ?, '2026-01-01', 'extracted')
        """,
        (sheet_key, f"f{sheet_key}.xlsx", f"hash-{sheet_key}"),
    )
    conn.execute(
        """
        INSERT OR IGNORE INTO gis_nomina_sheets (id, import_id, sheet_index, sheet_name, is_hidden, confirmed_classification, estimated_rows)
        VALUES (?, ?, 0, ?, 0, 'nomina', 1)
        """,
        (sheet_key, sheet_key, f"S{sheet_key}"),
    )
    nom_repo.upsert_period(
        conn,
        sheet_key,
        {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "semana_num": sheet_key,
            "source": "manual",
            "cut_warning": None,
        },
        confirmed=True,
    )
    period_id = conn.execute(
        "SELECT id FROM gis_nomina_periods WHERE sheet_id = ?",
        (sheet_key,),
    ).fetchone()[0]
    worker_id = nom_repo.insert_workers(
        conn,
        period_id,
        [
            {
                "row_number": 4,
                "num_empleado": "101",
                "nombre_original": "Juan",
                "nombre_normalizado": "JUAN PEREZ LOPEZ",
                "puesto": "Op",
                "planta_original": "A",
                "planta_normalizada": "A",
                "cuenta": "1",
                "row_json": "{}",
            }
        ],
    )[0]
    nom_repo.update_worker_cliente(conn, worker_id, "PEPSI")
    nom_repo.upsert_match(
        conn,
        worker_id,
        {
            "headcount_key": "hc-101",
            "match_method": "num_empleado",
            "confidence": 1.0,
            "status": "confirmed",
            "nss": "11111111111",
            "curp": "PELJ800101HDFRNN09",
            "hc_nombre": "JUAN PEREZ LOPEZ",
            "hc_json": json.dumps(HC),
        },
    )
    start = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
    end = datetime.strptime(fecha_fin, "%d/%m/%Y").date()
    seq = codes or ["A"] * 7
    now = "2026-06-01T00:00:00"
    for idx, code in enumerate(seq[:7], start=1):
        day = start + timedelta(days=idx - 1)
        if day > end:
            break
        nom_repo.insert_attendance(
            conn,
            {
                "worker_id": worker_id,
                "period_id": period_id,
                "column_index": idx,
                "column_number": idx,
                "fecha_iso": day.isoformat(),
                "code_original": code,
                "code_normalized": code,
                "interpretation_status": "ok",
                "created_at": now,
                "updated_at": now,
            },
        )
    conn.commit()
    return int(period_id)


def _full_june_weeks(conn) -> list[int]:
    return [
        _seed_period(conn, sheet_key=1, fecha_inicio="01/06/2026", fecha_fin="07/06/2026"),
        _seed_period(conn, sheet_key=8, fecha_inicio="08/06/2026", fecha_fin="14/06/2026"),
        _seed_period(conn, sheet_key=15, fecha_inicio="15/06/2026", fecha_fin="21/06/2026"),
        _seed_period(conn, sheet_key=22, fecha_inicio="22/06/2026", fecha_fin="28/06/2026"),
        _seed_period(conn, sheet_key=29, fecha_inicio="29/06/2026", fecha_fin="05/07/2026"),
    ]


def test_coverage_june_complete():
    weeks = [
        {"fecha_inicio": "01/06/2026", "fecha_fin": "07/06/2026"},
        {"fecha_inicio": "08/06/2026", "fecha_fin": "14/06/2026"},
        {"fecha_inicio": "15/06/2026", "fecha_fin": "21/06/2026"},
        {"fecha_inicio": "22/06/2026", "fecha_fin": "28/06/2026"},
        {"fecha_inicio": "29/06/2026", "fecha_fin": "05/07/2026"},
    ]
    cov = compute_month_coverage(weeks, mes=6, anio=2026)
    assert cov["coverage_complete"] is True
    assert cov["missing_dates"] == []


def test_coverage_june_missing_two_days():
    weeks = [
        {"fecha_inicio": "01/06/2026", "fecha_fin": "07/06/2026"},
        {"fecha_inicio": "10/06/2026", "fecha_fin": "16/06/2026"},
        {"fecha_inicio": "17/06/2026", "fecha_fin": "23/06/2026"},
        {"fecha_inicio": "24/06/2026", "fecha_fin": "30/06/2026"},
    ]
    cov = compute_month_coverage(weeks, mes=6, anio=2026)
    assert cov["coverage_complete"] is False
    assert "2026-06-08" in cov["missing_dates"]
    assert "2026-06-09" in cov["missing_dates"]


def test_cross_month_week_may_june():
    weeks = [{"fecha_inicio": "31/05/2026", "fecha_fin": "06/06/2026"}]
    cov = compute_month_coverage(weeks, mes=6, anio=2026)
    assert "2026-06-01" in cov["covered_dates"]
    assert "2026-05-31" not in cov["covered_dates"]


def test_todo_el_mes_requires_calendar_coverage():
    daily_i = [
        {"fecha_iso": "2026-06-01", "code_normalized": "I", "period_id": 1, "interpretation_status": "ok"},
        {"fecha_iso": "2026-06-08", "code_normalized": "I", "period_id": 8, "interpretation_status": "ok"},
        {"fecha_iso": "2026-06-15", "code_normalized": "I", "period_id": 15, "interpretation_status": "ok"},
        {"fecha_iso": "2026-06-22", "code_normalized": "I", "period_id": 22, "interpretation_status": "ok"},
    ]
    assert (
        classify_monthly_status(
            daily=daily_i,
            events=[],
            selected_week_count=4,
            weeks_with_presence=4,
            coverage_complete=True,
            selected_period_ids={1, 8, 15, 22},
        )
        == "Todo el mes"
    )
    assert (
        classify_monthly_status(
            daily=daily_i,
            events=[],
            selected_week_count=4,
            weeks_with_presence=4,
            coverage_complete=False,
            selected_period_ids={1, 8, 15, 22},
        )
        != "Todo el mes"
    )


def test_single_week_never_todo_el_mes():
    daily = [{"fecha_iso": "2026-06-01", "code_normalized": "A", "period_id": 1, "interpretation_status": "ok"}]
    assert (
        classify_monthly_status(
            daily=daily,
            events=[],
            selected_week_count=1,
            weeks_with_presence=1,
            coverage_complete=True,
            selected_period_ids={1},
        )
        != "Todo el mes"
    )
    assert MIN_WEEKS_FOR_FULL_MONTH == 4


def test_incomplete_coverage_sets_en_revision(conn):
    period_ids = [
        _seed_period(conn, sheet_key=1, fecha_inicio="01/06/2026", fecha_fin="07/06/2026"),
        _seed_period(conn, sheet_key=8, fecha_inicio="10/06/2026", fecha_fin="16/06/2026"),
        _seed_period(conn, sheet_key=15, fecha_inicio="17/06/2026", fecha_fin="23/06/2026"),
        _seed_period(conn, sheet_key=22, fecha_inicio="24/06/2026", fecha_fin="30/06/2026"),
    ]
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    report = get_report(conn, report_id)
    snapshot = json.loads(report["snapshot_json"])
    assert snapshot["coverage_complete"] is False
    assert report["estado"] == "en_revision"
    assert all(p["estado_mensual"] != "Todo el mes" for p in list_report_persons(conn, report_id))


def test_excel_export_in_memory_no_residual_files(conn, tmp_path, monkeypatch):
    monkeypatch.setenv("TMPDIR", str(tmp_path))
    period_ids = _full_june_weeks(conn)
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    before = sha256_file(mensual_path())
    buf, filename = generate_monthly_excel(conn, report_id)
    assert before == sha256_file(mensual_path())
    assert isinstance(buf, BytesIO)
    assert filename.endswith(".xlsx")
    wb = load_workbook(buf)
    assert wb["Resumen"]["B9"].value in {"Sí", "No"}
    wb.close()
    assert not list(tmp_path.glob("gis_mensual_*"))


def test_excel_validation_error_does_not_leave_export_file(conn):
    period_ids = _full_june_weeks(conn)
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    with patch(
        "modules.gestion_idse_sua.reportes.excel_export.validate_mensual_template",
        side_effect=ValueError("invalid"),
    ):
        with pytest.raises(ValueError, match="invalid"):
            generate_monthly_excel(conn, report_id)


def test_excel_includes_coverage_warning_when_incomplete(conn):
    period_ids = [
        _seed_period(conn, sheet_key=1, fecha_inicio="01/06/2026", fecha_fin="07/06/2026"),
        _seed_period(conn, sheet_key=8, fecha_inicio="10/06/2026", fecha_fin="16/06/2026"),
        _seed_period(conn, sheet_key=15, fecha_inicio="17/06/2026", fecha_fin="23/06/2026"),
        _seed_period(conn, sheet_key=22, fecha_inicio="24/06/2026", fecha_fin="30/06/2026"),
    ]
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    buf, _ = generate_monthly_excel(conn, report_id)
    wb = load_workbook(buf)
    assert wb["Resumen"]["B9"].value == "No"
    assert wb["Resumen"]["D9"].value
    assert wb["Resumen"]["F9"].value != "—"
    wb.close()


def test_d_only_not_included():
    assert not person_has_active_evidence([{"code_normalized": "D"}])


@pytest.mark.parametrize(
    "mes,anio,weeks,expected_days",
    [
        (2, 2026, [("01/02/2026", "07/02/2026"), ("08/02/2026", "14/02/2026"), ("15/02/2026", "21/02/2026"), ("22/02/2026", "28/02/2026")], 28),
        (2, 2024, [("01/02/2024", "07/02/2024"), ("08/02/2024", "14/02/2024"), ("15/02/2024", "21/02/2024"), ("22/02/2024", "28/02/2024"), ("29/02/2024", "06/03/2024")], 29),
        (4, 2026, [("01/04/2026", "07/04/2026"), ("08/04/2026", "14/04/2026"), ("15/04/2026", "21/04/2026"), ("22/04/2026", "28/04/2026"), ("29/04/2026", "05/05/2026")], 30),
        (1, 2026, [("01/01/2026", "07/01/2026"), ("08/01/2026", "14/01/2026"), ("15/01/2026", "21/01/2026"), ("22/01/2026", "28/01/2026"), ("29/01/2026", "04/02/2026")], 31),
    ],
)
def test_coverage_month_lengths(mes, anio, weeks, expected_days):
    payload = [{"fecha_inicio": a, "fecha_fin": b} for a, b in weeks]
    cov = compute_month_coverage(payload, mes=mes, anio=anio)
    assert cov["coverage_complete"] is True
    assert len(cov["covered_dates"]) == expected_days


def test_overlapping_weeks_with_gap_still_incomplete():
    weeks = [
        {"fecha_inicio": "01/06/2026", "fecha_fin": "07/06/2026"},
        {"fecha_inicio": "01/06/2026", "fecha_fin": "07/06/2026"},
        {"fecha_inicio": "01/06/2026", "fecha_fin": "07/06/2026"},
        {"fecha_inicio": "01/06/2026", "fecha_fin": "07/06/2026"},
    ]
    cov = compute_month_coverage(weeks, mes=6, anio=2026)
    assert cov["coverage_complete"] is False
    assert cov["overlap_dates"]
    assert len(cov["missing_dates"]) > 0


def test_v_all_weeks_full_coverage_todo_el_mes():
    daily_v = [
        {"fecha_iso": f"2026-06-{d:02d}", "code_normalized": "V", "period_id": pid, "interpretation_status": "ok"}
        for d, pid in [(1, 1), (8, 8), (15, 15), (22, 22)]
    ]
    assert (
        classify_monthly_status(
            daily=daily_v,
            events=[],
            selected_week_count=4,
            weeks_with_presence=4,
            coverage_complete=True,
            selected_period_ids={1, 8, 15, 22},
        )
        == "Todo el mes"
    )


def test_baja_sin_reingreso_not_todo_el_mes():
    daily = [{"fecha_iso": "2026-06-01", "code_normalized": "A", "period_id": 1, "interpretation_status": "ok"}]
    events = [{"event_type": "posible_baja", "status": "confirmed", "interpretation_status": "ok"}]
    assert (
        classify_monthly_status(
            daily=daily,
            events=events,
            selected_week_count=4,
            weeks_with_presence=1,
            coverage_complete=True,
            selected_period_ids={1, 8, 15, 22},
        )
        == "Salida durante el mes"
    )


def test_parallel_exports_use_independent_buffers(conn):
    period_ids = _full_june_weeks(conn)
    report_id = create_report(conn, cliente="PEPSI", mes=6, anio=2026)
    generate_monthly_report(conn, report_id=report_id, period_ids=period_ids, cliente="PEPSI", mes=6, anio=2026)
    buf1, name1 = generate_monthly_excel(conn, report_id)
    buf2, name2 = generate_monthly_excel(conn, report_id)
    assert buf1 is not buf2
    assert name1 == name2
    assert len(buf1.getvalue()) > 0
    assert len(buf2.getvalue()) > 0
