"""Banorte batch: employee substitution, reserved numbers, prepare perf."""

from __future__ import annotations

import io
import time
from pathlib import Path

import pytest
from openpyxl import Workbook

from modules.nomina.banorte.batch_service import confirm_batch, prepare_reporte_batch
from modules.nomina.banorte.beneficiary_service import list_beneficiaries
from modules.nomina.banorte.employee_number_service import (
    BANORTE_RESERVED_EMPLOYEE_NUMBERS,
    list_available_employee_numbers,
)
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.schema import ensure_banorte_tables
from modules.nomina.db import ensure_nomina_tables

SUBSTITUTION_COMMENT = (
    "El número de empleado ya existía se asignó el número de cuenta como tu número de Empleado"
)


def _db(tmp_path):
    path = str(tmp_path / "batch_sub.db")
    conn = connect(path)
    ensure_nomina_tables(conn)
    ensure_banorte_tables(conn)
    conn.commit()
    conn.close()
    return path


def _reporte_substitution_bytes(
    *,
    requested: str = "0000000616",
    account: str = "1377672164",
    nombre: str = "EMPLEADO SUSTITUIDO",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    headers = [
        "NUMERO DE EMPLEADO",
        "NOMBRE DEL EMPLEADO",
        "NUMERO DE CUENTA",
        "ESTATUS",
        "COMENTARIOS",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)
    ws.cell(2, 1, requested)
    ws.cell(2, 2, nombre)
    ws.cell(2, 3, account)
    ws.cell(2, 4, "EXITOSO")
    ws.cell(2, 5, SUBSTITUTION_COMMENT)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _reporte_many_bytes(n_rows: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    headers = [
        "NUMERO DE EMPLEADO",
        "NOMBRE DEL EMPLEADO",
        "NUMERO DE CUENTA",
        "ESTATUS",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)
    for i in range(n_rows):
        r = i + 2
        ws.cell(r, 1, f"{3000000000 + i:010d}")
        ws.cell(r, 2, f"EMP {i}")
        ws.cell(r, 3, f"{4000000000 + i:010d}")
        ws.cell(r, 4, "EXITOSO")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_reserved_employee_numbers_are_business_blocklist():
    assert BANORTE_RESERVED_EMPLOYEE_NUMBERS == frozenset(
        {
            "0000000154",
            "0000000155",
            "0000000156",
            "0000000157",
            "0000000616",
        }
    )


def test_available_numbers_never_suggest_reserved(tmp_path):
    db = _db(tmp_path)
    avail = list_available_employee_numbers(db, limit=20)
    nums = set(avail["numbers"])
    assert nums.isdisjoint(BANORTE_RESERVED_EMPLOYEE_NUMBERS)
    assert "0000000001" in nums


def test_available_numbers_skip_reserved_even_when_unoccupied(tmp_path):
    db = _db(tmp_path)
    avail = list_available_employee_numbers(db, limit=1)
    assert avail["numbers"][0] == "0000000001"


def test_available_numbers_exclude_each_business_reserved(tmp_path):
    db = _db(tmp_path)
    avail = list_available_employee_numbers(db, limit=200)
    nums = set(avail["numbers"])
    for reserved in (
        "0000000154",
        "0000000155",
        "0000000156",
        "0000000157",
        "0000000616",
    ):
        assert reserved not in nums
    # cursor after 153 should skip the 154-157 block
    after_block = list_available_employee_numbers(db, limit=5, after="0000000153")
    assert "0000000154" not in after_block["numbers"]
    assert "0000000158" in after_block["numbers"]


def test_reporte_batch_substitution_preserved_through_confirm(tmp_path):
    db = _db(tmp_path)
    raw = _reporte_substitution_bytes()
    out = prepare_reporte_batch(db, "u", raw, "Reporte_Detallado_subst.xlsx")
    assert out["ok"] is True
    batch = out["batch"]
    assert len(batch["rows"]) == 1
    row = batch["rows"][0]
    assert row["employee_number"] == "0000000616"
    assert row["cuenta"] == "1377672164"
    assert int(row["use_account_as_employee_number"]) == 1

    confirmed = confirm_batch(db, int(batch["id"]), "u", int(batch["revision"]))
    assert confirmed["status"] == "CONFIRMED"

    listing = list_beneficiaries(db, page=1)
    assert listing["total"] == 1
    ben = listing["rows"][0]
    assert ben["employee_number_requested"] == "0000000616"
    assert ben["employee_number_effective"] == "1377672164"
    assert ben["account_number"] == "1377672164"
    assert int(ben["banorte_employee_substituted"]) == 1


def test_reporte_real_file_substitution_case(tmp_path):
    real = Path(r"c:\Users\Yahir\Downloads\Reporte_Detallado_6705926070000000010.xlsx")
    if not real.is_file():
        pytest.skip("real reporte fixture not available on this machine")
    db = _db(tmp_path)
    raw = real.read_bytes()
    out = prepare_reporte_batch(db, "u", raw, real.name, confirm_reimport=True)
    assert out["ok"] is True
    substituted = [
        r
        for r in out["batch"]["rows"]
        if r.get("employee_number") == "0000000616" and r.get("cuenta") == "1377672164"
    ]
    assert substituted, "expected substitution row 0000000616 -> 1377672164"
    row = substituted[0]
    assert int(row["use_account_as_employee_number"]) == 1
    assert SUBSTITUTION_COMMENT in str(row.get("comment") or "")

    confirmed = confirm_batch(db, int(out["batch"]["id"]), "u", int(out["batch"]["revision"]))
    assert confirmed["status"] == "CONFIRMED"
    ben = next(
        b
        for b in list_beneficiaries(db, page=1, page_size=50)["rows"]
        if b["employee_number_effective"] == "1377672164"
    )
    assert ben["employee_number_requested"] == "0000000616"
    assert int(ben["banorte_employee_substituted"]) == 1


def test_prepare_reporte_batch_bulk_under_two_seconds_for_200_rows(tmp_path):
    db = _db(tmp_path)
    raw = _reporte_many_bytes(200)
    t0 = time.perf_counter()
    out = prepare_reporte_batch(db, "u", raw, "reporte_200.xlsx", confirm_reimport=True)
    elapsed = time.perf_counter() - t0
    assert out["ok"] is True
    assert len(out["batch"]["rows"]) == 200
    assert elapsed < 2.0, f"prepare_reporte_batch took {elapsed:.2f}s for 200 rows"


def test_manual_batch_confirm_still_fast_with_many_rows(tmp_path):
    db = _db(tmp_path)
    raw = _reporte_many_bytes(200)
    out = prepare_reporte_batch(db, "u", raw, "reporte_200.xlsx", confirm_reimport=True)
    batch = out["batch"]
    t0 = time.perf_counter()
    confirm_batch(db, int(batch["id"]), "u", int(batch["revision"]))
    elapsed = time.perf_counter() - t0
    assert elapsed < 2.0, f"confirm_batch took {elapsed:.2f}s for 200 rows"
    assert list_beneficiaries(db, page=1)["total"] == 200
