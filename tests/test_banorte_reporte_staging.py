"""Fase 2.3C — Reporte Detallado prepares staging batch without direct inserts."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from modules.nomina.banorte.batch_service import confirm_batch, prepare_reporte_batch
from modules.nomina.banorte.beneficiary_service import list_beneficiaries
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.schema import ensure_banorte_tables
from modules.nomina.db import ensure_nomina_tables


def _db(tmp_path):
    path = str(tmp_path / "rep.db")
    conn = connect(path)
    ensure_nomina_tables(conn)
    ensure_banorte_tables(conn)
    conn.commit()
    conn.close()
    return path


def _reporte_bytes(*, nombre="ANA REPORTE", emp="1234567890", acct="1234567890", estatus="EXITOSO"):
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE"
    headers = [
        "NUMERO DE EMPLEADO",
        "NOMBRE DEL EMPLEADO",
        "NUMERO DE CUENTA",
        "ESTATUS",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)
    ws.cell(2, 1, emp)
    ws.cell(2, 2, nombre)
    ws.cell(2, 3, acct)
    ws.cell(2, 4, estatus)
    ws.cell(3, 1, "9999999999")
    ws.cell(3, 2, "FALLIDO REPORTE")
    ws.cell(3, 3, "9999999999")
    ws.cell(3, 4, "FALLIDO")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_prepare_reporte_does_not_insert_until_confirm(tmp_path):
    db = _db(tmp_path)
    raw = _reporte_bytes()
    out = prepare_reporte_batch(db, "u", raw, "reporte.xlsx", confirm_reimport=False)
    assert out["ok"] is True
    batch = out["batch"]
    assert batch["status"] == "OPEN"
    assert batch["origin_kind"] == "REPORTE_DETALLADO"
    assert len(batch["rows"]) == 1
    assert batch["rows"][0]["nombre"] == "ANA REPORTE"
    assert list_beneficiaries(db, page=1)["total"] == 0

    confirmed = confirm_batch(db, int(batch["id"]), "u", int(batch["revision"]))
    assert confirmed["status"] == "CONFIRMED"
    listing = list_beneficiaries(db, page=1)
    assert listing["total"] == 1
    assert listing["rows"][0]["validation_status"] == "IMPORTADO_EXITOSO"
    conn = connect(db)
    src = conn.execute(
        "SELECT source_kind FROM nomina_banorte_beneficiaries WHERE id=?",
        (listing["rows"][0]["id"],),
    ).fetchone()
    conn.close()
    assert src["source_kind"] == "REPORTE_DETALLADO"


def test_prepare_reporte_duplicate_requires_confirmation(tmp_path):
    db = _db(tmp_path)
    raw = _reporte_bytes()
    first = prepare_reporte_batch(db, "u", raw, "reporte.xlsx")
    assert first["ok"] is True
    confirm_batch(db, int(first["batch"]["id"]), "u", int(first["batch"]["revision"]))

    dup = prepare_reporte_batch(db, "u", raw, "reporte.xlsx", confirm_reimport=False)
    assert dup["ok"] is False
    assert dup["code"] == "duplicate_file_confirmation_required"


def test_hub_sha_checkboxes_removed_and_js_uses_contextual_modal():
    repo = Path(__file__).resolve().parents[1]
    html = (repo / "templates/nomina/exportaciones_banorte.html").read_text(encoding="utf-8")
    js = (repo / "static/nomina/exportaciones_banorte_editor.js").read_text(encoding="utf-8")
    assert "reimport_confirmed" not in html
    assert 'name="confirm_reimport"' not in html
    assert "duplicate_file_confirmation_required" in js
    assert "prepare-batch" in js
    assert "banorte-import-altas" in js or "import/altas" in js
    assert "Este archivo de base ya fue procesado" in js or "ya fue procesado anteriormente" in js
