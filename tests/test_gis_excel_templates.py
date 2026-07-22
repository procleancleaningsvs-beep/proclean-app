"""Validación de plantillas Excel canónicas Gestión IDSE / SUA."""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook

from modules.gestion_idse_sua.template_contract import (
    COMPARATIVO_SHA256,
    MENSUAL_SHA256,
    comparativo_path,
    mensual_path,
)
from modules.gestion_idse_sua.template_validator import (
    TemplateValidationError,
    sha256_file,
    validate_all_canonical_templates,
    validate_comparativo_template,
    validate_mensual_template,
)


def test_canonical_files_exist_and_sha_stable():
    assert comparativo_path().is_file()
    assert mensual_path().is_file()
    assert sha256_file(comparativo_path()) == COMPARATIVO_SHA256
    assert sha256_file(mensual_path()) == MENSUAL_SHA256


def test_validate_canonical_templates_ok():
    result = validate_all_canonical_templates()
    assert result["comparativo"]["ok"] is True
    assert result["mensual"]["ok"] is True
    assert "Detalle Comparativo" in result["comparativo"]["headers"]
    assert "Personal Mensual" in result["mensual"]["headers"]


def test_validation_does_not_modify_bytes(tmp_path):
    src = comparativo_path()
    copy = tmp_path / src.name
    shutil.copy2(src, copy)
    before = copy.read_bytes()
    validate_comparativo_template(copy)
    after = copy.read_bytes()
    assert before == after
    assert sha256_file(copy) == COMPARATIVO_SHA256


def test_missing_sheet_raises(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    path = tmp_path / "incomplete.xlsx"
    wb.save(path)
    with pytest.raises(TemplateValidationError, match="faltan hojas"):
        validate_comparativo_template(path)


def test_missing_column_raises(tmp_path):
    src = comparativo_path()
    copy = tmp_path / src.name
    shutil.copy2(src, copy)
    from openpyxl import load_workbook

    wb = load_workbook(copy)
    ws = wb["Detalle Comparativo"]
    # Clear NSS header
    for c in range(1, 40):
        if ws.cell(6, c).value == "NSS":
            ws.cell(6, c).value = "NSS_REMOVED"
            break
    wb.save(copy)
    with pytest.raises(TemplateValidationError, match="faltan columnas"):
        validate_comparativo_template(copy)


def test_corrupt_workbook_raises(tmp_path):
    bad = tmp_path / "corrupt.xlsx"
    bad.write_bytes(b"not-an-xlsx")
    with pytest.raises(TemplateValidationError, match="No se pudo abrir"):
        validate_mensual_template(bad)
