from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from flask import Flask, g
from openpyxl import load_workbook

from modules.nomina.banorte.history_service import (
    build_historical_export_excel,
    historical_export_excel_filename,
)
from modules.nomina.banorte.repository import connect
from modules.nomina.blueprint import register_nomina
from tests.test_banorte_export_movements import _make_app, _seed_export


def test_historical_excel_uses_export_items_snapshot(tmp_path):
    app, db_path = _make_app(tmp_path)
    exported, beneficiary_ids = _seed_export(db_path)
    payload = build_historical_export_excel(db_path, exported.export_id)
    wb = load_workbook(BytesIO(payload["data"]))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
    assert rows[0][1] == "Nombre recibido uno"
    assert rows[0][2] == "0000000011"
    assert rows[0][3] == "1321431243"

    conn = connect(db_path)
    conn.execute(
        "UPDATE nomina_banorte_beneficiaries SET nombre_original='CAMBIO VIVO' WHERE id=?",
        (beneficiary_ids[0],),
    )
    conn.commit()
    conn.close()

    payload2 = build_historical_export_excel(db_path, exported.export_id)
    wb2 = load_workbook(BytesIO(payload2["data"]))
    assert list(wb2.active.iter_rows(min_row=2, max_row=2, values_only=True))[0][1] == "Nombre recibido uno"


def test_historical_excel_route_headers_and_permissions(tmp_path):
    admin_dir = tmp_path / "admin"
    admin_dir.mkdir(parents=True, exist_ok=True)
    app, db_path = _make_app(admin_dir)
    exported, _ = _seed_export(db_path)
    url = f"/nomina/exportaciones/banorte/historial/{exported.export_id}/movimientos.xlsx"
    ok = app.test_client().get(url)
    assert ok.status_code == 200
    assert ok.headers["Cache-Control"] == "private, no-store"
    assert "attachment" in ok.headers["Content-Disposition"]
    assert ok.data[:2] == b"PK"

    user_dir = tmp_path / "usuario"
    user_dir.mkdir(parents=True, exist_ok=True)
    forbidden_app, forbidden_db = _make_app(user_dir, role="usuario")
    exported2, _ = _seed_export(forbidden_db)
    denied = forbidden_app.test_client().get(
        f"/nomina/exportaciones/banorte/historial/{exported2.export_id}/movimientos.xlsx"
    )
    assert denied.status_code == 403


def test_historical_excel_filename_deterministic():
    assert historical_export_excel_filename(
        export_id=16,
        pag_filename="banorte_20260115_001.pag",
    ).endswith("-movimientos-historico.xlsx")
