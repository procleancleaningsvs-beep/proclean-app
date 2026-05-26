"""Performance helpers: cache TTL, pagination, stats without full materialization."""
from __future__ import annotations

import sqlite3
import time
from io import BytesIO

import pandas as pd
import pytest

from modules.comparativo import headcount_service
from modules.nomina.db import ensure_nomina_tables, upsert_empleado_parametros, save_parametros_import
from modules.nomina.parametros_consolidado import (
    RECORD_HEADCOUNT_CANONICAL,
    build_consolidado_view,
    compute_parametros_stats,
    count_consolidado_view,
)
from services.perf_logging import perf_log_enabled, perf_span


def _sample_hc() -> list[dict]:
    return [
        {
            "nombre_completo": "Empleado Uno",
            "nss": "11122233344",
            "cliente": "Carrier",
            "patron": "Planta A",
            "puesto": "Aux",
            "status_operacion": "ALTA",
            "status_imss": "ALTA",
            "fecha_ingreso": "2020-01-01",
        },
        {
            "nombre_completo": "Empleado Dos",
            "nss": "55566677788",
            "cliente": "Pepsi",
            "patron": "Planta B",
            "puesto": "Op",
            "status_operacion": "ALTA",
            "status_imss": "ALTA",
            "fecha_ingreso": "2021-02-02",
        },
    ]


def test_perf_log_disabled_by_default(monkeypatch):
    monkeypatch.delenv("PERF_LOG_ENABLED", raising=False)
    assert perf_log_enabled() is False


def test_perf_span_noop_when_disabled(monkeypatch, caplog):
    monkeypatch.delenv("PERF_LOG_ENABLED", raising=False)
    with perf_span("test.block"):
        pass
    assert not any("[PERF]" in r.message for r in caplog.records)


def test_headcount_cache_hit_avoids_download(monkeypatch):
    df = pd.DataFrame([["CLIENTE", "NSS"], ["Carrier", "111"]])
    calls = {"n": 0}

    def fake_download(_url: str) -> bytes:
        calls["n"] += 1
        bio = BytesIO()
        df.to_excel(bio, index=False, header=False)
        return bio.getvalue()

    monkeypatch.setenv("HEADCOUNT_ONEDRIVE_URL", "https://example.test/headcount.xlsx")
    monkeypatch.setenv("HEADCOUNT_CACHE_TTL_SECONDS", "300")
    headcount_service.actualizar_headcount()
    monkeypatch.setattr(headcount_service, "descargar_excel_desde_onedrive", fake_download)

    out1 = headcount_service.obtener_df_headcount()
    out2 = headcount_service.obtener_df_headcount()
    assert calls["n"] == 1
    assert len(out1.index) == len(out2.index)


def test_headcount_cache_expired_refreshes(monkeypatch):
    df = pd.DataFrame([["A"], ["B"]])
    calls = {"n": 0}

    def fake_download(_url: str) -> bytes:
        calls["n"] += 1
        bio = BytesIO()
        df.to_excel(bio, index=False, header=False)
        return bio.getvalue()

    monkeypatch.setenv("HEADCOUNT_ONEDRIVE_URL", "https://example.test/headcount.xlsx")
    monkeypatch.setenv("HEADCOUNT_CACHE_TTL_SECONDS", "1")
    headcount_service.actualizar_headcount()
    monkeypatch.setattr(headcount_service, "descargar_excel_desde_onedrive", fake_download)

    headcount_service.obtener_df_headcount()
    with headcount_service._cache_lock:
        headcount_service._cache_loaded_at = time.monotonic() - 999
    headcount_service.obtener_df_headcount()
    assert calls["n"] == 2


def test_headcount_stale_fallback_on_download_error(monkeypatch):
    df = pd.DataFrame([["cached"]])
    monkeypatch.setenv("HEADCOUNT_ONEDRIVE_URL", "https://example.test/headcount.xlsx")
    monkeypatch.setenv("HEADCOUNT_CACHE_TTL_SECONDS", "1")
    headcount_service.actualizar_headcount()

    def seed_then_fail(_url: str) -> bytes:
        bio = BytesIO()
        df.to_excel(bio, index=False, header=False)
        return bio.getvalue()

    monkeypatch.setattr(headcount_service, "descargar_excel_desde_onedrive", seed_then_fail)
    headcount_service.obtener_df_headcount()

    def boom(_url: str) -> bytes:
        raise ConnectionError("onedrive down")

    monkeypatch.setattr(headcount_service, "descargar_excel_desde_onedrive", boom)
    with headcount_service._cache_lock:
        headcount_service._cache_loaded_at = time.monotonic() - 10
    stale = headcount_service.obtener_df_headcount()
    assert stale.iloc[0, 0] == "cached"
    assert headcount_service.consume_headcount_cache_warning()


def test_pagination_preserves_total_count(tmp_path):
    db = str(tmp_path / "pag.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    iso = "2026-01-01 12:00:00"
    imp_id = save_parametros_import(
        db,
        {"tipo_importacion": "NOMINA_ACTUAL", "cliente": "Carrier", "source_filename": "n.xlsx", "total_rows": 1},
        created_by=None,
        now_iso=iso,
    )
    upsert_empleado_parametros(
        db,
        [
            {
                "nombre": "Empleado Uno",
                "nombre_normalizado": "EMPLEADO UNO",
                "nss": "11122233344",
                "cliente": "Carrier",
                "record_kind": RECORD_HEADCOUNT_CANONICAL,
                "headcount_match_status": "headcount_canonical",
                "salario_operativo": 100.0,
                "valor_x_he": 10.0,
                "warnings": [],
                "editable_json": {},
            }
        ],
        import_id=imp_id,
        now_iso=iso,
    )
    hc = _sample_hc()
    total = count_consolidado_view(db, hc)
    page1 = build_consolidado_view(db, hc, offset=0, limit=1)
    page2 = build_consolidado_view(db, hc, offset=1, limit=1)
    assert total == 2
    assert len(page1) == 1
    assert len(page2) == 1
    assert page1[0]["nss"] != page2[0]["nss"]


def test_compute_parametros_stats_legacy_mode_does_not_inflate_activos(tmp_path):
    db = str(tmp_path / "legacy_stats.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    stats = compute_parametros_stats(db, _sample_hc())
    assert stats["activos_headcount"] == 2
    assert stats["stats_mode"] == "headcount"


def test_nomina_index_get_does_not_load_headcount_completo(tmp_path, monkeypatch):
    calls = {"completo": 0, "activos": 0}

    def boom_completo():
        calls["completo"] += 1
        raise AssertionError("obtener_headcount_completo must not run on GET /nomina/")

    def boom_activos(*_a, **_k):
        calls["activos"] += 1
        raise AssertionError("obtener_activos must not run on GET /nomina/")

    monkeypatch.setattr("modules.nomina.headcount_bridge.obtener_headcount_completo", boom_completo)
    monkeypatch.setattr("modules.comparativo.headcount_service.obtener_activos", boom_activos)
    monkeypatch.setattr("threading.Thread.start", lambda self: None)

    app, _db = _perf_test_app(tmp_path, monkeypatch)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/", follow_redirects=True)
            assert resp.status_code == 200
            assert "Nóminas" in resp.data.decode("utf-8", errors="replace")
    assert calls["completo"] == 0
    assert calls["activos"] == 0


def test_get_nomina_dashboard_summary_fast_uses_local_stats_only(tmp_path):
    import sqlite3

    from modules.nomina.db import ensure_nomina_tables, get_nomina_dashboard_summary_fast

    db = str(tmp_path / "dash_fast.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    summary = get_nomina_dashboard_summary_fast(db, recent_limit=5)
    assert summary["headcount_source"] in {"snapshot", "snapshot_missing", "snapshot_invalid"}
    assert summary["param_stats"]["stats_mode"] in {"legacy", "headcount_meta", "invalid"}
    assert "dash" in summary and "vac_stats" in summary


def test_dashboard_does_not_load_snapshot_rows(tmp_path, monkeypatch):
    from modules.nomina.db import ensure_nomina_tables, get_nomina_dashboard_summary_fast

    db = str(tmp_path / "dash_no_rows.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()

    calls = {"load": 0}

    def boom_load(_path, **kwargs):
        calls["load"] += 1
        raise AssertionError("load_headcount_snapshot_rows must not run on dashboard")

    monkeypatch.setattr("modules.nomina.headcount_snapshot.load_headcount_snapshot_rows", boom_load)
    get_nomina_dashboard_summary_fast(db, recent_limit=3)
    assert calls["load"] == 0


def test_nomina_index_does_not_trigger_auto_refresh(tmp_path, monkeypatch):
    calls = {"trigger": 0}

    def track_trigger(*a, **k):
        calls["trigger"] += 1
        return {"triggered": False, "reason": "web_worker_disabled"}

    monkeypatch.setattr(
        "modules.nomina.headcount_snapshot.trigger_headcount_refresh_if_needed",
        track_trigger,
    )
    monkeypatch.setattr("threading.Thread.start", lambda self: None)
    app, _db = _perf_test_app(tmp_path, monkeypatch)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/", follow_redirects=True)
            assert resp.status_code == 200
    assert calls["trigger"] == 0


def test_headcount_bridge_uses_streaming_parser():
    import inspect

    from modules.nomina import headcount_bridge

    src = inspect.getsource(headcount_bridge.parse_headcount_excel_bytes)
    assert "iter_rows" in src
    assert "read_only=True" in src
    assert "df.iloc" not in src


def test_perf_startup_log_when_enabled(monkeypatch, caplog):
    monkeypatch.setenv("PERF_LOG_ENABLED", "1")
    from app import create_app

    with caplog.at_level("INFO"):
        create_app()
    assert any("[PERF] performance logging enabled" in r.message for r in caplog.records)


def _perf_test_app(tmp_path, monkeypatch):
    import sqlite3
    from pathlib import Path

    from werkzeug.security import generate_password_hash

    from modules.nomina.db import ensure_nomina_tables

    db = str(tmp_path / "perf_app.db")
    monkeypatch.setattr("app.DB_PATH", Path(db))
    conn = sqlite3.connect(db)
    conn.execute(
        """
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    ensure_nomina_tables(conn)
    conn.execute(
        "INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, 'admin', ?)",
        ("perfadmin", generate_password_hash("secret"), "2026-01-01 00:00:00"),
    )
    conn.commit()
    conn.close()

    monkeypatch.setenv("PERF_LOG_ENABLED", "1")

    from app import create_app

    app = create_app()
    app.config["TESTING"] = True
    app.config["DATABASE"] = db
    return app, db


def test_parametros_get_does_not_call_obtener_headcount_completo(tmp_path, monkeypatch):
    calls = {"completo": 0}

    def boom_completo():
        calls["completo"] += 1
        raise AssertionError("obtener_headcount_completo must not run on GET /nomina/parametros")

    monkeypatch.setattr("modules.nomina.headcount_bridge.obtener_headcount_completo", boom_completo)
    monkeypatch.setattr("threading.Thread.start", lambda self: None)
    app, _db = _perf_test_app(tmp_path, monkeypatch)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros", follow_redirects=True)
            assert resp.status_code == 200
            body = resp.data.decode("utf-8", errors="replace")
            assert "Headcount pendiente" in body or "Headcount actualizado:" in body or "Parámetros base" in body
    assert calls["completo"] == 0


def test_parametros_conciliacion_get_does_not_call_obtener_headcount_completo(tmp_path, monkeypatch):
    calls = {"completo": 0}

    def boom_completo():
        calls["completo"] += 1
        raise AssertionError("obtener_headcount_completo must not run on GET conciliacion")

    monkeypatch.setattr("modules.nomina.headcount_bridge.obtener_headcount_completo", boom_completo)
    monkeypatch.setattr("threading.Thread.start", lambda self: None)
    app, _db = _perf_test_app(tmp_path, monkeypatch)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros/conciliacion", follow_redirects=True)
            assert resp.status_code == 200
    assert calls["completo"] == 0


def test_parametros_uses_local_snapshot(tmp_path, monkeypatch):
    from modules.nomina.headcount_snapshot import refresh_headcount_snapshot

    app, db = _perf_test_app(tmp_path, monkeypatch)
    sample = _sample_hc()
    from modules.nomina.headcount_bridge import HeadcountParseResult

    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=sample,
            source_rows_scanned=len(sample),
            saved_rows=len(sample),
            skipped_empty_rows=0,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)

    refresh_headcount_snapshot(db, now_iso="2026-05-26 10:00:00")

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros", follow_redirects=True)
            assert resp.status_code == 200
            body = resp.data.decode("utf-8", errors="replace")
            assert "Headcount actualizado: 2026-05-26 10:00:00" in body
            assert "Activos: 2" in body


def test_refresh_headcount_snapshot_persists_metadata(tmp_path, monkeypatch):
    from modules.nomina.db import ensure_nomina_tables
    from modules.nomina.headcount_snapshot import get_headcount_snapshot_meta, refresh_headcount_snapshot

    db = str(tmp_path / "snap.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()

    sample = _sample_hc()
    from modules.nomina.headcount_bridge import HeadcountParseResult

    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=sample,
            source_rows_scanned=len(sample),
            saved_rows=len(sample),
            skipped_empty_rows=0,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)

    result = refresh_headcount_snapshot(db, now_iso="2026-05-26 11:00:00")
    assert result["ok"] is True
    assert result["total_rows"] == 2
    assert result["activos_count"] == 2

    meta = get_headcount_snapshot_meta(db)
    assert meta is not None
    assert meta["status"] == "ok"
    assert meta["total_rows"] == 2
    assert meta["activos_count"] == 2


def test_dashboard_kpis_use_snapshot_not_contpaq_inflation(tmp_path, monkeypatch):
    from modules.nomina.db import ensure_nomina_tables, get_nomina_dashboard_summary_fast, save_parametros_import, upsert_empleado_parametros
    from modules.nomina.headcount_snapshot import refresh_headcount_snapshot
    from modules.nomina.parametros_consolidado import RECORD_EXTERNAL_CONTPAQ

    db = str(tmp_path / "kpi_snap.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    iso = "2026-01-01 12:00:00"
    imp_id = save_parametros_import(
        db,
        {"tipo_importacion": "CONTPAQ", "cliente": "Carrier", "source_filename": "c.xlsx", "total_rows": 10},
        created_by=None,
        now_iso=iso,
    )
    upsert_empleado_parametros(
        db,
        [
            {
                "nombre": f"Externo {i}",
                "nombre_normalizado": f"EXTERNO {i}",
                "nss": f"9990000000{i}",
                "cliente": "Carrier",
                "record_kind": RECORD_EXTERNAL_CONTPAQ,
                "headcount_match_status": "no_match_contpaq",
                "salario_operativo": None,
                "valor_x_he": None,
                "warnings": [],
                "editable_json": {},
            }
            for i in range(5)
        ],
        import_id=imp_id,
        now_iso=iso,
    )

    sample = _sample_hc()
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: __import__(
            "modules.nomina.headcount_bridge", fromlist=["HeadcountParseResult"]
        ).HeadcountParseResult(
            rows=sample,
            source_rows_scanned=len(sample),
            saved_rows=len(sample),
            skipped_empty_rows=0,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)

    refresh_headcount_snapshot(db, now_iso="2026-05-26 12:00:00")

    summary = get_nomina_dashboard_summary_fast(db, recent_limit=5)
    assert summary["param_stats"]["activos_headcount"] == 2
    assert summary["param_stats"]["stats_mode"] == "headcount_meta"
    assert summary["headcount_source"] == "snapshot"


def test_perf_request_log_when_enabled(tmp_path, monkeypatch, caplog):
    monkeypatch.setattr("modules.nomina.headcount_bridge.obtener_headcount_completo", lambda: [])
    monkeypatch.setattr("threading.Thread.start", lambda self: None)
    app, _db = _perf_test_app(tmp_path, monkeypatch)

    with app.app_context():
        with caplog.at_level("INFO"):
            with app.test_client() as client:
                client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
                client.get("/nomina/parametros", follow_redirects=True)
    assert any(
        "[PERF] GET /nomina/parametros" in r.message and "duration_ms=" in r.message
        for r in caplog.records
    )


def _seed_snapshot(db: str, *, now_iso: str, activos: int = 2) -> None:
    from modules.nomina.db import ensure_nomina_tables
    from modules.nomina.headcount_bridge import HeadcountParseResult
    from modules.nomina.headcount_snapshot import refresh_headcount_snapshot

    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    sample = _sample_hc()
    import modules.comparativo.headcount_service as hs
    import modules.nomina.headcount_bridge as hb

    old_fetch = hb.fetch_and_parse_headcount
    old_actualizar = hs.actualizar_headcount

    def fake_fetch():
        return HeadcountParseResult(
            rows=sample,
            source_rows_scanned=len(sample),
            saved_rows=len(sample),
            skipped_empty_rows=0,
            guardrail_triggered=False,
        )

    hb.fetch_and_parse_headcount = fake_fetch
    hs.actualizar_headcount = lambda: None
    try:
        refresh_headcount_snapshot(db, now_iso=now_iso)
    finally:
        hb.fetch_and_parse_headcount = old_fetch
        hs.actualizar_headcount = old_actualizar


def test_stale_snapshot_loads_without_remote_on_get(tmp_path, monkeypatch):
    from modules.nomina.headcount_snapshot import get_headcount_snapshot_meta, is_headcount_snapshot_stale

    app, db = _perf_test_app(tmp_path, monkeypatch)
    _seed_snapshot(db, now_iso="2020-01-01 08:00:00")
    monkeypatch.setenv("HEADCOUNT_SNAPSHOT_TTL_SECONDS", "60")
    assert is_headcount_snapshot_stale(db) is True

    calls = {"completo": 0}

    def boom():
        calls["completo"] += 1
        raise AssertionError("remote headcount must not run on GET")

    monkeypatch.setattr("modules.nomina.headcount_bridge.obtener_headcount_completo", boom)
    monkeypatch.setattr("threading.Thread.start", lambda self: None)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros", follow_redirects=True)
            assert resp.status_code == 200
            body = resp.data.decode("utf-8", errors="replace")
            assert "Headcount actualizado: 2020-01-01 08:00:00" in body
            assert "Activos: 2" in body
    assert calls["completo"] == 0
    meta = get_headcount_snapshot_meta(db)
    assert int(meta.get("total_rows") or 0) == 2


def test_refresh_failure_preserves_last_valid_snapshot(tmp_path, monkeypatch):
    from modules.nomina.headcount_snapshot import load_headcount_snapshot_rows, refresh_headcount_snapshot

    db = str(tmp_path / "preserve.db")
    _seed_snapshot(db, now_iso="2026-05-26 10:00:00")
    assert len(load_headcount_snapshot_rows(db)) == 2

    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: (_ for _ in ()).throw(RuntimeError("onedrive down")),
    )
    result = refresh_headcount_snapshot(db, now_iso="2026-05-26 11:00:00")
    assert result["ok"] is False
    assert result.get("preserved") is True
    assert len(load_headcount_snapshot_rows(db)) == 2


def test_acquire_headcount_refresh_lock_prevents_duplicate(tmp_path):
    from modules.nomina.db import ensure_nomina_tables
    from modules.nomina.headcount_snapshot import acquire_headcount_refresh_lock

    db = str(tmp_path / "lock.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    assert acquire_headcount_refresh_lock(db, now_iso="2026-05-26 12:00:00") is True
    assert acquire_headcount_refresh_lock(db, now_iso="2026-05-26 12:00:01") is False


def test_manual_refresh_endpoint_updates_snapshot(tmp_path, monkeypatch):
    app, db = _perf_test_app(tmp_path, monkeypatch)
    from modules.nomina.headcount_bridge import HeadcountParseResult

    sample = _sample_hc()
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=sample,
            source_rows_scanned=len(sample),
            saved_rows=len(sample),
            skipped_empty_rows=0,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.post("/nomina/headcount/actualizar", follow_redirects=True)
            assert resp.status_code == 200
            body = resp.data.decode("utf-8", errors="replace")
            assert "Headcount actualizado:" in body or "activos" in body


def test_get_headcount_snapshot_meta_fields(tmp_path):
    db = str(tmp_path / "meta.db")
    _seed_snapshot(db, now_iso="2026-05-26 13:00:00")
    from modules.nomina.headcount_snapshot import get_headcount_snapshot_meta

    meta = get_headcount_snapshot_meta(db)
    assert meta["last_refresh_at"] == "2026-05-26 13:00:00"
    assert meta["activos_count"] == 2
    assert meta["status"] == "ok"


def test_no_snapshot_page_loads_fast_with_notice(tmp_path, monkeypatch):
    monkeypatch.setattr("threading.Thread.start", lambda self: None)
    app, _db = _perf_test_app(tmp_path, monkeypatch)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros", follow_redirects=True)
            assert resp.status_code == 200
            body = resp.data.decode("utf-8", errors="replace")
            assert "Headcount pendiente" in body or "Headcount actualizado:" in body


def test_get_headcount_snapshot_meta_skips_ddl_on_read(tmp_path, monkeypatch):
    db = str(tmp_path / "no_ddl_read.db")
    _seed_snapshot(db, now_iso="2026-05-26 14:00:00")
    from modules.nomina import headcount_snapshot as hs

    def boom_ddl(conn):
        raise AssertionError("ensure_headcount_snapshot_tables must not run on GET read")

    monkeypatch.setattr(hs, "ensure_headcount_snapshot_tables", boom_ddl)
    meta = hs.get_headcount_snapshot_meta(db)
    rows = hs.load_headcount_snapshot_rows(db)
    assert meta is not None
    assert len(rows) == 2


def test_parametros_get_survives_sqlite_locked(tmp_path, monkeypatch):
    from modules.nomina import headcount_snapshot as hs

    app, db = _perf_test_app(tmp_path, monkeypatch)
    _seed_snapshot(db, now_iso="2026-05-26 15:00:00")

    def locked_meta(_path):
        raise sqlite3.OperationalError("database is locked")

    def locked_rows(_path, **kwargs):
        raise sqlite3.OperationalError("database is locked")

    monkeypatch.setattr(hs, "get_headcount_snapshot_meta", locked_meta)
    monkeypatch.setattr(hs, "load_headcount_snapshot_rows", locked_rows)
    monkeypatch.setattr(hs, "is_headcount_snapshot_refreshing", lambda *a, **k: True)

    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros", follow_redirects=True)
            assert resp.status_code == 200
            body = resp.data.decode("utf-8", errors="replace")
            assert "Headcount se está actualizando" in body or "modo limitado" in body


def test_refresh_download_happens_before_db_write(tmp_path, monkeypatch):
    from modules.nomina.headcount_bridge import HeadcountParseResult
    from modules.nomina.headcount_snapshot import refresh_headcount_snapshot

    db = str(tmp_path / "order.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()

    order: list[str] = []
    sample = _sample_hc()

    def fake_fetch():
        order.append("download")
        return HeadcountParseResult(
            rows=sample,
            source_rows_scanned=len(sample),
            saved_rows=len(sample),
            skipped_empty_rows=0,
            guardrail_triggered=False,
        )

    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)
    monkeypatch.setattr("modules.nomina.headcount_bridge.fetch_and_parse_headcount", fake_fetch)
    monkeypatch.setattr(
        "modules.nomina.headcount_snapshot._atomic_write_snapshot",
        lambda *a, **k: order.append("db_write"),
    )

    result = refresh_headcount_snapshot(db, now_iso="2026-05-26 16:00:00")
    assert result["ok"] is True
    assert order.index("download") < order.index("db_write")


def test_trigger_does_not_start_refresh_in_web_worker(tmp_path, monkeypatch):
    from modules.nomina.db import ensure_nomina_tables
    from modules.nomina.headcount_snapshot import trigger_headcount_refresh_if_needed

    db = str(tmp_path / "trigger_skip.db")
    conn = sqlite3.connect(db)
    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    monkeypatch.setenv("HEADCOUNT_SNAPSHOT_TTL_SECONDS", "1")
    out = trigger_headcount_refresh_if_needed(db, now_iso="2026-05-26 18:00:00")
    assert out["triggered"] is False
    assert out["reason"] == "web_worker_disabled"


def test_get_headcount_snapshot_locked_fallback(tmp_path, monkeypatch):
    from modules.nomina import headcount_snapshot as hs

    db = str(tmp_path / "fallback.db")
    _seed_snapshot(db, now_iso="2026-05-26 19:00:00")

    def locked_rows(_path, **kwargs):
        raise sqlite3.OperationalError("database is locked")

    monkeypatch.setattr(hs, "load_headcount_snapshot_rows", locked_rows)
    monkeypatch.setattr(hs, "is_headcount_snapshot_refreshing", lambda *a, **k: True)
    ctx = hs.get_headcount_snapshot(db)
    assert ctx["status"] in {"refreshing", "locked", "limited"}
    assert "actualizando" in ctx["message"].lower()


def _make_headcount_xlsx(data_rows: list[list], trailing_empty: int = 0) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "X",
            "NOMBRE COMPLETO",
            "CLIENTE",
            "PATRON",
            "FECHA DE INGRESO",
            "SUELDO DIARIO",
            "PUESTO",
            "NSS",
            "STATUS OPERACIÓN",
            "STATUS IMSS",
        ]
    )
    for row in data_rows:
        ws.append(row)
    for _ in range(trailing_empty):
        ws.append([None] * 10)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_parser_ignores_empty_rows_and_stops_after_consecutive_blank():
    from modules.nomina.headcount_bridge import parse_headcount_excel_bytes

    raw = _make_headcount_xlsx(
        [
            ["", "Empleado Uno", "Carrier", "Planta A", "2020-01-01", 100, "Aux", "111", "ALTA", "ALTA"],
            ["", "Empleado Dos", "Pepsi", "Planta B", "2021-02-02", 100, "Op", "222", "ALTA", "ALTA"],
        ],
        trailing_empty=120,
    )
    result = parse_headcount_excel_bytes(raw)
    assert result.guardrail_triggered is False
    assert result.saved_rows == 2
    assert result.skipped_empty_rows >= 100


def test_parser_aborts_row_explosion_at_million(monkeypatch):
    from modules.nomina.headcount_bridge import parse_headcount_excel_bytes

    calls = {"n": 0}

    def fake_iter(self, values_only=True):
        if calls["n"] == 0:
            calls["n"] = 1
            yield (
                "X",
                "NOMBRE COMPLETO",
                "CLIENTE",
                "PATRON",
                "FECHA DE INGRESO",
                "SUELDO DIARIO",
                "PUESTO",
                "NSS",
                "STATUS OPERACIÓN",
                "STATUS IMSS",
            )
        while calls["n"] <= 1_000_000:
            calls["n"] += 1
            yield ("", "X", "C", "P", "", "", "", "1", "ALTA", "ALTA")

    class FakeWS:
        def iter_rows(self, values_only=True):
            return fake_iter(self, values_only=values_only)

    class FakeWB:
        def __init__(self):
            self.active = FakeWS()

        def close(self):
            pass

    monkeypatch.setattr("openpyxl.load_workbook", lambda **k: FakeWB())
    result = parse_headcount_excel_bytes(b"fake")
    assert result.guardrail_triggered is True


def test_refresh_metadata_stores_source_and_saved_rows(tmp_path, monkeypatch):
    from modules.nomina.headcount_bridge import HeadcountParseResult
    from modules.nomina.headcount_snapshot import get_headcount_snapshot_meta, refresh_headcount_snapshot

    db = str(tmp_path / "meta_fields.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    sample = _sample_hc()
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=sample,
            source_rows_scanned=5000,
            saved_rows=2,
            skipped_empty_rows=4998,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)
    refresh_headcount_snapshot(db, now_iso="2026-05-26 20:00:00")
    meta = get_headcount_snapshot_meta(db)
    assert meta["total_rows"] == 2
    assert meta["total_rows_source"] == 5000
    assert meta["skipped_empty_rows"] == 4998
    assert meta["activos_count"] == 2


def _seed_snapshot_meta(db: str, *, total_rows: int, activos: int = 452, status: str = "ok") -> None:
    conn = sqlite3.connect(db)
    from modules.nomina.headcount_snapshot import ensure_headcount_snapshot_tables

    ensure_headcount_snapshot_tables(conn)
    conn.execute(
        """
        UPDATE nomina_headcount_snapshot_meta SET
            status = ?, total_rows = ?, activos_count = ?, last_refresh_at = ?
        WHERE id = 1
        """,
        (status, total_rows, activos, "2026-05-20 10:00:00"),
    )
    conn.commit()
    conn.close()


def _insert_dummy_snapshot_rows(db: str, count: int) -> None:
    conn = sqlite3.connect(db)
    from modules.nomina.headcount_snapshot import ensure_headcount_snapshot_tables

    ensure_headcount_snapshot_tables(conn)
    rows = [
        (
            f"nss{i}",
            f"Empleado {i}",
            f"empleado {i}",
            "Carrier",
            "Planta",
            "Op",
            "2020-01-01",
            "ALTA",
            "ALTA",
            1,
            None,
            None,
            None,
            "{}",
            1,
            "2026-05-20 10:00:00",
        )
        for i in range(count)
    ]
    conn.executemany(
        """
        INSERT INTO nomina_headcount_snapshot (
            nss, nombre_completo, nombre_normalizado, cliente, planta, puesto,
            fecha_ingreso, status, status_imss, activo, sueldo, banco, cuenta,
            raw_json, snapshot_version, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    conn.commit()
    conn.close()


def test_corrupt_meta_total_saved_marks_invalid(tmp_path):
    from modules.nomina.headcount_snapshot import get_headcount_snapshot_meta_fast, validate_snapshot_state

    db = str(tmp_path / "corrupt_meta.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    _seed_snapshot_meta(db, total_rows=1_048_570, activos=452, status="ok")
    state = validate_snapshot_state(db)
    assert state["snapshot_valid"] is False
    assert state["invalid_reason"] == "too_many_rows"
    meta = get_headcount_snapshot_meta_fast(db)
    assert meta.get("snapshot_valid") is False
    assert meta.get("status") == "invalid"


def test_invalid_snapshot_skips_row_load(tmp_path, caplog):
    from modules.nomina.headcount_snapshot import load_headcount_snapshot_rows

    db = str(tmp_path / "skip_load.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    _seed_snapshot_meta(db, total_rows=1_048_570)
    _insert_dummy_snapshot_rows(db, 100)
    rows = load_headcount_snapshot_rows(db)
    assert rows == []
    assert any("snapshot_invalid_skip_load" in r.message for r in caplog.records)


def test_parametros_fast_with_invalid_snapshot(tmp_path, monkeypatch):
    app, db = _perf_test_app(tmp_path, monkeypatch)
    _seed_snapshot_meta(db, total_rows=1_048_570)
    load_calls = {"n": 0}
    real_load = __import__(
        "modules.nomina.headcount_snapshot", fromlist=["load_headcount_snapshot_rows"]
    ).load_headcount_snapshot_rows

    def track_load(path, **kwargs):
        load_calls["n"] += 1
        return real_load(path, **kwargs)

    monkeypatch.setattr(
        "modules.nomina.headcount_snapshot.load_headcount_snapshot_rows",
        track_load,
    )
    t0 = time.perf_counter()
    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/parametros", follow_redirects=True)
    elapsed_ms = int((time.perf_counter() - t0) * 1000)
    assert resp.status_code == 200
    body = resp.data.decode("utf-8", errors="replace")
    assert "requiere reconstrucción" in body.lower() or "Reconstruir Headcount" in body
    assert load_calls["n"] == 0
    assert elapsed_ms < 5000


def test_manual_refresh_calls_force_true(tmp_path, monkeypatch):
    calls = {}

    def fake_refresh(db_path, **kwargs):
        calls.update(kwargs)
        return {
            "ok": True,
            "activos_count": 2,
            "total_rows": 2,
            "last_refresh_at": "2026-05-26 12:00:00",
        }

    monkeypatch.setattr(
        "modules.nomina.headcount_snapshot.refresh_headcount_snapshot",
        fake_refresh,
    )
    app, db = _perf_test_app(tmp_path, monkeypatch)
    _seed_snapshot_meta(db, total_rows=1_048_570, status="refreshing")
    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.post("/nomina/headcount/actualizar", follow_redirects=True)
    assert resp.status_code == 200
    assert calls.get("force") is True
    assert calls.get("source") == "manual_button"


def test_manual_refresh_does_not_use_trigger(tmp_path, monkeypatch):
    trigger_calls = {"n": 0}

    def boom_trigger(*a, **k):
        trigger_calls["n"] += 1
        raise AssertionError("trigger must not run on manual POST")

    monkeypatch.setattr(
        "modules.nomina.headcount_snapshot.trigger_headcount_refresh_if_needed",
        boom_trigger,
    )
    monkeypatch.setattr(
        "modules.nomina.headcount_snapshot.refresh_headcount_snapshot",
        lambda *a, **k: {"ok": True, "activos_count": 1, "total_rows": 1, "last_refresh_at": "x"},
    )
    app, db = _perf_test_app(tmp_path, monkeypatch)
    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            client.post("/nomina/headcount/actualizar", follow_redirects=True)
    assert trigger_calls["n"] == 0


def test_force_refresh_replaces_corrupt_snapshot(tmp_path, monkeypatch):
    from modules.nomina.headcount_bridge import HeadcountParseResult
    from modules.nomina.headcount_snapshot import (
        load_headcount_snapshot_rows,
        refresh_headcount_snapshot,
    )

    db = str(tmp_path / "replace_corrupt.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    _seed_snapshot_meta(db, total_rows=1_048_570)
    _insert_dummy_snapshot_rows(db, 50)
    sample = _sample_hc()
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=sample,
            source_rows_scanned=500,
            saved_rows=2,
            skipped_empty_rows=498,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)
    result = refresh_headcount_snapshot(db, now_iso="2026-05-26 13:00:00", force=True, source="manual_button")
    assert result["ok"] is True
    assert len(load_headcount_snapshot_rows(db)) == 2


def test_refresh_does_not_append_over_old_rows(tmp_path, monkeypatch):
    from modules.nomina.headcount_bridge import HeadcountParseResult
    from modules.nomina.headcount_snapshot import refresh_headcount_snapshot

    db = str(tmp_path / "no_append.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    _insert_dummy_snapshot_rows(db, 30)
    _seed_snapshot_meta(db, total_rows=30)
    sample = _sample_hc()
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=sample,
            source_rows_scanned=100,
            saved_rows=2,
            skipped_empty_rows=98,
            guardrail_triggered=False,
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)
    refresh_headcount_snapshot(db, now_iso="2026-05-26 14:00:00", force=True)
    conn = sqlite3.connect(db)
    count = conn.execute("SELECT COUNT(*) FROM nomina_headcount_snapshot").fetchone()[0]
    conn.close()
    assert count == 2


def test_failed_refresh_keeps_corrupt_invalid(tmp_path, monkeypatch):
    from modules.nomina.headcount_bridge import HeadcountParseResult
    from modules.nomina.headcount_snapshot import (
        get_headcount_snapshot_meta,
        headcount_snapshot_available,
        refresh_headcount_snapshot,
    )

    db = str(tmp_path / "fail_refresh.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    _seed_snapshot_meta(db, total_rows=1_048_570)
    _insert_dummy_snapshot_rows(db, 25_000)
    monkeypatch.setattr(
        "modules.nomina.headcount_bridge.fetch_and_parse_headcount",
        lambda: HeadcountParseResult(
            rows=[],
            source_rows_scanned=1_048_570,
            saved_rows=0,
            skipped_empty_rows=0,
            guardrail_triggered=True,
            guardrail_reason="row_explosion",
        ),
    )
    monkeypatch.setattr("modules.comparativo.headcount_service.actualizar_headcount", lambda: None)
    result = refresh_headcount_snapshot(db, now_iso="2026-05-26 15:00:00", force=True)
    assert result["ok"] is False
    meta = get_headcount_snapshot_meta(db)
    assert meta["status"] == "invalid"
    assert headcount_snapshot_available(db) is False


def test_purge_snapshot_clears_rows_keeps_meta(tmp_path):
    from modules.nomina.headcount_snapshot import (
        get_headcount_snapshot_meta,
        purge_headcount_snapshot,
    )

    db = str(tmp_path / "purge.db")
    conn = sqlite3.connect(db)
    from modules.nomina.db import ensure_nomina_tables

    ensure_nomina_tables(conn)
    conn.commit()
    conn.close()
    _seed_snapshot_meta(db, total_rows=1_048_570)
    _insert_dummy_snapshot_rows(db, 10)
    result = purge_headcount_snapshot(db, now_iso="2026-05-26 16:00:00")
    assert result["ok"] is True
    conn = sqlite3.connect(db)
    row_count = conn.execute("SELECT COUNT(*) FROM nomina_headcount_snapshot").fetchone()[0]
    staging_count = conn.execute("SELECT COUNT(*) FROM nomina_headcount_snapshot_staging").fetchone()[0]
    conn.close()
    assert row_count == 0
    assert staging_count == 0
    meta = get_headcount_snapshot_meta(db)
    assert meta is not None
    assert meta["status"] == "purged"
    assert int(meta["total_rows"] or 0) == 0


def test_dashboard_shows_reconstruction_message(tmp_path, monkeypatch):
    app, db = _perf_test_app(tmp_path, monkeypatch)
    _seed_snapshot_meta(db, total_rows=1_048_570)
    with app.app_context():
        with app.test_client() as client:
            client.post("/login", data={"username": "perfadmin", "password": "secret"}, follow_redirects=True)
            resp = client.get("/nomina/", follow_redirects=True)
    body = resp.data.decode("utf-8", errors="replace")
    assert resp.status_code == 200
    assert "requiere reconstrucción" in body.lower() or "Reconstruir Headcount" in body
