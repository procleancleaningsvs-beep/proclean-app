"""GIS Nóminas — parser de asistencia y trayectoria."""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas.attendance_parser import (
    detect_attendance_block,
    extract_attendance_for_workers,
    normalize_attendance_code,
)
from modules.gestion_idse_sua.nominas.import_service import extract_sheet_workers, register_import
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables
from modules.gestion_idse_sua.nominas.trajectory_service import (
    count_consecutive_absences,
    detect_four_absence_event,
    merge_daily_records,
)


FIXTURE = Path("tests/fixtures/gis_nomina_asistencia_anon.xlsx")
CARRIER = Path("tests/fixtures/nomina_carrier_anon.xlsx")


def test_normalize_attendance_codes():
    assert normalize_attendance_code(" inc ") == {"original": "inc", "normalized": "I", "status": "ok"}
    assert normalize_attendance_code("a")["normalized"] == "A"
    assert normalize_attendance_code("")["status"] == "empty"
    assert normalize_attendance_code("X")["status"] == "review"


def test_detect_attendance_block_on_v4_fixture():
    wb = load_workbook(FIXTURE, data_only=True)
    ws = wb.active
    block = detect_attendance_block(ws, header_row=4, nombre_col=1)
    wb.close()
    assert block is not None
    assert block["start_col"] == 7
    assert len(block["headers"]) == 7


def test_pepsi_layout_block_starts_after_valor_he():
    pepsi = Path("tests/fixtures/gis_pepsi_nomina_anon.xlsx")
    if not pepsi.is_file():
        pytest.skip("Pepsi anon fixture not generated")
    wb = load_workbook(pepsi, data_only=True)
    ws = wb["Semana Normal Anon"]
    block = detect_attendance_block(ws, header_row=6, nombre_col=3)
    wb.close()
    assert block is not None
    assert block["start_col"] == 11


def test_carrier_fixture_has_no_attendance_block():
    wb = load_workbook(CARRIER, data_only=True)
    ws = wb.active
    block = detect_attendance_block(ws, header_row=3, nombre_col=3)
    wb.close()
    assert block is None


def test_assign_dates_from_confirmed_period():
    wb = load_workbook(FIXTURE, data_only=True)
    ws = wb.active
    payload = extract_attendance_for_workers(
        ws,
        header_row=4,
        nombre_col=1,
        fecha_inicio="01/06/2026",
        fecha_fin="07/06/2026",
        worker_rows=[5],
    )
    wb.close()
    days = payload["rows"][5]
    assert days[0]["fecha_iso"] == "2026-06-01"
    assert days[6]["fecha_iso"] == "2026-06-07"
    assert days[0]["code_normalized"] == "A"


def test_four_absences_ignore_descanso():
    assert count_consecutive_absences(["F", "F", "D", "F", "F"]) == 4
    assert detect_four_absence_event(["F", "F", "D", "F", "F"]) is True
    assert detect_four_absence_event(["F", "F", "D", "F"]) is False


def test_merge_daily_records_deduplicates_same_code():
    records = [
        {"fecha_iso": "2026-06-01", "code_normalized": "A", "period_id": 1},
        {"fecha_iso": "2026-06-01", "code_normalized": "A", "period_id": 2},
    ]
    merged, warnings = merge_daily_records(records)
    assert len(merged) == 1
    assert warnings == []


def test_merge_daily_records_flags_conflict():
    records = [
        {"fecha_iso": "2026-06-01", "code_normalized": "A", "period_id": 1},
        {"fecha_iso": "2026-06-01", "code_normalized": "F", "period_id": 2},
    ]
    merged, warnings = merge_daily_records(records)
    assert len(merged) == 1
    assert merged[0]["interpretation_status"] == "conflict"
    assert warnings


@pytest.fixture
def conn(tmp_path):
    connection = sqlite3.connect(tmp_path / "att.db")
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    yield connection
    connection.close()


def test_import_persists_attendance_rows(conn):
    data = FIXTURE.read_bytes()
    registered = register_import(conn, file_bytes=data, filename="asistencia.xlsx", uploaded_by="test")
    sheet = next(s for s in registered["sheets"] if s["sheet_name"] == "Asistencia")
    conn.execute(
        "UPDATE gis_nomina_sheets SET confirmed_classification = 'nomina' WHERE id = ?",
        (sheet["id"],),
    )
    conn.commit()
    from modules.gestion_idse_sua.nominas.import_service import confirm_period

    confirm_period(
        conn,
        int(sheet["id"]),
        fecha_inicio="01/06/2026",
        fecha_fin="07/06/2026",
        cliente="PEPSI",
        confirmed=True,
    )
    out = extract_sheet_workers(conn, file_bytes=data, sheet_id=int(sheet["id"]), headcount_rows=[])
    assert out["attendance"]["block_detected"] is True
    count = conn.execute("SELECT COUNT(*) FROM gis_nomina_attendance WHERE period_id = ?", (out["period_id"],)).fetchone()[0]
    assert count == 14
