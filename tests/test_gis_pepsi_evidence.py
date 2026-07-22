"""Local evidence tests against operational Pepsi workbook (skipped when unavailable)."""

from __future__ import annotations

import hashlib
import os
import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas.attendance_parser import detect_attendance_block
from modules.gestion_idse_sua.nominas.period_signals import collect_period_signals
from modules.gestion_idse_sua.nominas.sheet_inspector import inspect_sheet

EVIDENCE_SHA256 = "3f0042cf96ec8f35d496516b272e32d95b66e08249f8f60e18c1cda77f824227"
DEFAULT_EVIDENCE = Path(r"c:\Users\Yahir\Documents\Nomina de Pepsi Definitiva 2026.xlsm")


def _evidence_path() -> Path:
    return Path(os.environ.get("GIS_PEPSI_EVIDENCE_XLSM", str(DEFAULT_EVIDENCE)))


pytestmark = pytest.mark.skipif(not _evidence_path().is_file(), reason="Pepsi evidence xlsm not available locally")


@pytest.fixture
def evidence_workbook():
    src = _evidence_path()
    data = src.read_bytes()
    assert hashlib.sha256(data).hexdigest() == EVIDENCE_SHA256
    tmp = Path(tempfile.mkdtemp()) / "evidence.xlsm"
    shutil.copy2(src, tmp)
    wb = load_workbook(tmp, data_only=True)
    yield wb
    wb.close()
    tmp.unlink(missing_ok=True)
    shutil.rmtree(tmp.parent, ignore_errors=True)


def test_workbook_structure(evidence_workbook):
    assert len(evidence_workbook.sheetnames) == 27
    assert "CONTPAQi" in evidence_workbook.sheetnames


def test_weekly_sheet_detects_kq_block(evidence_workbook):
    ws = evidence_workbook["2 al 8 julio"]
    inspection = inspect_sheet(ws, sheet_name="2 al 8 julio", sheet_index=0, is_hidden=False)
    block = detect_attendance_block(
        ws,
        header_row=int(inspection["header_row"]),
        nombre_col=int(inspection["columns"]["nombre"]),
    )
    assert block is not None
    assert block["start_col"] == 11
    assert inspection["suggested_classification"] == "nomina"


def test_contpaq_classified_auxiliar(evidence_workbook):
    ws = evidence_workbook["CONTPAQi"]
    inspection = inspect_sheet(ws, sheet_name="CONTPAQi", sheet_index=0, is_hidden=False)
    assert inspection["suggested_classification"] == "auxiliar"


def test_period_conflict_signals(evidence_workbook):
    ws = evidence_workbook["26 al 1 de julio"]
    inspection = inspect_sheet(ws, sheet_name="26 al 1 de julio", sheet_index=0, is_hidden=False)
    payload = collect_period_signals(
        ws,
        sheet_name="26 al 1 de julio",
        header_row=inspection.get("header_row"),
        nombre_col=(inspection.get("columns") or {}).get("nombre"),
    )
    assert payload["warnings"]
    starts = {(s.get("fecha_inicio"), s.get("fecha_fin")) for s in payload["signals"] if s.get("fecha_inicio")}
    assert len(starts) >= 2
