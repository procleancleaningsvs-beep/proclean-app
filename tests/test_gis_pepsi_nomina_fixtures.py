"""GIS Nóminas — Pepsi anonymized fixtures."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas.attendance_parser import detect_attendance_block
from modules.gestion_idse_sua.nominas.import_service import confirm_period, extract_sheet_workers, register_import
from modules.gestion_idse_sua.nominas.period_signals import collect_period_signals
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables
from modules.gestion_idse_sua.nominas.sheet_inspector import inspect_sheet
from modules.gestion_idse_sua.nominas.trajectory_service import detect_four_absence_event, suggest_trajectory_events

PEPSI_FIXTURE = Path("tests/fixtures/gis_pepsi_nomina_anon.xlsx")


@pytest.fixture
def conn(tmp_path):
    connection = sqlite3.connect(tmp_path / "pepsi.db")
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    yield connection
    connection.close()


def test_pepsi_anon_fixture_detects_block_col_11():
    wb = load_workbook(PEPSI_FIXTURE, data_only=True)
    ws = wb["Semana Normal Anon"]
    inspection = inspect_sheet(ws, sheet_name=ws.title, sheet_index=0, is_hidden=False)
    block = detect_attendance_block(ws, header_row=6, nombre_col=3)
    wb.close()
    assert block is not None
    assert block["start_col"] == 11
    assert inspection["columns"]["planta"] == 4


def test_pepsi_conflict_sheet_warns():
    wb = load_workbook(PEPSI_FIXTURE, data_only=True)
    ws = wb["Conflicto Periodo Anon"]
    payload = collect_period_signals(ws, sheet_name=ws.title, header_row=6, nombre_col=3)
    wb.close()
    assert payload["warnings"]


def test_pepsi_import_persists_attendance(conn):
    data = PEPSI_FIXTURE.read_bytes()
    registered = register_import(conn, file_bytes=data, filename="pepsi_anon.xlsx", uploaded_by="test")
    sheet = next(s for s in registered["sheets"] if s["sheet_name"] == "Semana Normal Anon")
    conn.execute(
        "UPDATE gis_nomina_sheets SET confirmed_classification = 'nomina' WHERE id = ?",
        (sheet["id"],),
    )
    conn.commit()
    confirm_period(
        conn,
        int(sheet["id"]),
        fecha_inicio="02/07/2026",
        fecha_fin="08/07/2026",
        cliente="PEPSI",
        confirmed=True,
    )
    out = extract_sheet_workers(conn, file_bytes=data, sheet_id=int(sheet["id"]), headcount_rows=[])
    count = conn.execute(
        "SELECT COUNT(*) FROM gis_nomina_attendance WHERE period_id = ?",
        (out["period_id"],),
    ).fetchone()[0]
    assert out["attendance"]["block_detected"] is True
    assert count >= 14


def test_contpaq_sheet_classified_auxiliar():
    wb = load_workbook(PEPSI_FIXTURE, data_only=True)
    ws = wb["CONTPAQi"]
    inspection = inspect_sheet(ws, sheet_name="CONTPAQi", sheet_index=0, is_hidden=False)
    wb.close()
    assert inspection["suggested_classification"] == "auxiliar"


def test_trajectory_baja_blocked_by_vacation():
    daily = [
        {"fecha_iso": "2026-07-01", "code_normalized": "A", "interpretation_status": "ok"},
        {"fecha_iso": "2026-07-02", "code_normalized": "V", "interpretation_status": "ok"},
        {"fecha_iso": "2026-07-03", "code_normalized": "F", "interpretation_status": "ok"},
        {"fecha_iso": "2026-07-04", "code_normalized": "F", "interpretation_status": "ok"},
        {"fecha_iso": "2026-07-05", "code_normalized": "F", "interpretation_status": "ok"},
        {"fecha_iso": "2026-07-06", "code_normalized": "F", "interpretation_status": "ok"},
    ]
    events = suggest_trajectory_events(daily)
    baja = [e for e in events if e["event_type"] == "posible_baja"]
    assert baja
    assert baja[0]["status"] == "review"
    assert baja[0]["fecha_sugerida"] == ""


def test_four_absences_with_descanso():
    assert detect_four_absence_event(["F", "F", "D", "F", "F"]) is True
