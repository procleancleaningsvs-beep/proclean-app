"""GIS Nóminas — exportación Excel canónica."""

from __future__ import annotations

import json
import sqlite3
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas.excel_export import generate_comparative_excel
from modules.gestion_idse_sua.nominas.schema import ensure_gis_nominas_tables
from modules.gestion_idse_sua.template_contract import COMPARATIVO_SHA256, comparativo_path
from modules.gestion_idse_sua.template_validator import sha256_file, validate_comparativo_template


@pytest.fixture
def conn(tmp_path):
    connection = sqlite3.connect(tmp_path / "gis_xlsx.db")
    connection.row_factory = sqlite3.Row
    ensure_gis_nominas_tables(connection)
    imp = connection.execute(
        "INSERT INTO gis_nomina_imports (original_filename, file_hash, uploaded_at, status) VALUES (?,?,?,?)",
        ("t.xlsx", "hash", "2026-01-01", "compared"),
    )
    import_id = imp.lastrowid
    sheet = connection.execute(
        "INSERT INTO gis_nomina_sheets (import_id, sheet_index, sheet_name, is_hidden, confirmed_classification, estimated_rows) VALUES (?,?,?,?,?,?)",
        (import_id, 0, "Semana", 0, "nomina", 1),
    )
    sheet_id = sheet.lastrowid
    period = connection.execute(
        "INSERT INTO gis_nomina_periods (sheet_id, fecha_inicio, fecha_fin, semana_num, detection_source, user_confirmed) VALUES (?,?,?,?,?,?)",
        (sheet_id, "01/06/2026", "07/06/2026", 23, "manual", 1),
    )
    period_id = period.lastrowid
    worker = connection.execute(
        """
        INSERT INTO gis_nomina_workers
        (period_id, row_number, num_empleado, nombre_original, nombre_normalizado, planta_normalizada)
        VALUES (?,?,?,?,?,?)
        """,
        (period_id, 4, "101", "Juan", "JUAN PEREZ", "FLOTADO"),
    )
    worker_id = worker.lastrowid
    comp = connection.execute(
        "INSERT INTO gis_nomina_comparatives (period_id, cliente, generated_at, generated_by, status) VALUES (?,?,?,?,?)",
        (period_id, "VITROFLEX", "2026-06-08", "tester", "completed"),
    )
    comp_id = comp.lastrowid
    connection.execute(
        """
        INSERT INTO gis_nomina_results
        (comparative_id, worker_id, headcount_only, resultado, semaforo, tipo_sugerido, decision_final)
        VALUES (?,?,?,?,?,?,?)
        """,
        (comp_id, worker_id, 0, "Coincidencia", "azul", "", "Coincidencia"),
    )
    connection.commit()
    yield connection
    connection.close()


def test_canonical_template_hash_unchanged():
    assert sha256_file(comparativo_path()) == COMPARATIVO_SHA256


def test_generated_workbook_valid(conn):
    before = sha256_file(comparativo_path())
    buf, name = generate_comparative_excel(conn, 1, username="tester")
    assert isinstance(buf, BytesIO)
    assert name.endswith(".xlsx")
    validate_comparativo_template(buf)
    assert sha256_file(comparativo_path()) == before
    buf.seek(0)
    wb = load_workbook(buf)
    assert "Detalle Comparativo" in wb.sheetnames
    ws = wb["Detalle Comparativo"]
    assert ws.cell(7, 8).value == "JUAN PEREZ"
    wb.close()


def test_export_with_headcount_match_sbc(conn):
    hc = json.dumps({"sueldo_diario": 500, "nss": "11111111111"})
    conn.execute(
        """
        INSERT INTO gis_nomina_matches
        (worker_id, headcount_key, match_method, confidence, status, hc_json, hc_nombre)
        VALUES (?,?,?,?,?,?,?)
        """,
        (1, "hc-1", "auto", 1.0, "confirmed", hc, "JUAN PEREZ"),
    )
    conn.commit()
    buf, _ = generate_comparative_excel(conn, 1)
    buf.seek(0)
    wb = load_workbook(buf)
    sbc = wb["Detalle Comparativo"].cell(7, 17).value
    assert sbc not in (None, "")
    wb.close()


def test_parallel_exports_use_independent_buffers(conn):
    buf1, name1 = generate_comparative_excel(conn, 1)
    buf2, name2 = generate_comparative_excel(conn, 1)
    assert buf1 is not buf2
    assert name1 == name2
    assert len(buf1.getvalue()) > 0
