"""Importación Excel trabajadores Vitroflex."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

DEFAULT_ACTIVIDAD = "Aux. de limpieza"
DEFAULT_TEL = "81 2183 9413"

RowDict = dict[str, str]


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and float(v).is_integer():
        return str(int(v))
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _classify_header(cell: Any) -> str | None:
    h = _norm_cell(cell).lower()
    if not h:
        return None
    if ("nombre" in h and "trabaj" in h) or h == "nombre":
        return "nombre"
    if "nss" == h or h == "nss":
        return "imss"
    if "imss" in h or (h.startswith("no") and "imss" in h.replace(".", " ")):
        return "imss"
    if "actividad" in h:
        return "actividad"
    if "emergencia" in h or ("tel" in h and "emer" in h):
        return "tel"
    return None


def _map_headers(row_values: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(row_values):
        kind = _classify_header(raw)
        if kind and kind not in mapping:
            mapping[kind] = idx
    return mapping


def detectar_estructura_columnas(mapping: dict[str, int]) -> tuple[bool, bool]:
    tiene_nombre = "nombre" in mapping
    tiene_imss = "imss" in mapping
    tiene_act = "actividad" in mapping
    tiene_tel = "tel" in mapping
    esenciales = tiene_nombre and tiene_imss
    solo_dos = esenciales and not tiene_act and not tiene_tel
    return esenciales, solo_dos


def _fila_vacia(nombre: str, imss: str, act: str, tel: str) -> bool:
    return not any((nombre.strip(), imss.strip(), act.strip(), tel.strip()))


def parse_excel_bytes(data: bytes) -> tuple[list[RowDict], bool, str | None]:
    """
    Devuelve (filas, necesita_confirm_defaults, error).
    necesita_confirm_defaults True si el archivo solo define columnas nombre + imss.
    """
    try:
        wb = load_workbook(BytesIO(data), read_only=False, data_only=True)
    except Exception as exc:
        return [], False, f"No se pudo leer el Excel: {exc}"

    try:
        ws = wb.active
        all_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    header_idx = None
    mapping: dict[str, int] = {}
    for i, row in enumerate(all_rows[:40]):
        m = _map_headers(row)
        esenciales, _ = detectar_estructura_columnas(m)
        if esenciales:
            header_idx = i
            mapping = m
            break

    if header_idx is None:
        return (
            [],
            False,
            "No se encontraron columnas reconocibles (se espera al menos NOMBRE TRABAJADOR y NO. IMSS).",
        )

    esenciales, solo_dos = detectar_estructura_columnas(mapping)
    if not esenciales:
        return [], False, "Faltan columnas esenciales: NOMBRE TRABAJADOR y NO. IMSS."

    max_idx = max(mapping.values()) if mapping else 0
    out: list[RowDict] = []
    for row in all_rows[header_idx + 1 :]:
        cells = list(row) + [None] * (max_idx + 1)
        nombre = _norm_cell(cells[mapping["nombre"]]) if "nombre" in mapping else ""
        imss = _norm_cell(cells[mapping["imss"]]) if "imss" in mapping else ""
        act = _norm_cell(cells[mapping["actividad"]]) if "actividad" in mapping else ""
        tel = _norm_cell(cells[mapping["tel"]]) if "tel" in mapping else ""
        if _fila_vacia(nombre, imss, act, tel):
            continue
        out.append(
            {
                "nombre": nombre,
                "imss": imss,
                "actividad": act,
                "tel": tel,
            }
        )

    return out, solo_dos, None
