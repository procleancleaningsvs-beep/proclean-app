from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.nomina.asistencia_miercoles import (
    is_miercoles_v3_excel,
    parse_miercoles_v3_excel,
)
from modules.nomina.config import VALID_DAILY_KEYS

# v4 column layout (1-based)
COL_NOMBRE = 1
COL_CLIENTE = 2
COL_PLANTA = 3
COL_PUESTO = 4
COL_BANCO = 5
COL_CUENTA = 6
COL_DAY_START = 7
COL_DAY_END = 13
COL_HORAS_EXTRA = 14
COL_HORAS_EXTRA_NORMALES = 15
COL_DIAS_CUBIERTOS = 16
COL_VACACIONES_LABORADAS = 17
COL_PRIMA_VACACIONAL = 18
COL_BONO = 19
COL_DEDUCCIONES = 20
COL_OBSERVACIONES = 21
BASE_MAX_COL = 21
START_DATA_ROW = 5
HEADER_ROW = 4

NUMERIC_NON_NEGATIVE_COLUMNS = {
    "horas_extra": COL_HORAS_EXTRA,
    "horas_extra_normales": COL_HORAS_EXTRA_NORMALES,
    "dias_cubiertos_normales": COL_DIAS_CUBIERTOS,
    "vacaciones_laboradas": COL_VACACIONES_LABORADAS,
}


class ValidationError(Exception):
    pass


@dataclass
class HeaderMeta:
    semana: str
    cliente: str
    coordinador: str
    dias_headers: list[str]


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").replace("\n", " ").split()).strip()


def _norm_header(value: Any) -> str:
    raw = _norm_text(value).upper()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    return raw


def _as_text_or_empty(value: Any) -> str:
    if value is None:
        return ""
    return _norm_text(value)


def _normalize_cliente_cell(value: Any) -> str:
    """Trim, mayúsculas, colapsa espacios (detección automática de cliente)."""
    return " ".join(str(value or "").replace("\u00a0", " ").upper().split()).strip()


def _find_column_by_header(ws: Worksheet, expected: str) -> int | None:
    want = _norm_header(expected)
    for col in range(1, (ws.max_column or BASE_MAX_COL) + 1):
        if _norm_header(ws.cell(row=HEADER_ROW, column=col).value) == want:
            return col
    return None


def _nss_digits(value: Any) -> str:
    s = _as_text_or_empty(value)
    return "".join(ch for ch in s if ch.isdigit())


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _extract_label_value(ws: Worksheet, label: str) -> str:
    """v4 layout writes 'LABEL: VALUE' in the same merged top-block cell.
    Backward-compatible: also looks for a separate adjacent cell (v3 style)."""
    wanted = label.upper().rstrip(":") + ":"
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=row, column=col).value
            if raw is None:
                continue
            text = _norm_text(raw)
            up = text.upper()
            if up.startswith(wanted):
                value = text[len(wanted):].strip()
                if value:
                    return value
                label_end_col = col
                for merged in ws.merged_cells.ranges:
                    if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                        label_end_col = max(label_end_col, merged.max_col)
                candidate_col = label_end_col + 1
                for merged in ws.merged_cells.ranges:
                    if merged.min_row <= row <= merged.max_row and merged.min_col > label_end_col:
                        if merged.min_col < candidate_col:
                            candidate_col = merged.min_col
                return _as_text_or_empty(ws.cell(row=row, column=candidate_col).value)
    return ""


def _read_header_meta(ws: Worksheet) -> HeaderMeta:
    semana = _extract_label_value(ws, "SEMANA")
    cliente = _extract_label_value(ws, "CLIENTE")
    coordinador = _extract_label_value(ws, "COORDINADOR")
    dias_headers = [
        _as_text_or_empty(ws.cell(row=HEADER_ROW, column=col).value)
        for col in range(COL_DAY_START, COL_DAY_END + 1)
    ]
    return HeaderMeta(
        semana=semana,
        cliente=cliente,
        coordinador=coordinador,
        dias_headers=dias_headers,
    )


def _validate_header_structure(ws: Worksheet) -> None:
    expected_by_col = {
        COL_NOMBRE: "NOMBRE DE EMPLEADO",
        COL_CLIENTE: "CLIENTE",
        COL_PLANTA: "PLANTA",
        COL_PUESTO: "PUESTO",
        COL_BANCO: "BANCO",
        COL_CUENTA: "CUENTA",
        COL_HORAS_EXTRA: "HORAS EXTRA",
        COL_HORAS_EXTRA_NORMALES: "HORAS EXTRA NORMALES",
        COL_DIAS_CUBIERTOS: "DIAS CUBIERTOS NORMALES",
        COL_VACACIONES_LABORADAS: "VACACIONES LABORADAS",
        COL_PRIMA_VACACIONAL: "PRIMA VACACIONAL",
        COL_BONO: "BONO",
        COL_DEDUCCIONES: "DEDUCCIONES",
        COL_OBSERVACIONES: "OBSERVACIONES",
    }
    missing: list[str] = []
    for col, expected in expected_by_col.items():
        got = _norm_header(ws.cell(row=HEADER_ROW, column=col).value)
        if got != _norm_header(expected):
            missing.append(f"columna {col} ({expected})")
    if missing:
        raise ValidationError(f"Faltan encabezados obligatorios en Asistencia: {', '.join(missing)}")

    daily = [
        _as_text_or_empty(ws.cell(row=HEADER_ROW, column=col).value)
        for col in range(COL_DAY_START, COL_DAY_END + 1)
    ]
    if len(daily) != 7:
        raise ValidationError("Las columnas G:M deben contener exactamente 7 días de asistencia.")
    for value in daily:
        if not re.match(r"^[LMDJSV]\d{1,2}$", value.strip().upper()):
            raise ValidationError(
                "Las columnas G:M deben tener encabezados válidos con formato abreviatura+día (ej. V1, L4)."
            )


def _row_is_empty(ws: Worksheet, row_number: int, *, max_col: int) -> bool:
    for col in range(1, max_col + 1):
        if _as_text_or_empty(ws.cell(row=row_number, column=col).value):
            return False
    return True


def parse_and_validate_asistencia_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    if is_miercoles_v3_excel(file_bytes):
        try:
            return parse_miercoles_v3_excel(file_bytes, filename)
        except ValueError as exc:
            raise ValidationError(str(exc)) from exc

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if "Asistencia" not in wb.sheetnames:
        raise ValidationError("El archivo no contiene la hoja obligatoria 'Asistencia'.")
    ws = wb["Asistencia"]
    _validate_header_structure(ws)
    header_meta = _read_header_meta(ws)
    nss_col = _find_column_by_header(ws, "NSS")
    max_col_scan = max(BASE_MAX_COL, nss_col or 0)

    blocking_errors: list[str] = []
    rows: list[dict[str, Any]] = []

    for row_number in range(START_DATA_ROW, ws.max_row + 1):
        if _row_is_empty(ws, row_number, max_col=max_col_scan):
            continue

        nombre = _as_text_or_empty(ws.cell(row=row_number, column=COL_NOMBRE).value)
        nss_raw = ""
        if nss_col is not None:
            nss_raw = _nss_digits(ws.cell(row=row_number, column=nss_col).value)
        row_data: dict[str, Any] = {
            "row_number": row_number,
            "nombre_empleado": nombre,
            "cliente": _normalize_cliente_cell(ws.cell(row=row_number, column=COL_CLIENTE).value),
            "nss": nss_raw,
            "planta": _as_text_or_empty(ws.cell(row=row_number, column=COL_PLANTA).value),
            "puesto": _as_text_or_empty(ws.cell(row=row_number, column=COL_PUESTO).value),
            "banco": _as_text_or_empty(ws.cell(row=row_number, column=COL_BANCO).value),
            "cuenta": _as_text_or_empty(ws.cell(row=row_number, column=COL_CUENTA).value),
            "dia_1_header": header_meta.dias_headers[0],
            "dia_1_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START).value),
            "dia_2_header": header_meta.dias_headers[1],
            "dia_2_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START + 1).value),
            "dia_3_header": header_meta.dias_headers[2],
            "dia_3_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START + 2).value),
            "dia_4_header": header_meta.dias_headers[3],
            "dia_4_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START + 3).value),
            "dia_5_header": header_meta.dias_headers[4],
            "dia_5_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START + 4).value),
            "dia_6_header": header_meta.dias_headers[5],
            "dia_6_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START + 5).value),
            "dia_7_header": header_meta.dias_headers[6],
            "dia_7_value": _as_text_or_empty(ws.cell(row=row_number, column=COL_DAY_START + 6).value),
            "he": _as_text_or_empty(ws.cell(row=row_number, column=COL_HORAS_EXTRA).value),
            "horas_extra_normales": _as_text_or_empty(ws.cell(row=row_number, column=COL_HORAS_EXTRA_NORMALES).value),
            "dias_cubiertos_normales": _as_text_or_empty(ws.cell(row=row_number, column=COL_DIAS_CUBIERTOS).value),
            "vacaciones_laboradas": _as_text_or_empty(ws.cell(row=row_number, column=COL_VACACIONES_LABORADAS).value),
            "prima_vacacional": _as_text_or_empty(ws.cell(row=row_number, column=COL_PRIMA_VACACIONAL).value),
            "bono": _as_text_or_empty(ws.cell(row=row_number, column=COL_BONO).value),
            "deducciones": _as_text_or_empty(ws.cell(row=row_number, column=COL_DEDUCCIONES).value),
            "observaciones": _as_text_or_empty(ws.cell(row=row_number, column=COL_OBSERVACIONES).value),
            "errors": [],
            "warnings": [],
        }

        non_name_values = [
            row_data["cliente"], row_data["planta"], row_data["puesto"],
            row_data["banco"], row_data["cuenta"], nss_raw,
            row_data["dia_1_value"], row_data["dia_2_value"], row_data["dia_3_value"],
            row_data["dia_4_value"], row_data["dia_5_value"], row_data["dia_6_value"],
            row_data["dia_7_value"],
            row_data["he"], row_data["horas_extra_normales"], row_data["dias_cubiertos_normales"],
            row_data["vacaciones_laboradas"], row_data["prima_vacacional"],
            row_data["bono"], row_data["deducciones"], row_data["observaciones"],
        ]
        if not nombre:
            if any(non_name_values):
                row_data["errors"].append("Fila con datos pero sin nombre de empleado.")
            else:
                continue

        daily_values = [
            row_data["dia_1_value"], row_data["dia_2_value"], row_data["dia_3_value"],
            row_data["dia_4_value"], row_data["dia_5_value"], row_data["dia_6_value"],
            row_data["dia_7_value"],
        ]
        for idx, raw_daily in enumerate(daily_values, start=1):
            code = _norm_header(raw_daily)
            if code and code not in VALID_DAILY_KEYS:
                msg = f"Clave diaria inválida en día {idx}: '{raw_daily}'."
                row_data["errors"].append(msg)
                blocking_errors.append(f"Fila {row_number}: {msg}")
            if code == "R":
                row_data["warnings"].append(
                    f"Día {idx} marcado como R (retardo); revisar posible deducción operativa."
                )

        for field_key, col in NUMERIC_NON_NEGATIVE_COLUMNS.items():
            raw_value = ws.cell(row=row_number, column=col).value
            numeric = _to_float(raw_value)
            if numeric is not None and numeric < 0:
                msg = f"Valor negativo no permitido en '{field_key}'."
                row_data["errors"].append(msg)
                blocking_errors.append(f"Fila {row_number}: {msg}")

        prima = _norm_header(row_data["prima_vacacional"])
        if not prima:
            row_data["prima_vacacional"] = "N/A"
            prima = "N/A"
        if prima not in {"N/A", "SOLICITA"}:
            row_data["errors"].append("PRIMA VACACIONAL solo acepta N/A, SOLICITA o vacío.")
        elif prima == "SOLICITA":
            row_data["warnings"].append("PRIMA VACACIONAL en SOLICITA: revisar vacaciones en fase 2.")
            row_data["prima_vacacional"] = "SOLICITA"
        else:
            row_data["prima_vacacional"] = "N/A"

        any_ot = any(_norm_header(v) == "OT" for v in daily_values)
        if any_ot and not row_data["observaciones"]:
            row_data["warnings"].append("Tiene OT en asistencia y no incluye observaciones.")

        for key in ("bono", "deducciones"):
            value = row_data[key]
            if not value:
                continue
            numeric = _to_float(value)
            if numeric is None:
                row_data["warnings"].append(f"{key.upper()} contiene texto: revisar manualmente.")
            row_data["warnings"].append(f"{key.upper()} capturado: requiere revisión de Nómina.")

        rows.append(row_data)

    error_count = sum(len(row.get("errors") or []) for row in rows)
    warning_count = sum(len(row.get("warnings") or []) for row in rows)
    return {
        "filename": filename,
        "semana": header_meta.semana,
        "cliente": header_meta.cliente,
        "coordinador": header_meta.coordinador,
        "dias_headers": header_meta.dias_headers,
        "rows": rows,
        "total_rows": len(rows),
        "error_count": error_count,
        "warning_count": warning_count,
        "blocking_errors": blocking_errors,
    }
