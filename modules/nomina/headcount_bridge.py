"""Headcount remoto → filas reales (streaming; sin leer hoja Excel completa a memoria)."""
from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from modules.comparativo.headcount_service import (
    _is_empty as hc_is_empty,
    _normalize_header as hc_normalize_header,
    _normalize_name as hc_normalize_name,
    _normalize_spaces as hc_normalize_spaces,
)
from modules.finiquitos.excel_mirror_fecha_ingreso import descargar_excel_desde_onedrive
from modules.nomina.vacaciones_util import sanitize_display_value
from services.perf_logging import perf_headcount_log

_HEADER_MARKERS = frozenset({"STATUS OPERACIÓN", "STATUS OPERACION"})
_COLS = (
    ("nombre_completo", "NOMBRE COMPLETO"),
    ("nombre", "NOMBRE"),
    ("apellido_paterno", "APELLIDO PATERNO"),
    ("apellido_materno", "APELLIDO MATERNO"),
    ("cliente", "CLIENTE"),
    ("ubicacion", "UBICACION"),
    ("patron", "PATRON"),
    ("fecha_ingreso", "FECHA DE INGRESO"),
    ("sueldo_diario", "SUELDO DIARIO"),
    ("sueldo_semanal", "SUELDO SEMANAL"),
    ("puesto", "PUESTO"),
    ("nss", "NSS"),
    ("status_operacion", "STATUS OPERACIÓN"),
    ("status_imss", "STATUS IMSS"),
    ("rfc_homoclave", "RFC HOMOCLAVE"),
    ("cp_fiscal", "CP FISCAL"),
    ("curp", "CURP"),
    ("genero", "GENERO"),
    ("fecha_nacimiento", "FECHA DE NACIMIENTO"),
    ("lugar_nacimiento", "LUGAR DE NACIMIENTO"),
)
_KNOWN_STATUS = frozenset(
    {"ALTA", "BAJA", "ACTIVO", "INACTIVO", "SUSPENDIDO", "VACACIONES", "LICENCIA"}
)
_CONSECUTIVE_EMPTY_STOP = 100
_MAX_SAVED_ROWS = 20_000
_ROW_EXPLOSION_THRESHOLD = 1_000_000
_EXCEL_MAX_ROWS_HINT = 1_048_576


@dataclass
class HeadcountParseResult:
    rows: list[dict[str, Any]]
    source_rows_scanned: int
    saved_rows: int
    skipped_empty_rows: int
    guardrail_triggered: bool
    guardrail_reason: str | None = None


def _format_fecha(value: Any) -> str:
    if hc_is_empty(value):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _col_idx(header_map: dict[str, int], name: str) -> int | None:
    if name in header_map:
        return header_map[name]
    alt = name.replace("Ó", "O")
    if alt in header_map:
        return header_map[alt]
    return None


def _cell(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _row_all_key_empty(row: tuple[Any, ...], key_indices: list[int | None]) -> bool:
    indices = [i for i in key_indices if i is not None]
    if not indices:
        return True
    return all(hc_is_empty(_cell(row, i)) for i in indices)


def _is_real_employee(row: tuple[Any, ...], col_idx: dict[str, int | None]) -> bool:
    nombre = hc_normalize_name(_cell(row, col_idx.get("nombre_completo")) or "")
    nss = hc_normalize_spaces(str(_cell(row, col_idx.get("nss")) or "").strip())
    status_op = hc_normalize_spaces(
        sanitize_display_value(_cell(row, col_idx.get("status_operacion")) or "")
    ).upper()
    status_imss = hc_normalize_spaces(
        sanitize_display_value(_cell(row, col_idx.get("status_imss")) or "")
    ).upper()
    cliente = hc_normalize_spaces(str(_cell(row, col_idx.get("cliente")) or "").strip())
    patron = hc_normalize_spaces(str(_cell(row, col_idx.get("patron")) or "").strip())
    if nombre:
        return True
    if nss:
        return True
    if status_op in _KNOWN_STATUS or status_imss in _KNOWN_STATUS:
        return True
    if cliente or patron:
        return True
    return False


def _row_to_record(row: tuple[Any, ...], col_idx: dict[str, int | None]) -> dict[str, Any]:
    nombre = hc_normalize_name(_cell(row, col_idx.get("nombre_completo")) or "")
    status_op = hc_normalize_spaces(
        sanitize_display_value(_cell(row, col_idx.get("status_operacion")) or "")
    ).upper()
    status_imss = hc_normalize_spaces(
        sanitize_display_value(_cell(row, col_idx.get("status_imss")) or "")
    ).upper()
    sueldo_raw = _cell(row, col_idx.get("sueldo_diario"))
    return {
        "nombre_completo": nombre,
        "nombre": hc_normalize_spaces(str(_cell(row, col_idx.get("nombre")) or "").strip()),
        "apellido_paterno": hc_normalize_spaces(str(_cell(row, col_idx.get("apellido_paterno")) or "").strip()),
        "apellido_materno": hc_normalize_spaces(str(_cell(row, col_idx.get("apellido_materno")) or "").strip()),
        "cliente": hc_normalize_spaces(str(_cell(row, col_idx.get("cliente")) or "").strip()),
        "ubicacion": hc_normalize_spaces(str(_cell(row, col_idx.get("ubicacion")) or "").strip()),
        "patron": hc_normalize_spaces(str(_cell(row, col_idx.get("patron")) or "").strip()),
        "fecha_ingreso": _format_fecha(_cell(row, col_idx.get("fecha_ingreso"))),
        "sueldo_diario": None if hc_is_empty(sueldo_raw) else sueldo_raw,
        "sueldo_semanal": None
        if hc_is_empty(_cell(row, col_idx.get("sueldo_semanal")))
        else _cell(row, col_idx.get("sueldo_semanal")),
        "puesto": hc_normalize_spaces(str(_cell(row, col_idx.get("puesto")) or "").strip()),
        "nss": hc_normalize_spaces(str(_cell(row, col_idx.get("nss")) or "").strip()),
        "status_operacion": status_op or "DESCONOCIDO",
        "status_imss": status_imss or "DESCONOCIDO",
        "rfc_homoclave": hc_normalize_spaces(str(_cell(row, col_idx.get("rfc_homoclave")) or "").strip()),
        "cp_fiscal": hc_normalize_spaces(str(_cell(row, col_idx.get("cp_fiscal")) or "").strip()),
        "curp": hc_normalize_spaces(str(_cell(row, col_idx.get("curp")) or "").strip()).upper(),
        "genero": hc_normalize_spaces(str(_cell(row, col_idx.get("genero")) or "").strip()),
        "fecha_nacimiento": _format_fecha(_cell(row, col_idx.get("fecha_nacimiento"))),
        "lugar_nacimiento": hc_normalize_spaces(str(_cell(row, col_idx.get("lugar_nacimiento")) or "").strip()),
    }


def parse_headcount_excel_bytes(raw: bytes) -> HeadcountParseResult:
    """Parsea Headcount fila a fila; ignora rango formateado vacío de Excel."""
    from openpyxl import load_workbook

    perf_headcount_log("parse_started")
    wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        header_map: dict[str, int] | None = None
        col_idx: dict[str, int | None] = {}
        registros: list[dict[str, Any]] = []
        source_rows_scanned = 0
        skipped_empty_rows = 0
        consecutive_empty = 0
        key_indices: list[int | None] = []

        for row_vals in ws.iter_rows(values_only=True):
            row = tuple(row_vals or ())
            if header_map is None:
                normalized = [hc_normalize_header(v) for v in row]
                if _HEADER_MARKERS & set(normalized):
                    header_map = {normalized[j]: j for j in range(len(normalized))}
                    col_idx = {field: _col_idx(header_map, header) for field, header in _COLS}
                    key_indices = [
                        col_idx.get("nombre_completo"),
                        col_idx.get("nss"),
                        col_idx.get("cliente"),
                        col_idx.get("patron"),
                        col_idx.get("status_operacion"),
                    ]
                continue

            source_rows_scanned += 1
            if source_rows_scanned >= _ROW_EXPLOSION_THRESHOLD:
                perf_headcount_log(
                    "refresh_aborted",
                    reason="row_explosion",
                    source_rows=source_rows_scanned,
                )
                return HeadcountParseResult(
                    rows=[],
                    source_rows_scanned=source_rows_scanned,
                    saved_rows=0,
                    skipped_empty_rows=skipped_empty_rows,
                    guardrail_triggered=True,
                    guardrail_reason=(
                        "Headcount inválido: se detectaron demasiadas filas. "
                        "Posible rango formateado completo."
                    ),
                )

            if _row_all_key_empty(row, key_indices):
                skipped_empty_rows += 1
                consecutive_empty += 1
                if consecutive_empty >= _CONSECUTIVE_EMPTY_STOP:
                    break
                continue

            consecutive_empty = 0
            if not _is_real_employee(row, col_idx):
                skipped_empty_rows += 1
                continue

            registros.append(_row_to_record(row, col_idx))
            if len(registros) > _MAX_SAVED_ROWS:
                perf_headcount_log(
                    "refresh_aborted",
                    reason="too_many_real_rows",
                    saved_rows=len(registros),
                )
                return HeadcountParseResult(
                    rows=[],
                    source_rows_scanned=source_rows_scanned,
                    saved_rows=0,
                    skipped_empty_rows=skipped_empty_rows,
                    guardrail_triggered=True,
                    guardrail_reason=(
                        f"Headcount inválido: más de {_MAX_SAVED_ROWS} empleados detectados."
                    ),
                )

        if header_map is None:
            raise ValueError("No se encontró encabezado STATUS OPERACIÓN en Headcount.")

        if source_rows_scanned >= _EXCEL_MAX_ROWS_HINT - 1000 and len(registros) < 100:
            perf_headcount_log(
                "refresh_aborted",
                reason="row_explosion",
                source_rows=source_rows_scanned,
            )
            return HeadcountParseResult(
                rows=[],
                source_rows_scanned=source_rows_scanned,
                saved_rows=0,
                skipped_empty_rows=skipped_empty_rows,
                guardrail_triggered=True,
                guardrail_reason=(
                    "Headcount inválido: hoja Excel con rango masivo formateado y casi sin datos."
                ),
            )

        perf_headcount_log(
            "parse_finished",
            source_rows=source_rows_scanned,
            saved_rows=len(registros),
            skipped_empty=skipped_empty_rows,
        )
        return HeadcountParseResult(
            rows=registros,
            source_rows_scanned=source_rows_scanned,
            saved_rows=len(registros),
            skipped_empty_rows=skipped_empty_rows,
            guardrail_triggered=False,
        )
    finally:
        wb.close()


def fetch_and_parse_headcount() -> HeadcountParseResult:
    """Descarga OneDrive y parsea Headcount (solo refresh manual/cron/CLI)."""
    url = (os.environ.get("HEADCOUNT_ONEDRIVE_URL") or "").strip()
    if not url:
        raise ValueError("No está configurada la variable HEADCOUNT_ONEDRIVE_URL.")
    perf_headcount_log("refresh_download_started")
    t0 = __import__("time").perf_counter()
    raw = descargar_excel_desde_onedrive(url)
    elapsed_ms = int((__import__("time").perf_counter() - t0) * 1000)
    perf_headcount_log("refresh_download_finished", duration_ms=elapsed_ms)
    return parse_headcount_excel_bytes(raw)


def obtener_headcount_completo() -> list[dict[str, Any]]:
    """Compat: parsea Headcount remoto y devuelve solo filas reales (refresh/CLI only)."""
    from modules.headcount.remote_guard import assert_remote_headcount_allowed

    assert_remote_headcount_allowed("obtener_headcount_completo")
    result = fetch_and_parse_headcount()
    if result.guardrail_triggered:
        raise ValueError(result.guardrail_reason or "Headcount inválido.")
    return result.rows
