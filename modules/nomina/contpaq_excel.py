"""Parser de Excel exportado desde CONTPAQ Nómina.

Microfase 4.0. Aporta principalmente Código (número de empleado), Nombre, NSS
y fechas/estatus/zona. NO sobreescribe NSS de Headcount sin advertencia.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").replace("\n", " ").split()).strip()


def _norm_header(value: Any) -> str:
    raw = _norm_text(value).upper()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    raw = raw.replace(".", " ").replace(",", " ")
    return " ".join(raw.split())


def _norm_name_for_match(value: Any) -> str:
    raw = _norm_text(value).upper()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    raw = "".join(ch if (ch.isalnum() or ch == " ") else " " for ch in raw)
    return " ".join(raw.split())


def _norm_nss(value: Any) -> str:
    raw = _norm_text(value)
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    return digits


def _to_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        d = value.date() if isinstance(value, datetime) else value
        if d.year < 1900:
            return None
        return d.isoformat()
    s = _norm_text(value)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            if d.year < 1900:
                return None
            return d.isoformat()
        except ValueError:
            continue
    return None


HEADER_ALIASES = {
    "codigo": ["CODIGO", "CLAVE", "NUM EMPLEADO", "NUMERO EMPLEADO"],
    "nombre_completo": ["NOMBRE COMPLETO"],
    "apellido_paterno": ["APELLIDO PATERNO"],
    "apellido_materno": ["APELLIDO MATERNO"],
    "nombre": ["NOMBRE"],
    "fecha_alta": ["FECHA DE ALTA", "FECHA ALTA"],
    "fecha_baja": ["FECHA DE BAJA", "FECHA BAJA"],
    "fecha_reingreso": ["FECHA DE REINGRESO", "FECHA REINGRESO"],
    "estatus": ["ESTATUS", "ESTATUS EMPLEADO"],
    "departamento": ["DEPARTAMENTO"],
    "zona_salario": ["ZONA DE SALARIO", "ZONA SALARIO"],
    "puesto": ["PUESTO"],
    "salario_diario": ["SALARIO DIARIO"],
    "registro_patronal": [
        "REGISTRO PATRONAL DEL IMSS",
        "REGISTRO PATRONAL",
        "REGISTRO PATRONAL IMSS",
    ],
    "nss": [
        "NUM SEGURIDAD SOCIAL",
        "NUMERO DE SEGURIDAD SOCIAL",
        "NSS",
        "NSS IMSS",
    ],
    "rfc": ["RFC"],
    "curp": ["CURP"],
    "tipo_periodo": ["TIPO DE PERIODO", "PERIODO"],
}


@dataclass
class HeaderMap:
    row: int
    cols: dict[str, int] = field(default_factory=dict)


def _find_header_row(ws: Worksheet, max_scan_rows: int = 8) -> HeaderMap | None:
    best: HeaderMap | None = None
    best_hits = 0
    max_row = min(ws.max_row or max_scan_rows, max_scan_rows)
    for r in range(1, max_row + 1):
        cols: dict[str, int] = {}
        for c in range(1, (ws.max_column or 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            norm = _norm_header(v)
            for key, aliases in HEADER_ALIASES.items():
                if key in cols:
                    continue
                for alias in aliases:
                    if norm == _norm_header(alias):
                        cols[key] = c
                        break
        hits = sum(1 for k in {"codigo", "nombre", "apellido_paterno", "nombre_completo"} if k in cols)
        if hits > best_hits:
            best = HeaderMap(row=r, cols=cols)
            best_hits = hits
    if best and "codigo" in best.cols and (
        "nombre_completo" in best.cols
        or ("nombre" in best.cols and "apellido_paterno" in best.cols)
    ):
        return best
    return None


def _row_is_empty(ws: Worksheet, row: int, hm: HeaderMap) -> bool:
    for col in hm.cols.values():
        if _norm_text(ws.cell(row=row, column=col).value):
            return False
    return True


def parse_contpaq(file_bytes: bytes, *, filename: str) -> dict[str, Any]:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo abrir el Excel CONTPAQ: {exc}") from exc

    chosen_sheet: str | None = None
    chosen_hm: HeaderMap | None = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        hm = _find_header_row(ws)
        if hm is not None:
            chosen_sheet = sn
            chosen_hm = hm
            break

    if chosen_sheet is None or chosen_hm is None:
        raise ValueError(
            "No se detectó hoja CONTPAQ válida (se esperan columnas Código + Nombre completo / Apellido paterno)."
        )

    ws = wb[chosen_sheet]

    rows_out: list[dict[str, Any]] = []
    warnings: list[str] = []
    errors: list[str] = []
    seen_codigos: set[str] = set()
    seen_nss: dict[str, str] = {}

    def _get(cols: dict[str, int], r: int, key: str) -> Any:
        if key not in cols:
            return None
        return ws.cell(row=r, column=cols[key]).value

    for r in range(chosen_hm.row + 1, (ws.max_row or 0) + 1):
        if _row_is_empty(ws, r, chosen_hm):
            continue
        codigo = _norm_text(_get(chosen_hm.cols, r, "codigo"))
        if not codigo:
            continue
        nombre_completo = _norm_text(_get(chosen_hm.cols, r, "nombre_completo"))
        if not nombre_completo:
            ap = _norm_text(_get(chosen_hm.cols, r, "apellido_paterno"))
            am = _norm_text(_get(chosen_hm.cols, r, "apellido_materno"))
            nm = _norm_text(_get(chosen_hm.cols, r, "nombre"))
            nombre_completo = " ".join(p for p in [nm, ap, am] if p).strip()
        if not nombre_completo:
            warnings.append(f"Fila CONTPAQ {r}: sin nombre legible.")
            continue
        nombre_norm = _norm_name_for_match(nombre_completo)

        nss = _norm_nss(_get(chosen_hm.cols, r, "nss"))
        if nss and len(nss) != 11:
            warnings.append(f"Fila CONTPAQ {r}: NSS '{nss}' no tiene 11 dígitos.")
            nss = ""

        row_warnings: list[str] = []
        if codigo in seen_codigos:
            row_warnings.append(f"Código CONTPAQ duplicado: {codigo}.")
        else:
            seen_codigos.add(codigo)
        if nss:
            if nss in seen_nss and seen_nss[nss] != nombre_norm:
                row_warnings.append(
                    f"NSS {nss} ya usado por otro nombre en este archivo: '{seen_nss[nss]}'."
                )
            else:
                seen_nss[nss] = nombre_norm

        estatus = _norm_text(_get(chosen_hm.cols, r, "estatus")).upper()
        zona_salario_raw = _norm_text(_get(chosen_hm.cols, r, "zona_salario"))
        # CONTPAQ zona suele ser texto corto (ej. 'B', 'F'). No inferimos frontera
        # solo del código; lo dejamos como raw para revisión.

        rows_out.append({
            "row_number": r,
            "codigo_contpaq": codigo,
            "nombre": nombre_completo,
            "nombre_normalizado": nombre_norm,
            "nss": nss or None,
            "fecha_alta": _to_date(_get(chosen_hm.cols, r, "fecha_alta")),
            "fecha_baja": _to_date(_get(chosen_hm.cols, r, "fecha_baja")),
            "fecha_reingreso": _to_date(_get(chosen_hm.cols, r, "fecha_reingreso")),
            "estatus": estatus or None,
            "departamento": _norm_text(_get(chosen_hm.cols, r, "departamento")) or None,
            "zona_salario_raw": zona_salario_raw or None,
            "puesto": _norm_text(_get(chosen_hm.cols, r, "puesto")) or None,
            "salario_diario": _to_float_safe(_get(chosen_hm.cols, r, "salario_diario")),
            "registro_patronal": _norm_text(_get(chosen_hm.cols, r, "registro_patronal")) or None,
            "rfc": _norm_text(_get(chosen_hm.cols, r, "rfc")) or None,
            "curp": _norm_text(_get(chosen_hm.cols, r, "curp")) or None,
            "warnings": row_warnings,
        })

    if not rows_out:
        errors.append("No se detectaron empleados en el archivo CONTPAQ.")

    return {
        "tipo_importacion": "CONTPAQ",
        "filename": filename,
        "sheet": chosen_sheet,
        "rows": rows_out,
        "warnings": warnings,
        "errors": errors,
        "total_rows": len(rows_out),
    }


def _to_float_safe(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    s = _norm_text(value).replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None
