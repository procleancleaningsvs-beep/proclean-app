from __future__ import annotations

import hashlib
import io
import json
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from modules.nomina.banorte.models import ImportDecision, RecordStatus, SourceKind, ValidationStatus
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.schema import ensure_banorte_tables
from modules.nomina.banorte.validators import (
    extract_identifier_cell,
    is_banorte_employee_substituted_comment,
    normalize_header,
    normalize_name,
    safe_upload_filename,
)

TZ = ZoneInfo("America/Monterrey")


@dataclass
class ImportBatchResult:
    batch_id: int | None
    mutated: bool
    message: str
    count_exitosos: int = 0
    count_manuales: int = 0
    count_fallidos_estatus: int = 0
    count_fallidos_hoja_sin_estatus: int = 0
    count_excluidos_hoja_fallidos_total: int = 0
    count_duplicados_reemplazados: int = 0
    count_conflictos: int = 0
    count_omitidos: int = 0
    rows_processed: int = 0
    file_sha256: str = ""
    details: list[dict[str, Any]] = field(default_factory=list)


def _now_iso() -> str:
    return datetime.now(TZ).isoformat(timespec="seconds")


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _find_header_map(ws, required: set[str]) -> tuple[int, dict[str, int]]:
    for r in range(1, min(30, (ws.max_row or 1) + 1)):
        mapping: dict[str, int] = {}
        for c in range(1, (ws.max_column or 1) + 1):
            key = normalize_header(ws.cell(r, c).value)
            if key:
                mapping[key] = c
        keys = set(mapping)
        # Tolerate ESTATUS vs ESTATUS RESPUESTA
        if "ESTATUS RESPUESTA" in keys and "ESTATUS" not in keys:
            mapping["ESTATUS"] = mapping["ESTATUS RESPUESTA"]
            keys.add("ESTATUS")
        if required.issubset(keys) or (
            "NOMBRE DEL EMPLEADO" in keys
            and "NUMERO DE EMPLEADO" in keys
            and "NUMERO DE CUENTA" in keys
            and ("ESTATUS" in keys or "ESTATUS RESPUESTA" in keys)
        ):
            return r, mapping
    raise ValueError("headers_not_found")


def _cell(ws, row: int, col: int | None) -> Any:
    if col is None:
        return None
    return ws.cell(row, col).value


def _existing_sha_batch(conn: sqlite3.Connection, sha: str) -> int | None:
    row = conn.execute(
        "SELECT id FROM nomina_banorte_import_batches WHERE file_sha256=? ORDER BY id DESC LIMIT 1",
        (sha,),
    ).fetchone()
    return int(row[0]) if row else None


def _active_by_account(conn: sqlite3.Connection, account: str) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM nomina_banorte_beneficiaries WHERE account_number=? AND record_status='ACTIVO'",
        (account,),
    ).fetchone()


def _active_by_emp(conn: sqlite3.Connection, emp: str) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM nomina_banorte_beneficiaries WHERE employee_number_effective=? AND record_status='ACTIVO'",
        (emp,),
    ).fetchone()


def _insert_beneficiary(conn: sqlite3.Connection, payload: dict[str, Any]) -> int:
    cur = conn.execute(
        """
        INSERT INTO nomina_banorte_beneficiaries (
            nombre_original, nombre_normalizado, curp,
            employee_number_requested, employee_number_effective, account_number,
            source_kind, validation_status, record_status,
            banorte_employee_substituted, banorte_comment,
            source_filename, source_sheet, source_row, report_date,
            imported_at, imported_by, replaces_id, created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            payload["nombre_original"],
            payload["nombre_normalizado"],
            payload.get("curp"),
            payload.get("employee_number_requested"),
            payload["employee_number_effective"],
            payload["account_number"],
            payload["source_kind"],
            payload["validation_status"],
            payload["record_status"],
            payload.get("banorte_employee_substituted", 0),
            payload.get("banorte_comment"),
            payload.get("source_filename"),
            payload.get("source_sheet"),
            payload.get("source_row"),
            payload.get("report_date"),
            payload["imported_at"],
            payload["imported_by"],
            payload.get("replaces_id"),
            payload["imported_at"],
            payload["imported_at"],
        ),
    )
    return int(cur.lastrowid)


def _inactivate(conn: sqlite3.Connection, beneficiary_id: int, now: str) -> None:
    conn.execute(
        """
        UPDATE nomina_banorte_beneficiaries
        SET record_status='INACTIVO_REEMPLAZADO', updated_at=?
        WHERE id=?
        """,
        (now, beneficiary_id),
    )


def _same_material(existing: sqlite3.Row, payload: dict[str, Any]) -> bool:
    return (
        str(existing["nombre_normalizado"]) == payload["nombre_normalizado"]
        and str(existing["employee_number_effective"]) == payload["employee_number_effective"]
        and str(existing["account_number"]) == payload["account_number"]
        and (existing["curp"] or None) == (payload.get("curp") or None)
        and str(existing["validation_status"]) == payload["validation_status"]
        and int(existing["banorte_employee_substituted"] or 0) == int(payload.get("banorte_employee_substituted", 0))
    )


def _apply_identity(
    conn: sqlite3.Connection,
    payload: dict[str, Any],
    *,
    now: str,
) -> tuple[str, int | None, str]:
    """Return (decision, beneficiary_id, reason)."""
    account = payload["account_number"]
    emp = payload["employee_number_effective"]
    curp = payload.get("curp")
    nombre_norm = payload["nombre_normalizado"]

    by_acct = _active_by_account(conn, account)
    by_emp = _active_by_emp(conn, emp)

    if by_acct is not None:
        same_person = False
        if curp and by_acct["curp"] and str(by_acct["curp"]).upper() == curp:
            same_person = True
        elif str(by_acct["nombre_normalizado"]) == nombre_norm and (
            str(by_acct["employee_number_effective"]) == emp
            or (curp and by_acct["curp"] and str(by_acct["curp"]).upper() == curp)
        ):
            same_person = True
        elif str(by_acct["employee_number_effective"]) == emp and str(by_acct["nombre_normalizado"]) == nombre_norm:
            same_person = True

        if same_person:
            if _same_material(by_acct, payload):
                return ImportDecision.REIMPORT_NO_CHANGE.value, int(by_acct["id"]), "identical_active"
            # version update
            _inactivate(conn, int(by_acct["id"]), now)
            payload["replaces_id"] = int(by_acct["id"])
            new_id = _insert_beneficiary(conn, payload)
            return ImportDecision.REPLACED_DUPLICATE.value, new_id, "version_update_same_account"
        # different person same account
        payload["record_status"] = RecordStatus.CONFLICTO_CRITICO.value
        new_id = _insert_beneficiary(conn, payload)
        return ImportDecision.CONFLICT.value, new_id, "account_person_conflict"

    if by_emp is not None:
        if curp and by_emp["curp"] and str(by_emp["curp"]).upper() == curp:
            if _same_material(by_emp, payload):
                return ImportDecision.REIMPORT_NO_CHANGE.value, int(by_emp["id"]), "identical_active_emp"
            _inactivate(conn, int(by_emp["id"]), now)
            payload["replaces_id"] = int(by_emp["id"])
            # may need new account — unique on emp will free after inactivate
            new_id = _insert_beneficiary(conn, payload)
            return ImportDecision.REPLACED_DUPLICATE.value, new_id, "version_update_same_emp_curp"
        if str(by_emp["nombre_normalizado"]) == nombre_norm and str(by_emp["account_number"]) == account:
            if _same_material(by_emp, payload):
                return ImportDecision.REIMPORT_NO_CHANGE.value, int(by_emp["id"]), "identical"
            _inactivate(conn, int(by_emp["id"]), now)
            payload["replaces_id"] = int(by_emp["id"])
            new_id = _insert_beneficiary(conn, payload)
            return ImportDecision.REPLACED_DUPLICATE.value, new_id, "version_update_emp"
        # emp taken by different identity without enough evidence → conflict on new row
        payload["record_status"] = RecordStatus.CONFLICTO_CRITICO.value
        new_id = _insert_beneficiary(conn, payload)
        return ImportDecision.CONFLICT.value, new_id, "employee_conflict"

    # Name alone never merges; insert fresh active
    new_id = _insert_beneficiary(conn, payload)
    decision = (
        ImportDecision.IMPORTED_EXITOSO.value
        if payload["validation_status"] == ValidationStatus.IMPORTADO_EXITOSO.value
        else ImportDecision.IMPORTED_MANUAL.value
    )
    return decision, new_id, "created"


def _insert_batch(conn: sqlite3.Connection, meta: dict[str, Any]) -> int:
    cur = conn.execute(
        """
        INSERT INTO nomina_banorte_import_batches (
            file_name, file_sha256, file_size, detected_type, imported_by, imported_at,
            rows_processed, count_exitosos, count_manuales, count_fallidos_estatus,
            count_fallidos_hoja_sin_estatus, count_excluidos_hoja_fallidos_total,
            count_duplicados_reemplazados, count_conflictos, count_omitidos,
            summary_json, reimport_confirmed
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            meta["file_name"],
            meta["file_sha256"],
            meta["file_size"],
            meta["detected_type"],
            meta["imported_by"],
            meta["imported_at"],
            meta["rows_processed"],
            meta["count_exitosos"],
            meta["count_manuales"],
            meta["count_fallidos_estatus"],
            meta["count_fallidos_hoja_sin_estatus"],
            meta["count_excluidos_hoja_fallidos_total"],
            meta["count_duplicados_reemplazados"],
            meta["count_conflictos"],
            meta["count_omitidos"],
            meta["summary_json"],
            meta["reimport_confirmed"],
        ),
    )
    return int(cur.lastrowid)


def _insert_import_row(conn: sqlite3.Connection, batch_id: int, row: dict[str, Any]) -> None:
    conn.execute(
        """
        INSERT INTO nomina_banorte_import_rows (
            batch_id, sheet_name, row_number, decision, reason, nombre, curp,
            employee_number_requested, employee_number_effective, account_number,
            estatus_raw, comentarios_raw, beneficiary_id, payload_json
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            batch_id,
            row["sheet_name"],
            row["row_number"],
            row["decision"],
            row["reason"],
            row.get("nombre"),
            row.get("curp"),
            row.get("employee_number_requested"),
            row.get("employee_number_effective"),
            row.get("account_number"),
            row.get("estatus_raw"),
            row.get("comentarios_raw"),
            row.get("beneficiary_id"),
            row.get("payload_json"),
        ),
    )


def import_nomina_banorte_xlsx(
    db_path: str,
    file_bytes: bytes,
    filename: str,
    user: str,
    *,
    reimport_confirmed: bool = False,
) -> ImportBatchResult:
    safe_name = safe_upload_filename(filename)
    if not safe_name.lower().endswith(".xlsx"):
        raise ValueError("invalid_extension")
    sha = _sha256(file_bytes)
    now = _now_iso()

    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        prior = _existing_sha_batch(conn, sha)
        if prior is not None and not reimport_confirmed:
            return ImportBatchResult(
                batch_id=None,
                mutated=False,
                message="duplicate_sha_confirmation_required",
                file_sha256=sha,
            )

        wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
        if "ALTAS" not in wb.sheetnames:
            raise ValueError("missing_altas_sheet")

        result = ImportBatchResult(batch_id=None, mutated=True, message="ok", file_sha256=sha)
        audit_rows: list[dict[str, Any]] = []

        # FALLIDOS sheet: exclude entirely
        if "FALLIDOS" in wb.sheetnames:
            ws_f = wb["FALLIDOS"]
            try:
                header_row_f, fmap = _find_header_map(
                    ws_f, {"NUMERO DE EMPLEADO", "NOMBRE DEL EMPLEADO", "ESTATUS"}
                )
            except ValueError:
                header_row_f, fmap = 3, {}
            est_c = fmap.get("ESTATUS")
            for r in range(header_row_f + 1, (ws_f.max_row or header_row_f) + 1):
                vals = [ws_f.cell(r, c).value for c in range(1, (ws_f.max_column or 1) + 1)]
                if not any(v is not None and str(v).strip() for v in vals):
                    continue
                est = normalize_header(_cell(ws_f, r, est_c)) if est_c else ""
                if est == "FALLIDO":
                    result.count_fallidos_estatus += 1
                    decision = ImportDecision.EXCLUDED_FALLIDO.value
                    reason = "fallidos_sheet_fallido"
                else:
                    result.count_fallidos_hoja_sin_estatus += 1
                    decision = ImportDecision.EXCLUDED_FALLIDOS_SHEET_EMPTY_STATUS.value
                    reason = "fallidos_sheet_empty_or_other_status"
                result.count_excluidos_hoja_fallidos_total += 1
                result.rows_processed += 1
                audit_rows.append(
                    {
                        "sheet_name": "FALLIDOS",
                        "row_number": r,
                        "decision": decision,
                        "reason": reason,
                        "estatus_raw": str(_cell(ws_f, r, est_c) or ""),
                    }
                )

        ws = wb["ALTAS"]
        header_row, colmap = _find_header_map(
            ws, {"NUMERO DE EMPLEADO", "NOMBRE DEL EMPLEADO", "NUMERO DE CUENTA", "ESTATUS"}
        )
        # Process top-to-bottom; later duplicates win via versioning.
        pending_payloads: list[dict[str, Any]] = []
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            nombre = _cell(ws, r, colmap.get("NOMBRE DEL EMPLEADO"))
            emp_cell = ws.cell(r, colmap["NUMERO DE EMPLEADO"])
            acct_cell = ws.cell(r, colmap["NUMERO DE CUENTA"])
            estatus = _cell(ws, r, colmap.get("ESTATUS"))
            curp_raw = _cell(ws, r, colmap.get("CURP"))
            comment = _cell(ws, r, colmap.get("COMENTARIOS"))
            fecha = _cell(ws, r, colmap.get("FECHA DE ALTA SOLICITUD"))

            if not any(
                [
                    nombre and str(nombre).strip(),
                    emp_cell.value is not None and str(emp_cell.value).strip() != "",
                    acct_cell.value is not None and str(acct_cell.value).strip() != "",
                    estatus and str(estatus).strip(),
                ]
            ):
                continue

            # repeated header
            if normalize_header(nombre) == "NOMBRE DEL EMPLEADO":
                result.count_omitidos += 1
                result.rows_processed += 1
                audit_rows.append(
                    {
                        "sheet_name": "ALTAS",
                        "row_number": r,
                        "decision": ImportDecision.EXCLUDED_HEADER.value,
                        "reason": "repeated_header",
                    }
                )
                continue

            est_n = normalize_header(estatus)
            emp, emp_err = extract_identifier_cell(emp_cell.value, number_format=emp_cell.number_format)
            acct, acct_err = extract_identifier_cell(acct_cell.value, number_format=acct_cell.number_format)
            curp = None
            if curp_raw is not None and str(curp_raw).strip():
                curp = str(curp_raw).strip().upper()

            if emp_err == "PRECISION_RISK" or acct_err == "PRECISION_RISK":
                result.count_omitidos += 1
                result.rows_processed += 1
                audit_rows.append(
                    {
                        "sheet_name": "ALTAS",
                        "row_number": r,
                        "decision": ImportDecision.PRECISION_REVIEW.value,
                        "reason": "excel_precision_risk",
                        "nombre": str(nombre or ""),
                    }
                )
                continue

            if est_n == "FALLIDO":
                result.count_fallidos_estatus += 1
                result.rows_processed += 1
                audit_rows.append(
                    {
                        "sheet_name": "ALTAS",
                        "row_number": r,
                        "decision": ImportDecision.EXCLUDED_FALLIDO.value,
                        "reason": "altas_fallido",
                        "estatus_raw": str(estatus or ""),
                    }
                )
                continue

            has_n = bool(nombre and str(nombre).strip())
            has_e = bool(emp)
            has_a = bool(acct)
            if est_n == "EXITOSO":
                if not (has_n and has_e and has_a):
                    result.count_omitidos += 1
                    result.rows_processed += 1
                    audit_rows.append(
                        {
                            "sheet_name": "ALTAS",
                            "row_number": r,
                            "decision": ImportDecision.EXCLUDED_INCOMPLETE.value,
                            "reason": "exitoso_incomplete",
                        }
                    )
                    continue
                validation = ValidationStatus.IMPORTADO_EXITOSO.value
            elif est_n == "":
                if has_n and has_e and has_a:
                    validation = ValidationStatus.MANUAL_PENDIENTE_VALIDACION.value
                else:
                    result.count_omitidos += 1
                    result.rows_processed += 1
                    audit_rows.append(
                        {
                            "sheet_name": "ALTAS",
                            "row_number": r,
                            "decision": ImportDecision.EXCLUDED_INCOMPLETE.value,
                            "reason": "manual_incomplete",
                            "nombre": str(nombre or ""),
                        }
                    )
                    continue
            else:
                result.count_omitidos += 1
                result.rows_processed += 1
                audit_rows.append(
                    {
                        "sheet_name": "ALTAS",
                        "row_number": r,
                        "decision": ImportDecision.EXCLUDED_EMPTY.value,
                        "reason": f"unknown_status:{est_n}",
                    }
                )
                continue

            requested = emp
            effective = emp
            substituted = 0
            if is_banorte_employee_substituted_comment(comment):
                effective = acct
                substituted = 1

            payload = {
                "nombre_original": str(nombre).strip(),
                "nombre_normalizado": normalize_name(nombre),
                "curp": curp,
                "employee_number_requested": requested,
                "employee_number_effective": effective,
                "account_number": acct,
                "source_kind": SourceKind.ALTAS_NOMINA_BANORTE.value,
                "validation_status": validation,
                "record_status": RecordStatus.ACTIVO.value,
                "banorte_employee_substituted": substituted,
                "banorte_comment": str(comment) if comment is not None else None,
                "source_filename": safe_name,
                "source_sheet": "ALTAS",
                "source_row": r,
                "report_date": str(fecha)[:10] if fecha else None,
                "imported_at": now,
                "imported_by": user,
            }
            pending_payloads.append(payload)

        # Apply later rows last so lower rows win
        for payload in pending_payloads:
            decision, ben_id, reason = _apply_identity(conn, payload, now=now)
            result.rows_processed += 1
            if decision == ImportDecision.REIMPORT_NO_CHANGE.value:
                result.count_omitidos += 0
            elif decision == ImportDecision.CONFLICT.value:
                result.count_conflictos += 1
            elif decision == ImportDecision.REPLACED_DUPLICATE.value:
                result.count_duplicados_reemplazados += 1
                if payload["validation_status"] == ValidationStatus.IMPORTADO_EXITOSO.value:
                    result.count_exitosos += 1
                else:
                    result.count_manuales += 1
            elif payload["validation_status"] == ValidationStatus.IMPORTADO_EXITOSO.value:
                result.count_exitosos += 1
            else:
                result.count_manuales += 1
            audit_rows.append(
                {
                    "sheet_name": "ALTAS",
                    "row_number": payload["source_row"],
                    "decision": decision,
                    "reason": reason,
                    "nombre": payload["nombre_original"],
                    "curp": payload.get("curp"),
                    "employee_number_requested": payload.get("employee_number_requested"),
                    "employee_number_effective": payload.get("employee_number_effective"),
                    "account_number": payload.get("account_number"),
                    "estatus_raw": payload["validation_status"],
                    "comentarios_raw": payload.get("banorte_comment"),
                    "beneficiary_id": ben_id,
                }
            )

        summary = {
            "file": safe_name,
            "sha256": sha,
            "reimport_confirmed": bool(reimport_confirmed and prior is not None),
            "counts": {
                "exitosos": result.count_exitosos,
                "manuales": result.count_manuales,
                "fallidos_estatus": result.count_fallidos_estatus,
                "fallidos_hoja_sin_estatus": result.count_fallidos_hoja_sin_estatus,
                "excluidos_hoja_fallidos_total": result.count_excluidos_hoja_fallidos_total,
                "duplicados": result.count_duplicados_reemplazados,
                "conflictos": result.count_conflictos,
                "omitidos": result.count_omitidos,
            },
        }
        batch_id = _insert_batch(
            conn,
            {
                "file_name": safe_name,
                "file_sha256": sha,
                "file_size": len(file_bytes),
                "detected_type": "ALTAS_NOMINA_BANORTE",
                "imported_by": user,
                "imported_at": now,
                "rows_processed": result.rows_processed,
                "count_exitosos": result.count_exitosos,
                "count_manuales": result.count_manuales,
                "count_fallidos_estatus": result.count_fallidos_estatus,
                "count_fallidos_hoja_sin_estatus": result.count_fallidos_hoja_sin_estatus,
                "count_excluidos_hoja_fallidos_total": result.count_excluidos_hoja_fallidos_total,
                "count_duplicados_reemplazados": result.count_duplicados_reemplazados,
                "count_conflictos": result.count_conflictos,
                "count_omitidos": result.count_omitidos,
                "summary_json": json.dumps(summary, ensure_ascii=False),
                "reimport_confirmed": 1 if reimport_confirmed and prior is not None else 0,
            },
        )
        for row in audit_rows:
            row.setdefault("payload_json", None)
            _insert_import_row(conn, batch_id, row)
        conn.commit()
        result.batch_id = batch_id
        result.details = audit_rows
        return result
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def import_reporte_detallado_xlsx(
    db_path: str,
    file_bytes: bytes,
    filename: str,
    user: str,
    *,
    reimport_confirmed: bool = False,
) -> ImportBatchResult:
    safe_name = safe_upload_filename(filename)
    if not safe_name.lower().endswith(".xlsx"):
        raise ValueError("invalid_extension")
    sha = _sha256(file_bytes)
    now = _now_iso()
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        prior = _existing_sha_batch(conn, sha)
        if prior is not None and not reimport_confirmed:
            return ImportBatchResult(
                batch_id=None,
                mutated=False,
                message="duplicate_sha_confirmation_required",
                file_sha256=sha,
            )

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row, colmap = _find_header_map(
            ws, {"NUMERO DE EMPLEADO", "NOMBRE DEL EMPLEADO", "NUMERO DE CUENTA"}
        )
        result = ImportBatchResult(batch_id=None, mutated=True, message="ok", file_sha256=sha)
        audit_rows: list[dict[str, Any]] = []

        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            if not any(ws.cell(r, c).value for c in range(1, (ws.max_column or 1) + 1)):
                continue
            estatus = _cell(ws, r, colmap.get("ESTATUS"))
            est_n = normalize_header(estatus)
            result.rows_processed += 1
            if est_n != "EXITOSO":
                if est_n == "FALLIDO":
                    result.count_fallidos_estatus += 1
                else:
                    result.count_omitidos += 1
                audit_rows.append(
                    {
                        "sheet_name": ws.title,
                        "row_number": r,
                        "decision": ImportDecision.EXCLUDED_FALLIDO.value
                        if est_n == "FALLIDO"
                        else ImportDecision.EXCLUDED_EMPTY.value,
                        "reason": f"reporte_status:{est_n or 'EMPTY'}",
                        "estatus_raw": str(estatus or ""),
                    }
                )
                continue

            nombre = _cell(ws, r, colmap.get("NOMBRE DEL EMPLEADO"))
            emp_cell = ws.cell(r, colmap["NUMERO DE EMPLEADO"])
            acct_cell = ws.cell(r, colmap["NUMERO DE CUENTA"])
            emp, emp_err = extract_identifier_cell(emp_cell.value, number_format=emp_cell.number_format)
            acct, acct_err = extract_identifier_cell(acct_cell.value, number_format=acct_cell.number_format)
            if emp_err or acct_err or not emp or not acct or not nombre:
                result.count_omitidos += 1
                audit_rows.append(
                    {
                        "sheet_name": ws.title,
                        "row_number": r,
                        "decision": ImportDecision.EXCLUDED_INCOMPLETE.value,
                        "reason": emp_err or acct_err or "incomplete",
                    }
                )
                continue
            curp_raw = _cell(ws, r, colmap.get("CURP"))
            curp = str(curp_raw).strip().upper() if curp_raw else None
            comment = _cell(ws, r, colmap.get("COMENTARIOS"))
            fecha = _cell(ws, r, colmap.get("FECHA DE ALTA SOLICITUD"))
            requested = emp
            effective = emp
            substituted = 0
            if is_banorte_employee_substituted_comment(comment):
                effective = acct
                substituted = 1

            # Try link manual → validated
            manual = conn.execute(
                """
                SELECT * FROM nomina_banorte_beneficiaries
                WHERE record_status='ACTIVO'
                  AND validation_status='MANUAL_PENDIENTE_VALIDACION'
                  AND (
                    employee_number_effective=? OR employee_number_requested=?
                    OR (? IS NOT NULL AND curp=?)
                    OR account_number=?
                  )
                ORDER BY id DESC LIMIT 1
                """,
                (requested, requested, curp, curp, acct),
            ).fetchone()

            payload = {
                "nombre_original": str(nombre).strip(),
                "nombre_normalizado": normalize_name(nombre),
                "curp": curp,
                "employee_number_requested": requested,
                "employee_number_effective": effective,
                "account_number": acct,
                "source_kind": SourceKind.REPORTE_DETALLADO.value,
                "validation_status": ValidationStatus.IMPORTADO_EXITOSO.value,
                "record_status": RecordStatus.ACTIVO.value,
                "banorte_employee_substituted": substituted,
                "banorte_comment": str(comment) if comment is not None else None,
                "source_filename": safe_name,
                "source_sheet": ws.title,
                "source_row": r,
                "report_date": str(fecha)[:10] if fecha else None,
                "imported_at": now,
                "imported_by": user,
            }

            if manual is not None:
                _inactivate(conn, int(manual["id"]), now)
                payload["replaces_id"] = int(manual["id"])
                # clear active emp/account uniqueness by inactivation already done
                new_id = _insert_beneficiary(conn, payload)
                decision = ImportDecision.LINKED_VALIDATED.value
                reason = "manual_to_validated"
                result.count_exitosos += 1
            else:
                decision, new_id, reason = _apply_identity(conn, payload, now=now)
                if decision == ImportDecision.CONFLICT.value:
                    result.count_conflictos += 1
                elif decision == ImportDecision.REPLACED_DUPLICATE.value:
                    result.count_duplicados_reemplazados += 1
                    result.count_exitosos += 1
                elif decision != ImportDecision.REIMPORT_NO_CHANGE.value:
                    result.count_exitosos += 1

            audit_rows.append(
                {
                    "sheet_name": ws.title,
                    "row_number": r,
                    "decision": decision,
                    "reason": reason,
                    "nombre": payload["nombre_original"],
                    "curp": curp,
                    "employee_number_requested": requested,
                    "employee_number_effective": effective,
                    "account_number": acct,
                    "estatus_raw": str(estatus or ""),
                    "comentarios_raw": str(comment) if comment is not None else None,
                    "beneficiary_id": new_id,
                }
            )

        summary = {"file": safe_name, "sha256": sha, "type": "REPORTE_DETALLADO"}
        batch_id = _insert_batch(
            conn,
            {
                "file_name": safe_name,
                "file_sha256": sha,
                "file_size": len(file_bytes),
                "detected_type": "REPORTE_DETALLADO",
                "imported_by": user,
                "imported_at": now,
                "rows_processed": result.rows_processed,
                "count_exitosos": result.count_exitosos,
                "count_manuales": result.count_manuales,
                "count_fallidos_estatus": result.count_fallidos_estatus,
                "count_fallidos_hoja_sin_estatus": result.count_fallidos_hoja_sin_estatus,
                "count_excluidos_hoja_fallidos_total": result.count_excluidos_hoja_fallidos_total,
                "count_duplicados_reemplazados": result.count_duplicados_reemplazados,
                "count_conflictos": result.count_conflictos,
                "count_omitidos": result.count_omitidos,
                "summary_json": json.dumps(summary, ensure_ascii=False),
                "reimport_confirmed": 1 if reimport_confirmed and prior is not None else 0,
            },
        )
        for row in audit_rows:
            row.setdefault("payload_json", None)
            _insert_import_row(conn, batch_id, row)
        conn.commit()
        result.batch_id = batch_id
        result.details = audit_rows
        return result
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
