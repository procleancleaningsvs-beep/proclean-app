"""Import payroll Excel (.xlsx/.xlsm) into Banorte drafts (EXCEL_NOMINA)."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from modules.nomina.banorte.calculo_adapter import origin_hash_for_manual_capture
from modules.nomina.banorte.draft_repository import save_draft_rows
from modules.nomina.banorte.excel_token import issue_excel_token, verify_excel_token
from modules.nomina.banorte.money import parse_money, to_cents
from modules.nomina.banorte.prepare_service import prepare_draft_rows
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.schema import ensure_banorte_tables

MAX_FILE_BYTES = 25 * 1024 * 1024
MAX_DATA_ROWS = 10_000
ALLOWED_EXT = {".xlsx", ".xlsm"}


class ExcelNominaError(ValueError):
    def __init__(self, code: str):
        super().__init__(code)
        self.code = code


@dataclass
class ParsedExcelRow:
    excel_row: int
    nombre: str
    banco: str
    neto_raw: Any
    neto_cents: int
    included: bool
    reason: str | None = None


@dataclass
class ExcelPreview:
    sheet: str
    header_row: int
    total_rows_scanned: int
    banorte_count: int
    excluded_hidden_count: int
    excluded_other_bank_count: int
    blocked_formula_count: int
    total_banorte_cents: int
    sample: list[dict[str, Any]]
    warnings: list[str]


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _norm_header(value: Any) -> str:
    return str(value or "").strip().upper()


def _norm_banco(value: Any) -> str:
    return str(value or "").strip().casefold()


def _find_header_row(ws, scan_limit: int = 30) -> tuple[int, dict[str, int]]:
    need = {"NOMBRE DE EMPLEADO", "BANCO", "NETO A PAGAR"}
    for row_idx in range(1, scan_limit + 1):
        cols: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            label = _norm_header(ws.cell(row=row_idx, column=col_idx).value)
            if label in need:
                cols[label] = col_idx
        if need.issubset(cols.keys()):
            return row_idx, cols
    raise ExcelNominaError("header_not_found")


def _cell_net_value(ws, ws_formula, row_idx: int, col_idx: int) -> tuple[Any, str | None]:
    val = ws.cell(row=row_idx, column=col_idx).value
    if val is not None:
        return val, None
    fcell = ws_formula.cell(row=row_idx, column=col_idx)
    if getattr(fcell, "data_type", None) == "f" or str(fcell.value or "").startswith("="):
        return None, "formula_without_cache"
    return None, "empty_net"


def inspect_excel(file_bytes: bytes, filename: str, *, secret_key: str, user: str) -> dict[str, Any]:
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ExcelNominaError("file_too_large")
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_EXT:
        raise ExcelNominaError("unsupported_extension")
    digest = sha256_bytes(file_bytes)
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
    finally:
        wb.close()
    token = issue_excel_token(secret_key, user=user, sha256=digest, size=len(file_bytes))
    return {
        "sheets": sheets,
        "sha256": digest,
        "sha256_prefix": digest[:12],
        "size": len(file_bytes),
        "token": token,
    }


def _parse_sheet_rows(file_bytes: bytes, sheet: str) -> tuple[list[ParsedExcelRow], ExcelPreview]:
    wb = load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
    wb_formula = load_workbook(BytesIO(file_bytes), read_only=False, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            raise ExcelNominaError("sheet_not_found")
        ws = wb[sheet]
        ws_f = wb_formula[sheet]
        header_row, cols = _find_header_row(ws)
        name_col = cols["NOMBRE DE EMPLEADO"]
        bank_col = cols["BANCO"]
        net_col = cols["NETO A PAGAR"]
        parsed: list[ParsedExcelRow] = []
        hidden_count = 0
        other_bank = 0
        formula_blocked = 0
        banorte_count = 0
        total_cents = 0
        data_rows = 0
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            if data_rows >= MAX_DATA_ROWS:
                break
            data_rows += 1
            if ws.row_dimensions[row_idx].hidden:
                hidden_count += 1
                continue
            nombre = str(ws.cell(row=row_idx, column=name_col).value or "").strip()
            banco = str(ws.cell(row=row_idx, column=bank_col).value or "").strip()
            if not nombre and not banco:
                continue
            net_raw, net_reason = _cell_net_value(ws, ws_f, row_idx, net_col)
            banco_norm = _norm_banco(banco)
            if banco_norm != "banorte":
                other_bank += 1
                parsed.append(
                    ParsedExcelRow(row_idx, nombre, banco, net_raw, 0, False, "other_bank")
                )
                continue
            if net_reason == "formula_without_cache":
                formula_blocked += 1
                parsed.append(
                    ParsedExcelRow(row_idx, nombre, banco, net_raw, 0, False, net_reason)
                )
                continue
            money = parse_money(str(net_raw) if net_raw is not None else "")
            if not money.ok or money.amount is None:
                parsed.append(
                    ParsedExcelRow(row_idx, nombre, banco, net_raw, 0, False, "invalid_amount")
                )
                continue
            cents = to_cents(money.amount)
            if cents <= 0:
                parsed.append(
                    ParsedExcelRow(row_idx, nombre, banco, net_raw, 0, False, "non_positive")
                )
                continue
            banorte_count += 1
            total_cents += cents
            parsed.append(ParsedExcelRow(row_idx, nombre, banco, net_raw, cents, True, None))
        warnings: list[str] = []
        if hidden_count:
            warnings.append(f"hidden_rows_excluded:{hidden_count}")
        if other_bank:
            warnings.append(f"other_bank_excluded:{other_bank}")
        if formula_blocked:
            warnings.append(f"formula_without_cache:{formula_blocked}")
        preview = ExcelPreview(
            sheet=sheet,
            header_row=header_row,
            total_rows_scanned=data_rows,
            banorte_count=banorte_count,
            excluded_hidden_count=hidden_count,
            excluded_other_bank_count=other_bank,
            blocked_formula_count=formula_blocked,
            total_banorte_cents=total_cents,
            sample=[
                {
                    "row": r.excel_row,
                    "nombre": r.nombre,
                    "banco": r.banco,
                    "neto_cents": r.neto_cents,
                    "included": r.included,
                    "reason": r.reason,
                }
                for r in parsed[:10]
            ],
            warnings=warnings,
        )
        return parsed, preview
    finally:
        wb.close()
        wb_formula.close()


def preview_excel(
    file_bytes: bytes,
    *,
    filename: str,
    sheet: str,
    token: str,
    secret_key: str,
    user: str,
) -> ExcelPreview:
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ExcelNominaError("file_too_large")
    digest = sha256_bytes(file_bytes)
    verify_excel_token(secret_key, token, user=user, sha256=digest, size=len(file_bytes))
    _, preview = _parse_sheet_rows(file_bytes, sheet)
    return preview


def prepare_excel_draft(
    db_path: str,
    user: str,
    file_bytes: bytes,
    *,
    filename: str,
    sheet: str,
    token: str,
    secret_key: str,
) -> dict[str, Any]:
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ExcelNominaError("file_too_large")
    digest = sha256_bytes(file_bytes)
    verify_excel_token(secret_key, token, user=user, sha256=digest, size=len(file_bytes))
    parsed, preview = _parse_sheet_rows(file_bytes, sheet)
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        from datetime import datetime
        from zoneinfo import ZoneInfo

        now = datetime.now(ZoneInfo("America/Monterrey")).isoformat(timespec="seconds")
        origin_hash = origin_hash_for_manual_capture(digest, sheet)
        cur = conn.execute(
            """
            INSERT INTO nomina_banorte_export_drafts (
                created_by, updated_by, created_at, updated_at, origin_kind, calculo_id,
                origin_updated_at, origin_hash, status, revision,
                source_filename, source_sha256, source_sheet, source_file_size
            ) VALUES (?,?,?,?, 'EXCEL_NOMINA', NULL, ?, ?, 'OPEN', 1, ?, ?, ?, ?)
            """,
            (user, user, now, now, now, origin_hash, filename, digest, sheet, len(file_bytes)),
        )
        draft_id = int(cur.lastrowid)
        conn.commit()
    finally:
        conn.close()
    draft_shell = {"id": draft_id, "revision": 1}
    rows = []
    pos = 0
    for item in parsed:
        if not item.included:
            continue
        pos += 1
        rows.append(
            {
                "position": pos,
                "nombre_recibido": item.nombre,
                "banco_snapshot": item.banco,
                "amount_original_cents": item.neto_cents,
                "amount_final_cents": item.neto_cents,
                "included": 1,
                "match_kind": "NONE",
                "row_state": "NEEDS_REVIEW",
                "warnings": [],
                "user_decision": {
                    "excel_row": item.excel_row,
                    "source_sheet": sheet,
                    "source_sha256": digest,
                    "source_filename": filename,
                },
            }
        )
    prepared = prepare_draft_rows(db_path, rows, origin_kind="EXCEL_NOMINA")
    draft = save_draft_rows(db_path, draft_id, user, int(draft_shell["revision"]), prepared)
    draft["excel_preview"] = preview.__dict__
    return draft
