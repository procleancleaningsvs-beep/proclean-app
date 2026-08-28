"""Persistent Banorte beneficiary staging batches (MANUAL / REPORTE_DETALLADO)."""

from __future__ import annotations

import hashlib
import io
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from modules.nomina.banorte.employee_number_service import collect_occupied_employee_numbers
from modules.nomina.banorte.repository import connect
from modules.nomina.banorte.schema import ensure_banorte_tables
from modules.nomina.banorte.validators import (
    digits_only,
    extract_identifier_cell,
    is_banorte_employee_substituted_comment,
    normalize_header,
    normalize_name,
    safe_upload_filename,
)

TZ = ZoneInfo("America/Monterrey")


class BatchStaleError(Exception):
    def __init__(self, batch_id: int, current_revision: int):
        super().__init__("batch_stale")
        self.batch_id = batch_id
        self.current_revision = current_revision
        self.code = "batch_stale"


class BatchAccessError(Exception):
    def __init__(self, code: str = "batch_not_owned"):
        super().__init__(code)
        self.code = code


class ManualBatchValidationError(Exception):
    def __init__(self, errors: list[dict[str, Any]]):
        super().__init__("beneficiary_rows_invalid")
        self.code = "beneficiary_rows_invalid"
        self.errors = errors


def _now() -> str:
    return datetime.now(TZ).isoformat(timespec="seconds")


def _digits(value: Any) -> str:
    return digits_only(value)


def get_batch(db_path: str, batch_id: int) -> dict[str, Any] | None:
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        b = conn.execute(
            "SELECT * FROM nomina_banorte_beneficiary_batches WHERE id=?",
            (int(batch_id),),
        ).fetchone()
        if b is None:
            return None
        rows = conn.execute(
            """
            SELECT * FROM nomina_banorte_beneficiary_batch_rows
            WHERE batch_id=? ORDER BY position ASC
            """,
            (int(batch_id),),
        ).fetchall()
        payload = dict(b)
        payload["rows"] = [dict(r) for r in rows]
        return payload
    finally:
        conn.close()


def find_open_manual_batch(db_path: str, user: str) -> dict[str, Any] | None:
    """Return only the authenticated actor's transitional MANUAL batch."""
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        row = conn.execute(
            """
            SELECT id FROM nomina_banorte_beneficiary_batches
            WHERE created_by=? AND origin_kind='MANUAL' AND status='OPEN'
            ORDER BY id DESC LIMIT 1
            """,
            (user,),
        ).fetchone()
    finally:
        conn.close()
    return get_batch(db_path, int(row["id"])) if row is not None else None


def create_batch(
    db_path: str,
    user: str,
    *,
    origin_kind: str,
    source_filename: str | None = None,
    source_sha256: str | None = None,
    prior_batch_id: int | None = None,
) -> dict[str, Any]:
    if origin_kind not in {"MANUAL", "REPORTE_DETALLADO"}:
        raise ValueError("invalid_origin_kind")
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        existing = conn.execute(
            """
            SELECT id FROM nomina_banorte_beneficiary_batches
            WHERE created_by=? AND origin_kind=? AND status='OPEN'
            """,
            (user, origin_kind),
        ).fetchone()
        if existing is not None:
            return get_batch(db_path, int(existing["id"]))  # type: ignore[return-value]
        now = _now()
        cur = conn.execute(
            """
            INSERT INTO nomina_banorte_beneficiary_batches (
                origin_kind, status, revision, source_filename, source_sha256,
                prior_batch_id, created_by, created_at, updated_at
            ) VALUES (?, 'OPEN', 1, ?, ?, ?, ?, ?, ?)
            """,
            (origin_kind, source_filename, source_sha256, prior_batch_id, user, now, now),
        )
        conn.commit()
        return get_batch(db_path, int(cur.lastrowid))  # type: ignore[return-value]
    finally:
        conn.close()


def _bump_batch(conn, batch_id: int, expected_revision: int, user: str) -> None:
    now = _now()
    cur = conn.execute(
        """
        UPDATE nomina_banorte_beneficiary_batches
        SET revision = revision + 1, updated_at=?, created_by=created_by
        WHERE id=? AND revision=? AND status='OPEN'
        """,
        (now, int(batch_id), int(expected_revision)),
    )
    # touch updated_by-less: keep created_by; only revision/updated_at
    if cur.rowcount != 1:
        row = conn.execute(
            "SELECT revision FROM nomina_banorte_beneficiary_batches WHERE id=?",
            (int(batch_id),),
        ).fetchone()
        raise BatchStaleError(int(batch_id), int(row["revision"]) if row else expected_revision)
    _ = user  # reserved for future audit column


def _batch_requested_employee(row: dict[str, Any]) -> str:
    return _digits(row.get("employee_number"))


def _batch_effective_employee(row: dict[str, Any]) -> str:
    acct = _digits(row.get("cuenta"))
    requested = _batch_requested_employee(row)
    use_acct = int(row.get("use_account_as_employee_number") or 0) == 1
    if use_acct:
        return acct
    return requested


def _insert_batch_rows_conn(
    conn,
    batch_id: int,
    *,
    start_position: int,
    rows: list[dict[str, Any]],
    now: str,
) -> None:
    if not rows:
        return
    values: list[tuple[Any, ...]] = []
    for i, row in enumerate(rows):
        name = (row.get("nombre") or "").strip()
        acct = _digits(row.get("cuenta"))
        use_acct = bool(row.get("use_account_as_employee_number"))
        if use_acct:
            requested = _digits(row.get("employee_number") or "")
            emp = requested or acct
        else:
            emp = _digits(row.get("employee_number") or "")
        state = "OK" if name and acct and emp else "DRAFT"
        values.append(
            (
                int(batch_id),
                start_position + i,
                name or None,
                acct or None,
                emp or None,
                1 if use_acct else 0,
                row.get("comment"),
                state,
                row.get("source_row"),
                now,
                now,
            )
        )
    conn.executemany(
        """
        INSERT INTO nomina_banorte_beneficiary_batch_rows (
            batch_id, position, nombre, cuenta, employee_number,
            use_account_as_employee_number, comment, row_state,
            source_row, created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """,
        values,
    )


def add_batch_rows_bulk(
    db_path: str,
    batch_id: int,
    user: str,
    expected_revision: int,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    if not rows:
        return get_batch(db_path, batch_id)  # type: ignore[return-value]
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        _bump_batch(conn, batch_id, expected_revision, user)
        occupied = collect_occupied_employee_numbers(conn)
        accepted_effective: set[str] = set()
        for row in rows:
            effective = _batch_effective_employee(row)
            if len(effective) == 10 and effective != "0000000000":
                if effective in occupied or effective in accepted_effective:
                    raise ValueError("duplicate_employee_number")
                accepted_effective.add(effective)
        pos = int(
            conn.execute(
                "SELECT COALESCE(MAX(position), 0) + 1 AS p FROM nomina_banorte_beneficiary_batch_rows WHERE batch_id=?",
                (int(batch_id),),
            ).fetchone()["p"]
        )
        now = _now()
        _insert_batch_rows_conn(conn, batch_id, start_position=pos, rows=rows, now=now)
        conn.commit()
    except BatchStaleError:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_batch(db_path, batch_id)  # type: ignore[return-value]


def add_batch_row(
    db_path: str,
    batch_id: int,
    user: str,
    expected_revision: int,
    *,
    nombre: str,
    cuenta: str,
    employee_number: str | None = None,
    use_account_as_employee_number: bool = False,
    comment: str | None = None,
    source_row: int | None = None,
) -> dict[str, Any]:
    name = (nombre or "").strip()
    acct = _digits(cuenta)
    use_acct = bool(use_account_as_employee_number)
    if use_acct:
        if len(acct) != 10:
            raise ValueError("account_must_be_exactly_10_digits")
        requested = _digits(employee_number or "")
        if requested and (len(requested) != 10 or requested == "0000000000"):
            raise ValueError("employee_number_must_be_exactly_10_digits")
        emp = requested or acct
    else:
        emp = _digits(employee_number or "")
        if emp and (len(emp) != 10 or emp == "0000000000"):
            raise ValueError("employee_number_must_be_exactly_10_digits")
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        _bump_batch(conn, batch_id, expected_revision, user)
        effective = acct if use_acct else emp
        if effective and effective in collect_occupied_employee_numbers(conn):
            raise ValueError("duplicate_employee_number")
        pos = int(
            conn.execute(
                "SELECT COALESCE(MAX(position), 0) + 1 AS p FROM nomina_banorte_beneficiary_batch_rows WHERE batch_id=?",
                (int(batch_id),),
            ).fetchone()["p"]
        )
        now = _now()
        state = "OK" if name and acct and emp else "DRAFT"
        conn.execute(
            """
            INSERT INTO nomina_banorte_beneficiary_batch_rows (
                batch_id, position, nombre, cuenta, employee_number,
                use_account_as_employee_number, comment, row_state,
                source_row, created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                int(batch_id),
                pos,
                name or None,
                acct or None,
                emp or None,
                1 if use_acct else 0,
                comment,
                state,
                source_row,
                now,
                now,
            ),
        )
        conn.commit()
    except BatchStaleError:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_batch(db_path, batch_id)  # type: ignore[return-value]


def delete_batch_row(
    db_path: str,
    batch_id: int,
    row_id: int,
    user: str,
    expected_revision: int,
) -> dict[str, Any]:
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        _bump_batch(conn, batch_id, expected_revision, user)
        cur = conn.execute(
            "DELETE FROM nomina_banorte_beneficiary_batch_rows WHERE id=? AND batch_id=?",
            (int(row_id), int(batch_id)),
        )
        if cur.rowcount != 1:
            raise ValueError("batch_row_not_found")
        conn.commit()
    except BatchStaleError:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_batch(db_path, batch_id)  # type: ignore[return-value]


def abandon_batch(
    db_path: str,
    batch_id: int,
    user: str,
    expected_revision: int,
) -> dict[str, Any]:
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        now = _now()
        cur = conn.execute(
            """
            UPDATE nomina_banorte_beneficiary_batches
            SET status='ABANDONED', revision = revision + 1, updated_at=?
            WHERE id=? AND revision=? AND status='OPEN'
            """,
            (now, int(batch_id), int(expected_revision)),
        )
        if cur.rowcount != 1:
            row = conn.execute(
                "SELECT revision FROM nomina_banorte_beneficiary_batches WHERE id=?",
                (int(batch_id),),
            ).fetchone()
            raise BatchStaleError(int(batch_id), int(row["revision"]) if row else expected_revision)
        conn.commit()
        _ = user
    except BatchStaleError:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_batch(db_path, batch_id)  # type: ignore[return-value]


def _canonicalize_batch_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    canonical: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        account = _digits(row.get("cuenta") if "cuenta" in row else row.get("account"))
        use_account = bool(row.get("use_account_as_employee_number"))
        requested = _digits(row.get("employee_number"))
        if use_account and not requested:
            requested = account
        canonical.append(
            {
                "row_id": int(row["id"]) if row.get("id") is not None else None,
                "row_index": index,
                "client_row_key": str(row.get("client_row_key") or f"row-{index + 1}"),
                "nombre": str(row.get("nombre") or "").strip(),
                "cuenta": account,
                "employee_number": requested,
                "employee_number_effective": account if use_account else requested,
                "use_account_as_employee_number": 1 if use_account else 0,
                "comment": row.get("comment"),
                "source_row": row.get("source_row"),
            }
        )
    return canonical


def _validate_batch_rows_conn(
    conn,
    rows: list[dict[str, Any]],
    *,
    exclude_batch_id: int | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    canonical = _canonicalize_batch_rows(rows)
    occupied = collect_occupied_employee_numbers(conn, exclude_batch_id=exclude_batch_id)
    active_accounts = {
        str(r["account_number"])
        for r in conn.execute(
            "SELECT account_number FROM nomina_banorte_beneficiaries WHERE record_status='ACTIVO'"
        )
    }
    seen_employees: set[str] = set()
    seen_accounts: set[str] = set()
    errors: list[dict[str, Any]] = []
    for row in canonical:
        name = row["nombre"]
        account = row["cuenta"]
        requested = row["employee_number"]
        effective = row["employee_number_effective"]
        use_account = bool(row["use_account_as_employee_number"])
        code = ""
        message = ""
        field = ""
        if not name:
            field, code, message = "nombre", "nombre_required", "Nombre obligatorio."
        elif not account:
            field, code, message = "account", "account_required", "Cuenta obligatoria."
        elif len(account) > 18:
            field, code, message = "account", "account_too_long", "La cuenta no puede exceder 18 dígitos."
        elif use_account and len(account) != 10:
            field, code, message = (
                "account",
                "account_must_be_exactly_10_digits",
                "La cuenta debe tener exactamente 10 dígitos.",
            )
        elif len(requested) != 10 or requested == "0000000000":
            field, code, message = (
                "employee_number",
                "employee_number_must_be_exactly_10_digits",
                "El número debe tener exactamente 10 dígitos.",
            )
        elif len(effective) != 10 or effective == "0000000000":
            field, code, message = (
                "employee_number",
                "employee_number_must_be_exactly_10_digits",
                "El número efectivo debe tener exactamente 10 dígitos.",
            )
        elif effective in seen_employees or effective in occupied:
            field, code, message = (
                "employee_number",
                "duplicate_employee_number",
                "Número de empleado no disponible.",
            )
        elif account in seen_accounts:
            field, code, message = (
                "account",
                "duplicate_account_in_batch",
                "Cuenta duplicada en el lote.",
            )
        elif account in active_accounts:
            field, code, message = (
                "account",
                "duplicate_active_account",
                "Ya existe una cuenta activa igual.",
            )
        if code:
            errors.append(
                {
                    "row_id": row["row_id"],
                    "row_index": row["row_index"],
                    "client_row_key": row["client_row_key"],
                    "field": field,
                    "error_code": code,
                    "message": message,
                    "error_message": message,
                }
            )
            continue
        seen_employees.add(effective)
        seen_accounts.add(account)
        occupied.add(effective)
    return canonical, errors


def _persist_confirmed_beneficiaries_conn(
    conn,
    batch: dict[str, Any],
    rows: list[dict[str, Any]],
    *,
    user: str,
    now: str,
) -> None:
    batch_id = int(batch["id"])
    origin = str(batch.get("origin_kind") or "MANUAL")
    source_kind = "ALTA_MANUAL" if origin == "MANUAL" else "REPORTE_DETALLADO"
    validation = (
        "MANUAL_PENDIENTE_VALIDACION" if origin == "MANUAL" else "IMPORTADO_EXITOSO"
    )
    for row in rows:
        requested = row["employee_number"]
        effective = row["employee_number_effective"]
        use_account = bool(row["use_account_as_employee_number"])
        substituted = 1 if use_account and requested != effective else 0
        manual_effective = 1 if use_account and substituted == 0 else 0
        comment = row.get("comment") or f"batch:{batch_id}"
        conn.execute(
            """
            INSERT INTO nomina_banorte_beneficiaries (
                nombre_original, nombre_normalizado, curp,
                employee_number_requested, employee_number_effective, account_number,
                source_kind, validation_status, record_status,
                banorte_employee_substituted, manual_effective_from_account,
                banorte_comment, imported_at, imported_by, created_at, updated_at
            ) VALUES (?,?,NULL,?,?,?,?,?,'ACTIVO',?,?,?,?,?,?,?)
            """,
            (
                row["nombre"],
                normalize_name(row["nombre"]),
                requested,
                effective,
                row["cuenta"],
                source_kind,
                validation,
                substituted,
                manual_effective,
                comment,
                now,
                user,
                now,
                now,
            ),
        )
    if origin == "REPORTE_DETALLADO" and batch.get("source_sha256"):
        conn.execute(
            """
            INSERT INTO nomina_banorte_import_batches (
                file_name, file_sha256, file_size, detected_type, imported_by, imported_at,
                rows_processed, count_exitosos, count_manuales, count_fallidos_estatus,
                count_fallidos_hoja_sin_estatus, count_excluidos_hoja_fallidos_total,
                count_duplicados_reemplazados, count_conflictos, count_omitidos,
                summary_json, reimport_confirmed
            ) VALUES (?,?,?,?,?,?,?,?,0,0,0,0,0,0,0,?,0)
            """,
            (
                batch.get("source_filename") or f"batch-{batch_id}.xlsx",
                batch["source_sha256"],
                0,
                "REPORTE_DETALLADO",
                user,
                now,
                len(rows),
                len(rows),
                '{"via":"beneficiary_batch"}',
            ),
        )


def save_manual_beneficiaries(
    db_path: str,
    user: str,
    rows: list[dict[str, Any]],
    *,
    batch_id: int | None = None,
    expected_revision: int | None = None,
) -> dict[str, Any]:
    """Persist one complete MANUAL snapshot in one all-or-nothing transaction."""
    if not isinstance(rows, list) or not rows:
        raise ManualBatchValidationError(
            [{
                "row_id": None,
                "row_index": 0,
                "client_row_key": "",
                "field": "rows",
                "error_code": "beneficiary_rows_required",
                "message": "Añada al menos un beneficiario.",
            }]
        )
    if len(rows) > 200:
        raise ManualBatchValidationError(
            [{
                "row_id": None,
                "row_index": 200,
                "client_row_key": "",
                "field": "rows",
                "error_code": "beneficiary_rows_limit",
                "message": "El lote excede el máximo de 200 beneficiarios.",
            }]
        )

    conn = connect(db_path)
    created_batch_id: int | None = None
    try:
        ensure_banorte_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        batch: dict[str, Any] | None = None
        if batch_id is not None:
            raw_batch = conn.execute(
                "SELECT * FROM nomina_banorte_beneficiary_batches WHERE id=?",
                (int(batch_id),),
            ).fetchone()
            if raw_batch is None:
                raise BatchStaleError(int(batch_id), int(expected_revision or 0))
            batch = dict(raw_batch)
            if str(batch["created_by"]) != user:
                raise BatchAccessError()
            if str(batch["origin_kind"]) != "MANUAL":
                raise BatchAccessError("batch_not_manual")
            if str(batch["status"]) != "OPEN":
                raise BatchStaleError(int(batch_id), int(batch["revision"]))
            if expected_revision is None or int(batch["revision"]) != int(expected_revision):
                raise BatchStaleError(int(batch_id), int(batch["revision"]))
        else:
            existing = conn.execute(
                """
                SELECT id, revision FROM nomina_banorte_beneficiary_batches
                WHERE created_by=? AND origin_kind='MANUAL' AND status='OPEN'
                ORDER BY id DESC LIMIT 1
                """,
                (user,),
            ).fetchone()
            if existing is not None:
                raise BatchStaleError(int(existing["id"]), int(existing["revision"]))

        canonical, errors = _validate_batch_rows_conn(
            conn, rows, exclude_batch_id=int(batch_id) if batch_id is not None else None
        )
        if errors:
            raise ManualBatchValidationError(errors)

        now = _now()
        if batch is None:
            cur = conn.execute(
                """
                INSERT INTO nomina_banorte_beneficiary_batches (
                    origin_kind, status, revision, created_by, created_at, updated_at
                ) VALUES ('MANUAL', 'OPEN', 1, ?, ?, ?)
                """,
                (user, now, now),
            )
            created_batch_id = int(cur.lastrowid)
            batch = dict(
                conn.execute(
                    "SELECT * FROM nomina_banorte_beneficiary_batches WHERE id=?",
                    (created_batch_id,),
                ).fetchone()
            )
        else:
            conn.execute(
                "DELETE FROM nomina_banorte_beneficiary_batch_rows WHERE batch_id=?",
                (int(batch["id"]),),
            )

        _insert_batch_rows_conn(
            conn,
            int(batch["id"]),
            start_position=1,
            rows=canonical,
            now=now,
        )
        _persist_confirmed_beneficiaries_conn(conn, batch, canonical, user=user, now=now)
        revision = int(batch["revision"])
        cur = conn.execute(
            """
            UPDATE nomina_banorte_beneficiary_batches
            SET status='CONFIRMED', revision=revision+1, updated_at=?, confirmed_at=?
            WHERE id=? AND revision=? AND status='OPEN'
            """,
            (now, now, int(batch["id"]), revision),
        )
        if cur.rowcount != 1:
            raise BatchStaleError(int(batch["id"]), revision)
        conn.commit()
        created_batch_id = int(batch["id"])
    except (BatchAccessError, BatchStaleError, ManualBatchValidationError):
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_batch(db_path, int(created_batch_id))  # type: ignore[return-value]


def confirm_batch(
    db_path: str,
    batch_id: int,
    user: str,
    expected_revision: int,
) -> dict[str, Any]:
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        conn.execute("BEGIN IMMEDIATE")
        batch = conn.execute(
            "SELECT * FROM nomina_banorte_beneficiary_batches WHERE id=?",
            (int(batch_id),),
        ).fetchone()
        if batch is None:
            raise ValueError("batch_not_found")
        if batch["status"] != "OPEN":
            raise ValueError("batch_not_open")
        if int(batch["revision"]) != int(expected_revision):
            raise BatchStaleError(int(batch_id), int(batch["revision"]))
        rows = [
            dict(r)
            for r in conn.execute(
                "SELECT * FROM nomina_banorte_beneficiary_batch_rows WHERE batch_id=? ORDER BY position",
                (int(batch_id),),
            )
        ]
        if not rows:
            raise ValueError("batch_empty")
        canonical, errors = _validate_batch_rows_conn(
            conn, rows, exclude_batch_id=int(batch_id)
        )
        if errors:
            for e in errors:
                conn.execute(
                    """
                    UPDATE nomina_banorte_beneficiary_batch_rows
                    SET row_state='ERROR', error_code=?, error_message=?, updated_at=?
                    WHERE id=?
                    """,
                    (e["error_code"], e["error_message"], _now(), e["row_id"]),
                )
            conn.commit()
            raise ValueError("batch_row_errors:" + ",".join(e["error_code"] for e in errors))

        now = _now()
        _persist_confirmed_beneficiaries_conn(
            conn, dict(batch), canonical, user=user, now=now
        )
        cur = conn.execute(
            """
            UPDATE nomina_banorte_beneficiary_batches
            SET status='CONFIRMED', revision = revision + 1, updated_at=?, confirmed_at=?
            WHERE id=? AND revision=? AND status='OPEN'
            """,
            (now, now, int(batch_id), int(expected_revision)),
        )
        if cur.rowcount != 1:
            raise BatchStaleError(int(batch_id), int(expected_revision))
        conn.commit()
    except BatchStaleError:
        conn.rollback()
        raise
    except ValueError:
        # row errors already committed with ERROR markers for UI
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_batch(db_path, batch_id)  # type: ignore[return-value]


def prepare_reporte_batch(
    db_path: str,
    user: str,
    file_bytes: bytes,
    filename: str,
    *,
    confirm_reimport: bool = False,
) -> dict[str, Any]:
    """Parse Reporte Detallado into a staging batch — does not insert beneficiaries."""
    from modules.nomina.banorte.import_service import _cell, _existing_sha_batch, _find_header_map

    safe_name = safe_upload_filename(filename)
    if not safe_name.lower().endswith(".xlsx"):
        raise ValueError("invalid_extension")
    sha = hashlib.sha256(file_bytes).hexdigest()
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        prior_id = _existing_sha_batch(conn, sha)
        prior_at = None
        if prior_id is not None:
            prow = conn.execute(
                "SELECT imported_at FROM nomina_banorte_import_batches WHERE id=?",
                (int(prior_id),),
            ).fetchone()
            prior_at = prow["imported_at"] if prow else None
        if prior_id is not None and not confirm_reimport:
            return {
                "ok": False,
                "code": "duplicate_file_confirmation_required",
                "message": (
                    "Este reporte ya fue procesado anteriormente. "
                    "¿Deseas preparar un nuevo lote con el mismo archivo?"
                ),
                "prior": {"imported_at": prior_at, "batch_ref": f"IMP-{prior_id}"},
            }
    finally:
        conn.close()

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, colmap = _find_header_map(
        ws, {"NUMERO DE EMPLEADO", "NOMBRE DEL EMPLEADO", "NUMERO DE CUENTA"}
    )

    if confirm_reimport:
        existing = create_batch(db_path, user, origin_kind="REPORTE_DETALLADO")
        if existing and existing.get("status") == "OPEN":
            abandon_batch(db_path, int(existing["id"]), user, int(existing["revision"]))

    batch = create_batch(
        db_path,
        user,
        origin_kind="REPORTE_DETALLADO",
        source_filename=safe_name,
        source_sha256=sha,
        prior_batch_id=prior_id,
    )
    if (
        not confirm_reimport
        and batch.get("source_sha256") == sha
        and batch.get("rows")
    ):
        return {"ok": True, "batch": batch, "idempotent": True}

    # If reusing empty OPEN batch, stamp source metadata
    conn = connect(db_path)
    try:
        ensure_banorte_tables(conn)
        conn.execute(
            """
            UPDATE nomina_banorte_beneficiary_batches
            SET source_filename=?, source_sha256=?, prior_batch_id=?, updated_at=?
            WHERE id=? AND status='OPEN'
            """,
            (safe_name, sha, prior_id, _now(), int(batch["id"])),
        )
        conn.commit()
    finally:
        conn.close()
    batch = get_batch(db_path, int(batch["id"]))  # type: ignore[assignment]
    assert batch is not None

    pending_rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        if not any(ws.cell(r, c).value for c in range(1, (ws.max_column or 1) + 1)):
            continue
        estatus = _cell(ws, r, colmap.get("ESTATUS"))
        if normalize_header(estatus) != "EXITOSO":
            continue
        nombre = _cell(ws, r, colmap.get("NOMBRE DEL EMPLEADO"))
        emp_cell = ws.cell(r, colmap["NUMERO DE EMPLEADO"])
        acct_cell = ws.cell(r, colmap["NUMERO DE CUENTA"])
        emp, _emp_err = extract_identifier_cell(emp_cell.value, number_format=emp_cell.number_format)
        acct, _acct_err = extract_identifier_cell(acct_cell.value, number_format=acct_cell.number_format)
        if not nombre or not emp or not acct:
            continue
        comment = _cell(ws, r, colmap.get("COMENTARIOS"))
        substituted = is_banorte_employee_substituted_comment(comment)
        pending_rows.append(
            {
                "nombre": str(nombre),
                "cuenta": str(acct),
                "employee_number": str(emp),
                "use_account_as_employee_number": substituted,
                "comment": str(comment) if comment is not None else None,
                "source_row": r,
            }
        )
    batch = add_batch_rows_bulk(
        db_path,
        int(batch["id"]),
        user,
        int(batch["revision"]),
        pending_rows,
    )
    return {"ok": True, "batch": batch, "idempotent": False}
