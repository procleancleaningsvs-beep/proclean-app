"""Paleta unificada Master de Asistencia (plantilla Excel + visualizer web)."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Font, PatternFill


@dataclass(frozen=True)
class DailyKeyStyle:
    """Estilos para una clave diaria."""

    fill_hex: str  # ARGB sin #, mayúsculas
    font_bold: bool = False


# Colores alineados al requerimiento operativo (Excel 2010+ ARGB)
ATTENDANCE_KEY_STYLES: dict[str, DailyKeyStyle] = {
    "A": DailyKeyStyle("C6EFCE"),  # verde asistencia
    "F": DailyKeyStyle("FFC7CE"),  # rojo falta
    "D": DailyKeyStyle("9DC3E6"),  # azul descanso
    "DL": DailyKeyStyle("5B9BD5", font_bold=True),  # azul destacado domingo/descanso laborado
    "FL": DailyKeyStyle("F4B084", font_bold=True),  # naranja festivo laborado
    "FE": DailyKeyStyle("FCE4D6"),  # naranja suave festivo
    "V": DailyKeyStyle("FFEB9C"),  # amarillo vacaciones
    "I": DailyKeyStyle("D9D9D9"),  # gris incapacidad
    "PSS": DailyKeyStyle("FFC7CE"),  # rojo inasistencia
    "PCS": DailyKeyStyle("F4CCCC"),  # rojo suave permiso con goce
    "NI": DailyKeyStyle("D9D9D9"),
    "B": DailyKeyStyle("FF6B6B"),  # rojo baja
    "R": DailyKeyStyle("FFE699"),  # advertencia retardo
    "S": DailyKeyStyle("E2BCBC"),  # suspensión
    "OT": DailyKeyStyle("D7C6E8"),  # lavanda (distinto de verde/rojo/azul/naranja/amarillo/gris)
}

# Orden de reglas CF en Excel (todas son EXACT mutuamente excluyentes).
CF_RULE_KEY_ORDER: tuple[str, ...] = (
    "PSS",
    "PCS",
    "FL",
    "DL",
    "FE",
    "NI",
    "OT",
    "A",
    "D",
    "F",
    "V",
    "I",
    "B",
    "R",
    "S",
)


def pattern_fill_for_key(code: str) -> PatternFill:
    st = ATTENDANCE_KEY_STYLES.get(code)
    if st is None:
        return PatternFill("solid", fgColor="F3F4F6")
    return PatternFill("solid", fgColor=st.fill_hex)


def font_for_key(code: str) -> Font | None:
    st = ATTENDANCE_KEY_STYLES.get(code)
    if st is None:
        return None
    if st.font_bold:
        return Font(bold=True)
    return None


def css_vars_for_json() -> dict[str, dict[str, str | bool]]:
    """Serializable para JSON en plantillas (bgcolor, bold)."""
    out: dict[str, dict[str, str | bool]] = {}
    for k, st in ATTENDANCE_KEY_STYLES.items():
        out[k] = {"bg": f"#{st.fill_hex}", "bold": st.font_bold}
    out["_neutral"] = {"bg": "#f3f4f6", "bold": False}
    return out
