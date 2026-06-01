from __future__ import annotations

import re
import subprocess
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.nomina.asistencia_excel import build_daily_headers, format_period_label
from modules.nomina.db import NominaBaseRow
from modules.vitroflex_docs.libreoffice_pdf import resolve_soffice_path

BASE_DIR = Path(__file__).resolve().parent
MIERCOLES_TEMPLATE_CANDIDATES = (
    BASE_DIR / "templates_excel" / "Plantilla_Asistencia_Miercoles_v3.xlsx",
    BASE_DIR / "templates_excel" / "Plantilla_asistencia_miercoles_v3.xlsx",
)
MIERCOLES_NOMINA_LABEL = "MIERCOLES / AURIGA"

ASISTENCIA_SECTION_ORDER = (
    "ARMIDA",
    "AURIGA",
    "AURORA",
    "CITICA",
    "GM",
    "TORRE BALZAK",
    "TORRE OLMA",
    "OTROS",
)

_SECTION_ALIAS = {
    "ARMIDA": "ARMIDA",
    "AURIGA": "AURIGA",
    "AURORA": "AURORA",
    "CITICA": "CITICA",
    "GM": "GM",
    "TORRE BALZAK": "TORRE BALZAK",
    "TORRE OLMA": "TORRE OLMA",
    "OTROS": "OTROS",
}


@dataclass
class SectionBlock:
    name: str
    header_row: int
    start_row: int
    end_row: int


def _norm_text(value: Any) -> str:
    text = " ".join(str(value or "").replace("\u00a0", " ").replace("\n", " ").upper().split()).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Z0-9 /:_-]+", " ", text)
    return " ".join(text.split()).strip()


def _resolve_miercoles_template_path() -> Path:
    for path in MIERCOLES_TEMPLATE_CANDIDATES:
        if path.exists():
            return path
    raise FileNotFoundError(f"No existe plantilla base de miercoles: {MIERCOLES_TEMPLATE_CANDIDATES[0]}")


def _write_label_value_in_top(ws: Worksheet, label: str, value: str) -> None:
    wanted = _norm_text(label).rstrip(":") + ":"
    for row in range(1, 8):
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=row, column=col).value
            text = str(raw or "").strip()
            if not text:
                continue
            head = _norm_text(text).split(" ")[0].rstrip(":") + ":"
            if head == wanted:
                ws.cell(row=row, column=col, value=f"{wanted} {value}".strip())
                return


def _extract_label_value_in_top(ws: Worksheet, label: str) -> str:
    wanted = _norm_text(label).rstrip(":") + ":"
    for row in range(1, 8):
        for col in range(1, min(18, ws.max_column) + 1):
            raw = ws.cell(row=row, column=col).value
            text = str(raw or "").strip()
            if not text:
                continue
            norm = _norm_text(text)
            if norm.startswith(wanted):
                return text[len(label) :].replace(":", "", 1).strip() if ":" in text else text
    return ""


def _find_asistencia_header_rows(ws: Worksheet) -> list[int]:
    out: list[int] = []
    for row in range(1, ws.max_row + 1):
        a = _norm_text(ws.cell(row=row, column=1).value)
        b = _norm_text(ws.cell(row=row, column=2).value)
        c = _norm_text(ws.cell(row=row, column=3).value)
        d = _norm_text(ws.cell(row=row, column=4).value)
        if a == "NOMBRE DE EMPLEADO" and b == "CLIENTE" and c == "PLANTA" and d == "TURNO":
            out.append(row)
    return out


def _find_compilado_header_row(ws: Worksheet) -> int | None:
    for row in range(1, 20):
        a = _norm_text(ws.cell(row=row, column=1).value)
        e = _norm_text(ws.cell(row=row, column=5).value)
        if "NOMBRE" in a and "CLIENTES / TURNOS DONDE LABORO" in e:
            return row
    return None


def _read_asistencia_section_blocks(ws: Worksheet) -> list[SectionBlock]:
    labels: list[tuple[str, int]] = []
    for row in range(1, ws.max_row + 1):
        a = str(ws.cell(row=row, column=1).value or "").strip()
        if not a.upper().startswith("CLIENTE:"):
            continue
        section = _norm_text(a.split(":", 1)[1] if ":" in a else a)
        if not section:
            continue
        labels.append((section, row))
    out: list[SectionBlock] = []
    for idx, (section, label_row) in enumerate(labels):
        next_label_row = labels[idx + 1][1] if idx + 1 < len(labels) else ws.max_row + 1
        header_row = label_row + 1
        start_row = label_row + 2
        end_row = next_label_row - 1
        if start_row <= end_row:
            out.append(SectionBlock(name=section, header_row=header_row, start_row=start_row, end_row=end_row))
    return out


def _map_row_to_section(cliente: str) -> str:
    norm = _norm_text(cliente)
    for alias, section in _SECTION_ALIAS.items():
        if alias and alias in norm:
            return section
    return "OTROS"


def _primary_cliente_from_clientes_turnos(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "AURIGA"
    first = text.split("/", 1)[0].strip()
    first = first.split("-", 1)[0].strip()
    mapped = _map_row_to_section(first)
    if mapped in _SECTION_ALIAS.values():
        return mapped
    return "AURIGA"


def _update_dynamic_headers(
    ws_asistencia: Worksheet,
    ws_compilado: Worksheet,
    *,
    fecha_inicio: date,
) -> None:
    headers = build_daily_headers(fecha_inicio)
    for row in _find_asistencia_header_rows(ws_asistencia):
        for idx, header in enumerate(headers):
            ws_asistencia.cell(row=row, column=6 + idx, value=header)
    comp_row = _find_compilado_header_row(ws_compilado)
    if comp_row is not None:
        for idx, header in enumerate(headers):
            ws_compilado.cell(row=comp_row, column=6 + idx, value=header)


def _clear_section_weekly_values(ws: Worksheet, section: SectionBlock) -> None:
    for row in range(section.start_row, section.end_row + 1):
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value="")
        for col in range(6, 18):
            ws.cell(row=row, column=col, value="")


def _preload_asistencia_sections(ws: Worksheet, base_rows: list[NominaBaseRow]) -> None:
    sections = _read_asistencia_section_blocks(ws)
    if not sections:
        return
    by_name = {section.name: section for section in sections}
    for section in sections:
        _clear_section_weekly_values(ws, section)

    grouped: dict[str, list[NominaBaseRow]] = {name: [] for name in by_name.keys()}
    grouped.setdefault("OTROS", [])
    for item in base_rows:
        section_name = _map_row_to_section(item.cliente)
        if section_name not in grouped:
            section_name = "OTROS"
        grouped.setdefault(section_name, []).append(item)

    for section_name in ASISTENCIA_SECTION_ORDER:
        section = by_name.get(section_name)
        if section is None:
            continue
        items = grouped.get(section_name) or []
        for idx, item in enumerate(items):
            row = section.start_row + idx
            if row > section.end_row:
                break
            ws.cell(row=row, column=1, value=item.nombre_empleado or "")
            ws.cell(row=row, column=2, value=item.cliente or section_name)
            ws.cell(row=row, column=3, value=item.planta or "")
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value=item.puesto or "")
            for col in range(6, 18):
                ws.cell(row=row, column=col, value="")


def build_miercoles_template_file(
    *,
    fecha_inicio: date,
    fecha_fin: date,
    coordinador: str,
    base_rows: list[NominaBaseRow] | None = None,
) -> bytes:
    wb = load_workbook(_resolve_miercoles_template_path())
    if "Asistencia" not in wb.sheetnames or "Compilado" not in wb.sheetnames:
        raise ValueError("La plantilla de miercoles requiere hojas 'Asistencia' y 'Compilado'.")
    ws_asistencia = wb["Asistencia"]
    ws_compilado = wb["Compilado"]

    periodo = format_period_label(fecha_inicio, fecha_fin)
    _write_label_value_in_top(ws_asistencia, "SEMANA", periodo)
    _write_label_value_in_top(ws_asistencia, "NOMINA", MIERCOLES_NOMINA_LABEL)
    _write_label_value_in_top(ws_asistencia, "COORDINADOR", coordinador or "")
    _write_label_value_in_top(ws_compilado, "SEMANA", periodo)
    _write_label_value_in_top(ws_compilado, "NOMINA", MIERCOLES_NOMINA_LABEL)
    _write_label_value_in_top(ws_compilado, "COORDINADOR", coordinador or "")
    _update_dynamic_headers(ws_asistencia, ws_compilado, fecha_inicio=fecha_inicio)
    _preload_asistencia_sections(ws_asistencia, list(base_rows or []))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def is_miercoles_v3_excel(file_bytes: bytes) -> bool:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
    except Exception:
        return False
    if "Asistencia" not in wb.sheetnames or "Compilado" not in wb.sheetnames:
        return False
    ws_asistencia = wb["Asistencia"]
    ws_compilado = wb["Compilado"]

    has_turno = False
    for row in range(1, min(ws_asistencia.max_row, 30) + 1):
        for col in range(1, min(ws_asistencia.max_column, 20) + 1):
            if _norm_text(ws_asistencia.cell(row=row, column=col).value) == "TURNO":
                has_turno = True
                break
        if has_turno:
            break
    if not has_turno:
        return False

    has_clientes_turnos = False
    for row in range(1, min(ws_compilado.max_row, 15) + 1):
        for col in range(1, min(ws_compilado.max_column, 20) + 1):
            if _norm_text(ws_compilado.cell(row=row, column=col).value) == "CLIENTES / TURNOS DONDE LABORO":
                has_clientes_turnos = True
                break
        if has_clientes_turnos:
            break
    return has_clientes_turnos


def _try_recalculate_with_libreoffice(file_bytes: bytes, *, timeout_sec: int = 180) -> tuple[bytes, str | None]:
    soffice = resolve_soffice_path()
    if not soffice:
        return file_bytes, "No se pudo recalcular el Excel; se usaron valores guardados."
    with tempfile.TemporaryDirectory(prefix="proclean_nomina_miercoles_") as tmp:
        tmpdir = Path(tmp)
        source = tmpdir / "entrada.xlsx"
        source.write_bytes(file_bytes)
        cmd = [
            soffice,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--norestore",
            "--nodefault",
            "--nolockcheck",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(tmpdir),
            str(source),
        ]
        try:
            subprocess.run(cmd, check=True, timeout=timeout_sec, capture_output=True, text=True)
        except Exception:
            return file_bytes, "No se pudo recalcular el Excel; se usaron valores guardados."
        recalculated = tmpdir / "entrada.xlsx"
        if not recalculated.exists():
            return file_bytes, "No se pudo recalcular el Excel; se usaron valores guardados."
        got = recalculated.read_bytes()
        if not got:
            return file_bytes, "No se pudo recalcular el Excel; se usaron valores guardados."
        return got, None


def _numeric_like_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace(",", "")
    return text


def _is_valid_number(value: Any) -> bool:
    text = _numeric_like_text(value)
    if not text:
        return True
    try:
        float(text)
        return True
    except ValueError:
        return False


def _header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _norm_text(ws.cell(row=header_row, column=col).value)
        if key and key not in out:
            out[key] = col
    return out


def _extract_day_columns(headers: dict[str, int]) -> list[tuple[int, str]]:
    pattern = re.compile(r"^[LMDJSV]\d{1,2}$")
    out: list[tuple[int, str]] = []
    for key, col in headers.items():
        if pattern.match(key):
            out.append((col, key))
    out.sort(key=lambda x: x[0])
    return out[:7]


def _get_first_col(headers: dict[str, int], *options: str) -> int | None:
    for option in options:
        norm = _norm_text(option)
        if norm in headers:
            return headers[norm]
    return None


def parse_miercoles_v3_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    parse_bytes, recalc_warning = _try_recalculate_with_libreoffice(file_bytes)
    wb = load_workbook(BytesIO(parse_bytes), data_only=True)
    if "Compilado" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Compilado' en la plantilla de miercoles.")
    ws = wb["Compilado"]
    header_row = _find_compilado_header_row(ws)
    if header_row is None:
        raise ValueError("No se detectaron encabezados minimos en la hoja Compilado.")
    headers = _header_map(ws, header_row)

    col_nombre = _get_first_col(headers, "NOMBRE COMPLETO", "NOMBRE")
    col_puesto = _get_first_col(headers, "PUESTO")
    col_banco = _get_first_col(headers, "BANCO")
    col_cuenta = _get_first_col(headers, "CUENTA")
    col_clientes_turnos = _get_first_col(headers, "CLIENTES / TURNOS DONDE LABORO")
    col_dias_adicionales = _get_first_col(headers, "DIAS ADICIONALES DETECTADOS")
    col_eventos = _get_first_col(headers, "EVENTOS ADICIONALES MISMO DIA")
    col_he = _get_first_col(headers, "HORAS EXTRA")
    col_he_norm = _get_first_col(headers, "HORAS EXTRA NORMALES")
    col_dias_cub = _get_first_col(headers, "DIAS CUBIERTOS NORMALES")
    col_vac = _get_first_col(headers, "VACACIONES LABORADAS")
    col_prima = _get_first_col(headers, "PRIMA VACACIONAL")
    col_bono = _get_first_col(headers, "BONO")
    col_deducciones = _get_first_col(headers, "DEDUCCIONES")
    col_obs = _get_first_col(headers, "OBSERVACIONES")
    day_cols = _extract_day_columns(headers)

    if col_nombre is None or col_clientes_turnos is None or not day_cols:
        raise ValueError("No se detectaron encabezados minimos en la hoja Compilado.")

    rows: list[dict[str, Any]] = []
    blocking_errors: list[str] = []
    global_warnings: list[str] = []
    if recalc_warning:
        global_warnings.append(recalc_warning)

    for row_number in range(header_row + 1, ws.max_row + 1):
        name_raw = ws.cell(row=row_number, column=col_nombre).value
        nombre = str(name_raw or "").strip()
        has_other = any(
            str(ws.cell(row=row_number, column=col).value or "").strip()
            for col, _ in day_cols
        ) or bool(str(ws.cell(row=row_number, column=col_clientes_turnos).value or "").strip())
        if not nombre:
            if has_other:
                global_warnings.append(f"Fila {row_number} sin nombre; se omitio.")
            continue

        clientes_turnos = str(ws.cell(row=row_number, column=col_clientes_turnos).value or "").strip()
        days_headers = [head for _, head in day_cols]
        days_values = [str(ws.cell(row=row_number, column=col).value or "").strip() for col, _ in day_cols]
        while len(days_headers) < 7:
            days_headers.append(f"D{len(days_headers) + 1}")
            days_values.append("")

        warnings = []
        errors = []

        def cell_text(col: int | None) -> str:
            if col is None:
                return ""
            return str(ws.cell(row=row_number, column=col).value or "").strip()

        for label, col in (
            ("HORAS EXTRA", col_he),
            ("HORAS EXTRA NORMALES", col_he_norm),
            ("DIAS CUBIERTOS NORMALES", col_dias_cub),
            ("VACACIONES LABORADAS", col_vac),
            ("BONO", col_bono),
            ("DEDUCCIONES", col_deducciones),
        ):
            raw = cell_text(col)
            if raw and not _is_valid_number(raw):
                warnings.append(f"{label} no numerico; revisar.")

        if col_dias_adicionales is not None:
            txt = cell_text(col_dias_adicionales)
            try:
                if txt and float(txt) > 0:
                    warnings.append("Dias adicionales detectados > 0.")
            except ValueError:
                warnings.append("Dias adicionales detectados no numerico; revisar.")
        if col_eventos is not None:
            txt = cell_text(col_eventos)
            try:
                if txt and float(txt) > 0:
                    warnings.append("Eventos adicionales mismo dia > 0.")
            except ValueError:
                warnings.append("Eventos adicionales mismo dia no numerico; revisar.")

        banco = cell_text(col_banco)
        cuenta = cell_text(col_cuenta)
        if not banco or not cuenta:
            warnings.append("Banco/cuenta vacios; revisar para pago.")

        observaciones = cell_text(col_obs)
        obs_norm = _norm_text(observaciones)
        if "DOBLE" in obs_norm or "FL" in obs_norm or "CODIGOS" in obs_norm:
            warnings.append("Observaciones con posibles eventos multiples; revisar.")
        if clientes_turnos:
            if observaciones:
                observaciones = f"{observaciones} | CLIENTES/TURNOS: {clientes_turnos}"
            else:
                observaciones = f"CLIENTES/TURNOS: {clientes_turnos}"

        prima = cell_text(col_prima) or "N/A"

        row_data: dict[str, Any] = {
            "row_number": row_number,
            "nombre_empleado": nombre,
            "cliente": _primary_cliente_from_clientes_turnos(clientes_turnos),
            "planta": "",
            "puesto": cell_text(col_puesto),
            "banco": banco,
            "cuenta": cuenta,
            "dia_1_header": days_headers[0],
            "dia_1_value": days_values[0],
            "dia_2_header": days_headers[1],
            "dia_2_value": days_values[1],
            "dia_3_header": days_headers[2],
            "dia_3_value": days_values[2],
            "dia_4_header": days_headers[3],
            "dia_4_value": days_values[3],
            "dia_5_header": days_headers[4],
            "dia_5_value": days_values[4],
            "dia_6_header": days_headers[5],
            "dia_6_value": days_values[5],
            "dia_7_header": days_headers[6],
            "dia_7_value": days_values[6],
            "he": cell_text(col_he),
            "horas_extra_normales": cell_text(col_he_norm),
            "dias_cubiertos_normales": cell_text(col_dias_cub),
            "vacaciones_laboradas": cell_text(col_vac),
            "prima_vacacional": prima,
            "bono": cell_text(col_bono),
            "deducciones": cell_text(col_deducciones),
            "observaciones": observaciones,
            "errors": errors,
            "warnings": warnings,
            "nss": "",
            "clientes_turnos_compilado": clientes_turnos,
        }
        rows.append(row_data)

    if not rows:
        msg = (
            "Compilado vacio o sin filas validas con nombre. "
            "Abre y guarda el archivo en Excel/LibreOffice para refrescar formulas y vuelve a importar."
        )
        if recalc_warning:
            msg = recalc_warning + " " + msg
        raise ValueError(msg)

    error_count = sum(len(row.get("errors") or []) for row in rows)
    warning_count = sum(len(row.get("warnings") or []) for row in rows) + len(global_warnings)

    semana = _extract_label_value_in_top(ws, "SEMANA")
    coordinador = _extract_label_value_in_top(ws, "COORDINADOR")
    dias_headers = [h for _, h in day_cols]
    while len(dias_headers) < 7:
        dias_headers.append(f"D{len(dias_headers) + 1}")
    return {
        "filename": filename,
        "semana": semana,
        "cliente": "AURIGA",
        "coordinador": coordinador,
        "dias_headers": dias_headers[:7],
        "rows": rows,
        "total_rows": len(rows),
        "error_count": error_count,
        "warning_count": warning_count,
        "blocking_errors": blocking_errors,
        "template_kind": "miercoles_v3",
        "template_warnings": global_warnings,
        "source_sheet": "Compilado",
    }
