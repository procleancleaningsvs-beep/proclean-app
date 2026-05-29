from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.nomina.vacaciones_util import PENDIENTE_REVISION


IMPORT_KIND_SIMPLE = "SIMPLE_TEMPLATE"
IMPORT_KIND_NORMALIZED = "NORMALIZED_HISTORY"
IMPORT_KIND_LEGACY = "LEGACY_HISTORY"


@dataclass
class ParsedVacaciones:
    rows: list[dict[str, Any]]
    warnings: list[str]
    errors: list[str]
    cliente: str
    weekly_events_total: int = 0
    import_kind: str = IMPORT_KIND_LEGACY
    preview: dict[str, Any] = field(default_factory=dict)


def normalize_sheet_name(value: Any) -> str:
    s = str(value or "").strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def normalize_column_name(value: Any) -> str:
    s = str(value or "").replace("\n", " ").strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def normalize_name(value: str) -> str:
    s = " ".join(str(value or "").replace("\u00a0", " ").split()).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return " ".join(s.split()).strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value):
            return None
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s or s.lower() in {"nan", "none", "null", "na", "n/a"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_bool(value: Any) -> bool:
    norm = normalize_column_name(value)
    if norm in {"1", "SI", "S", "TRUE", "YES", "PAGADA", "X"}:
        return True
    n = _to_float(value)
    return bool(n and n >= 1)


def _to_days(value: Any) -> float:
    n = _to_float(value)
    if n is not None:
        return float(n)
    return 1.0 if _to_bool(value) else 0.0


def _excel_serial_to_date(serial: float) -> date | None:
    if serial <= 0:
        return None
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(serial))).date()
    except (OverflowError, ValueError):
        return None


def _to_date_iso(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        parsed = _excel_serial_to_date(float(value))
        return parsed.isoformat() if parsed else str(value).strip()
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return ""
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        dd = int(m.group(1))
        mm = int(m.group(2))
        yy = int(m.group(3))
        if yy < 100:
            yy += 2000
        return f"{yy:04d}-{mm:02d}-{dd:02d}"
    m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m2:
        return f"{int(m2.group(1)):04d}-{int(m2.group(2)):02d}-{int(m2.group(3)):02d}"
    n = _to_float(s)
    if n is not None and n > 1000:
        parsed = _excel_serial_to_date(n)
        if parsed:
            return parsed.isoformat()
    return s


def detect_sheet_by_name(workbook, aliases: list[str]) -> Worksheet | None:
    alias_norm = {normalize_sheet_name(a) for a in aliases if a}
    by_norm = {normalize_sheet_name(name): name for name in workbook.sheetnames}
    for candidate in alias_norm:
        if candidate in by_norm:
            return workbook[by_norm[candidate]]
    for ws_name in workbook.sheetnames:
        norm = normalize_sheet_name(ws_name)
        if any(candidate in norm or norm in candidate for candidate in alias_norm):
            return workbook[ws_name]
    return None


def detect_header_row(ws: Worksheet, expected_columns: list[str], *, scan_rows: int = 20) -> int:
    expected = {normalize_column_name(v) for v in expected_columns if v}
    if not expected:
        return 1
    best_row = 1
    best_hits = -1
    max_row = min(scan_rows, ws.max_row)
    for row_idx in range(1, max_row + 1):
        row_norm = {
            normalize_column_name(ws.cell(row=row_idx, column=col).value)
            for col in range(1, ws.max_column + 1)
        }
        row_norm.discard("")
        if not row_norm:
            continue
        hits = len(expected.intersection(row_norm))
        if hits > best_hits:
            best_hits = hits
            best_row = row_idx
        if hits >= max(1, min(2, len(expected))):
            return row_idx
    return best_row


def build_column_index(ws: Worksheet, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        norm = normalize_column_name(ws.cell(row=header_row, column=col).value)
        if norm and norm not in out:
            out[norm] = col
    return out


def resolve_column(
    columns: dict[str, int],
    aliases: list[str],
    *,
    required: bool,
    warnings: list[str],
    sheet_label: str,
) -> int | None:
    alias_norm = [normalize_column_name(a) for a in aliases if a]
    for candidate in alias_norm:
        if candidate in columns:
            return columns[candidate]
    for key, idx in columns.items():
        if any(candidate in key or key in candidate for candidate in alias_norm):
            return idx
    expected = aliases[0] if aliases else "(sin alias)"
    msg = f"Hoja '{sheet_label}': Columna no encontrada -> {expected}"
    if required:
        raise ValueError(msg)
    warnings.append(msg)
    return None


def _sheet_rows_as_dicts(ws: Worksheet, header_row: int, limit_cols: int = 120) -> list[dict[str, Any]]:
    columns = build_column_index(ws, header_row)
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data: dict[str, Any] = {}
        non_empty = False
        for norm_key, col_idx in columns.items():
            if col_idx > limit_cols:
                continue
            val = ws.cell(row=row_idx, column=col_idx).value
            row_data[norm_key] = val
            if val not in (None, ""):
                non_empty = True
        if non_empty:
            rows.append(row_data)
    return rows


def conciliar_dias(
    *,
    dias_utilizados_excel: float,
    vacaciones_laboradas_excel: float,
    dias_pagados_excel: float,
    suma_movimientos_detectada: float,
    dias_restantes_excel: float | None,
) -> dict[str, Any]:
    warnings: list[str] = []
    du = float(dias_utilizados_excel or 0.0)
    vl = float(vacaciones_laboradas_excel or 0.0)
    dp = float(dias_pagados_excel or 0.0)
    sm = float(suma_movimientos_detectada or 0.0)

    if abs(sm - du) <= 0.01:
        clasificacion = "CONCILIADO_COMPLETO"
    elif abs((sm + vl) - du) <= 0.01:
        clasificacion = "LABORADAS_PROBABLEMENTE_FUERA_DEL_DESGLOSE"
    elif sm < du:
        clasificacion = "DIFERENCIA_CON_AJUSTE_HISTORICO"
    else:
        clasificacion = "POSIBLE_DUPLICIDAD_O_EXCESO_EN_SEMANAS"

    if dias_restantes_excel is not None and float(dias_restantes_excel) < 0:
        warnings.append("SALDO_NEGATIVO_VALIDO")
    if vl > du and du > 0:
        warnings.append("VACACIONES_LABORADAS_SUPERAN_DIAS_UTILIZADOS")
    if dp > du and du > 0:
        warnings.append("DIAS_PAGADOS_SUPERAN_DIAS_UTILIZADOS")

    return {
        "dias_utilizados_excel": du,
        "vacaciones_laboradas_excel": vl,
        "dias_pagados_excel": dp,
        "suma_movimientos_detectada": sm,
        "diferencia_detectada": round(sm - du, 4),
        "clasificacion_conciliacion": clasificacion,
        "warnings": warnings,
    }


def _simple_template_sheets(wb) -> tuple[Worksheet, Worksheet | None, Worksheet | None]:
    saldos = detect_sheet_by_name(wb, ["1_CAPTURA_SALDOS", "CAPTURA_SALDOS", "SALDOS"])
    if saldos is None:
        raise ValueError("Hoja principal no encontrada: 1_CAPTURA_SALDOS")
    movimientos = detect_sheet_by_name(wb, ["2_MOVIMIENTOS", "MOVIMIENTOS"])
    primas = detect_sheet_by_name(wb, ["3_PRIMA_VACACIONAL", "PRIMA_VACACIONAL"])
    return saldos, movimientos, primas


def _normalized_history_sheets(wb) -> tuple[Worksheet, Worksheet, Worksheet | None, Worksheet | None]:
    saldos = detect_sheet_by_name(wb, ["VACACIONES_SALDOS"])
    if saldos is None:
        raise ValueError("Hoja requerida no encontrada: VACACIONES_SALDOS")
    movimientos = detect_sheet_by_name(wb, ["VACACIONES_MOVIMIENTOS"])
    if movimientos is None:
        raise ValueError("Hoja requerida no encontrada: VACACIONES_MOVIMIENTOS")
    primas = detect_sheet_by_name(wb, ["PRIMA_VACACIONAL_HISTORICO"])
    diferencias = detect_sheet_by_name(wb, ["DIFERENCIAS_Y_WARNINGS"])
    return saldos, movimientos, primas, diferencias


def _build_base_row(
    *,
    source_row_number: int,
    nombre: str,
    fecha_ingreso_excel: str,
    sueldo_excel: float | None,
    planta_excel: str,
    dias_vacaciones_historico: float,
    dias_utilizados_excel: float,
    dias_utilizados_semanal: float,
    vacaciones_laboradas: float,
    dias_pagados: float,
    dias_restantes_excel: float | None,
    monto_total_historico: float | None,
    comentarios: str,
    prima_2025_pagada: bool,
    semana_pago_prima_2025: str,
    prima_2026_pagada: bool,
    fecha_pago_prima_2026: str,
    desglose_semanal: list[dict[str, Any]],
    clasificacion_conciliacion: str,
    diferencia_detectada: float,
    fuente_fecha_ingreso: str,
    sueldo_headcount_in: float | None = None,
    salario_parametros_nomina: float | None = None,
    fuente_salario_preliminar: str | None = None,
    warnings: list[str] | None = None,
    extra_editable: dict[str, Any] | None = None,
) -> dict[str, Any]:
    row_warnings = list(warnings or [])
    nombre_norm = normalize_name(nombre)

    salario_usado = None
    fuente_salario = "SIN_DATO"
    if salario_parametros_nomina not in (None, ""):
        salario_usado = float(salario_parametros_nomina)
        fuente_salario = "PARAMETROS_NOMINA"
    elif sueldo_headcount_in not in (None, ""):
        salario_usado = float(sueldo_headcount_in)
        fuente_salario = "HEADCOUNT"
    elif sueldo_excel not in (None, ""):
        salario_usado = float(sueldo_excel)
        fuente_salario = "EXCEL"
    if fuente_salario_preliminar:
        fuente_salario = fuente_salario_preliminar

    if dias_restantes_excel is None:
        dias_restantes_calc = float(dias_vacaciones_historico or 0.0) - float(dias_utilizados_excel or 0.0)
    else:
        dias_restantes_calc = float(dias_restantes_excel)

    editable_json = {
        "revision_status": "pending_revision",
        "desglose_semanal": desglose_semanal,
        "excel_resumen": {
            "dias_vacaciones_excel": dias_vacaciones_historico,
            "dias_utilizados_excel_resumen": dias_utilizados_excel,
            "dias_utilizados_calculado_semanal": dias_utilizados_semanal,
            "vacaciones_laboradas_excel": vacaciones_laboradas,
            "dias_pagados_excel": dias_pagados,
            "dias_restantes_excel": dias_restantes_excel,
            "comentarios_excel": comentarios,
            "clasificacion_conciliacion": clasificacion_conciliacion,
            "diferencia_detectada": diferencia_detectada,
            "fuente_fecha_ingreso": fuente_fecha_ingreso,
            "fuente_salario": fuente_salario,
            "salario_parametros_nomina": salario_parametros_nomina,
        },
    }
    if extra_editable:
        editable_json.update(extra_editable)

    return {
        "source_row_number": source_row_number,
        "nss": "",
        "excel_nombre_original": nombre,
        "nombre_historico": nombre,
        "nombre_normalizado": nombre_norm,
        "headcount_nombre_original": "",
        "headcount_nombre_normalizado": "",
        "nombre_headcount": "",
        "cliente": "Carrier",
        "planta_historica": planta_excel,
        "planta_headcount": "",
        "fecha_ingreso_historica": fecha_ingreso_excel,
        "fecha_ingreso_headcount": "",
        "fecha_ingreso_usada": fecha_ingreso_excel,
        "fuente_fecha_ingreso": fuente_fecha_ingreso,
        "status_headcount": "SIN STATUS HEADCOUNT",
        "estatus_headcount": "SIN STATUS HEADCOUNT",
        "sueldo_historico": sueldo_excel,
        "sueldo_headcount": sueldo_headcount_in,
        "sueldo_usado": salario_usado,
        "salario_parametros_nomina": salario_parametros_nomina,
        "fuente_salario": fuente_salario,
        "dias_vacaciones_historico": dias_vacaciones_historico,
        "dias_utilizados": dias_utilizados_excel,
        "dias_utilizados_excel_resumen": dias_utilizados_excel,
        "dias_utilizados_calculado_semanal": dias_utilizados_semanal,
        "vacaciones_laboradas": vacaciones_laboradas,
        "dias_pagados": dias_pagados,
        "dias_restantes_historico": dias_restantes_excel,
        "dias_restantes_calculado": dias_restantes_calc,
        "prima_2025_pagada": prima_2025_pagada,
        "semana_pago_prima_2025": semana_pago_prima_2025,
        "prima_2026_pagada": prima_2026_pagada,
        "fecha_pago_prima_2026": fecha_pago_prima_2026,
        "monto_total_historico": monto_total_historico,
        "monto_total_recalculado": None,
        "comentarios": comentarios,
        "match_status": PENDIENTE_REVISION,
        "match_method": "",
        "match_notes": "",
        "match_score": None,
        "desglose_semanal": desglose_semanal,
        "clasificacion_conciliacion": clasificacion_conciliacion,
        "diferencia_detectada": diferencia_detectada,
        "warnings": row_warnings,
        "editable_json": editable_json,
        "is_active": 0,
    }


def parse_simple_template(wb, filename: str) -> ParsedVacaciones:
    saldos_ws, movimientos_ws, primas_ws = _simple_template_sheets(wb)
    warnings: list[str] = []
    errors: list[str] = []

    # 2_MOVIMIENTOS
    movimientos_by_name: dict[str, list[dict[str, Any]]] = {}
    if movimientos_ws is not None:
        m_header = detect_header_row(movimientos_ws, ["NOMBRE COMPLETO", "DIAS CAPTURADOS", "SEMANA/PERIODO ORIGINAL"])
        m_cols = build_column_index(movimientos_ws, m_header)
        c_nombre = resolve_column(m_cols, ["NOMBRE COMPLETO"], required=True, warnings=warnings, sheet_label=movimientos_ws.title)
        c_anio = resolve_column(m_cols, ["A?O", "ANIO"], required=False, warnings=warnings, sheet_label=movimientos_ws.title)
        c_periodo = resolve_column(m_cols, ["SEMANA/PERIODO ORIGINAL", "SEMANA", "ENCABEZADO ORIGINAL"], required=False, warnings=warnings, sheet_label=movimientos_ws.title)
        c_dias = resolve_column(m_cols, ["DIAS CAPTURADOS", "DIAS", "DIAS TOMADOS"], required=True, warnings=warnings, sheet_label=movimientos_ws.title)
        c_tipo = resolve_column(m_cols, ["TIPO MOVIMIENTO"], required=False, warnings=warnings, sheet_label=movimientos_ws.title)
        c_obs = resolve_column(m_cols, ["COMENTARIOS", "OBSERVACIONES"], required=False, warnings=warnings, sheet_label=movimientos_ws.title)

        for row_idx in range(m_header + 1, movimientos_ws.max_row + 1):
            nombre = str(movimientos_ws.cell(row=row_idx, column=c_nombre).value or "").strip() if c_nombre else ""
            if not nombre:
                continue
            dias = _to_days(movimientos_ws.cell(row=row_idx, column=c_dias).value if c_dias else None)
            if dias <= 0:
                continue
            anio_val = movimientos_ws.cell(row=row_idx, column=c_anio).value if c_anio else ""
            periodo = str(movimientos_ws.cell(row=row_idx, column=c_periodo).value or "").strip() if c_periodo else ""
            tipo = str(movimientos_ws.cell(row=row_idx, column=c_tipo).value or "VACACIONES").strip() if c_tipo else "VACACIONES"
            obs = str(movimientos_ws.cell(row=row_idx, column=c_obs).value or "").strip() if c_obs else ""
            key = normalize_name(nombre)
            movimientos_by_name.setdefault(key, []).append(
                {
                    "anio": str(anio_val or "").strip(),
                    "semana": periodo,
                    "period_label": periodo,
                    "days": dias,
                    "tipo_movimiento": tipo,
                    "source": "simple_template",
                    "observaciones": obs,
                    "excel_row": row_idx,
                }
            )

    # 3_PRIMA_VACACIONAL
    primas_by_name: dict[str, list[dict[str, Any]]] = {}
    if primas_ws is not None:
        p_header = detect_header_row(primas_ws, ["NOMBRE COMPLETO", "A?O PRIMA", "PRIMA PAGADA"])
        p_cols = build_column_index(primas_ws, p_header)
        c_nombre = resolve_column(p_cols, ["NOMBRE COMPLETO"], required=True, warnings=warnings, sheet_label=primas_ws.title)
        c_anio = resolve_column(p_cols, ["A?O PRIMA", "ANIO PRIMA"], required=False, warnings=warnings, sheet_label=primas_ws.title)
        c_pagada = resolve_column(p_cols, ["PRIMA PAGADA"], required=False, warnings=warnings, sheet_label=primas_ws.title)
        c_sueldo = resolve_column(p_cols, ["SUELDO DIARIO BASE HISTORICO"], required=False, warnings=warnings, sheet_label=primas_ws.title)
        c_monto = resolve_column(p_cols, ["MONTO HISTORICO SUGERIDO", "MONTO PAGADO REAL"], required=False, warnings=warnings, sheet_label=primas_ws.title)
        c_semana = resolve_column(p_cols, ["SEMANA/FECHA DE PAGO", "SEMANA PAGO"], required=False, warnings=warnings, sheet_label=primas_ws.title)
        c_com = resolve_column(p_cols, ["COMENTARIOS"], required=False, warnings=warnings, sheet_label=primas_ws.title)

        for row_idx in range(p_header + 1, primas_ws.max_row + 1):
            nombre = str(primas_ws.cell(row=row_idx, column=c_nombre).value or "").strip() if c_nombre else ""
            if not nombre:
                continue
            key = normalize_name(nombre)
            anio = str(primas_ws.cell(row=row_idx, column=c_anio).value or "").strip() if c_anio else ""
            pagada = _to_bool(primas_ws.cell(row=row_idx, column=c_pagada).value if c_pagada else None)
            sueldo_base = _to_float(primas_ws.cell(row=row_idx, column=c_sueldo).value if c_sueldo else None)
            monto = _to_float(primas_ws.cell(row=row_idx, column=c_monto).value if c_monto else None)
            semana_pago = str(primas_ws.cell(row=row_idx, column=c_semana).value or "").strip() if c_semana else ""
            comentarios = str(primas_ws.cell(row=row_idx, column=c_com).value or "").strip() if c_com else ""
            primas_by_name.setdefault(key, []).append(
                {
                    "anio": anio,
                    "prima_pagada": pagada,
                    "indicador_original": primas_ws.cell(row=row_idx, column=c_pagada).value if c_pagada else None,
                    "sueldo_diario_base_excel": sueldo_base,
                    "monto_estimado_historico": monto,
                    "semana_pago": semana_pago,
                    "fecha_pago": _to_date_iso(semana_pago),
                    "formula_historica": "sueldo_diario_x_3",
                    "comentarios": comentarios,
                }
            )

    # 1_CAPTURA_SALDOS
    header_row = detect_header_row(saldos_ws, ["NOMBRE COMPLETO", "DIAS UTILIZADOS TOTAL", "DIAS RESTANTES HISTORICO"])
    cols = build_column_index(saldos_ws, header_row)

    c_nombre = resolve_column(cols, ["NOMBRE COMPLETO", "NOMBRE"], required=True, warnings=warnings, sheet_label=saldos_ws.title)
    c_fecha = resolve_column(cols, ["FECHA INGRESO EXCEL", "FECHA DE INGRESO"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_sueldo = resolve_column(cols, ["SUELDO DIARIO HISTORICO", "SUELDO"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_planta = resolve_column(cols, ["PLANTA HISTORICA", "PLANTA"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_du = resolve_column(cols, ["DIAS UTILIZADOS TOTAL", "DIAS UTILIZADOS"], required=True, warnings=warnings, sheet_label=saldos_ws.title)
    c_vl = resolve_column(cols, ["VACACIONES LABORADAS INCLUIDAS", "VACACIONES LABORADAS"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_dp = resolve_column(cols, ["DIAS PAGADOS/LABORADAS", "DIAS PAGADOS"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_dr = resolve_column(cols, ["DIAS RESTANTES HISTORICO", "DIAS RESTANTES"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_p25 = resolve_column(cols, ["PRIMA 2025 PAGADA", "PRIMA VACACIONAL 2025"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_sp25 = resolve_column(cols, ["SEMANA/FECHA PAGO PRIMA 2025", "SEMANA PAGO PRIMA 2025"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_p26 = resolve_column(cols, ["PRIMA 2026 PAGADA", "PRIMA VACACIONAL 2026"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_fp26 = resolve_column(cols, ["FECHA PAGO PRIMA 2026", "FECHA DE PAGO"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_monto = resolve_column(cols, ["MONTO TOTAL HISTORICO", "MONTO TOTAL"], required=False, warnings=warnings, sheet_label=saldos_ws.title)
    c_com = resolve_column(cols, ["COMENTARIOS"], required=False, warnings=warnings, sheet_label=saldos_ws.title)

    rows: list[dict[str, Any]] = []
    total_movs = 0
    total_primas = 0

    for row_idx in range(header_row + 1, saldos_ws.max_row + 1):
        nombre = str(saldos_ws.cell(row=row_idx, column=c_nombre).value or "").strip() if c_nombre else ""
        if not nombre:
            has_data = False
            for check_col in [c_fecha, c_sueldo, c_planta, c_du, c_dr, c_com]:
                if check_col and saldos_ws.cell(row=row_idx, column=check_col).value not in (None, ""):
                    has_data = True
                    break
            if has_data:
                errors.append(f"Fila {row_idx}: contiene datos pero falta nombre completo.")
            continue

        nombre_norm = normalize_name(nombre)
        fecha_excel = _to_date_iso(saldos_ws.cell(row=row_idx, column=c_fecha).value if c_fecha else None)
        sueldo_excel = _to_float(saldos_ws.cell(row=row_idx, column=c_sueldo).value if c_sueldo else None)
        planta_excel = str(saldos_ws.cell(row=row_idx, column=c_planta).value or "").strip() if c_planta else ""
        dias_utilizados_excel = _to_days(saldos_ws.cell(row=row_idx, column=c_du).value if c_du else None)
        vacaciones_laboradas = _to_days(saldos_ws.cell(row=row_idx, column=c_vl).value if c_vl else None)
        dias_pagados = _to_days(saldos_ws.cell(row=row_idx, column=c_dp).value if c_dp else None)
        dias_restantes_excel = _to_float(saldos_ws.cell(row=row_idx, column=c_dr).value if c_dr else None)
        monto_total = _to_float(saldos_ws.cell(row=row_idx, column=c_monto).value if c_monto else None)
        comentarios = str(saldos_ws.cell(row=row_idx, column=c_com).value or "").strip() if c_com else ""

        desglose = movimientos_by_name.get(nombre_norm, [])
        suma_movs = round(sum(float(item.get("days") or 0.0) for item in desglose), 4)
        total_movs += len(desglose)

        prima_det = primas_by_name.get(nombre_norm, [])
        total_primas += len(prima_det)

        prima_2025 = _to_bool(saldos_ws.cell(row=row_idx, column=c_p25).value if c_p25 else None)
        prima_2026 = _to_bool(saldos_ws.cell(row=row_idx, column=c_p26).value if c_p26 else None)
        semana_25 = str(saldos_ws.cell(row=row_idx, column=c_sp25).value or "").strip() if c_sp25 else ""
        fecha_26 = _to_date_iso(saldos_ws.cell(row=row_idx, column=c_fp26).value if c_fp26 else None)

        for pr in prima_det:
            anio = str(pr.get("anio") or "")
            if anio == "2025" and pr.get("prima_pagada"):
                prima_2025 = True
                if not semana_25:
                    semana_25 = str(pr.get("semana_pago") or "")
            if anio == "2026" and pr.get("prima_pagada"):
                prima_2026 = True
                if not fecha_26:
                    fecha_26 = _to_date_iso(pr.get("fecha_pago"))

        conc = conciliar_dias(
            dias_utilizados_excel=dias_utilizados_excel,
            vacaciones_laboradas_excel=vacaciones_laboradas,
            dias_pagados_excel=dias_pagados,
            suma_movimientos_detectada=suma_movs,
            dias_restantes_excel=dias_restantes_excel,
        )

        row_warnings = []
        if conc["clasificacion_conciliacion"] != "CONCILIADO_COMPLETO":
            row_warnings.append(f"Conciliaci?n: {conc['clasificacion_conciliacion']}")
        for w in conc["warnings"]:
            row_warnings.append(w)
        if not desglose and dias_utilizados_excel > 0:
            row_warnings.append("Sin movimientos semanales para respaldar DIAS UTILIZADOS total")

        dias_vac_historico = max(0.0, dias_utilizados_excel + (dias_restantes_excel or 0.0))

        row = _build_base_row(
            source_row_number=row_idx,
            nombre=nombre,
            fecha_ingreso_excel=fecha_excel,
            sueldo_excel=sueldo_excel,
            planta_excel=planta_excel,
            dias_vacaciones_historico=dias_vac_historico,
            dias_utilizados_excel=dias_utilizados_excel,
            dias_utilizados_semanal=suma_movs,
            vacaciones_laboradas=vacaciones_laboradas,
            dias_pagados=dias_pagados,
            dias_restantes_excel=dias_restantes_excel,
            monto_total_historico=monto_total,
            comentarios=comentarios,
            prima_2025_pagada=prima_2025,
            semana_pago_prima_2025=semana_25,
            prima_2026_pagada=prima_2026,
            fecha_pago_prima_2026=fecha_26,
            desglose_semanal=desglose,
            clasificacion_conciliacion=conc["clasificacion_conciliacion"],
            diferencia_detectada=conc["diferencia_detectada"],
            fuente_fecha_ingreso="EXCEL",
            warnings=row_warnings,
            extra_editable={
                "prima_historica_detalle": prima_det,
                "warnings_detalle": [
                    {
                        "severidad": "warning" if w != "SALDO_NEGATIVO_VALIDO" else "info",
                        "campo_afectado": "conciliacion",
                        "detalle": w,
                        "accion_recomendada": "Revisar detalle de saldos y movimientos",
                    }
                    for w in row_warnings
                ],
                "import_kind": IMPORT_KIND_SIMPLE,
            },
        )
        rows.append(row)
        warnings.extend([f"Fila {row_idx}: {w}" for w in row_warnings])

    preview = {
        "tipo_archivo": IMPORT_KIND_SIMPLE,
        "total_trabajadores": len(rows),
        "total_movimientos": total_movs,
        "total_primas": total_primas,
        "total_warnings": len(warnings),
    }

    return ParsedVacaciones(
        rows=rows,
        warnings=warnings,
        errors=errors,
        cliente="Carrier",
        weekly_events_total=total_movs,
        import_kind=IMPORT_KIND_SIMPLE,
        preview=preview,
    )


def parse_normalized_history(wb, filename: str) -> ParsedVacaciones:
    saldos_ws, movimientos_ws, primas_ws, diferencias_ws = _normalized_history_sheets(wb)
    warnings: list[str] = []
    errors: list[str] = []

    # Movimientos
    m_header = detect_header_row(movimientos_ws, ["NOMBRE_COMPLETO_NORMALIZADO", "DIAS_CAPTURADOS"])
    m_rows = _sheet_rows_as_dicts(movimientos_ws, m_header)
    movimientos_by_name: dict[str, list[dict[str, Any]]] = {}
    for item in m_rows:
        name_norm = normalize_name(item.get("NOMBRE_COMPLETO_NORMALIZADO") or item.get("NOMBRE_COMPLETO_ORIGINAL") or "")
        if not name_norm:
            continue
        days = _to_days(item.get("DIAS_CAPTURADOS"))
        if days <= 0:
            continue
        movimientos_by_name.setdefault(name_norm, []).append(
            {
                "anio": str(item.get("ANIO_DETECTADO") or "").strip(),
                "semana": str(item.get("SEMANA_DETECTADA") or item.get("ENCABEZADO_ORIGINAL_EXCEL") or "").strip(),
                "period_label": str(item.get("ENCABEZADO_ORIGINAL_EXCEL") or item.get("SEMANA_DETECTADA") or "").strip(),
                "days": days,
                "tipo_movimiento": str(item.get("TIPO_MOVIMIENTO") or "VACACIONES").strip(),
                "source": str(item.get("ORIGEN") or "normalized_history").strip() or "normalized_history",
                "observaciones": str(item.get("OBSERVACIONES") or "").strip(),
            }
        )

    # Prima historica
    primas_by_name: dict[str, list[dict[str, Any]]] = {}
    if primas_ws is not None:
        p_header = detect_header_row(primas_ws, ["NOMBRE_COMPLETO_NORMALIZADO", "ANIO_PRIMA", "PRIMA_PAGADA"])
        p_rows = _sheet_rows_as_dicts(primas_ws, p_header)
        for item in p_rows:
            name_norm = normalize_name(item.get("NOMBRE_COMPLETO_NORMALIZADO") or item.get("NOMBRE_COMPLETO_ORIGINAL") or "")
            if not name_norm:
                continue
            primas_by_name.setdefault(name_norm, []).append(
                {
                    "anio": str(item.get("ANIO_PRIMA") or "").strip(),
                    "prima_pagada": _to_bool(item.get("PRIMA_PAGADA")),
                    "indicador_original": item.get("INDICADOR_ORIGINAL"),
                    "formula_historica": str(item.get("FORMULA_HISTORICA") or "sueldo_diario_x_3").strip(),
                    "sueldo_diario_base_excel": _to_float(item.get("SUELDO_DIARIO_BASE_EXCEL")),
                    "monto_estimado_historico": _to_float(item.get("MONTO_ESTIMADO_HISTORICO")),
                    "semana_pago": str(item.get("SEMANA_PAGO") or "").strip(),
                    "fecha_pago": _to_date_iso(item.get("FECHA_PAGO")),
                    "comentarios": str(item.get("COMENTARIOS") or "").strip(),
                }
            )

    # DIFERENCIAS_Y_WARNINGS
    warning_detail_by_name: dict[str, list[dict[str, Any]]] = {}
    if diferencias_ws is not None:
        d_header = detect_header_row(diferencias_ws, ["NOMBRE_COMPLETO_ORIGINAL", "SEVERIDAD", "DETALLE"])
        d_rows = _sheet_rows_as_dicts(diferencias_ws, d_header)
        for item in d_rows:
            nombre = str(item.get("NOMBRE_COMPLETO_ORIGINAL") or "").strip()
            if not nombre:
                continue
            key = normalize_name(nombre)
            warning_detail_by_name.setdefault(key, []).append(
                {
                    "severidad": str(item.get("SEVERIDAD") or "warning").strip().lower(),
                    "campo_afectado": str(item.get("CAMPO_AFECTADO") or "").strip(),
                    "detalle": str(item.get("DETALLE") or item.get("TIPO_WARNING") or "").strip(),
                    "accion_recomendada": str(item.get("ACCION_RECOMENDADA") or "").strip(),
                }
            )

    # SALDOS
    s_header = detect_header_row(saldos_ws, ["NOMBRE_COMPLETO_ORIGINAL", "DIAS_UTILIZADOS_EXCEL", "DIAS_RESTANTES_EXCEL"])
    s_rows = _sheet_rows_as_dicts(saldos_ws, s_header)
    if not s_rows:
        raise ValueError("VACACIONES_SALDOS no contiene registros v?lidos")

    rows: list[dict[str, Any]] = []
    total_movs = 0
    total_primas = 0

    for idx, item in enumerate(s_rows, start=s_header + 1):
        nombre = str(item.get("NOMBRE_COMPLETO_ORIGINAL") or "").strip()
        if not nombre:
            errors.append(f"Fila {idx}: columna cr?tica 'nombre_completo_original' vac?a")
            continue

        nombre_norm = normalize_name(item.get("NOMBRE_COMPLETO_NORMALIZADO") or nombre)
        desglose = movimientos_by_name.get(nombre_norm, [])
        suma_movs = round(sum(float(m.get("days") or 0.0) for m in desglose), 4)
        total_movs += len(desglose)

        dias_utilizados_excel = _to_days(item.get("DIAS_UTILIZADOS_EXCEL"))
        vacaciones_laboradas = _to_days(item.get("VACACIONES_LABORADAS_EXCEL"))
        dias_pagados = _to_days(item.get("DIAS_PAGADOS_EXCEL"))
        dias_restantes_excel = _to_float(item.get("DIAS_RESTANTES_EXCEL"))

        conc = conciliar_dias(
            dias_utilizados_excel=dias_utilizados_excel,
            vacaciones_laboradas_excel=vacaciones_laboradas,
            dias_pagados_excel=dias_pagados,
            suma_movimientos_detectada=suma_movs,
            dias_restantes_excel=dias_restantes_excel,
        )

        row_warnings = []
        warning_resumen = str(item.get("WARNING_RESUMEN") or "").strip()
        if warning_resumen:
            row_warnings.append(warning_resumen)
        for d in warning_detail_by_name.get(nombre_norm, []):
            detalle = str(d.get("detalle") or "").strip()
            if detalle:
                row_warnings.append(detalle)
        row_warnings.extend(conc["warnings"])
        if conc["clasificacion_conciliacion"] != "CONCILIADO_COMPLETO":
            row_warnings.append(f"Conciliaci?n: {conc['clasificacion_conciliacion']}")

        fecha_excel = _to_date_iso(item.get("FECHA_INGRESO_EXCEL_CONVERTIDA") or item.get("FECHA_INGRESO_EXCEL_ORIGINAL"))
        fecha_hc = _to_date_iso(item.get("FECHA_INGRESO_HEADCOUNT"))
        fecha_usada = _to_date_iso(item.get("FECHA_INGRESO_USADA")) or (fecha_hc or fecha_excel)
        fuente_fecha = str(item.get("FUENTE_FECHA_INGRESO") or ("HEADCOUNT" if fecha_hc else "EXCEL")).strip() or "EXCEL"

        sueldo_excel = _to_float(item.get("SUELDO_DIARIO_HISTORICO_EXCEL") or item.get("SUELDO_EXCEL_ORIGINAL"))
        sueldo_hc = _to_float(item.get("SUELDO_HEADCOUNT"))
        salario_param = _to_float(item.get("SALARIO_PARAMETROS_NOMINA"))

        prima_det = primas_by_name.get(nombre_norm, [])
        total_primas += len(prima_det)
        prima_2025 = any(str(p.get("anio") or "") == "2025" and bool(p.get("prima_pagada")) for p in prima_det)
        prima_2026 = any(str(p.get("anio") or "") == "2026" and bool(p.get("prima_pagada")) for p in prima_det)
        sem_25 = next((str(p.get("semana_pago") or "") for p in prima_det if str(p.get("anio") or "") == "2025"), "")
        fecha_26 = next((_to_date_iso(p.get("fecha_pago")) for p in prima_det if str(p.get("anio") or "") == "2026"), "")

        row = _build_base_row(
            source_row_number=idx,
            nombre=nombre,
            fecha_ingreso_excel=fecha_excel,
            sueldo_excel=sueldo_excel,
            planta_excel=str(item.get("PLANTA_EXCEL_ORIGINAL") or "").strip(),
            dias_vacaciones_historico=max(0.0, _to_float(item.get("DIAS_DE_VACACIONES_EXCEL")) or 0.0),
            dias_utilizados_excel=dias_utilizados_excel,
            dias_utilizados_semanal=suma_movs,
            vacaciones_laboradas=vacaciones_laboradas,
            dias_pagados=dias_pagados,
            dias_restantes_excel=dias_restantes_excel,
            monto_total_historico=_to_float(item.get("MONTO_TOTAL_EXCEL")),
            comentarios=str(item.get("COMENTARIOS_ORIGINALES") or "").strip(),
            prima_2025_pagada=prima_2025,
            semana_pago_prima_2025=sem_25,
            prima_2026_pagada=prima_2026,
            fecha_pago_prima_2026=fecha_26,
            desglose_semanal=desglose,
            clasificacion_conciliacion=conc["clasificacion_conciliacion"],
            diferencia_detectada=conc["diferencia_detectada"],
            fuente_fecha_ingreso=fuente_fecha,
            sueldo_headcount_in=sueldo_hc,
            salario_parametros_nomina=salario_param,
            warnings=row_warnings,
            extra_editable={
                "prima_historica_detalle": prima_det,
                "warnings_detalle": warning_detail_by_name.get(nombre_norm, []),
                "import_kind": IMPORT_KIND_NORMALIZED,
                "origen_normalizado": {
                    "id_importacion": item.get("ID_IMPORTACION"),
                    "clasificacion_conciliacion_origen": item.get("CLASIFICACION_CONCILIACION"),
                    "requiere_revision_manual": item.get("REQUIERE_REVISION_MANUAL"),
                    "fuente_salario_origen": item.get("FUENTE_SALARIO"),
                },
            },
        )

        # conservar estatus/planta de HC si ya vienen normalizados
        row["planta_headcount"] = str(item.get("PLANTA_HEADCOUNT") or "").strip()
        row["status_headcount"] = str(item.get("ESTATUS_HEADCOUNT") or "SIN STATUS HEADCOUNT").strip() or "SIN STATUS HEADCOUNT"
        row["estatus_headcount"] = row["status_headcount"]
        row["fecha_ingreso_headcount"] = fecha_hc
        row["fecha_ingreso_usada"] = fecha_usada
        rows.append(row)
        warnings.extend([f"Fila {idx}: {w}" for w in row_warnings])

    preview = {
        "tipo_archivo": IMPORT_KIND_NORMALIZED,
        "total_trabajadores": len(rows),
        "total_movimientos": total_movs,
        "total_primas": total_primas,
        "total_warnings": len(warnings),
    }

    return ParsedVacaciones(
        rows=rows,
        warnings=warnings,
        errors=errors,
        cliente="Carrier",
        weekly_events_total=total_movs,
        import_kind=IMPORT_KIND_NORMALIZED,
        preview=preview,
    )


def parse_legacy_history(wb, filename: str) -> ParsedVacaciones:
    # Compatibilidad m?nima con el layout hist?rico previo.
    ws = detect_sheet_by_name(wb, ["VACACIONES", "VACACIONES_HISTORICAS"]) or wb[wb.sheetnames[0]]
    warnings: list[str] = []
    errors: list[str] = []

    header_row = detect_header_row(ws, ["NOMBRE", "FECHA DE INGRESO", "DIAS UTILIZADOS"])
    cols = build_column_index(ws, header_row)

    c_nombre = resolve_column(cols, ["NOMBRE", "NOMBRE COMPLETO"], required=True, warnings=warnings, sheet_label=ws.title)
    c_fecha = resolve_column(cols, ["FECHA DE INGRESO"], required=False, warnings=warnings, sheet_label=ws.title)
    c_sueldo = resolve_column(cols, ["SUELDO", "SUELDO DIARIO"], required=False, warnings=warnings, sheet_label=ws.title)
    c_planta = resolve_column(cols, ["PLANTA"], required=False, warnings=warnings, sheet_label=ws.title)
    c_du = resolve_column(cols, ["DIAS UTILIZADOS"], required=False, warnings=warnings, sheet_label=ws.title)
    c_vl = resolve_column(cols, ["VACACIONES LABORADAS"], required=False, warnings=warnings, sheet_label=ws.title)
    c_dp = resolve_column(cols, ["DIAS PAGADOS"], required=False, warnings=warnings, sheet_label=ws.title)
    c_dr = resolve_column(cols, ["DIAS RESTANTES"], required=False, warnings=warnings, sheet_label=ws.title)
    c_monto = resolve_column(cols, ["MONTO TOTAL"], required=False, warnings=warnings, sheet_label=ws.title)
    c_com = resolve_column(cols, ["COMENTARIOS"], required=False, warnings=warnings, sheet_label=ws.title)

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        nombre = str(ws.cell(row=row_idx, column=c_nombre).value or "").strip() if c_nombre else ""
        if not nombre:
            continue

        dias_utilizados_excel = _to_days(ws.cell(row=row_idx, column=c_du).value if c_du else None)
        vacaciones_laboradas = _to_days(ws.cell(row=row_idx, column=c_vl).value if c_vl else None)
        dias_pagados = _to_days(ws.cell(row=row_idx, column=c_dp).value if c_dp else None)
        dias_restantes_excel = _to_float(ws.cell(row=row_idx, column=c_dr).value if c_dr else None)

        conc = conciliar_dias(
            dias_utilizados_excel=dias_utilizados_excel,
            vacaciones_laboradas_excel=vacaciones_laboradas,
            dias_pagados_excel=dias_pagados,
            suma_movimientos_detectada=0.0,
            dias_restantes_excel=dias_restantes_excel,
        )

        row = _build_base_row(
            source_row_number=row_idx,
            nombre=nombre,
            fecha_ingreso_excel=_to_date_iso(ws.cell(row=row_idx, column=c_fecha).value if c_fecha else None),
            sueldo_excel=_to_float(ws.cell(row=row_idx, column=c_sueldo).value if c_sueldo else None),
            planta_excel=str(ws.cell(row=row_idx, column=c_planta).value or "").strip() if c_planta else "",
            dias_vacaciones_historico=max(0.0, dias_utilizados_excel + (dias_restantes_excel or 0.0)),
            dias_utilizados_excel=dias_utilizados_excel,
            dias_utilizados_semanal=0.0,
            vacaciones_laboradas=vacaciones_laboradas,
            dias_pagados=dias_pagados,
            dias_restantes_excel=dias_restantes_excel,
            monto_total_historico=_to_float(ws.cell(row=row_idx, column=c_monto).value if c_monto else None),
            comentarios=str(ws.cell(row=row_idx, column=c_com).value or "").strip() if c_com else "",
            prima_2025_pagada=False,
            semana_pago_prima_2025="",
            prima_2026_pagada=False,
            fecha_pago_prima_2026="",
            desglose_semanal=[],
            clasificacion_conciliacion=conc["clasificacion_conciliacion"],
            diferencia_detectada=conc["diferencia_detectada"],
            fuente_fecha_ingreso="EXCEL",
            warnings=conc["warnings"],
            extra_editable={"import_kind": IMPORT_KIND_LEGACY},
        )
        rows.append(row)

    return ParsedVacaciones(
        rows=rows,
        warnings=warnings,
        errors=errors,
        cliente="Carrier",
        weekly_events_total=0,
        import_kind=IMPORT_KIND_LEGACY,
        preview={
            "tipo_archivo": IMPORT_KIND_LEGACY,
            "total_trabajadores": len(rows),
            "total_movimientos": 0,
            "total_primas": 0,
            "total_warnings": len(warnings),
        },
    )


def _detect_import_kind(wb) -> str:
    names = {normalize_sheet_name(name) for name in wb.sheetnames}
    if "VACACIONES_SALDOS" in names and "VACACIONES_MOVIMIENTOS" in names:
        return IMPORT_KIND_NORMALIZED
    if "1_CAPTURA_SALDOS" in names:
        return IMPORT_KIND_SIMPLE
    return IMPORT_KIND_LEGACY


def parse_vacaciones_historico_excel(
    file_bytes: bytes,
    filename: str,
    import_kind: str | None = None,
) -> ParsedVacaciones:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    kind = (import_kind or "AUTO").strip().upper()
    if kind == "AUTO":
        kind = _detect_import_kind(wb)

    if kind == IMPORT_KIND_SIMPLE:
        return parse_simple_template(wb, filename)
    if kind == IMPORT_KIND_NORMALIZED:
        return parse_normalized_history(wb, filename)
    return parse_legacy_history(wb, filename)
