from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from typing import Iterable
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.nomina.config import get_holidays_for_year
from modules.nomina.db import NominaBaseRow

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_CANDIDATES = (
    BASE_DIR / "templates_excel" / "Plantilla_asistencia_ProC.xlsx",
    BASE_DIR / "templates_excel" / "Plantilla_asistencia_v4.xlsx",
)

# --- v4 column layout (1-based) ---
COL_NOMBRE = 1   # A
COL_CLIENTE = 2  # B
COL_PLANTA = 3   # C
COL_PUESTO = 4   # D
COL_BANCO = 5    # E
COL_CUENTA = 6   # F
COL_DAY_START = 7  # G..M = 7 days (cols 7..13)
COL_DAY_END = 13
COL_HORAS_EXTRA = 14            # N
COL_HORAS_EXTRA_NORMALES = 15   # O
COL_DIAS_CUBIERTOS = 16         # P
COL_VACACIONES_LABORADAS = 17   # Q
COL_PRIMA_VACACIONAL = 18       # R
COL_BONO = 19                   # S
COL_DEDUCCIONES = 20            # T
COL_OBSERVACIONES = 21          # U

START_DATA_ROW = 5
DAILY_HEADER_ROW = 4
MONTH_ABBR_ES = {
    1: "ene",
    2: "feb",
    3: "mar",
    4: "abr",
    5: "may",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "sep",
    10: "oct",
    11: "nov",
    12: "dic",
}


def _resolve_template_xlsx_path() -> Path:
    for path in TEMPLATE_CANDIDATES:
        if path.exists():
            return path
    first = TEMPLATE_CANDIDATES[0]
    raise FileNotFoundError(f"No existe plantilla base: {first}")


def _norm_header(value: str) -> str:
    text = " ".join(str(value or "").replace("\u00a0", " ").replace("\n", " ").upper().split()).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return " ".join(text.split()).strip()


def _build_header_index(ws: Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, (ws.max_column or COL_OBSERVACIONES) + 1):
        key = _norm_header(ws.cell(row=DAILY_HEADER_ROW, column=col).value)
        if key and key not in out:
            out[key] = col
    return out


def _resolve_columns(ws: Worksheet) -> dict[str, int]:
    headers = _build_header_index(ws)

    def first(names: tuple[str, ...], fallback: int | None = None) -> int | None:
        for name in names:
            col = headers.get(_norm_header(name))
            if col is not None:
                return col
        return fallback

    resolved: dict[str, int] = {}
    for key, names, fallback in (
        ("nombre_empleado", ("NOMBRE DE EMPLEADO",), COL_NOMBRE),
        ("cliente", ("CLIENTE",), COL_CLIENTE),
        ("planta", ("PLANTA",), COL_PLANTA),
        ("puesto", ("PUESTO",), COL_PUESTO),
        ("banco", ("BANCO",), COL_BANCO),
        ("cuenta", ("CUENTA",), COL_CUENTA),
        ("nss", ("NSS",), None),
        ("numero_empleado", ("NUMERO DE EMPLEADO", "NUM EMPLEADO", "NO EMPLEADO", "NUMERO EMPLEADO"), None),
        ("horas_extra", ("HORAS EXTRA",), COL_HORAS_EXTRA),
        ("horas_extra_normales", ("HORAS EXTRA NORMALES",), COL_HORAS_EXTRA_NORMALES),
        ("dias_cubiertos_normales", ("DIAS CUBIERTOS NORMALES",), COL_DIAS_CUBIERTOS),
        ("vacaciones_laboradas", ("VACACIONES LABORADAS",), COL_VACACIONES_LABORADAS),
        ("prima_vacacional", ("PRIMA VACACIONAL",), COL_PRIMA_VACACIONAL),
        ("bono", ("BONO",), COL_BONO),
        ("deducciones", ("DEDUCCIONES",), COL_DEDUCCIONES),
        ("observaciones", ("OBSERVACIONES",), COL_OBSERVACIONES),
    ):
        col = first(names, fallback)
        if col is not None:
            resolved[key] = col
    return resolved


def format_period_label(fecha_inicio: date, fecha_fin: date) -> str:
    if fecha_inicio.year == fecha_fin.year and fecha_inicio.month == fecha_fin.month:
        return f"{fecha_inicio.day} al {fecha_fin.day} {MONTH_ABBR_ES[fecha_inicio.month]} {fecha_inicio.year}"
    if fecha_inicio.year == fecha_fin.year:
        return (
            f"{fecha_inicio.day:02d} {MONTH_ABBR_ES[fecha_inicio.month]} "
            f"al {fecha_fin.day:02d} {MONTH_ABBR_ES[fecha_fin.month]} {fecha_inicio.year}"
        )
    return (
        f"{fecha_inicio.day:02d} {MONTH_ABBR_ES[fecha_inicio.month]} {fecha_inicio.year} "
        f"al {fecha_fin.day:02d} {MONTH_ABBR_ES[fecha_fin.month]} {fecha_fin.year}"
    )


def format_period_slug(fecha_inicio: date, fecha_fin: date) -> str:
    if fecha_inicio.year == fecha_fin.year and fecha_inicio.month == fecha_fin.month:
        return f"{fecha_inicio.day}_al_{fecha_fin.day}_{MONTH_ABBR_ES[fecha_inicio.month]}_{fecha_inicio.year}"
    if fecha_inicio.year == fecha_fin.year:
        return (
            f"{fecha_inicio.day:02d}_{MONTH_ABBR_ES[fecha_inicio.month]}_al_"
            f"{fecha_fin.day:02d}_{MONTH_ABBR_ES[fecha_fin.month]}_{fecha_inicio.year}"
        )
    return (
        f"{fecha_inicio.day:02d}_{MONTH_ABBR_ES[fecha_inicio.month]}_{fecha_inicio.year}_al_"
        f"{fecha_fin.day:02d}_{MONTH_ABBR_ES[fecha_fin.month]}_{fecha_fin.year}"
    )


def week_label(fecha_inicio: date, fecha_fin: date) -> str:
    return format_period_label(fecha_inicio, fecha_fin)


def build_daily_headers(fecha_inicio: date) -> list[str]:
    day_codes = {0: "L", 1: "M", 2: "M", 3: "J", 4: "V", 5: "S", 6: "D"}
    out: list[str] = []
    for i in range(7):
        d = fecha_inicio + timedelta(days=i)
        out.append(f"{day_codes[d.weekday()]}{d.day}")
    return out


def _holidays_in_range(fecha_inicio: date) -> dict[int, date]:
    """Return mapping {day_index_in_period (0..6) -> date} for official holidays."""
    out: dict[int, date] = {}
    for i in range(7):
        d = fecha_inicio + timedelta(days=i)
        holidays = get_holidays_for_year(d.year)
        if d in holidays:
            out[i] = d
    return out


def _write_label_with_value(ws: Worksheet, label: str, value: str) -> None:
    """v4: SEMANA:/CLIENTE:/COORDINADOR: labels live in merged top cell.
    Write 'LABEL: VALUE' into the same cell so it renders inside the merged box.
    """
    target = (label or "").strip()
    if not target.endswith(":"):
        target += ":"
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            if text.split()[0].upper().rstrip(":") + ":" == target.upper():
                ws.cell(row=row, column=col, value=f"{target} {value}")
                return
            if text.upper() == target.upper():
                ws.cell(row=row, column=col, value=f"{target} {value}")
                return


def _write_superior_fields(ws: Worksheet, semana: str, cliente: str, coordinador: str) -> None:
    _write_label_with_value(ws, "SEMANA", semana)
    _write_label_with_value(ws, "CLIENTE", cliente)
    _write_label_with_value(ws, "COORDINADOR", coordinador)


def _write_daily_headers(ws: Worksheet, headers: list[str]) -> None:
    for idx, header in enumerate(headers):
        ws.cell(row=DAILY_HEADER_ROW, column=COL_DAY_START + idx, value=header)


def _write_base_rows(
    ws: Worksheet,
    base_rows: Iterable[NominaBaseRow],
    holiday_day_indices: list[int],
) -> int:
    cols = _resolve_columns(ws)
    written = 0
    for i, item in enumerate(base_rows):
        row = START_DATA_ROW + i
        ws.cell(row=row, column=cols["nombre_empleado"], value=item.nombre_empleado or "")
        ws.cell(row=row, column=cols["cliente"], value=item.cliente or "")
        ws.cell(row=row, column=cols["planta"], value=item.planta or "")
        ws.cell(row=row, column=cols["puesto"], value=item.puesto or "")
        ws.cell(row=row, column=cols["banco"], value=item.banco or "")
        ws.cell(row=row, column=cols["cuenta"], value=item.cuenta or "")
        nss_col = cols.get("nss")
        if nss_col is not None:
            ws.cell(row=row, column=nss_col, value=item.nss or "")
        num_emp_col = cols.get("numero_empleado")
        if num_emp_col is not None:
            ws.cell(row=row, column=num_emp_col, value=item.numero_empleado or "")
        for idx in range(7):
            col = COL_DAY_START + idx
            ws.cell(row=row, column=col, value=("FE" if idx in holiday_day_indices else ""))
        for variable_col in (
            cols.get("horas_extra"),
            cols.get("horas_extra_normales"),
            cols.get("dias_cubiertos_normales"),
            cols.get("vacaciones_laboradas"),
            cols.get("prima_vacacional"),
            cols.get("bono"),
            cols.get("deducciones"),
            cols.get("observaciones"),
        ):
            if variable_col is not None:
                ws.cell(row=row, column=variable_col, value="")
        written += 1
    return written


def build_asistencia_template_file(
    *,
    fecha_inicio: date,
    fecha_fin: date,
    cliente: str,
    coordinador: str,
    base_rows: Iterable[NominaBaseRow],
) -> bytes:
    wb = load_workbook(_resolve_template_xlsx_path())
    if "Asistencia" not in wb.sheetnames:
        raise ValueError("La plantilla base no contiene la hoja 'Asistencia'.")
    ws = wb["Asistencia"]
    _write_superior_fields(ws, week_label(fecha_inicio, fecha_fin), cliente.strip(), coordinador.strip())
    _write_daily_headers(ws, build_daily_headers(fecha_inicio))
    holiday_map = _holidays_in_range(fecha_inicio)
    _write_base_rows(ws, base_rows, sorted(holiday_map.keys()))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
