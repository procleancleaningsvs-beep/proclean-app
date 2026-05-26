"""Parser de Excel de nómina actual (tipo Carrier / Pepsi) para extraer
parámetros base por empleado. Microfase 4.0.

NO leer ni guardar: ISR, NETO, BONO TPT, PRIMA EFICIENCIA, BASE GRAVADA,
INFONAVIT ni fórmulas de cálculo. Solo parámetros base + LOCALIDAD/FRONTERA.
"""
from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").replace("\n", " ").split()).strip()


def _norm_header(value: Any) -> str:
    raw = _norm_text(value).upper()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    raw = raw.replace(".", " ").replace(",", " ")
    return " ".join(raw.split())


def _norm_name_for_match(value: Any) -> str:
    """Stable comparable name: uppercased, no accents, alpha-num-space only."""
    raw = _norm_text(value).upper()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    raw = "".join(ch if (ch.isalnum() or ch == " ") else " " for ch in raw)
    return " ".join(raw.split())


def _norm_locality(value: Any) -> str:
    """Locality normalized: lowercase, no accents, single spaces."""
    raw = _norm_text(value)
    if not raw:
        return ""
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    return raw.lower().strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    s = str(value).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_bool_frontera(value: Any) -> bool | None:
    """Interpret VERDADERO/FALSO/TRUE/FALSE/SI/NO. Returns None if unrecognized."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        try:
            return bool(int(value))
        except (TypeError, ValueError):
            return None
    s = _norm_header(value)
    if not s:
        return None
    if s in {"VERDADERO", "TRUE", "SI", "SI ", "1", "FRONTERA"}:
        return True
    if s in {"FALSO", "FALSE", "NO", "0", "GENERAL"}:
        return False
    return None


HEADER_ALIASES: dict[str, list[str]] = {
    "numero_empleado": ["NO", "NO.", "NUMERO", "NUMERO EMPLEADO", "NUM EMPLEADO", "CODIGO", "CLAVE"],
    "nombre_empleado": ["NOMBRE", "NOMBRE DE EMPLEADO", "NOMBRE EMPLEADO", "NOMBRE COMPLETO"],
    "cliente": ["CLIENTE"],
    "planta": ["PLANTA"],
    "puesto": ["PUESTO"],
    "banco": ["BANCO"],
    "cuenta": ["CUENTA", "CTA", "NO CUENTA", "CUENTA BANCO"],
    "salario_operativo": ["SALARIO OPERATIVO", "SUELDO OPERATIVO"],
    "valor_x_he": [
        "VALOR X HE",
        "VALOR X HE2",
        "VALOR HE",
        "VALOR DE HE",
        "VALOR DE HORA EXTRA",
    ],
    "horas_extra": ["HORAS EXTRA", "HE", "HORAS EXTRAS"],
    "localidad": ["LOCALIDAD", "LOCALIDAD/PLANTA"],
    "frontera": ["FRONTERA", "ZONA FRONTERA", "ES FRONTERA"],
    "nss": ["NSS", "NUM SEGURIDAD SOCIAL", "NUMERO DE SEGURO SOCIAL"],
}


KNOWN_CLIENTE_PATTERNS: list[tuple[str, str]] = [
    ("carrier", "Carrier"),
    ("pepsi", "Pepsi"),
    ("gepp", "GEPP"),
    ("auriga", "Auriga"),
    ("general motors", "GM"),
    (" gm", "GM"),
    ("gm ", "GM"),
]


def detect_cliente_from_import(
    *,
    filename: str,
    sheet_name: str | None,
    row_clientes: list[str],
    fallback: str = "",
) -> tuple[str | None, str | None]:
    """Detecta cliente desde filas, hoja, archivo o patrones conocidos.

    Returns (cliente, detection_source) or (None, None) if undetected.
    """
    unique_rows = sorted({c.strip() for c in row_clientes if c and c.strip()}, key=str.lower)
    if len(unique_rows) == 1:
        return unique_rows[0], "columna_excel"
    if len(unique_rows) > 1:
        return None, "multicliente"

    haystack = " ".join(
        part for part in (filename or "", sheet_name or "", fallback or "") if part
    ).lower()
    for pattern, label in KNOWN_CLIENTE_PATTERNS:
        if pattern in haystack:
            return label, "nombre_archivo_o_hoja"
    if fallback.strip():
        return fallback.strip(), "seleccion_manual"
    return None, None


@dataclass
class HeaderMap:
    row: int
    cols: dict[str, int] = field(default_factory=dict)


def _find_header_row(ws: Worksheet, max_scan_rows: int = 12) -> HeaderMap | None:
    """Find the most-likely header row by counting aliases hits per row."""
    target_aliases = {
        "nombre_empleado", "salario_operativo", "valor_x_he", "puesto",
    }
    best: HeaderMap | None = None
    best_hits = 0
    max_row = min(ws.max_row or max_scan_rows, max_scan_rows)
    for r in range(1, max_row + 1):
        cols: dict[str, int] = {}
        for c in range(1, (ws.max_column or 40) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            normalized = _norm_header(v)
            for field_key, aliases in HEADER_ALIASES.items():
                if field_key in cols:
                    continue
                for alias in aliases:
                    if normalized == _norm_header(alias):
                        cols[field_key] = c
                        break
        hits = sum(1 for k in target_aliases if k in cols)
        if hits > best_hits:
            best = HeaderMap(row=r, cols=cols)
            best_hits = hits
            if hits >= 3 and "nombre_empleado" in cols and ("salario_operativo" in cols or "valor_x_he" in cols):
                pass  # keep scanning, we may find a richer one later
    if best and "nombre_empleado" in best.cols and ("salario_operativo" in best.cols or "valor_x_he" in best.cols):
        return best
    return None


def _row_is_empty(ws: Worksheet, row: int, hm: HeaderMap) -> bool:
    relevant_cols = list(hm.cols.values()) or [1, 2, 3]
    for col in relevant_cols:
        if _norm_text(ws.cell(row=row, column=col).value):
            return False
    return True


def parse_nomina_actual(
    file_bytes: bytes,
    *,
    filename: str,
    cliente_hint: str = "",
    sheet_hint: str | None = None,
) -> dict[str, Any]:
    """Parse a current payroll Excel (one or multiple sheets) and return base params."""
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo abrir el Excel: {exc}") from exc

    warnings: list[str] = []
    errors: list[str] = []
    rows_out: list[dict[str, Any]] = []
    localidades_frontera_out: list[dict[str, Any]] = []
    sheets_used: list[str] = []

    candidate_sheets: list[str]
    if sheet_hint and sheet_hint in wb.sheetnames:
        candidate_sheets = [sheet_hint]
    else:
        candidate_sheets = list(wb.sheetnames)

    chosen_sheet: str | None = None
    chosen_hm: HeaderMap | None = None
    for sn in candidate_sheets:
        ws = wb[sn]
        hm = _find_header_row(ws)
        if hm is None:
            continue
        chosen_sheet = sn
        chosen_hm = hm
        break

    if chosen_sheet is None or chosen_hm is None:
        raise ValueError(
            "No se detectó hoja con encabezados mínimos (NOMBRE + SALARIO OPERATIVO o VALOR X HE)."
        )

    ws = wb[chosen_sheet]
    sheets_used.append(chosen_sheet)

    missing_required: list[str] = []
    if "nombre_empleado" not in chosen_hm.cols:
        missing_required.append("NOMBRE DE EMPLEADO")
    if "salario_operativo" not in chosen_hm.cols and "valor_x_he" not in chosen_hm.cols:
        missing_required.append("SALARIO OPERATIVO o VALOR X HE")
    if missing_required:
        raise ValueError(
            "Faltan columnas mínimas obligatorias: " + ", ".join(missing_required)
        )

    if "salario_operativo" not in chosen_hm.cols:
        warnings.append("Hoja no incluye SALARIO OPERATIVO; se importará sin salario.")
    if "valor_x_he" not in chosen_hm.cols:
        warnings.append("Hoja no incluye VALOR X HE; se importará sin valor de HE.")

    prelim_cliente, _ = detect_cliente_from_import(
        filename=filename,
        sheet_name=chosen_sheet,
        row_clientes=[],
        fallback=cliente_hint,
    )
    if "localidad" not in chosen_hm.cols and prelim_cliente and "pepsi" in prelim_cliente.lower():
        warnings.append("Cliente Pepsi sin columna LOCALIDAD detectada.")

    nss_seen: set[str] = set()
    numero_seen: dict[str, str] = {}
    nss_to_names: dict[str, set[str]] = {}
    name_to_clients: dict[str, set[str]] = {}

    for r in range(chosen_hm.row + 1, (ws.max_row or 0) + 1):
        if _row_is_empty(ws, r, chosen_hm):
            continue
        nombre_raw = _norm_text(ws.cell(row=r, column=chosen_hm.cols["nombre_empleado"]).value)
        if not nombre_raw:
            continue
        # Skip section headers like 'PERSONAL ACTIVO' etc.
        if _norm_header(nombre_raw) in {"PERSONAL ACTIVO", "PERSONAL BAJA", "INFORMACION DEL EMPLEADO", "TOTAL"}:
            continue

        def _get(field_key: str) -> Any:
            if field_key not in chosen_hm.cols:
                return None
            return ws.cell(row=r, column=chosen_hm.cols[field_key]).value

        row_warnings: list[str] = []
        nombre_norm = _norm_name_for_match(nombre_raw)
        numero = _norm_text(_get("numero_empleado"))
        nss = _norm_text(_get("nss"))
        cliente_val = _norm_text(_get("cliente"))
        planta = _norm_text(_get("planta"))
        puesto = _norm_text(_get("puesto"))
        banco = _norm_text(_get("banco"))
        cuenta = _norm_text(_get("cuenta"))
        salario_op = _to_float(_get("salario_operativo"))
        valor_he = _to_float(_get("valor_x_he"))
        horas_extra_periodo = _to_float(_get("horas_extra"))
        localidad_raw = _norm_text(_get("localidad"))
        localidad_norm = _norm_locality(localidad_raw)
        frontera_raw_val = _get("frontera")
        frontera_bool = _to_bool_frontera(frontera_raw_val)

        if salario_op is None:
            row_warnings.append("Empleado sin SALARIO OPERATIVO.")
        elif salario_op <= 0:
            row_warnings.append("SALARIO OPERATIVO cero o negativo.")
        if valor_he is None:
            row_warnings.append("Empleado sin VALOR X HE.")
        elif valor_he <= 0:
            row_warnings.append("VALOR X HE cero o negativo.")
        if numero:
            if numero in numero_seen and numero_seen[numero] != nombre_norm:
                row_warnings.append(
                    f"Número de empleado duplicado con otra persona: {numero} ya usado por '{numero_seen[numero]}'."
                )
            else:
                numero_seen[numero] = nombre_norm
        if nss:
            nss_to_names.setdefault(nss, set()).add(nombre_norm)
            if len(nss_to_names[nss]) > 1:
                row_warnings.append(f"NSS {nss} aparece con varios nombres: {sorted(nss_to_names[nss])}.")
            nss_seen.add(nss)
        if cliente_val:
            name_to_clients.setdefault(nombre_norm, set()).add(_norm_header(cliente_val))
            if len(name_to_clients[nombre_norm]) > 1:
                row_warnings.append(
                    f"Empleado aparece en múltiples clientes: {sorted(name_to_clients[nombre_norm])}."
                )

        if cliente_val and cliente_val.strip().lower() == "pepsi" and not localidad_raw:
            row_warnings.append("Cliente Pepsi sin LOCALIDAD en este renglón.")

        if frontera_bool is True and not localidad_raw:
            row_warnings.append("FRONTERA TRUE pero LOCALIDAD vacía.")
        elif frontera_raw_val is not None and frontera_bool is None:
            row_warnings.append(f"Valor de FRONTERA no reconocido: {frontera_raw_val!r}.")

        if localidad_raw:
            localidades_frontera_out.append({
                "cliente": cliente_val or (prelim_cliente or ""),
                "localidad": localidad_raw,
                "localidad_normalizada": localidad_norm,
                "es_frontera": bool(frontera_bool) if frontera_bool is not None else None,
                "source_filename": filename,
            })

        rows_out.append({
            "row_number": r,
            "nombre": nombre_raw,
            "nombre_normalizado": nombre_norm,
            "numero_empleado": numero or None,
            "nss": nss or None,
            "cliente": cliente_val or None,
            "planta": planta or None,
            "puesto": puesto or None,
            "banco": banco or None,
            "cuenta": cuenta or None,
            "salario_operativo": salario_op,
            "valor_x_he": valor_he,
            "horas_extra_periodo": horas_extra_periodo,
            "localidad": localidad_raw or None,
            "localidad_normalizada": localidad_norm or None,
            "frontera_raw": (
                "VERDADERO" if frontera_bool is True
                else ("FALSO" if frontera_bool is False else (str(frontera_raw_val) if frontera_raw_val is not None else None))
            ),
            "es_frontera": frontera_bool,
            "warnings": row_warnings,
        })

    # Cross-row deduped warnings already accumulated. Add file-level warnings.
    if not rows_out:
        errors.append("No se detectaron filas de empleados con datos válidos.")

    row_clientes = [str(r.get("cliente") or "") for r in rows_out]
    detected_cliente, detection_source = detect_cliente_from_import(
        filename=filename,
        sheet_name=chosen_sheet,
        row_clientes=row_clientes,
        fallback=cliente_hint,
    )
    if detection_source == "multicliente":
        warnings.append("Archivo con múltiples clientes detectados en columna CLIENTE.")
    elif detected_cliente is None:
        warnings.append("cliente_no_detectado")
    else:
        for row in rows_out:
            if not row.get("cliente"):
                row["cliente"] = detected_cliente

    return {
        "tipo_importacion": "NOMINA_ACTUAL",
        "filename": filename,
        "cliente": detected_cliente or cliente_hint or "",
        "cliente_detectado": detected_cliente,
        "cliente_detection_source": detection_source,
        "cliente_requiere_seleccion": detected_cliente is None and detection_source != "multicliente",
        "sheet": chosen_sheet,
        "rows": rows_out,
        "localidades": localidades_frontera_out,
        "warnings": warnings,
        "errors": errors,
        "total_rows": len(rows_out),
    }
