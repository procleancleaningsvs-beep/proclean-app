"""Validador de plantillas Excel canónicas (sin escritura de datos laborales)."""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from modules.gestion_idse_sua.template_contract import (
    COMPARATIVO_REQUIRED_HEADERS,
    COMPARATIVO_SHA256,
    COMPARATIVO_SHEETS,
    HEADER_ROW_BY_SHEET,
    MENSUAL_REQUIRED_HEADERS,
    MENSUAL_SHA256,
    MENSUAL_SHEETS,
    comparativo_path,
    mensual_path,
)


class TemplateValidationError(ValueError):
    """Plantilla Excel inválida o incompleta."""


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _row_headers(ws, row: int) -> list[str]:
    vals = [ws.cell(row, c).value for c in range(1, (ws.max_column or 1) + 1)]
    while vals and vals[-1] is None:
        vals.pop()
    return [str(v).strip() if v is not None else "" for v in vals]


def _scan_ref_errors(ws, max_rows: int = 120, max_cols: int = 50) -> int:
    count = 0
    for row in ws.iter_rows(max_row=min(ws.max_row or 1, max_rows), max_col=min(ws.max_column or 1, max_cols)):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and "#REF!" in value:
                count += 1
    return count


def validate_workbook(
    path: Path,
    *,
    expected_sheets: tuple[str, ...],
    required_headers: dict[str, tuple[str, ...]],
    expected_sha256: str | None = None,
) -> dict[str, Any]:
    if not path.is_file():
        raise TemplateValidationError(f"No existe la plantilla: {path}")

    digest = sha256_file(path)
    if expected_sha256 and digest != expected_sha256:
        raise TemplateValidationError(
            f"SHA-256 inesperado para {path.name}: {digest} (esperado {expected_sha256})"
        )

    try:
        wb = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001 — superficie de archivo corrupto
        raise TemplateValidationError(f"No se pudo abrir {path.name}: {exc}") from exc

    missing_sheets = [name for name in expected_sheets if name not in wb.sheetnames]
    if missing_sheets:
        raise TemplateValidationError(
            f"{path.name}: faltan hojas {missing_sheets}. Encontradas: {wb.sheetnames}"
        )

    header_map: dict[str, list[str]] = {}
    for sheet_name, expected_cols in required_headers.items():
        ws = wb[sheet_name]
        row = HEADER_ROW_BY_SHEET.get(sheet_name, 6)
        headers = _row_headers(ws, row)
        header_map[sheet_name] = headers
        missing = [col for col in expected_cols if col not in headers]
        if missing:
            raise TemplateValidationError(
                f"{path.name} / {sheet_name}: faltan columnas {missing}"
            )
        refs = _scan_ref_errors(ws)
        if refs:
            raise TemplateValidationError(
                f"{path.name} / {sheet_name}: se encontraron {refs} fórmulas con #REF!"
            )

    return {
        "path": str(path),
        "sha256": digest,
        "sheets": list(wb.sheetnames),
        "headers": header_map,
        "ok": True,
    }


def validate_comparativo_template(path: Path | None = None) -> dict[str, Any]:
    return validate_workbook(
        path or comparativo_path(),
        expected_sheets=COMPARATIVO_SHEETS,
        required_headers=COMPARATIVO_REQUIRED_HEADERS,
        expected_sha256=COMPARATIVO_SHA256 if path is None else None,
    )


def validate_mensual_template(path: Path | None = None) -> dict[str, Any]:
    return validate_workbook(
        path or mensual_path(),
        expected_sheets=MENSUAL_SHEETS,
        required_headers=MENSUAL_REQUIRED_HEADERS,
        expected_sha256=MENSUAL_SHA256 if path is None else None,
    )


def validate_all_canonical_templates() -> dict[str, Any]:
    return {
        "comparativo": validate_comparativo_template(),
        "mensual": validate_mensual_template(),
    }
