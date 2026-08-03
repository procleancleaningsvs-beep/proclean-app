from __future__ import annotations

import sqlite3
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from modules.gestion_idse_sua.nominas import repository as repo
from modules.gestion_idse_sua.nominas.attendance_parser import extract_attendance_for_workers
from modules.gestion_idse_sua.nominas.attendance_service import persist_attendance
from modules.gestion_idse_sua.nominas.period_parser import merge_cut_warning, parse_manual_period
from modules.gestion_idse_sua.nominas.planta_cliente_service import suggest_cliente_for_planta
from modules.gestion_idse_sua.nominas.sheet_inspector import inspect_sheet
from modules.gestion_idse_sua.nominas.text_utils import file_sha256, json_dumps
from modules.gestion_idse_sua.nominas.worker_extractor import extract_workers


def inspect_workbook(file_bytes: bytes, *, filename: str) -> dict[str, Any]:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError(f"Archivo Excel inválido o corrupto: {exc}") from exc

    sheets: list[dict[str, Any]] = []
    for index, name in enumerate(wb.sheetnames):
        ws = wb[name]
        state = getattr(ws, "sheet_state", "visible")
        is_hidden = state != "visible"
        sheets.append(inspect_sheet(ws, sheet_name=name, sheet_index=index, is_hidden=is_hidden))
    wb.close()
    return {
        "filename": filename,
        "file_hash": file_sha256(file_bytes),
        "sheets": sheets,
    }


def register_import(
    conn: sqlite3.Connection,
    *,
    file_bytes: bytes,
    filename: str,
    uploaded_by: str | None,
) -> dict[str, Any]:
    inspection = inspect_workbook(file_bytes, filename=filename)
    import_id = repo.create_import(
        conn,
        original_filename=filename,
        file_hash=inspection["file_hash"],
        uploaded_by=uploaded_by,
        file_content=file_bytes,
    )
    sheet_ids: list[int] = []
    for sheet in inspection["sheets"]:
        sheet_ids.append(repo.add_sheet(conn, import_id, sheet))
    repo.set_import_status(conn, import_id, "uploaded")
    conn.commit()
    return {
        "import_id": import_id,
        "filename": filename,
        "file_hash": inspection["file_hash"],
        "sheets": repo.list_sheets(conn, import_id),
    }


def confirm_classifications(
    conn: sqlite3.Connection,
    import_id: int,
    classifications: dict[int, str],
) -> None:
    for sheet_id, klass in classifications.items():
        if klass not in {"nomina", "auxiliar", "ignorada"}:
            raise ValueError(f"Clasificación inválida: {klass}")
        repo.update_sheet_classification(conn, int(sheet_id), klass)
    repo.set_import_status(conn, import_id, "classified")
    conn.commit()


def confirm_period(
    conn: sqlite3.Connection,
    sheet_id: int,
    *,
    fecha_inicio: str,
    fecha_fin: str,
    cliente: str | None = None,
    confirmed: bool = True,
    extra_warnings: list[str] | None = None,
) -> dict[str, Any]:
    period = parse_manual_period(fecha_inicio, fecha_fin)
    period = merge_cut_warning(period, cliente, conn)
    if extra_warnings:
        from modules.gestion_idse_sua.nominas.period_signals import merge_signal_warnings

        period["cut_warning"] = merge_signal_warnings(period.get("cut_warning"), extra_warnings)
    conflicts = repo.find_conflicting_periods(
        conn,
        fecha_inicio=period["fecha_inicio"],
        fecha_fin=period["fecha_fin"],
        exclude_sheet_id=sheet_id,
    )
    if conflicts:
        names = ", ".join(f"{c['original_filename']}/{c['sheet_name']}" for c in conflicts[:3])
        extra = f" Advertencia: ya existe un periodo confirmado ({names})."
        period["cut_warning"] = (period.get("cut_warning") or "") + extra
    period_id = repo.upsert_period(conn, sheet_id, period, confirmed=confirmed)
    conn.commit()
    return {"period_id": period_id, "conflicts": conflicts, **period}


def extract_sheet_workers(
    conn: sqlite3.Connection,
    *,
    file_bytes: bytes,
    sheet_id: int,
    headcount_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    sheet = conn.execute("SELECT * FROM gis_nomina_sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet is None:
        raise ValueError("Hoja no encontrada.")
    if sheet["confirmed_classification"] != "nomina":
        raise ValueError("La hoja seleccionada no está clasificada como nómina.")

    period = repo.get_period_for_sheet(conn, sheet_id)
    if period is None or not period.get("user_confirmed"):
        raise ValueError("Confirme el periodo antes de extraer trabajadores.")

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[int(sheet["sheet_index"])]
    payload = extract_workers(
        ws,
        sheet_name=str(sheet["sheet_name"]),
        sheet_index=int(sheet["sheet_index"]),
        is_hidden=bool(sheet["is_hidden"]),
    )
    wb.close()

    workers = payload["workers"]
    inspection = payload.get("inspection") or {}
    header_row = int(inspection.get("header_row") or 0)
    nombre_col = (inspection.get("columns") or {}).get("nombre")
    attendance_payload = {"block": None, "rows": {}, "warnings": []}
    if header_row and nombre_col:
        attendance_payload = extract_attendance_for_workers(
            ws,
            header_row=header_row,
            nombre_col=int(nombre_col),
            fecha_inicio=str(period["fecha_inicio"]),
            fecha_fin=str(period["fecha_fin"]),
            worker_rows=[int(w["row_number"]) for w in workers],
        )
        for worker in workers:
            import json

            try:
                row_data = json.loads(worker.get("row_json") or "{}")
            except json.JSONDecodeError:
                row_data = {}
            row_data["attendance"] = attendance_payload["rows"].get(int(worker["row_number"]), [])
            if attendance_payload.get("block"):
                row_data["attendance_block"] = attendance_payload["block"]
            worker["row_json"] = json_dumps(row_data)

    for worker in workers:
        planta = worker.get("planta_normalizada") or worker.get("planta_original") or ""
        suggestion = suggest_cliente_for_planta(conn, planta, headcount_rows)
        worker["cliente_sugerido"] = suggestion.get("cliente")
        worker["suggestion_source"] = suggestion.get("source")
        worker["suggestion_confidence"] = suggestion.get("confidence")

    conn.execute("DELETE FROM gis_nomina_workers WHERE period_id = ?", (period["id"],))
    worker_ids = repo.insert_workers(conn, int(period["id"]), workers)
    persist_attendance(
        conn,
        period_id=int(period["id"]),
        worker_ids=worker_ids,
        worker_row_numbers=[int(w["row_number"]) for w in workers],
        attendance_payload=attendance_payload,
    )
    import_row = conn.execute(
        "SELECT import_id FROM gis_nomina_sheets WHERE id = ?",
        (sheet_id,),
    ).fetchone()
    if import_row:
        repo.set_import_status(conn, int(import_row["import_id"]), "extracted")
    conn.commit()
    return {
        "period_id": period["id"],
        "worker_ids": worker_ids,
        "workers_count": len(workers),
        "discarded": payload["discarded"],
        "attendance": {
            "block_detected": attendance_payload.get("block") is not None,
            "warnings": attendance_payload.get("warnings") or [],
        },
    }
