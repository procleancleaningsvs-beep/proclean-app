from __future__ import annotations

import sqlite3
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from modules.exportacion_imss.exportacion_service import mapear_headcount_a_movimiento
from modules.gestion_idse_sua.nominas import repository as repo
from modules.gestion_idse_sua.nominas.comparative_service import summarize_results
from modules.gestion_idse_sua.nominas.text_utils import normalize_upper
from modules.gestion_idse_sua.template_contract import HEADER_ROW_BY_SHEET, comparativo_path
from modules.gestion_idse_sua.template_validator import validate_comparativo_template


def _safe_filename(cliente: str, periodo_inicio: str, periodo_fin: str) -> str:
    def clean(value: str) -> str:
        return "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in value)

    return f"comparativo_{clean(cliente)}_{clean(periodo_inicio)}_{clean(periodo_fin)}.xlsx"


def _sbc_for_row(row: dict[str, Any]) -> str:
    import json

    raw = row.get("hc_json")
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, list) and data:
                data = data[0]
            if isinstance(data, dict):
                return str(mapear_headcount_a_movimiento(data).get("sbc") or "")
        except json.JSONDecodeError:
            pass
    return ""


def generate_comparative_excel(
    conn: sqlite3.Connection,
    comparative_id: int,
    *,
    username: str | None = None,
    selected_result_ids: list[int] | None = None,
) -> tuple[BytesIO, str]:
    comp = repo.get_comparative(conn, comparative_id)
    if comp is None:
        raise ValueError("Comparativo no encontrado.")

    period = conn.execute(
        """
        SELECT p.*, s.sheet_name, i.original_filename
        FROM gis_nomina_periods p
        JOIN gis_nomina_sheets s ON s.id = p.sheet_id
        JOIN gis_nomina_imports i ON i.id = s.import_id
        WHERE p.id = ?
        """,
        (comp["period_id"],),
    ).fetchone()
    if period is None:
        raise ValueError("Periodo no encontrado.")

    results = repo.list_results(conn, comparative_id)
    totals = summarize_results(results)
    plantas = sorted(
        {
            normalize_upper(r.get("planta_normalizada") or "")
            for r in results
            if r.get("planta_normalizada")
        }
    )

    filename = _safe_filename(
        str(comp["cliente"]),
        str(period["fecha_inicio"]),
        str(period["fecha_fin"]),
    )
    wb = load_workbook(BytesIO(comparativo_path().read_bytes()))
    ws_res = wb["Resumen"]
    ws_res["B6"] = comp["cliente"]
    ws_res["D6"] = ", ".join(plantas)
    ws_res["F6"] = period["original_filename"]
    ws_res["B7"] = period["fecha_inicio"]
    ws_res["D7"] = period["fecha_fin"]
    ws_res["F7"] = period["semana_num"]
    ws_res["B8"] = period["sheet_name"]
    ws_res["D8"] = "Sí"
    ws_res["F8"] = comp["generated_at"]
    ws_res["B9"] = username or comp["generated_by"] or ""
    ws_res["D9"] = comp["status"]
    ws_res["F9"] = "GIS 1.0"

    ws_det = wb["Detalle Comparativo"]
    start_row = HEADER_ROW_BY_SHEET["Detalle Comparativo"] + 1
    for idx, row in enumerate(results):
        r = start_row + idx
        ws_det.cell(r, 1, row.get("id"))
        ws_det.cell(r, 2, comp["cliente"])
        ws_det.cell(r, 3, row.get("planta_normalizada") or "")
        ws_det.cell(r, 4, period["fecha_inicio"])
        ws_det.cell(r, 5, period["fecha_fin"])
        ws_det.cell(r, 6, period["semana_num"])
        ws_det.cell(r, 7, row.get("num_empleado") or "")
        ws_det.cell(r, 8, row.get("nombre_normalizado") or row.get("hc_nombre") or "")
        ws_det.cell(r, 9, row.get("match_hc_nombre") or row.get("hc_nombre") or "")
        ws_det.cell(r, 10, row.get("match_method") or "")
        ws_det.cell(r, 11, row.get("match_status") or "")
        ws_det.cell(r, 12, row.get("confidence") or "")
        ws_det.cell(r, 13, row.get("puesto") or "")
        ws_det.cell(r, 14, row.get("nss") or "")
        ws_det.cell(r, 15, row.get("rfc") or "")
        ws_det.cell(r, 16, row.get("curp") or "")
        ws_det.cell(r, 17, _sbc_for_row(row))
        ws_det.cell(r, 18, row.get("resultado") or "")
        ws_det.cell(r, 19, row.get("semaforo") or "")
        ws_det.cell(r, 20, row.get("fecha_sugerida") or "")
        ws_det.cell(r, 21, row.get("decision_final") or row.get("tipo_sugerido") or "")
        ws_det.cell(r, 22, "")
        ws_det.cell(r, 23, row.get("observaciones") or "")

    ws_asist = wb["Asistencia Semanal"]
    asist_start = HEADER_ROW_BY_SHEET["Asistencia Semanal"] + 1
    from datetime import datetime

    attendance_rows = repo.list_attendance_for_period(conn, int(comp["period_id"]))
    by_worker: dict[int, list[dict[str, Any]]] = {}
    for row in attendance_rows:
        by_worker.setdefault(int(row["worker_id"]), []).append(row)

    worker_results = {int(r["worker_id"]): r for r in results if r.get("worker_id")}
    asist_idx = 0
    for worker_id, days in sorted(by_worker.items(), key=lambda item: item[1][0].get("nombre_normalizado") or ""):
        result = worker_results.get(worker_id, {})
        r = asist_start + asist_idx
        asist_idx += 1
        day_map = {int(d["column_index"]): d for d in days}
        totals = {"A": 0, "F": 0, "I": 0, "V": 0, "D": 0}
        first_a = ""
        last_a = ""
        for code_key in totals:
            totals[code_key] = sum(1 for d in days if d.get("code_normalized") == code_key)
        for d in sorted(days, key=lambda x: int(x["column_index"])):
            if d.get("code_normalized") == "A":
                if not first_a:
                    first_a = d["fecha_iso"]
                last_a = d["fecha_iso"]

        ws_asist.cell(r, 1, worker_id)
        ws_asist.cell(r, 2, comp["cliente"])
        ws_asist.cell(r, 3, result.get("planta_normalizada") or "")
        ws_asist.cell(r, 4, result.get("num_empleado") or days[0].get("num_empleado") or "")
        ws_asist.cell(r, 5, result.get("nombre_normalizado") or days[0].get("nombre_normalizado") or "")
        ws_asist.cell(r, 6, result.get("puesto") or "")
        status_bits = [str(result.get("match_status") or "")]
        review_days = [d for d in days if str(d.get("interpretation_status") or "") not in {"ok", "empty", "corrected"}]
        if review_days:
            status_bits.append("revisión asistencia")
        ws_asist.cell(r, 7, " · ".join(bit for bit in status_bits if bit))
        for day_idx in range(1, 8):
            cell = day_map.get(day_idx)
            display = ""
            if cell:
                code = cell.get("code_normalized") or cell.get("code_original") or ""
                fecha = cell.get("fecha_iso") or ""
                display = f"{code} ({fecha})" if code and fecha else code
            ws_asist.cell(r, 7 + day_idx, display)
        ws_asist.cell(r, 15, totals["A"])
        ws_asist.cell(r, 16, totals["F"])
        ws_asist.cell(r, 17, totals["I"])
        ws_asist.cell(r, 18, totals["V"])
        ws_asist.cell(r, 19, totals["D"])
        ws_asist.cell(r, 20, first_a)
        ws_asist.cell(r, 21, last_a)

    if not by_worker:
        ws_asist["A3"] = "No se detectó asistencia semanal interpretada para este periodo."

    ws_mov = wb["Movimientos Seleccionados"]
    mov_start = HEADER_ROW_BY_SHEET["Movimientos Seleccionados"] + 1
    selected = selected_result_ids or [
        int(r["id"])
        for r in results
        if str(r.get("conversion_status") or "") in {"pending", "converted"}
        and str(r.get("tipo_sugerido") or "") in {"ALTA", "BAJA"}
    ]
    mov_idx = 0
    for row in results:
        if int(row["id"]) not in selected:
            continue
        tipo = str(row.get("tipo_sugerido") or row.get("decision_final") or "").upper()
        if tipo not in {"ALTA", "BAJA"}:
            continue
        r = mov_start + mov_idx
        mov_idx += 1
        ws_mov.cell(r, 1, "Sí")
        ws_mov.cell(r, 2, tipo)
        ws_mov.cell(r, 5, row.get("fecha_sugerida") or "")
        ws_mov.cell(r, 6, row.get("nss") or "")
        ws_mov.cell(r, 7, row.get("rfc") or "")
        ws_mov.cell(r, 8, row.get("curp") or "")
        ws_mov.cell(r, 13, comp["cliente"])
        ws_mov.cell(r, 14, row.get("planta_normalizada") or "")
        ws_mov.cell(r, 15, "GIS nominas")
        ws_mov.cell(r, 16, row.get("observaciones") or "")

    out_buf = BytesIO()
    try:
        wb.save(out_buf)
    finally:
        wb.close()
    out_buf.seek(0)
    validate_comparativo_template(out_buf)
    out_buf.seek(0)
    return out_buf, filename
