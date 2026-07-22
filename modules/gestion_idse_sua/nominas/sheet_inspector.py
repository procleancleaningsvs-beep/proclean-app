from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from modules.gestion_idse_sua.nominas.constants import (
    CUENTA_HEADERS,
    NOMBRE_HEADERS,
    NUM_EMPLEADO_HEADERS,
    PLANTA_HEADERS,
    PUESTO_HEADERS,
    TOTAL_MARKERS,
)
from modules.gestion_idse_sua.nominas.period_parser import detect_period
from modules.gestion_idse_sua.nominas.text_utils import normalize_header, normalize_upper


def _header_map(row_values: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, raw in enumerate(row_values, start=1):
        key = normalize_header(raw)
        if key:
            out[key] = idx
    return out


def _find_col(header: dict[str, int], aliases: frozenset[str]) -> int | None:
    for alias in aliases:
        if alias in header:
            return header[alias]
    return None


def inspect_sheet(ws: Worksheet, *, sheet_name: str, sheet_index: int, is_hidden: bool) -> dict[str, Any]:
    header_row_idx: int | None = None
    header: dict[str, int] = {}
    preview_lines: list[str] = []

    max_scan = min(ws.max_row or 1, 40)
    for row_idx in range(1, max_scan + 1):
        row_values = [ws.cell(row_idx, col).value for col in range(1, (ws.max_column or 1) + 1)]
        row_header = _header_map(row_values)
        if _find_col(row_header, NOMBRE_HEADERS):
            header_row_idx = row_idx
            header = row_header
            break
        text_bits = [normalize_upper(v) for v in row_values if v is not None and str(v).strip()]
        if text_bits:
            preview_lines.append(" ".join(text_bits[:6]))

    nombre_col = _find_col(header, NOMBRE_HEADERS) if header else None
    num_col = _find_col(header, NUM_EMPLEADO_HEADERS) if header else None
    puesto_col = _find_col(header, PUESTO_HEADERS) if header else None
    planta_col = _find_col(header, PLANTA_HEADERS) if header else None
    cuenta_col = _find_col(header, CUENTA_HEADERS) if header else None

    evidence_score = 0
    if nombre_col:
        evidence_score += 3
    if num_col:
        evidence_score += 1
    if puesto_col:
        evidence_score += 1
    if planta_col:
        evidence_score += 1
    if cuenta_col:
        evidence_score += 1

    estimated_rows = 0
    if header_row_idx and nombre_col:
        for row_idx in range(header_row_idx + 1, (ws.max_row or 0) + 1):
            name_val = ws.cell(row_idx, nombre_col).value
            if name_val is None or str(name_val).strip() == "":
                continue
            marker = normalize_upper(name_val)
            if marker in TOTAL_MARKERS:
                break
            if isinstance(name_val, (int, float)):
                continue
            estimated_rows += 1

    period_text = " ".join([sheet_name, *preview_lines[:3]])
    period = detect_period(period_text)

    if evidence_score >= 3 and estimated_rows >= 1:
        suggested = "nomina"
    elif evidence_score >= 1 or "AUX" in normalize_upper(sheet_name):
        suggested = "auxiliar"
    else:
        suggested = "ignorada"

    return {
        "sheet_index": sheet_index,
        "sheet_name": sheet_name,
        "is_hidden": is_hidden,
        "suggested_classification": suggested,
        "estimated_rows": estimated_rows,
        "suggested_period": period,
        "header_row": header_row_idx,
        "columns": {
            "nombre": nombre_col,
            "num_empleado": num_col,
            "puesto": puesto_col,
            "planta": planta_col,
            "cuenta": cuenta_col,
        },
    }
