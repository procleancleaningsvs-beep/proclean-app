from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from modules.gestion_idse_sua.nominas.constants import TOTAL_MARKERS
from modules.gestion_idse_sua.nominas.sheet_inspector import inspect_sheet
from modules.gestion_idse_sua.nominas.text_utils import json_dumps, normalize_name, normalize_planta, normalize_upper


def extract_workers(ws: Worksheet, *, sheet_name: str, sheet_index: int, is_hidden: bool) -> dict[str, Any]:
    inspection = inspect_sheet(ws, sheet_name=sheet_name, sheet_index=sheet_index, is_hidden=is_hidden)
    header_row = inspection.get("header_row")
    cols = inspection.get("columns") or {}
    nombre_col = cols.get("nombre")
    if not header_row or not nombre_col:
        raise ValueError("No se encontró encabezado de nómina en la hoja seleccionada.")

    workers: list[dict[str, Any]] = []
    discarded: list[dict[str, Any]] = []

    for row_idx in range(int(header_row) + 1, (ws.max_row or 0) + 1):
        name_val = ws.cell(row_idx, nombre_col).value
        if name_val is None or str(name_val).strip() == "":
            continue
        marker = normalize_upper(name_val)
        if marker in TOTAL_MARKERS:
            break
        if isinstance(name_val, (int, float)):
            discarded.append({"row": row_idx, "reason": "nombre_numérico"})
            continue

        nombre_original = str(name_val).strip()
        nombre_norm = normalize_name(nombre_original)
        if not nombre_norm:
            discarded.append({"row": row_idx, "reason": "nombre_vacio"})
            continue

        def _cell(col_key: str) -> str:
            col = cols.get(col_key)
            if not col:
                return ""
            val = ws.cell(row_idx, col).value
            if val is None:
                return ""
            return str(val).strip()

        row_payload = {
            "num_empleado": _cell("num_empleado"),
            "puesto": _cell("puesto"),
            "cliente": _cell("cliente"),
            "planta": _cell("planta"),
            "cuenta": _cell("cuenta"),
        }
        workers.append(
            {
                "row_number": row_idx,
                "num_empleado": row_payload["num_empleado"],
                "nombre_original": nombre_original,
                "nombre_normalizado": nombre_norm,
                "puesto": row_payload["puesto"],
                "planta_original": row_payload["planta"],
                "planta_normalizada": normalize_planta(row_payload["planta"]) if row_payload["planta"] else "",
                "cuenta": row_payload["cuenta"],
                "row_json": json_dumps(row_payload),
                "cliente_sugerido": normalize_upper(row_payload["cliente"]),
                "suggestion_source": "payroll" if row_payload["cliente"] else "",
                "suggestion_confidence": 1.0 if row_payload["cliente"] else 0.0,
            }
        )

    return {
        "workers": workers,
        "discarded": discarded,
        "inspection": inspection,
    }
