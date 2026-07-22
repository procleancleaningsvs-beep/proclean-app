from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from modules.gestion_idse_sua.nominas.constants import PAYROLL_HEADER_MARKERS, TOTAL_MARKERS, VALOR_HE_HEADERS
from modules.gestion_idse_sua.nominas.text_utils import normalize_header, normalize_upper

DAY_HEADER_RE = re.compile(r"^[LMDJSV]\d{1,2}$", re.I)
ATTENDANCE_CODES = frozenset({"A", "F", "I", "INC", "V", "D"})
MIN_BLOCK_SCORE = 6


def _header_texts(ws: Worksheet, header_row: int, start_col: int) -> list[str]:
    return [str(ws.cell(header_row, start_col + i).value or "").strip() for i in range(7)]


def _block_from_headers(start_col: int, headers: list[str]) -> dict[str, Any]:
    header_hits = sum(1 for h in headers if DAY_HEADER_RE.match(h))
    payroll_headers = sum(
        1
        for h in headers
        if normalize_upper(h) in PAYROLL_HEADER_MARKERS and not DAY_HEADER_RE.match(h)
    )
    if payroll_headers > 0 or header_hits < 4:
        return {}
    score = header_hits * 3 + 6
    return {
        "start_col": start_col,
        "end_col": start_col + 6,
        "headers": headers,
        "confidence": min(1.0, score / 20.0),
        "status": "ok" if header_hits >= 5 else "review",
        "score": score,
        "header_hits": header_hits,
    }


def _find_valor_he_column(ws: Worksheet, header_row: int) -> int | None:
    for col_idx in range(1, (ws.max_column or 0) + 1):
        key = normalize_header(ws.cell(header_row, col_idx).value)
        if key in VALOR_HE_HEADERS:
            return col_idx
    return None


def _detect_anchored_block(ws: Worksheet, *, header_row: int) -> dict[str, Any] | None:
    anchor = _find_valor_he_column(ws, header_row)
    if anchor is None:
        return None
    start_col = anchor + 1
    headers = _header_texts(ws, header_row, start_col)
    block = _block_from_headers(start_col, headers)
    return block or None


def normalize_attendance_code(raw: Any) -> dict[str, str]:
    if raw is None:
        return {"original": "", "normalized": "", "status": "empty"}
    original = str(raw).strip()
    if not original:
        return {"original": "", "normalized": "", "status": "empty"}
    upper = original.upper()
    if upper == "INC":
        return {"original": original, "normalized": "I", "status": "ok"}
    if upper in ATTENDANCE_CODES:
        return {"original": original, "normalized": upper, "status": "ok"}
    return {"original": original, "normalized": "", "status": "review"}


def assign_period_dates(fecha_inicio: str, column_index: int) -> str:
    start = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
    day = start + timedelta(days=int(column_index) - 1)
    return day.isoformat()


def header_day_number(header: str) -> int | None:
    match = DAY_HEADER_RE.match(str(header or "").strip())
    if not match:
        return None
    digits = re.sub(r"^\D+", "", match.group(0))
    return int(digits) if digits.isdigit() else None


def _row_attendance_values(ws: Worksheet, row_idx: int, start_col: int) -> list[str]:
    values: list[str] = []
    for offset in range(7):
        val = ws.cell(row_idx, start_col + offset).value
        if val is None:
            values.append("")
        else:
            values.append(str(val).strip())
    return values


def _score_candidate_block(
    ws: Worksheet,
    *,
    header_row: int,
    start_col: int,
    nombre_col: int,
    sample_rows: list[int],
) -> tuple[int, dict[str, Any]]:
    headers = [ws.cell(header_row, start_col + i).value for i in range(7)]
    header_texts = [str(h or "").strip() for h in headers]
    header_hits = sum(1 for h in header_texts if DAY_HEADER_RE.match(h))
    payroll_headers = sum(
        1
        for h in header_texts
        if (normalize_upper(h) in PAYROLL_HEADER_MARKERS or "EMPLEADO" in normalize_upper(h))
        and not DAY_HEADER_RE.match(h)
    )
    if payroll_headers > 0:
        return 0, {"start_col": start_col, "headers": header_texts, "header_hits": header_hits, "code_hits": 0, "proximity": 0}
    code_hits = 0
    blank_rows = 0
    for row_idx in sample_rows:
        row_codes = _row_attendance_values(ws, row_idx, start_col)
        if not any(row_codes):
            blank_rows += 1
            continue
        code_hits += sum(
            1 for value in row_codes if value and normalize_attendance_code(value)["status"] == "ok"
        )
    proximity = max(0, 8 - abs(start_col - int(nombre_col)))
    score = header_hits * 2 + min(code_hits, 14) + min(proximity, 6)
    return score, {
        "start_col": start_col,
        "headers": header_texts,
        "header_hits": header_hits,
        "code_hits": code_hits,
        "proximity": proximity,
    }


def detect_attendance_block(
    ws: Worksheet,
    *,
    header_row: int,
    nombre_col: int,
) -> dict[str, Any] | None:
    max_col = ws.max_column or 0
    if max_col < 7:
        return None

    anchored = _detect_anchored_block(ws, header_row=header_row)
    if anchored is not None:
        return anchored

    sample_rows: list[int] = []
    for row_idx in range(header_row + 1, min((ws.max_row or header_row) + 1, header_row + 25)):
        name_val = ws.cell(row_idx, nombre_col).value
        if name_val is None or str(name_val).strip() == "":
            continue
        if normalize_upper(name_val) in TOTAL_MARKERS:
            break
        if isinstance(name_val, (int, float)):
            continue
        sample_rows.append(row_idx)
        if len(sample_rows) >= 8:
            break

    best_score = 0
    best_meta: dict[str, Any] | None = None
    for start_col in range(1, max_col - 5):
        score, meta = _score_candidate_block(
            ws,
            header_row=header_row,
            start_col=start_col,
            nombre_col=nombre_col,
            sample_rows=sample_rows,
        )
        if score > best_score:
            best_score = score
            best_meta = meta

    if best_meta is None or best_score < MIN_BLOCK_SCORE:
        return None

    header_hits = int(best_meta.get("header_hits") or 0)
    code_hits = int(best_meta.get("code_hits") or 0)
    if header_hits < 3 and code_hits < 6:
        return None

    confidence = min(1.0, best_score / 20.0)
    status = "ok" if best_score >= 8 and best_meta["header_hits"] >= 3 else "review"
    return {
        "start_col": best_meta["start_col"],
        "end_col": best_meta["start_col"] + 6,
        "headers": best_meta["headers"],
        "confidence": round(confidence, 2),
        "status": status,
        "score": best_score,
    }


def parse_worker_attendance_row(
    ws: Worksheet,
    *,
    row_idx: int,
    block: dict[str, Any],
    fecha_inicio: str,
    fecha_fin: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    start_col = int(block["start_col"])
    headers = block.get("headers") or []
    start_date = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
    end_date = datetime.strptime(fecha_fin, "%d/%m/%Y").date()
    warnings: list[str] = []
    days: list[dict[str, Any]] = []

    for idx in range(7):
        col_index = idx + 1
        header_original = headers[idx] if idx < len(headers) else ""
        raw = ws.cell(row_idx, start_col + idx).value
        parsed = normalize_attendance_code(raw)
        fecha = start_date + timedelta(days=idx)
        fecha_iso = fecha.isoformat()

        header_day = header_day_number(header_original)
        if header_day is not None and header_day != fecha.day:
            warnings.append(
                f"Encabezado {header_original} no coincide con el día esperado ({fecha.day}) del periodo confirmado."
            )

        if fecha < start_date or fecha > end_date:
            warnings.append(f"Fecha derivada {fecha_iso} fuera del periodo confirmado.")

        days.append(
            {
                "column_index": col_index,
                "column_number": start_col + idx,
                "fecha_iso": fecha_iso,
                "header_original": header_original,
                "code_original": parsed["original"],
                "code_normalized": parsed["normalized"],
                "interpretation_status": parsed["status"],
                "warning": None,
            }
        )

    return days, warnings


def extract_attendance_for_workers(
    ws: Worksheet,
    *,
    header_row: int,
    nombre_col: int,
    fecha_inicio: str,
    fecha_fin: str,
    worker_rows: list[int],
) -> dict[str, Any]:
    block = detect_attendance_block(ws, header_row=header_row, nombre_col=nombre_col)
    if block is None:
        return {"block": None, "rows": {}, "warnings": ["No se detectó bloque de asistencia de siete columnas."]}

    parsed_rows: dict[int, list[dict[str, Any]]] = {}
    global_warnings: list[str] = []
    if block["status"] == "review":
        global_warnings.append("Bloque de asistencia detectado con confianza baja; requiere revisión.")

    for row_idx in worker_rows:
        days, row_warnings = parse_worker_attendance_row(
            ws,
            row_idx=row_idx,
            block=block,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        )
        if row_warnings:
            global_warnings.extend(row_warnings[:1])
        parsed_rows[row_idx] = days

    return {"block": block, "rows": parsed_rows, "warnings": global_warnings}
