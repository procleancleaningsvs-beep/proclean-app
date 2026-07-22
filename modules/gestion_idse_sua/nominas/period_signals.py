from __future__ import annotations

from datetime import date, datetime, timedelta
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from modules.gestion_idse_sua.nominas.attendance_parser import DAY_HEADER_RE, detect_attendance_block
from modules.gestion_idse_sua.nominas.period_parser import detect_period
from modules.gestion_idse_sua.nominas.text_utils import normalize_spaces, normalize_upper


def _period_key(period: dict[str, Any]) -> tuple[str, str] | None:
    if not period.get("detected"):
        return None
    start = str(period.get("fecha_inicio") or "")
    end = str(period.get("fecha_fin") or "")
    if not start or not end:
        return None
    return start, end


def _scan_sheet_text_periods(ws: Worksheet, *, max_rows: int = 8, max_cols: int = 18) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row_idx in range(1, min(max_rows, ws.max_row or 1) + 1):
        for col_idx in range(1, min(max_cols, ws.max_column or 1) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            if isinstance(value, (int, float)):
                serial = float(value)
                if 40000 < serial < 60000:
                    day = datetime(1899, 12, 30) + timedelta(days=serial)
                    key = (day.date().isoformat(), day.date().isoformat())
                    if key not in seen:
                        seen.add(key)
                        found.append(
                            {
                                "source": "excel_serial",
                                "row": row_idx,
                                "column": col_idx,
                                "fecha_inicio": day.strftime("%d/%m/%Y"),
                                "fecha_fin": day.strftime("%d/%m/%Y"),
                            }
                        )
                continue
            text = normalize_spaces(str(value))
            if not text:
                continue
            upper = normalize_upper(text)
            if "PERIODO" not in upper and "NOMINA" not in upper and " AL " not in upper:
                continue
            period = detect_period(text)
            key = _period_key(period)
            if key and key not in seen:
                seen.add(key)
                found.append(
                    {
                        "source": "sheet_text",
                        "row": row_idx,
                        "column": col_idx,
                        "fecha_inicio": period["fecha_inicio"],
                        "fecha_fin": period["fecha_fin"],
                    }
                )
    return found


def _header_block_period(header_row: int | None, block: dict[str, Any] | None) -> dict[str, Any] | None:
    if not header_row or not block:
        return None
    headers = block.get("headers") or []
    day_nums: list[int] = []
    for header in headers:
        match = DAY_HEADER_RE.match(str(header or "").strip())
        if not match:
            continue
        digits = "".join(ch for ch in match.group(0) if ch.isdigit())
        if digits.isdigit():
            day_nums.append(int(digits))
    if len(day_nums) < 2:
        return None
    return {
        "source": "attendance_headers",
        "day_numbers": day_nums,
        "headers": headers,
    }


def collect_period_signals(
    ws: Worksheet,
    *,
    sheet_name: str,
    header_row: int | None,
    nombre_col: int | None,
) -> dict[str, Any]:
    signals: list[dict[str, Any]] = []

    sheet_period = detect_period(sheet_name)
    if sheet_period.get("detected"):
        signals.append({"source": "sheet_name", **{k: sheet_period[k] for k in ("fecha_inicio", "fecha_fin", "days")}})

    for item in _scan_sheet_text_periods(ws):
        signals.append(item)

    block = None
    if header_row and nombre_col:
        block = detect_attendance_block(ws, header_row=header_row, nombre_col=nombre_col)
        header_signal = _header_block_period(header_row, block)
        if header_signal:
            signals.append(header_signal)

    return {
        "signals": signals,
        "attendance_block": block,
        "warnings": compare_period_signals(signals, sheet_name=sheet_name),
    }


def compare_period_signals(signals: list[dict[str, Any]], *, sheet_name: str | None = None) -> list[str]:
    ranges: set[tuple[str, str]] = set()
    for signal in signals:
        start = signal.get("fecha_inicio")
        end = signal.get("fecha_fin")
        if start and end:
            ranges.add((str(start), str(end)))

    warnings: list[str] = []
    if len(ranges) > 1:
        warnings.append(
            "Las señales de periodo (nombre de hoja, texto interno, fecha serial o encabezados) no coinciden; "
            "confirme manualmente el periodo correcto."
        )
    if sheet_name:
        sheet_period = detect_period(sheet_name)
        if sheet_period.get("detected"):
            key = (sheet_period["fecha_inicio"], sheet_period["fecha_fin"])
            if ranges and key not in ranges:
                warnings.append(
                    "El periodo sugerido por el nombre de la hoja difiere de otras señales internas; confirme manualmente."
                )
        elif any(s.get("source") == "sheet_text" for s in signals):
            warnings.append(
                "El nombre de la hoja no aporta un periodo claro y el texto interno sugiere otro rango; confirme manualmente."
            )
    return warnings


def merge_signal_warnings(existing: str | None, warnings: list[str]) -> str | None:
    parts = [w for w in [existing, *(warnings or [])] if w]
    if not parts:
        return None
    deduped: list[str] = []
    for part in parts:
        if part not in deduped:
            deduped.append(part)
    return " | ".join(deduped)
