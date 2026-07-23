from __future__ import annotations

import json
import sqlite3
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from modules.gestion_idse_sua.reportes import repository as repo
from modules.gestion_idse_sua.template_contract import HEADER_ROW_BY_SHEET, mensual_path
from modules.gestion_idse_sua.template_validator import validate_mensual_template


def _safe_filename(cliente: str, mes: int, anio: int) -> str:
    clean = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in cliente)
    return f"reporte_mensual_{clean}_{anio:04d}-{mes:02d}.xlsx"


def _json_list(raw: str | None) -> list[Any]:
    if not raw:
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []
    return data if isinstance(data, list) else []


def _coverage_labels(snapshot: dict[str, Any]) -> tuple[str, str, str]:
    complete = "Sí" if snapshot.get("coverage_complete") else "No"
    missing = ", ".join(snapshot.get("missing_dates") or []) or "—"
    warning = "; ".join(snapshot.get("coverage_warnings") or []) or "—"
    return complete, missing, warning


def _fill_workbook(
    wb,
    *,
    report: dict[str, Any],
    weeks: list[dict[str, Any]],
    persons: list[dict[str, Any]],
    events: list[dict[str, Any]],
    snapshot: dict[str, Any],
    username: str | None,
) -> None:
    ws_res = wb["Resumen"]
    semanas_label = ", ".join(f"{w.get('fecha_inicio')}–{w.get('fecha_fin')}" for w in weeks)
    archivos = ", ".join(sorted({str(w.get("original_filename") or "") for w in weeks if w.get("original_filename")}))
    coverage_complete, missing_dates, coverage_warning = _coverage_labels(snapshot)
    ws_res["B6"] = report["cliente"]
    ws_res["D6"] = int(report["mes"])
    ws_res["F6"] = int(report["anio"])
    ws_res["B7"] = semanas_label
    ws_res["D7"] = report["updated_at"]
    ws_res["F7"] = username or report.get("created_by") or ""
    ws_res["B8"] = archivos
    ws_res["D8"] = report["estado"]
    ws_res["F8"] = report.get("version") or "1.0"
    ws_res["A9"] = "Cobertura completa"
    ws_res["B9"] = coverage_complete
    ws_res["C9"] = "Fechas faltantes"
    ws_res["D9"] = missing_dates
    ws_res["E9"] = "Advertencia cobertura"
    ws_res["F9"] = coverage_warning

    ws_personal = wb["Personal Mensual"]
    p_start = HEADER_ROW_BY_SHEET["Personal Mensual"] + 1
    for idx, person in enumerate(persons):
        r = p_start + idx
        ws_personal.cell(r, 1, person["id"])
        ws_personal.cell(r, 2, person.get("num_empleado") or "")
        ws_personal.cell(r, 3, person.get("nombre_nomina") or "")
        ws_personal.cell(r, 4, person.get("nombre_hc") or "")
        ws_personal.cell(r, 5, person.get("match_method") or "")
        ws_personal.cell(r, 6, person.get("match_status") or "")
        ws_personal.cell(r, 7, person.get("nss") or "")
        ws_personal.cell(r, 8, person.get("rfc") or "")
        ws_personal.cell(r, 9, person.get("curp") or "")
        ws_personal.cell(r, 10, person.get("sbc") or "")
        ws_personal.cell(r, 11, ", ".join(_json_list(person.get("clientes_json"))))
        ws_personal.cell(r, 12, ", ".join(_json_list(person.get("plantas_json"))))
        ws_personal.cell(r, 13, ", ".join(str(s.get("period_id") or "") for s in _json_list(person.get("semanas_json"))))
        ws_personal.cell(r, 14, person.get("primera_a") or "")
        ws_personal.cell(r, 15, person.get("ultima_a") or "")
        totals = {}
        if person.get("totals_json"):
            try:
                totals = json.loads(person["totals_json"])
            except json.JSONDecodeError:
                totals = {}
        ws_personal.cell(r, 16, totals.get("A", 0))
        ws_personal.cell(r, 17, totals.get("F", 0))
        ws_personal.cell(r, 18, totals.get("I", 0))
        ws_personal.cell(r, 19, totals.get("V", 0))
        ws_personal.cell(r, 20, totals.get("D", 0))
        ws_personal.cell(r, 21, person.get("estado_mensual") or "")
        confirmed = [
            e for e in events
            if int(e["person_id"]) == int(person["id"]) and str(e.get("estado") or "") == "confirmado"
        ]
        proposed = [
            e for e in events
            if int(e["person_id"]) == int(person["id"]) and str(e.get("estado") or "") in {"propuesto", "incompleto"}
        ]
        ws_personal.cell(r, 22, ", ".join(f"{e.get('event_type_suggested')} {e.get('fecha_suggested') or '?'}" for e in proposed))
        ws_personal.cell(r, 23, ", ".join(f"{e.get('event_type_confirmed') or e.get('event_type_suggested')} {e.get('fecha_confirmed') or e.get('fecha_suggested') or '?'}" for e in confirmed))
        ws_personal.cell(r, 24, ", ".join(_json_list(person.get("warnings_json"))))

    ws_asist = wb["Asistencia Mensual"]
    a_start = HEADER_ROW_BY_SHEET["Asistencia Mensual"] + 1
    for idx, person in enumerate(persons):
        r = a_start + idx
        daily = _json_list(person.get("daily_json"))
        ws_asist.cell(r, 1, person["id"])
        ws_asist.cell(r, 2, person.get("num_empleado") or "")
        ws_asist.cell(r, 3, person.get("nombre_nomina") or "")
        ws_asist.cell(r, 4, ", ".join(_json_list(person.get("clientes_json"))))
        ws_asist.cell(r, 5, ", ".join(_json_list(person.get("plantas_json"))))
        ws_asist.cell(r, 6, person.get("estado_mensual") or "")
        ws_asist.cell(r, 7, person.get("primera_a") or "")
        ws_asist.cell(r, 8, person.get("ultima_a") or "")
        totals = {}
        if person.get("totals_json"):
            try:
                totals = json.loads(person["totals_json"])
            except json.JSONDecodeError:
                totals = {}
        ws_asist.cell(r, 9, totals.get("A", 0))
        ws_asist.cell(r, 10, totals.get("F", 0))
        ws_asist.cell(r, 11, ", ".join(_json_list(person.get("warnings_json"))))
        for day_idx in range(31):
            code = ""
            if day_idx < len(daily):
                code = str(daily[day_idx].get("code") or "")
            ws_asist.cell(r, 12 + day_idx, code)
        ws_asist.cell(r, 43, totals.get("I", 0))
        ws_asist.cell(r, 44, totals.get("V", 0))
        ws_asist.cell(r, 45, totals.get("D", 0))

    ws_traj = wb["Trayectoria Semanal"]
    t_start = HEADER_ROW_BY_SHEET["Trayectoria Semanal"] + 1
    t_idx = 0
    for person in persons:
        semanas = _json_list(person.get("semanas_json"))
        trajectory = {}
        if person.get("trajectory_json"):
            try:
                trajectory = json.loads(person["trajectory_json"])
            except json.JSONDecodeError:
                trajectory = {}
        person_events = [e for e in events if int(e["person_id"]) == int(person["id"])]
        for semana in semanas:
            r = t_start + t_idx
            t_idx += 1
            ws_traj.cell(r, 1, person["id"])
            ws_traj.cell(r, 2, person.get("num_empleado") or "")
            ws_traj.cell(r, 3, person.get("nombre_nomina") or "")
            ws_traj.cell(r, 4, ", ".join(_json_list(person.get("clientes_json"))))
            ws_traj.cell(r, 5, ", ".join(_json_list(person.get("plantas_json"))))
            ws_traj.cell(r, 6, semana.get("period_id") or "")
            ws_traj.cell(r, 7, semana.get("fecha_inicio") or "")
            ws_traj.cell(r, 8, "")
            ws_traj.cell(r, 9, " ".join(trajectory.get("sequence") or []))
            totals = trajectory.get("totals") or {}
            ws_traj.cell(r, 10, totals.get("A", 0))
            ws_traj.cell(r, 11, totals.get("F", 0))
            ws_traj.cell(r, 12, totals.get("I", 0))
            ws_traj.cell(r, 13, totals.get("V", 0))
            ws_traj.cell(r, 14, totals.get("D", 0))
            ws_traj.cell(r, 15, person.get("estado_mensual") or "")
            seg_event = next((e for e in person_events if not e.get("period_id")), person_events[0] if person_events else None)
            if seg_event:
                ws_traj.cell(r, 16, seg_event.get("event_type_suggested") or "")
                ws_traj.cell(r, 17, seg_event.get("fecha_suggested") or "")
                ws_traj.cell(r, 18, seg_event.get("estado") or "")
                ws_traj.cell(r, 19, seg_event.get("motivo") or seg_event.get("observaciones") or "")

    ws_mov = wb["Movimientos Seleccionados"]
    m_start = HEADER_ROW_BY_SHEET["Movimientos Seleccionados"] + 1
    m_idx = 0
    for event in events:
        if str(event.get("estado") or "") != "confirmado":
            continue
        tipo = str(event.get("event_type_confirmed") or event.get("event_type_suggested") or "").upper()
        if tipo not in {"ALTA", "BAJA"}:
            continue
        r = m_start + m_idx
        m_idx += 1
        ws_mov.cell(r, 1, "Sí")
        ws_mov.cell(r, 2, tipo)
        ws_mov.cell(r, 5, event.get("fecha_confirmed") or event.get("fecha_suggested") or "")
        ws_mov.cell(r, 6, event.get("nss") or "")
        ws_mov.cell(r, 7, event.get("rfc") or "")
        ws_mov.cell(r, 8, event.get("curp") or "")
        ws_mov.cell(r, 13, ", ".join(_json_list(event.get("clientes_json"))))
        ws_mov.cell(r, 14, ", ".join(_json_list(event.get("plantas_json"))))
        ws_mov.cell(r, 15, "GIS reporte mensual")
        ws_mov.cell(r, 16, event.get("observaciones") or event.get("motivo") or "")
        ws_mov.cell(r, 17, "Sí" if tipo in {"ALTA", "BAJA"} else "No")
        ws_mov.cell(r, 18, "Sí" if tipo == "ALTA" else "No")

    ws_pend = wb["Pendientes"]
    pend_start = HEADER_ROW_BY_SHEET["Pendientes"] + 1
    pendientes = snapshot.get("pendientes") or []
    for idx, item in enumerate(pendientes):
        r = pend_start + idx
        ws_pend.cell(r, 1, item.get("tipo") or "")
        ws_pend.cell(r, 3, item.get("num_empleado") or "")
        ws_pend.cell(r, 4, item.get("nombre") or "")
        ws_pend.cell(r, 8, item.get("detalle") or "")


def generate_monthly_excel(
    conn: sqlite3.Connection,
    report_id: int,
    *,
    username: str | None = None,
) -> tuple[BytesIO, str]:
    report = repo.get_report(conn, report_id)
    if report is None:
        raise ValueError("Reporte no encontrado.")
    weeks = repo.list_report_weeks(conn, report_id)
    persons = repo.list_report_persons(conn, report_id)
    events = repo.list_report_events(conn, report_id)
    snapshot: dict[str, Any] = {}
    if report.get("snapshot_json"):
        try:
            snapshot = json.loads(report["snapshot_json"])
        except json.JSONDecodeError:
            snapshot = {}

    template_bytes = mensual_path().read_bytes()
    wb = load_workbook(BytesIO(template_bytes))
    try:
        _fill_workbook(
            wb,
            report=report,
            weeks=weeks,
            persons=persons,
            events=events,
            snapshot=snapshot,
            username=username,
        )
        out_buf = BytesIO()
        wb.save(out_buf)
    finally:
        wb.close()

    out_buf.seek(0)
    validate_mensual_template(out_buf)
    out_buf.seek(0)
    return out_buf, _safe_filename(str(report["cliente"]), int(report["mes"]), int(report["anio"]))
