from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from modules.headcount.matching import info_estado_label, warning_label
from modules.headcount.ui_format import display_cell, display_cliente, display_ubicacion

_WARNINGS_OMIT_EXCEL = frozenset({"STATUS_IMSS_INCONSISTENTE"})

_DET_HEADERS = [
    "No.",
    "NSS SUA",
    "CURP",
    "Nombre SUA",
    "Estado SUA al corte",
    "Movimiento SUA",
    "Fecha mov. SUA",
    "Activo al corte",
    "Cliente HC",
    "Ubicación HC",
    "St. Op.",
    "Patrón HC",
    "Match por",
    "Match status",
    "Días SUA",
    "SDI SUA",
    "Warnings",
    "Info estado",
]


def _write_sheet_headers(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row, col, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(42, max(12, len(str(ws.cell(1, col).value or "")) + 4))


def _freeze_and_filter(ws, last_col: int, last_row: int) -> None:
    ws.freeze_panes = "A2"
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def _warning_labels_for_export(codes: list[str] | None) -> str:
    labels = []
    for code in codes or []:
        if code in _WARNINGS_OMIT_EXCEL:
            continue
        labels.append(warning_label(code))
    return "; ".join(labels)


def _nss_export(row: dict[str, Any]) -> str:
    nss = str(row.get("nss_normalizado") or "").strip()
    if nss:
        return nss
    from modules.headcount.matching import normalize_nss

    return normalize_nss(row.get("nss_sua_original", ""))


def _write_detalle_rows(ws, rows: list[dict[str, Any]], start_row: int = 2) -> None:
    for ri, row in enumerate(rows, start=start_row):
        ws.cell(ri, 1, row.get("registro_no", ri - 1))
        ws.cell(ri, 2, _nss_export(row))
        ws.cell(ri, 3, display_cell(row.get("curp", "")))
        ws.cell(ri, 4, display_cell(row.get("nombre_sua_original", "")))
        ws.cell(ri, 5, display_cell(row.get("estado_sua_al_corte", "")))
        ws.cell(ri, 6, display_cell(row.get("sua_movimiento_clave", "")))
        ws.cell(ri, 7, display_cell(row.get("movimiento_fecha", "")))
        ws.cell(ri, 8, display_cell(row.get("es_activo_al_corte_label", "")))
        ws.cell(ri, 9, display_cliente(row.get("cliente_headcount", "")))
        ws.cell(ri, 10, display_ubicacion(row.get("ubicacion_headcount", "")))
        ws.cell(ri, 11, display_cell(row.get("status_operacion_headcount", "")))
        ws.cell(ri, 12, display_cell(row.get("patron_headcount", "")))
        ws.cell(ri, 13, display_cell(row.get("match_por", "")))
        ws.cell(ri, 14, display_cell(row.get("match_status", "")))
        ws.cell(ri, 15, row.get("dias", ""))
        ws.cell(ri, 16, row.get("sdi", ""))
        ws.cell(ri, 17, _warning_labels_for_export(row.get("warnings")))
        info = row.get("info_estado") or ""
        ws.cell(ri, 18, info_estado_label(info) if info else "")


def build_auditoria_excel_bytes(payload: dict[str, Any]) -> bytes:
    resumen = payload.get("resumen") or {}
    detalle = payload.get("detalle") or []
    resumen_clientes = payload.get("resumen_clientes") or []
    hc_sin_sua = payload.get("headcount_sin_sua") or []
    sua_activos = payload.get("sua_activos") or [r for r in detalle if r.get("sua_es_activo_al_corte")]
    sua_bajas = payload.get("sua_bajas") or [r for r in detalle if r.get("sua_tiene_baja")]

    wb = openpyxl.Workbook()

    ws_r = wb.active
    ws_r.title = "Resumen"
    rows_resumen = [
        ("Registro patronal SUA", display_cell(resumen.get("registro_patronal_sua", ""), empty="No detectado")),
        ("Razón social SUA", display_cell(resumen.get("razon_social_sua", ""))),
        ("Periodo SUA", display_cell(resumen.get("periodo_proceso_sua", ""))),
        ("Fecha proceso SUA", display_cell(resumen.get("fecha_proceso_sua", ""))),
        ("Fecha corte reportada", display_cell(resumen.get("fecha_corte_sua", ""))),
        ("Total cotizantes SUA", resumen.get("total_cotizantes_sua", "")),
        ("Activos SUA al corte", resumen.get("total_sua_activos_al_corte", "")),
        ("Bajas SUA del periodo", resumen.get("total_sua_bajas_periodo", "")),
        ("Headcount RAFAEL activo", resumen.get("headcount_rafael_activo", "")),
        ("Diferencia activa SUA vs HC", resumen.get("diferencia_activa_sua_vs_headcount", "")),
        ("SUA activos sin match HC", resumen.get("sua_activos_sin_match_headcount", "")),
        ("HC activos no en SUA", resumen.get("headcount_activos_no_en_sua", "")),
        ("HC activos con Baja en SUA", resumen.get("headcount_activos_con_baja_en_sua", "")),
        ("Bajas conciliadas", resumen.get("bajas_conciliadas", "")),
        ("Warnings críticos", resumen.get("warnings_criticos", "")),
        ("Patrón filtro", resumen.get("patron_filtro", "RAFAEL")),
    ]
    for i, (k, v) in enumerate(rows_resumen, start=1):
        ws_r.cell(i, 1, k).font = Font(bold=True)
        ws_r.cell(i, 2, v)
    ws_r.column_dimensions["A"].width = 32
    ws_r.column_dimensions["B"].width = 48

    ws_d = wb.create_sheet("Detalle SUA vs Headcount")
    _write_sheet_headers(ws_d, _DET_HEADERS)
    _write_detalle_rows(ws_d, detalle)
    _freeze_and_filter(ws_d, len(_DET_HEADERS), max(len(detalle) + 1, 1))
    _autosize(ws_d, len(_DET_HEADERS))

    ws_a = wb.create_sheet("Activos SUA")
    _write_sheet_headers(ws_a, _DET_HEADERS)
    _write_detalle_rows(ws_a, sua_activos)
    _freeze_and_filter(ws_a, len(_DET_HEADERS), max(len(sua_activos) + 1, 1))

    ws_bj = wb.create_sheet("Bajas SUA del periodo")
    _write_sheet_headers(ws_bj, _DET_HEADERS)
    _write_detalle_rows(ws_bj, sua_bajas)
    _freeze_and_filter(ws_bj, len(_DET_HEADERS), max(len(sua_bajas) + 1, 1))

    ws_asm = wb.create_sheet("Activos SUA sin match")
    activos_sin = [r for r in detalle if "SUA_ACTIVO_SIN_MATCH_HEADCOUNT" in (r.get("warnings") or [])]
    _write_sheet_headers(ws_asm, _DET_HEADERS)
    _write_detalle_rows(ws_asm, activos_sin)
    _freeze_and_filter(ws_asm, len(_DET_HEADERS), max(len(activos_sin) + 1, 1))

    ws_hc = wb.create_sheet("HC activos no en SUA")
    hc_headers = ["Nombre", "NSS", "CURP", "Cliente", "Ubicación", "St. Op."]
    _write_sheet_headers(ws_hc, hc_headers)
    for ri, row in enumerate(hc_sin_sua, start=2):
        ws_hc.cell(ri, 1, display_cell(row.get("nombre_completo", "")))
        from modules.headcount.matching import normalize_nss

        ws_hc.cell(ri, 2, normalize_nss(row.get("nss", "")))
        ws_hc.cell(ri, 3, display_cell(row.get("curp", "")))
        ws_hc.cell(ri, 4, display_cliente(row.get("cliente", "")))
        ws_hc.cell(ri, 5, display_ubicacion(row.get("ubicacion", "")))
        ws_hc.cell(ri, 6, display_cell(row.get("status_operacion", "")))
    _freeze_and_filter(ws_hc, len(hc_headers), max(len(hc_sin_sua) + 1, 1))

    ws_hcb = wb.create_sheet("HC activos Baja en SUA")
    hc_baja_sua = [r for r in detalle if "HEADCOUNT_ACTIVO_APARECE_BAJA_EN_SUA" in (r.get("warnings") or [])]
    _write_sheet_headers(ws_hcb, _DET_HEADERS)
    _write_detalle_rows(ws_hcb, hc_baja_sua)
    _freeze_and_filter(ws_hcb, len(_DET_HEADERS), max(len(hc_baja_sua) + 1, 1))

    ws_bc = wb.create_sheet("Bajas conciliadas")
    bajas_ok = [r for r in detalle if r.get("info_estado") == "BAJA_CONCILIADA"]
    _write_sheet_headers(ws_bc, _DET_HEADERS)
    _write_detalle_rows(ws_bc, bajas_ok)
    _freeze_and_filter(ws_bc, len(_DET_HEADERS), max(len(bajas_ok) + 1, 1))

    ws_g = wb.create_sheet("Resumen por cliente")
    g_headers = [
        "Cliente",
        "Activos SUA",
        "Bajas SUA",
        "Match activos",
        "Activos sin match",
        "Bajas conciliadas",
        "Warnings",
        "Ubicaciones",
    ]
    _write_sheet_headers(ws_g, g_headers)
    for ri, g in enumerate(resumen_clientes, start=2):
        ubics = ", ".join(
            f"{u.get('ubicacion_label', '')} ({u.get('total', 0)})" for u in (g.get("ubicaciones_list") or [])
        )
        ws_g.cell(ri, 1, g.get("cliente_label", ""))
        ws_g.cell(ri, 2, g.get("activos_sua", 0))
        ws_g.cell(ri, 3, g.get("bajas_sua", 0))
        ws_g.cell(ri, 4, g.get("match_activos", 0))
        ws_g.cell(ri, 5, g.get("activos_sin_match", 0))
        ws_g.cell(ri, 6, g.get("bajas_conciliadas", 0))
        ws_g.cell(ri, 7, g.get("warnings", 0))
        ws_g.cell(ri, 8, ubics)
    _freeze_and_filter(ws_g, len(g_headers), max(len(resumen_clientes) + 1, 1))

    ws_w = wb.create_sheet("Warnings")
    w_headers = ["Tipo", "NSS", "CURP", "Nombre", "Cliente", "Ubicación", "Descripción"]
    _write_sheet_headers(ws_w, w_headers)
    wr = 2
    for row in detalle:
        for code in row.get("warnings") or []:
            if code in _WARNINGS_OMIT_EXCEL:
                continue
            ws_w.cell(wr, 1, code)
            ws_w.cell(wr, 2, _nss_export(row))
            ws_w.cell(wr, 3, display_cell(row.get("curp", "")))
            ws_w.cell(wr, 4, display_cell(row.get("nombre_sua_original", "")))
            ws_w.cell(wr, 5, display_cliente(row.get("cliente_headcount", "")))
            ws_w.cell(wr, 6, display_ubicacion(row.get("ubicacion_headcount", "")))
            ws_w.cell(wr, 7, warning_label(code))
            wr += 1
    for row in hc_sin_sua:
        for code in row.get("warnings") or []:
            if code in _WARNINGS_OMIT_EXCEL:
                continue
            from modules.headcount.matching import normalize_nss

            ws_w.cell(wr, 1, code)
            ws_w.cell(wr, 2, normalize_nss(row.get("nss", "")))
            ws_w.cell(wr, 3, display_cell(row.get("curp", "")))
            ws_w.cell(wr, 4, display_cell(row.get("nombre_completo", "")))
            ws_w.cell(wr, 5, display_cliente(row.get("cliente", "")))
            ws_w.cell(wr, 6, display_ubicacion(row.get("ubicacion", "")))
            ws_w.cell(wr, 7, warning_label(code))
            wr += 1
    _freeze_and_filter(ws_w, len(w_headers), max(wr - 1, 1))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
