from __future__ import annotations

import json
import os
import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from modules.comparativo.comparativo_service import DATA_DIR, REPORTES_MENSUALES_DIR, obtener_historial_reportes
from modules.comparativo.headcount_service import obtener_activos

MOVIMIENTOS_DIR = os.path.join(DATA_DIR, "movimientos_imss")
EXPORTACIONES_DIR = os.path.join(DATA_DIR, "exportaciones_imss")
PATRONES_PATH = os.path.join(DATA_DIR, "patrones_extra.json")

PATRONES_FIJOS = {
    "Y3752430102": "VIGR82040319A",
    "Y6673578107": "DCI241115CD1",
}

SBC_OPCIONES = {
    "330.57": "Normal",
    "462.61": "Fronterizo",
}


def _ensure_dirs() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(MOVIMIENTOS_DIR, exist_ok=True)
    os.makedirs(EXPORTACIONES_DIR, exist_ok=True)


def _normalize_spaces(value: str) -> str:
    return " ".join((value or "").split())


def _norm_str(value: Any) -> str:
    return _normalize_spaces(str(value or "").strip())


def _norm_upper(value: Any) -> str:
    return _norm_str(value).upper()


def _safe_cliente_slug(cliente: str) -> str:
    return _normalize_spaces(str(cliente or "")).replace(" ", "_").replace("/", "-")


def _parse_date_ddmmyyyy(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _to_ddmmyyyy(value: Any) -> str:
    parsed = _parse_date_ddmmyyyy(value)
    if parsed is not None:
        return parsed.strftime("%d/%m/%Y")
    return _norm_str(value)


def _to_ddmmyyyy_compact(value: Any) -> str:
    parsed = _parse_date_ddmmyyyy(value)
    if parsed is not None:
        return parsed.strftime("%d%m%Y")
    return _norm_str(value).replace("/", "").replace("-", "")[:8]


def _to_float(value: Any) -> float:
    try:
        s = str(value or "").strip().replace(",", "")
        return float(s) if s else 0.0
    except (TypeError, ValueError):
        return 0.0


def _fmt_sbc_str(value: Any) -> str:
    return f"{_to_float(value):.2f}"


def _read_json(path: str) -> dict[str, Any] | None:
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def _write_json(path: str, data: dict[str, Any]) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


def _movimiento_path(movimiento_id: str) -> str:
    return os.path.join(MOVIMIENTOS_DIR, f"{movimiento_id}.json")


def _export_json_path(exportacion_id: str) -> str:
    return os.path.join(EXPORTACIONES_DIR, f"{exportacion_id}.json")


def _export_txt_path(exportacion_id: str) -> str:
    return os.path.join(EXPORTACIONES_DIR, f"{exportacion_id}.txt")


def _all_movimientos() -> list[dict[str, Any]]:
    _ensure_dirs()
    out: list[dict[str, Any]] = []
    for name in os.listdir(MOVIMIENTOS_DIR):
        if not name.lower().endswith(".json"):
            continue
        item = _read_json(os.path.join(MOVIMIENTOS_DIR, name))
        if isinstance(item, dict):
            out.append(item)
    return out


def _read_movimiento(movimiento_id: str) -> dict[str, Any] | None:
    return _read_json(_movimiento_path(movimiento_id))


def _coerce_tipo(value: Any) -> str:
    tipo = _norm_upper(value)
    if tipo not in {"ALTA", "BAJA"}:
        raise ValueError("tipo_movimiento debe ser ALTA o BAJA.")
    return tipo


def _base_movimiento(payload: dict[str, Any], now_iso: str, movimiento_id: str | None = None) -> dict[str, Any]:
    tipo = _coerce_tipo(payload.get("tipo_movimiento"))
    rp = _norm_str(payload.get("rp"))[:11]
    patrones = obtener_patrones()
    rfc_patron = _norm_upper(patrones.get(rp, ""))
    return {
        "id": movimiento_id or str(uuid.uuid4()),
        "tipo_movimiento": tipo,
        "rp": rp,
        "rfc_patron": rfc_patron,
        "fecha_movimiento": _to_ddmmyyyy(payload.get("fecha_movimiento")),
        "nss": _norm_str(payload.get("nss"))[:11],
        "rfc": _norm_upper(payload.get("rfc")) or None,
        "curp": _norm_upper(payload.get("curp"))[:18],
        "apellido_paterno": _norm_upper(payload.get("apellido_paterno")),
        "apellido_materno": _norm_upper(payload.get("apellido_materno")),
        "nombres": _norm_upper(payload.get("nombres")),
        "sbc": _fmt_sbc_str(payload.get("sbc")),
        "alerta": _norm_str(payload.get("alerta")) or None,
        "origen": _norm_str(payload.get("origen") or "manual") or "manual",
        "fecha_captura": now_iso,
        "fecha_actualizacion": now_iso,
    }


def _validate_required(payload: dict[str, Any]) -> None:
    required = [
        "tipo_movimiento",
        "rp",
        "fecha_movimiento",
        "nss",
        "curp",
        "apellido_paterno",
        "apellido_materno",
        "nombres",
        "sbc",
    ]
    missing = [k for k in required if not _norm_str(payload.get(k))]
    if missing:
        raise ValueError(f"Faltan campos obligatorios: {', '.join(missing)}")


def _headcount_match(value: Any, query: str) -> bool:
    return _norm_upper(value) == _norm_upper(query)


def obtener_patrones() -> dict[str, str]:
    try:
        _ensure_dirs()
        extra = _read_json(PATRONES_PATH) or {}
        merged = {str(k): _norm_upper(v) for k, v in PATRONES_FIJOS.items()}
        for rp, rfc_patron in extra.items():
            merged[_norm_str(rp)] = _norm_upper(rfc_patron)
        return merged
    except Exception as exc:
        raise ValueError(f"No se pudieron obtener patrones: {exc}") from exc


def guardar_patron_extra(rp: str, rfc_patron: str) -> dict[str, str]:
    try:
        _ensure_dirs()
        rp_norm = _norm_str(rp)
        rfc_norm = _norm_upper(rfc_patron)
        if not rp_norm or not rfc_norm:
            raise ValueError("rp y rfc_patron son obligatorios.")
        extra = _read_json(PATRONES_PATH) or {}
        extra[rp_norm] = rfc_norm
        _write_json(PATRONES_PATH, extra)
        return obtener_patrones()
    except Exception as exc:
        raise ValueError(f"No se pudo guardar patrón extra: {exc}") from exc


def guardar_movimiento(data_dict: dict[str, Any]) -> dict[str, Any]:
    try:
        _ensure_dirs()
        payload = dict(data_dict or {})
        _validate_required(payload)
        now_iso = datetime.now().isoformat()
        movimiento = _base_movimiento(payload, now_iso)
        path = _movimiento_path(movimiento["id"])
        _write_json(path, movimiento)
        return movimiento
    except Exception as exc:
        raise ValueError(f"No se pudo guardar movimiento: {exc}") from exc


def actualizar_movimiento(movimiento_id: str, data_dict: dict[str, Any]) -> dict[str, Any]:
    try:
        _ensure_dirs()
        path = _movimiento_path(movimiento_id)
        existente = _read_json(path)
        if not isinstance(existente, dict):
            raise ValueError("Movimiento no encontrado.")

        updated = dict(existente)
        for key, value in (data_dict or {}).items():
            if key == "tipo_movimiento":
                updated[key] = _coerce_tipo(value)
            elif key == "fecha_movimiento":
                updated[key] = _to_ddmmyyyy(value)
            elif key == "sbc":
                updated[key] = _fmt_sbc_str(value)
            elif key in {"rp", "nss"}:
                updated[key] = _norm_str(value)[:11]
            elif key in {"rfc", "curp", "apellido_paterno", "apellido_materno", "nombres"}:
                normalized = _norm_upper(value)
                updated[key] = normalized if key != "rfc" else (normalized or None)
            elif key in {"alerta", "origen"}:
                updated[key] = _norm_str(value) or None
            elif key not in {"id", "fecha_captura", "fecha_actualizacion", "rfc_patron"}:
                updated[key] = value

        if "rp" in (data_dict or {}):
            updated["rfc_patron"] = _norm_upper(obtener_patrones().get(_norm_str(updated.get("rp")), ""))
        updated["fecha_actualizacion"] = datetime.now().isoformat()
        _write_json(path, updated)
        return updated
    except Exception as exc:
        raise ValueError(f"No se pudo actualizar movimiento: {exc}") from exc


def eliminar_movimiento(movimiento_id: str) -> dict[str, bool]:
    try:
        _ensure_dirs()
        path = _movimiento_path(movimiento_id)
        if not os.path.exists(path):
            raise ValueError("Movimiento no encontrado.")
        os.remove(path)
        return {"ok": True}
    except Exception as exc:
        raise ValueError(f"No se pudo eliminar movimiento: {exc}") from exc


def obtener_movimientos(tipo: str | None = None) -> list[dict[str, Any]]:
    try:
        out = _all_movimientos()
        if tipo:
            tipo_obj = _norm_upper(tipo)
            out = [m for m in out if _norm_upper(m.get("tipo_movimiento")) == tipo_obj]
        out.sort(key=lambda x: str(x.get("fecha_captura", "")), reverse=True)
        return out
    except Exception as exc:
        raise ValueError(f"No se pudieron obtener movimientos: {exc}") from exc


def buscar_en_headcount(query: str, campo: str) -> dict[str, Any]:
    try:
        campo_obj = _norm_str(campo)
        if campo_obj not in {"nss", "rfc_homoclave", "curp", "nombre_completo"}:
            raise ValueError("campo inválido para búsqueda.")
        q = _norm_upper(query)
        if not q:
            return {"encontrado": False}
        coincidencias = [t for t in obtener_activos() if _headcount_match(t.get(campo_obj, ""), q)]
        if not coincidencias:
            return {"encontrado": False}
        if len(coincidencias) == 1:
            return {"encontrado": True, "duplicado": False, "datos": coincidencias[0]}
        return {"encontrado": True, "duplicado": True, "opciones": coincidencias}
    except Exception as exc:
        raise ValueError(f"No se pudo buscar en headcount: {exc}") from exc


def mapear_headcount_a_movimiento(trabajador_dict: dict[str, Any]) -> dict[str, Any]:
    try:
        t = trabajador_dict or {}
        return {
            "nss": _norm_str(t.get("nss")),
            "rfc": _norm_upper(t.get("rfc_homoclave")) or None,
            "curp": _norm_upper(t.get("curp")),
            "apellido_paterno": _norm_upper(t.get("apellido_paterno")),
            "apellido_materno": _norm_upper(t.get("apellido_materno")),
            "nombres": _norm_upper(t.get("nombre")),
            "sbc": _fmt_sbc_str(t.get("sueldo_diario")),
        }
    except Exception as exc:
        raise ValueError(f"No se pudo mapear headcount a movimiento: {exc}") from exc


def cargar_desde_excel(file) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        header_row = None
        header_map: dict[str, int] = {}
        aliases = {
            "NSS": "nss",
            "RFC": "rfc",
            "CURP": "curp",
            "APELLIDO PATERNO": "apellido_paterno",
            "APELLIDO MATERNO": "apellido_materno",
            "NOMBRES": "nombres",
            "SBC": "sbc",
            "TIPO MOVIMIENTO": "tipo_movimiento",
            "RP": "rp",
            "FECHA MOVIMIENTO": "fecha_movimiento",
        }
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            normalized = [_norm_upper(cell.value) for cell in row]
            if "NSS" in normalized:
                header_row = row[0].row
                for idx, val in enumerate(normalized):
                    if val in aliases:
                        header_map[aliases[val]] = idx
                break
        if header_row is None:
            raise ValueError("No se encontró fila de encabezados con NSS.")

        required_fields = [
            "nss",
            "curp",
            "apellido_paterno",
            "apellido_materno",
            "nombres",
            "sbc",
            "tipo_movimiento",
            "rp",
            "fecha_movimiento",
        ]
        out: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            values: dict[str, Any] = {}
            empty_row = True
            for key, col_idx in header_map.items():
                cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                if cell_value not in (None, ""):
                    empty_row = False
                values[key] = cell_value
            if empty_row:
                continue

            tipo = _norm_upper(values.get("tipo_movimiento"))
            if tipo in {"ALTA", "BAJA"}:
                tipo_mov = tipo
            else:
                tipo_mov = tipo
            fecha_mov = _to_ddmmyyyy(values.get("fecha_movimiento"))
            rp = _norm_str(values.get("rp"))[:11]
            base = {
                "id": str(uuid.uuid4()),
                "tipo_movimiento": tipo_mov,
                "rp": rp,
                "rfc_patron": _norm_upper(obtener_patrones().get(rp, "")),
                "fecha_movimiento": fecha_mov,
                "nss": _norm_str(values.get("nss"))[:11],
                "rfc": _norm_upper(values.get("rfc")) or None,
                "curp": _norm_upper(values.get("curp"))[:18],
                "apellido_paterno": _norm_upper(values.get("apellido_paterno")),
                "apellido_materno": _norm_upper(values.get("apellido_materno")),
                "nombres": _norm_upper(values.get("nombres")),
                "sbc": _fmt_sbc_str(values.get("sbc")),
                "alerta": None,
                "origen": "excel",
                "fecha_captura": datetime.now().isoformat(),
                "fecha_actualizacion": datetime.now().isoformat(),
            }
            faltantes = [k for k in required_fields if not _norm_str(values.get(k))]
            if tipo_mov not in {"ALTA", "BAJA"}:
                faltantes.append("tipo_movimiento inválido")
            if faltantes:
                base["alerta"] = f"Datos incompletos: {', '.join(faltantes)}"
            out.append(base)
        return out
    except Exception as exc:
        raise ValueError(f"No se pudo cargar desde Excel: {exc}") from exc


def cargar_desde_reporte_mensual(
    cliente: str,
    mes: int,
    anio: int,
    incluir_fijos: bool = False,
) -> list[dict[str, Any]]:
    try:
        cliente_slug = _safe_cliente_slug(cliente)
        path = os.path.join(REPORTES_MENSUALES_DIR, f"{cliente_slug}_{int(anio):04d}-{int(mes):02d}.json")
        reporte = _read_json(path)
        if not isinstance(reporte, dict):
            raise ValueError("Reporte no encontrado")

        mes_inicio = date(int(anio), int(mes), 1)
        if int(mes) == 12:
            mes_fin = date(int(anio) + 1, 1, 1).fromordinal(date(int(anio) + 1, 1, 1).toordinal() - 1)
        else:
            mes_fin = date(int(anio), int(mes) + 1, 1).fromordinal(date(int(anio), int(mes) + 1, 1).toordinal() - 1)

        patrones = obtener_patrones()
        out: list[dict[str, Any]] = []

        def _build_mov(nombre: str, tipo: str, fecha_mov: str, extra_alerta: str | None = None) -> dict[str, Any]:
            match = buscar_en_headcount(nombre, "nombre_completo")
            hc_data: dict[str, Any] | None = None
            alerta = extra_alerta
            if bool(match.get("encontrado")):
                if bool(match.get("duplicado")):
                    opciones = match.get("opciones") or []
                    hc_data = opciones[0] if isinstance(opciones, list) and opciones else None
                    alerta = "Coincidencia duplicada en headcount"
                else:
                    hc_data = match.get("datos")
            if hc_data:
                mapped = mapear_headcount_a_movimiento(hc_data)
                rp_val = _norm_str(hc_data.get("patron"))[:11]
            else:
                mapped = {
                    "nss": "",
                    "rfc": None,
                    "curp": "",
                    "apellido_paterno": "",
                    "apellido_materno": "",
                    "nombres": _norm_upper(nombre),
                    "sbc": "0.00",
                }
                rp_val = ""
                alerta = alerta or "No encontrado en headcount"
            if not _norm_str(fecha_mov):
                alerta = "Sin fecha de movimiento"
            mov = {
                "id": str(uuid.uuid4()),
                "tipo_movimiento": tipo,
                "rp": rp_val,
                "rfc_patron": _norm_upper(patrones.get(rp_val, "")),
                "fecha_movimiento": _to_ddmmyyyy(fecha_mov),
                "nss": mapped.get("nss", ""),
                "rfc": mapped.get("rfc"),
                "curp": mapped.get("curp", ""),
                "apellido_paterno": mapped.get("apellido_paterno", ""),
                "apellido_materno": mapped.get("apellido_materno", ""),
                "nombres": mapped.get("nombres", _norm_upper(nombre)),
                "sbc": _fmt_sbc_str(mapped.get("sbc")),
                "alerta": alerta or None,
                "origen": "reporte_mensual",
                "fecha_captura": datetime.now().isoformat(),
                "fecha_actualizacion": datetime.now().isoformat(),
            }
            return mov

        for rot in reporte.get("rotativos", []):
            if not isinstance(rot, dict):
                continue
            nombre = _norm_upper(rot.get("nombre"))
            if not nombre:
                continue
            altas = rot.get("fechas_alta") if isinstance(rot.get("fechas_alta"), list) else []
            bajas = rot.get("fechas_baja") if isinstance(rot.get("fechas_baja"), list) else []
            for fecha_alta in altas:
                dt = _parse_date_ddmmyyyy(fecha_alta)
                if dt and mes_inicio <= dt <= mes_fin:
                    out.append(_build_mov(nombre, "ALTA", _to_ddmmyyyy(fecha_alta)))
            for fecha_baja in bajas:
                dt = _parse_date_ddmmyyyy(fecha_baja)
                if dt and mes_inicio <= dt <= mes_fin:
                    out.append(_build_mov(nombre, "BAJA", _to_ddmmyyyy(fecha_baja)))
            if not altas and not bajas:
                out.append(_build_mov(nombre, "ALTA", "", "Sin fecha de movimiento"))

        if incluir_fijos:
            for fijo in reporte.get("fijos", []):
                if not isinstance(fijo, dict):
                    continue
                nombre = _norm_upper(fijo.get("nombre"))
                if not nombre:
                    continue
                fecha_alta = ""
                head = buscar_en_headcount(nombre, "nombre_completo")
                if bool(head.get("encontrado")) and not bool(head.get("duplicado")):
                    datos = head.get("datos") or {}
                    fecha_alta = _to_ddmmyyyy(datos.get("fecha_ingreso"))
                alerta = None if _norm_str(fecha_alta) else "Sin fecha de alta"
                out.append(_build_mov(nombre, "ALTA", fecha_alta, alerta))
        return out
    except Exception as exc:
        raise ValueError(f"No se pudo cargar desde reporte mensual: {exc}") from exc


def generar_txt_idse(movimientos_ids: list[str]) -> str:
    try:
        lineas: list[str] = []
        consecutivo = 1
        for mov_id in movimientos_ids:
            mov = _read_movimiento(_norm_str(mov_id))
            if not isinstance(mov, dict):
                continue
            row = [
                str(consecutivo),
                _norm_str(mov.get("rfc_patron")),
                _norm_str(mov.get("rp")),
                _norm_str(mov.get("nss")),
                _norm_upper(mov.get("apellido_paterno")),
                _norm_upper(mov.get("apellido_materno")),
                _norm_upper(mov.get("nombres")),
                _fmt_sbc_str(mov.get("sbc")),
                "1",
                "0",
                "0",
                _to_ddmmyyyy_compact(mov.get("fecha_movimiento")),
                _norm_upper(mov.get("curp")),
            ]
            lineas.append("\t".join(row))
            consecutivo += 1
        return "\n".join(lineas)
    except Exception as exc:
        raise ValueError(f"No se pudo generar TXT IDSE: {exc}") from exc


def generar_txt_sua(movimientos_ids: list[str]) -> str:
    try:
        lineas: list[str] = []
        for mov_id in movimientos_ids:
            mov = _read_movimiento(_norm_str(mov_id))
            if not isinstance(mov, dict):
                continue
            if _norm_upper(mov.get("tipo_movimiento")) != "ALTA":
                continue
            rp = _norm_str(mov.get("rp"))[:11].ljust(11)
            nss = _norm_str(mov.get("nss"))[:11].ljust(11)
            rfc = (_norm_upper(mov.get("rfc")) + (" " * 13))[:13]
            curp = (_norm_upper(mov.get("curp")) + (" " * 18))[:18]
            nombre_comp = f"{_norm_upper(mov.get('apellido_paterno'))}${_norm_upper(mov.get('apellido_materno'))}${_norm_upper(mov.get('nombres'))}"
            nombre_fmt = (nombre_comp + (" " * 50))[:50]
            fecha = _to_ddmmyyyy_compact(mov.get("fecha_movimiento")).zfill(8)[:8]
            sbc_cent = int(round(_to_float(mov.get("sbc")) * 100))
            sbc_fmt = str(sbc_cent).zfill(7)[-7:]
            linea = (
                rp
                + nss
                + rfc
                + curp
                + nombre_fmt
                + "1"
                + "0"
                + fecha
                + sbc_fmt
                + (" " * 17)
                + (" " * 10)
                + "00000000"
                + "0"
                + "00000000"
                + "00000000"
            )
            lineas.append(linea[:164].ljust(164))
        return "\n".join(lineas)
    except Exception as exc:
        raise ValueError(f"No se pudo generar TXT SUA: {exc}") from exc


def guardar_exportacion(
    movimientos_ids: list[str],
    tipo_export: str,
    txt_content: str,
    rp: str,
) -> dict[str, Any]:
    try:
        _ensure_dirs()
        export_id = str(uuid.uuid4())
        now_iso = datetime.now().isoformat()
        rp_norm = _norm_str(rp)
        patrones = obtener_patrones()
        movimientos = [_read_movimiento(mid) for mid in movimientos_ids]
        movimientos_validos = [m for m in movimientos if isinstance(m, dict)]
        altas = sum(1 for m in movimientos_validos if _norm_upper(m.get("tipo_movimiento")) == "ALTA")
        bajas = sum(1 for m in movimientos_validos if _norm_upper(m.get("tipo_movimiento")) == "BAJA")
        txt_filename = f"{export_id}.txt"
        txt_path = _export_txt_path(export_id)
        with open(txt_path, "w", encoding="utf-8") as fh:
            fh.write(str(txt_content or ""))
        metadata = {
            "id": export_id,
            "rp": rp_norm,
            "rfc_patron": _norm_upper(patrones.get(rp_norm, "")),
            "tipo_export": _norm_upper(tipo_export),
            "fecha_exportacion": now_iso,
            "total_movimientos": len(movimientos_ids),
            "altas": altas,
            "bajas": bajas,
            "movimientos_ids": [str(m) for m in movimientos_ids],
            "txt_filename": txt_filename,
        }
        _write_json(_export_json_path(export_id), metadata)
        return metadata
    except Exception as exc:
        raise ValueError(f"No se pudo guardar exportación: {exc}") from exc


def obtener_historial_exportaciones() -> list[dict[str, Any]]:
    try:
        _ensure_dirs()
        out: list[dict[str, Any]] = []
        for name in os.listdir(EXPORTACIONES_DIR):
            if not name.lower().endswith(".json"):
                continue
            item = _read_json(os.path.join(EXPORTACIONES_DIR, name))
            if isinstance(item, dict):
                out.append(item)
        out.sort(key=lambda x: str(x.get("fecha_exportacion", "")), reverse=True)
        return out
    except Exception as exc:
        raise ValueError(f"No se pudo obtener historial de exportaciones: {exc}") from exc


def obtener_txt_exportacion(exportacion_id: str, formato: str | None = None) -> str:
    try:
        _ensure_dirs()
        meta = _read_json(_export_json_path(exportacion_id))
        if not isinstance(meta, dict):
            raise ValueError("Exportación no encontrada.")
        txt_name = _norm_str(meta.get("txt_filename")) or f"{exportacion_id}.txt"
        txt_path = os.path.join(EXPORTACIONES_DIR, txt_name)
        if not os.path.exists(txt_path):
            raise ValueError("TXT de exportación no encontrado.")
        with open(txt_path, "r", encoding="utf-8") as fh:
            content = fh.read()
        fmt = _norm_upper(formato)
        if meta.get("tipo_export") == "AMBOS" and fmt in {"IDSE", "SUA"}:
            sep = "\n---SUA---\n"
            if sep in content:
                idse_txt, sua_txt = content.split(sep, 1)
                return idse_txt if fmt == "IDSE" else sua_txt
        return content
    except Exception as exc:
        raise ValueError(f"No se pudo obtener TXT de exportación: {exc}") from exc


def obtener_reporte_mensuales_disponibles() -> list[dict[str, Any]]:
    try:
        return obtener_historial_reportes()
    except Exception as exc:
        raise ValueError(f"No se pudo obtener reportes mensuales disponibles: {exc}") from exc


def eliminar_exportacion(exportacion_id: str) -> dict[str, bool]:
    try:
        _ensure_dirs()
        meta = _read_json(_export_json_path(exportacion_id))
        if not isinstance(meta, dict):
            raise ValueError("Exportación no encontrada.")
        txt_name = _norm_str(meta.get("txt_filename")) or f"{exportacion_id}.txt"
        json_path = _export_json_path(exportacion_id)
        txt_path = os.path.join(EXPORTACIONES_DIR, txt_name)
        if os.path.exists(json_path):
            os.remove(json_path)
        if os.path.exists(txt_path):
            os.remove(txt_path)
        return {"ok": True}
    except Exception as exc:
        raise ValueError(f"No se pudo eliminar exportación: {exc}") from exc
