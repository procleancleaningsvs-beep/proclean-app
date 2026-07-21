from __future__ import annotations

import io
import os
from functools import wraps

from flask import Blueprint, Response, g, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from modules.comparativo.comparativo_service import obtener_historial as obtener_historial_comparativos
from modules.exportacion_imss.exportacion_service import (
    SBC_OPCIONES,
    actualizar_movimiento,
    buscar_en_headcount,
    cargar_desde_comparativo_semanal,
    cargar_desde_excel,
    cargar_desde_reporte_mensual,
    eliminar_todos_movimientos,
    eliminar_exportacion,
    eliminar_movimiento,
    generar_txt_idse,
    generar_txt_sua,
    guardar_exportacion,
    guardar_movimiento,
    guardar_movimientos_bulk,
    guardar_patron_extra,
    obtener_historial_exportaciones,
    obtener_movimientos,
    obtener_patrones,
    obtener_reporte_mensuales_disponibles,
    obtener_txt_exportacion,
)

_BASE = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
_TEMPLATE_DIR = os.path.join(_BASE, "templates", "exportacion_imss")

exportacion_imss_bp = Blueprint(
    "exportacion_imss",
    __name__,
    url_prefix="/exportacion-imss",
    template_folder=_TEMPLATE_DIR,
)


def _login_required_page(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return jsonify({"ok": False, "error": "No autenticado."}), 401
        return view(*args, **kwargs)

    return wrapped


def _is_admin() -> bool:
    user = g.user
    if user is None:
        return False
    if isinstance(user, dict):
        role = str(user.get("rol") or user.get("role") or "").strip().lower()
        return bool(user.get("is_admin")) or role in {"admin", "administrador"}
    role = str(getattr(user, "rol", "") or getattr(user, "role", "")).strip().lower()
    return bool(getattr(user, "is_admin", False)) or role in {"admin", "administrador"}


@exportacion_imss_bp.get("/")
@_login_required_page
def index():
    return render_template("exportacion_imss/index.html")


@exportacion_imss_bp.get("/patrones")
@_login_required_page
def patrones_get():
    try:
        return jsonify({"ok": True, "patrones": obtener_patrones(), "sbc_opciones": SBC_OPCIONES})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/patrones")
@_login_required_page
def patrones_post():
    data = request.get_json(silent=True) or {}
    rp = str(data.get("rp") or "").strip()
    rfc_patron = str(data.get("rfc_patron") or "").strip()
    try:
        updated = guardar_patron_extra(rp, rfc_patron)
        return jsonify({"ok": True, "patrones": updated})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.get("/movimientos")
@_login_required_page
def movimientos_get():
    try:
        tipo = (request.args.get("tipo") or "").strip() or None
        return jsonify({"ok": True, "movimientos": obtener_movimientos(tipo=tipo)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/movimientos")
@_login_required_page
def movimientos_post():
    data = request.get_json(silent=True) or {}
    try:
        mov = guardar_movimiento(data)
        return jsonify({"ok": True, "movimiento": mov})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/movimientos/bulk")
@_login_required_page
def movimientos_bulk_post():
    data = request.get_json(silent=True) or {}
    movimientos = data.get("movimientos")
    if not isinstance(movimientos, list):
        return jsonify({"ok": False, "error": "movimientos debe ser una lista."}), 400
    try:
        result = guardar_movimientos_bulk(movimientos)
        return jsonify({"ok": True, **result})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.put("/movimientos/<movimiento_id>")
@_login_required_page
def movimientos_put(movimiento_id: str):
    data = request.get_json(silent=True) or {}
    try:
        mov = actualizar_movimiento(movimiento_id, data)
        return jsonify({"ok": True, "movimiento": mov})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.delete("/movimientos/<movimiento_id>")
@_login_required_page
def movimientos_delete(movimiento_id: str):
    try:
        return jsonify(eliminar_movimiento(movimiento_id))
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.delete("/movimientos/todos")
@_login_required_page
def movimientos_delete_all():
    try:
        res = eliminar_todos_movimientos()
        return jsonify({"ok": True, "eliminados": int(res.get("eliminados", 0))})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/buscar-headcount")
@_login_required_page
def buscar_headcount():
    data = request.get_json(silent=True) or {}
    query = str(data.get("query") or "").strip()
    campo = str(data.get("campo") or "").strip()
    try:
        return jsonify({"ok": True, **buscar_en_headcount(query, campo)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/cargar-excel")
@_login_required_page
def cargar_excel():
    excel_file = request.files.get("excel_file")
    if excel_file is None:
        return jsonify({"ok": False, "error": "excel_file es obligatorio."}), 400
    try:
        movimientos = cargar_desde_excel(excel_file)
        con_alertas = sum(1 for m in movimientos if str(m.get("alerta") or "").strip())
        return jsonify({"ok": True, "movimientos": movimientos, "total": len(movimientos), "con_alertas": con_alertas})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.get("/reportes-mensuales-disponibles")
@_login_required_page
def reportes_mensuales_disponibles():
    try:
        return jsonify({"ok": True, "reportes": obtener_reporte_mensuales_disponibles()})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.get("/comparativos-disponibles")
@_login_required_page
def comparativos_disponibles():
    try:
        hist = obtener_historial_comparativos()
        items = []
        for c in hist:
            if not isinstance(c, dict):
                continue
            items.append(
                {
                    "id": str(c.get("id") or "").strip(),
                    "cliente": str(c.get("cliente") or "").strip(),
                    "periodo_inicio": str(c.get("periodo_inicio") or "").strip(),
                    "periodo_fin": str(c.get("periodo_fin") or "").strip(),
                    "altas": len(c.get("altas")) if isinstance(c.get("altas"), list) else 0,
                    "bajas": len(c.get("bajas")) if isinstance(c.get("bajas"), list) else 0,
                }
            )
        return jsonify({"ok": True, "items": items})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/cargar-reporte-mensual")
@_login_required_page
def cargar_reporte_mensual():
    data = request.get_json(silent=True) or {}
    cliente = str(data.get("cliente") or "").strip()
    mes = data.get("mes")
    anio = data.get("anio")
    incluir_fijos = bool(data.get("incluir_fijos", False))
    if not cliente or mes is None or anio is None:
        return jsonify({"ok": False, "error": "cliente, mes y anio son obligatorios."}), 400
    try:
        movimientos = cargar_desde_reporte_mensual(cliente, int(mes), int(anio), incluir_fijos=incluir_fijos)
        con_alertas = sum(1 for m in movimientos if str(m.get("alerta") or "").strip())
        fijos_incluidos = sum(1 for m in movimientos if m.get("origen") == "reporte_mensual" and m.get("tipo_movimiento") == "ALTA") if incluir_fijos else 0
        return jsonify(
            {
                "ok": True,
                "movimientos": movimientos,
                "total": len(movimientos),
                "con_alertas": con_alertas,
                "fijos_incluidos": fijos_incluidos,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.post("/cargar-comparativo-semanal")
@_login_required_page
def cargar_comparativo_semanal():
    data = request.get_json(silent=True) or {}
    comparativo_id = str(data.get("comparativo_id") or "").strip()
    rp_general = str(data.get("rp_general") or "").strip() or None
    sbc_general = str(data.get("sbc_general") or "").strip() or None
    if not comparativo_id:
        return jsonify({"ok": False, "error": "comparativo_id es obligatorio."}), 400
    try:
        movimientos = cargar_desde_comparativo_semanal(
            comparativo_id,
            rp_general=rp_general,
            sbc_general=sbc_general,
        )
        con_alertas = sum(1 for m in movimientos if str(m.get("alerta") or "").strip())
        return jsonify({"ok": True, "movimientos": movimientos, "total": len(movimientos), "con_alertas": con_alertas})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.route("/debug-reporte/<cliente>/<int:anio>/<int:mes>")
@_login_required_page
def debug_reporte(cliente: str, anio: int, mes: int):
    import json
    import os

    from modules.comparativo.comparativo_service import DATA_DIR

    slug = cliente.replace(" ", "_").replace("/", "-")
    path = os.path.join(DATA_DIR, "reportes_mensuales", f"{slug}_{anio:04d}-{mes:02d}.json")
    if not os.path.exists(path):
        path2 = os.path.join(DATA_DIR, "reportes_mensuales", f"{cliente}_{anio:04d}-{mes:02d}.json")
        if os.path.exists(path2):
            path = path2
        else:
            return {"error": "No existe", "path": path}
    with open(path, encoding="utf-8") as f:
        r = json.load(f)
    fijos = r.get("fijos", [])
    rotativos = r.get("rotativos", [])
    return {
        "total_fijos": len(fijos),
        "total_rotativos": len(rotativos),
        "keys_reporte": list(r.keys()),
        "fijo_ejemplo": fijos[0] if fijos else None,
        "rotativo_ejemplo": rotativos[0] if rotativos else None,
    }


@exportacion_imss_bp.post("/exportar")
@_login_required_page
def exportar():
    data = request.get_json(silent=True) or {}
    movimientos_ids = data.get("movimientos_ids") or []
    tipo_export = str(data.get("tipo_export") or "").strip().upper()
    rp = str(data.get("rp") or "").strip()
    if not isinstance(movimientos_ids, list) or not movimientos_ids:
        return jsonify({"ok": False, "error": "movimientos_ids debe ser lista no vacía."}), 400
    if tipo_export not in {"IDSE", "SUA", "AMBOS"}:
        return jsonify({"ok": False, "error": "tipo_export inválido."}), 400
    if not rp:
        return jsonify({"ok": False, "error": "rp es obligatorio."}), 400
    try:
        existentes = {m.get("id"): m for m in obtener_movimientos()}
        faltantes = [mid for mid in movimientos_ids if mid not in existentes]
        if faltantes:
            return jsonify({"ok": False, "error": f"IDs no encontrados: {', '.join(faltantes)}"}), 400
        txt_idse = generar_txt_idse(movimientos_ids) if tipo_export in {"IDSE", "AMBOS"} else None
        txt_sua = generar_txt_sua(movimientos_ids) if tipo_export in {"SUA", "AMBOS"} else None
        if tipo_export == "AMBOS":
            txt_content = f"{txt_idse or ''}\n---SUA---\n{txt_sua or ''}"
        else:
            txt_content = txt_idse if tipo_export == "IDSE" else txt_sua
        meta = guardar_exportacion(movimientos_ids, tipo_export, txt_content or "", rp)
        return jsonify({"ok": True, "exportacion_id": meta.get("id"), "txt_idse": txt_idse, "txt_sua": txt_sua})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.get("/exportar/<exportacion_id>/descargar")
@_login_required_page
def descargar_exportacion(exportacion_id: str):
    formato = str(request.args.get("formato") or "").strip().upper()
    if formato not in {"IDSE", "SUA"}:
        return jsonify({"ok": False, "error": "formato inválido, usa IDSE o SUA."}), 400
    try:
        txt = obtener_txt_exportacion(exportacion_id, formato=formato)
        filename = f"exportacion_{exportacion_id}_{formato}.txt"
        return Response(
            txt,
            mimetype="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.get("/historial-exportaciones")
@_login_required_page
def historial_exportaciones():
    try:
        return jsonify({"ok": True, "items": obtener_historial_exportaciones()})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.delete("/historial-exportaciones/<exportacion_id>")
@_login_required_page
def historial_exportaciones_delete(exportacion_id: str):
    if not _is_admin():
        return jsonify({"ok": False, "error": "Solo admin puede eliminar exportaciones."}), 403
    try:
        return jsonify(eliminar_exportacion(exportacion_id))
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@exportacion_imss_bp.get("/plantilla-excel")
@_login_required_page
def plantilla_excel():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla IMSS"
        headers = [
            "NSS",
            "RFC",
            "CURP",
            "APELLIDO PATERNO",
            "APELLIDO MATERNO",
            "NOMBRES",
            "SBC",
            "TIPO MOVIMIENTO",
            "RP",
            "FECHA MOVIMIENTO",
        ]
        ws.append(headers)
        ws.append(
            [
                "12345678901",
                "XAXX010101000",
                "XAXX010101HDFRRL01",
                "PEREZ",
                "LOPEZ",
                "JUAN",
                "330.57",
                "ALTA",
                "Y3752430102",
                "15/04/2026",
            ]
        )
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
        gray_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
        for cell in ws[2]:
            cell.fill = gray_fill
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="plantilla_exportacion_imss.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
