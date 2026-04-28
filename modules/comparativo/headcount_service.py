from __future__ import annotations

import json
import os
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

DATA_DIR = os.environ.get("DATA_DIR", "./data")
HEADCOUNT_PATH = os.path.join(DATA_DIR, "headcount.json")


def _normalize_spaces(value: str) -> str:
    return " ".join((value or "").split())


def _normalize_name(value: Any) -> str:
    return _normalize_spaces(str(value or "").upper().strip())


def _normalize_header(value: Any) -> str:
    return _normalize_spaces(str(value or "").strip().upper())


def _format_fecha(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s


def _load_headcount() -> list[dict[str, Any]]:
    if not os.path.exists(HEADCOUNT_PATH):
        return []
    with open(HEADCOUNT_PATH, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    return data if isinstance(data, list) else []


def actualizar_headcount(file) -> dict[str, Any]:
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        header_row_idx = None
        header_map: dict[str, int] = {}
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            normalized = [_normalize_header(cell) for cell in row]
            if "STATUS OPERACIÓN" in normalized or "STATUS OPERACION" in normalized:
                header_row_idx = row_idx
                header_map = {normalized[idx]: idx for idx in range(len(normalized))}
                break

        if header_row_idx is None:
            raise ValueError("No se encontró la fila de encabezados con 'STATUS OPERACIÓN'.")

        def col(name: str) -> int | None:
            if name in header_map:
                return header_map[name]
            alt = name.replace("Ó", "O")
            if alt in header_map:
                return header_map[alt]
            return None

        required = [
            "STATUS OPERACIÓN",
            "NOMBRE COMPLETO",
            "APELLIDO PATERNO",
            "APELLIDO MATERNO",
            "NOMBRE",
            "CLIENTE",
            "PATRON",
            "FECHA DE INGRESO",
            "SUELDO DIARIO",
            "PUESTO",
            "NSS",
            "RFC HOMOCLAVE",
            "CURP",
            "CP FISCAL",
            "STATUS IMSS",
            "GENERO",
        ]
        missing = [key for key in required if col(key) is None]
        if missing:
            raise ValueError(f"Faltan columnas requeridas en headcount: {', '.join(missing)}")

        activos: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            status_raw = row[col("STATUS OPERACIÓN")] if col("STATUS OPERACIÓN") is not None else ""
            status = _normalize_spaces(str(status_raw or "").strip().upper())
            if status != "ALTA":
                continue

            record = {
                "nombre_completo": _normalize_name(row[col("NOMBRE COMPLETO")]),
                "apellido_paterno": _normalize_spaces(str(row[col("APELLIDO PATERNO")] or "").strip()),
                "apellido_materno": _normalize_spaces(str(row[col("APELLIDO MATERNO")] or "").strip()),
                "nombre": _normalize_spaces(str(row[col("NOMBRE")] or "").strip()),
                "cliente": _normalize_spaces(str(row[col("CLIENTE")] or "").strip()),
                "patron": _normalize_spaces(str(row[col("PATRON")] or "").strip()),
                "fecha_ingreso": _format_fecha(row[col("FECHA DE INGRESO")]),
                "sueldo_diario": row[col("SUELDO DIARIO")],
                "puesto": _normalize_spaces(str(row[col("PUESTO")] or "").strip()),
                "nss": _normalize_spaces(str(row[col("NSS")] or "").strip()),
                "rfc_homoclave": _normalize_spaces(str(row[col("RFC HOMOCLAVE")] or "").strip()),
                "curp": _normalize_spaces(str(row[col("CURP")] or "").strip()),
                "cp_fiscal": _normalize_spaces(str(row[col("CP FISCAL")] or "").strip()),
                "status_imss": _normalize_spaces(str(row[col("STATUS IMSS")] or "").strip()),
                "genero": _normalize_spaces(str(row[col("GENERO")] or "").strip()),
            }
            if not record["nombre_completo"]:
                continue
            activos.append(record)

        with open(HEADCOUNT_PATH, "w", encoding="utf-8") as fh:
            json.dump(activos, fh, ensure_ascii=False, indent=2)

        clientes = sorted({item.get("cliente", "") for item in activos if item.get("cliente")})
        return {
            "total_activos": len(activos),
            "fecha_actualizacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "clientes_detectados": clientes,
        }
    except Exception as exc:
        raise ValueError(f"No se pudo actualizar headcount: {exc}") from exc


def obtener_activos(cliente: str | None = None) -> list[dict[str, Any]]:
    try:
        activos = _load_headcount()
        if not cliente:
            return activos
        filtro = str(cliente).strip().casefold()
        return [item for item in activos if str(item.get("cliente", "")).strip().casefold() == filtro]
    except Exception as exc:
        raise ValueError(f"No se pudo leer headcount: {exc}") from exc


def buscar_trabajador(nombre_completo: str) -> dict[str, Any] | None:
    try:
        if not nombre_completo:
            return None
        objetivo = _normalize_name(nombre_completo)
        for item in _load_headcount():
            if _normalize_name(item.get("nombre_completo", "")) == objetivo:
                return item
        return None
    except Exception as exc:
        raise ValueError(f"No se pudo buscar trabajador: {exc}") from exc
