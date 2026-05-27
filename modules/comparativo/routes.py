from __future__ import annotations

import io
import json
import logging
import math
import os
from functools import wraps
from urllib.parse import unquote

from flask import Blueprint, abort, g, jsonify, redirect, render_template, request, send_file, flash, url_for
from openpyxl import Workbook

from modules.comparativo import alias_service
from modules.comparativo.comparativo_service import (
    aplicar_aliases_y_comparar,
    guardar_reporte_mensual,
    generar_reporte_mensual_v2,
    obtener_historial_reportes,
    obtener_resumen_nominas_por_cliente,
    REPORTES_MENSUALES_DIR,
    detectar_similitudes,
    generar_reporte_mensual,
    guardar_comparativo_semanal,
    guardar_nomina_semana,
    obtener_historial,
    obtener_nominas_guardadas,
    parsear_nomina,
)
from modules.comparativo.headcount_service import actualizar_headcount, obtener_activos, obtener_metadata_headcount
from modules.roles_access import can_access_comparativo, normalized_role

DATA_DIR = os.environ.get("DATA_DIR", "./data")
COMPARATIVOS_DIR = os.path.join(DATA_DIR, "comparativos")
_BASE = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
_TEMPLATE_DIR = os.path.join(_BASE, "templates", "comparativo")

comparativo_bp = Blueprint(
    "comparativo",
    __name__,
    url_prefix="/comparativo",
    template_folder=_TEMPLATE_DIR,
)

logger = logging.getLogger(__name__)


@comparativo_bp.before_request
def _comparativo_role_guard() -> None:
    if g.user is None:
        return
    if not can_access_comparativo(normalized_role(g.user)):
        abort(403)


def _login_required_page(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def _ensure_dirs() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(COMPARATIVOS_DIR, exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, "nominas"), exist_ok=True)
    os.makedirs(REPORTES_MENSUALES_DIR, exist_ok=True)
    try:
        os.makedirs("/app/data/comparativos", exist_ok=True)
    except OSError:
        pass


def _clientes_disponibles() -> list[str]:
    try:
        return sorted({a.get("cliente", "") for a in obtener_activos() if a.get("cliente")})
    except Exception as exc:
        logger.warning("comparativo: no se pudieron listar clientes desde headcount: %s", exc)
        return []


def _resolver_activos_por_selector(cliente_selector: str) -> list[dict]:
    try:
        agrupaciones = alias_service.obtener_agrupaciones()
        if cliente_selector in agrupaciones:
            activos: list[dict] = []
            for c in agrupaciones.get(cliente_selector, []):
                activos.extend(obtener_activos(cliente=str(c)))
            dedup: dict[str, dict] = {}
            for item in activos:
                key = str(item.get("nombre_completo", "")).strip().upper()
                if key:
                    dedup[key] = item
            return list(dedup.values())
        return obtener_activos(cliente=cliente_selector)
    except Exception as exc:
        logger.warning("comparativo: no se pudieron resolver activos (%s): %s", cliente_selector, exc)
        return []


def _buscar_comparativo_duplicado(cliente: str, periodo_inicio: str, periodo_fin: str) -> tuple[dict | None, str | None]:
    for name in os.listdir(COMPARATIVOS_DIR):
        if not name.lower().endswith(".json"):
            continue
        path = os.path.join(COMPARATIVOS_DIR, name)
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if not isinstance(data, dict):
                continue
            same_cliente = str(data.get("cliente", "")).strip() == str(cliente).strip()
            same_inicio = str(data.get("periodo_inicio", "")).strip() == str(periodo_inicio).strip()
            same_fin = str(data.get("periodo_fin", "")).strip() == str(periodo_fin).strip()
            if same_cliente and same_inicio and same_fin:
                return data, path
        except Exception:
            continue
    return None, None


def _headcount_meta() -> dict:
    meta_base = obtener_metadata_headcount()
    try:
        activos = obtener_activos()
    except Exception as exc:
        logger.warning("comparativo: headcount no disponible: %s", exc)
        return {
            "exists": bool(meta_base.get("url_configurada")),
            "total_activos": 0,
            "fecha_actualizacion": meta_base.get("fecha_actualizacion"),
            "clientes_detectados": [],
            "activos_por_cliente": {},
            "headcount_error": str(exc),
        }
    clientes = sorted({a.get("cliente", "") for a in activos if a.get("cliente")})
    por_cliente: dict[str, int] = {}
    for item in activos:
        cliente = str(item.get("cliente", "")).strip()
        if not cliente:
            continue
        por_cliente[cliente] = por_cliente.get(cliente, 0) + 1
    return {
        "exists": bool(meta_base.get("url_configurada")),
        "total_activos": len(activos),
        "fecha_actualizacion": meta_base.get("fecha_actualizacion"),
        "clientes_detectados": clientes,
        "activos_por_cliente": por_cliente,
        "headcount_error": None,
    }


def _reporte_path(cliente: str, anio: int, mes: int) -> str:
    safe = " ".join(str(cliente or "").split()).replace(" ", "_").replace("/", "-")
    return os.path.join(REPORTES_MENSUALES_DIR, f"{safe}_{int(anio):04d}-{int(mes):02d}.json")


def _load_reporte_guardado(cliente: str, anio: int, mes: int) -> dict | None:
    path = _reporte_path(cliente, anio, mes)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if isinstance(data, dict):
            return data
    except Exception:
        return None
    return None


def _find_comparativo_file(comparativo_id: str) -> tuple[str | None, dict | None]:
    for name in os.listdir(COMPARATIVOS_DIR):
        if not name.lower().endswith(".json"):
            continue
        path = os.path.join(COMPARATIVOS_DIR, name)
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict) and str(data.get("id", "")).strip() == str(comparativo_id).strip():
                return path, data
        except Exception:
            continue
    return None, None


@comparativo_bp.get("/")
@_login_required_page
def index():
    _ensure_dirs()
    try:
        meta = _headcount_meta()
        clientes = meta.get("clientes_detectados") or _clientes_disponibles()
        headcount_warning = meta.get("headcount_error")
        return render_template(
            "comparativo/index.html",
            clientes=clientes,
            headcount=meta,
            headcount_warning=headcount_warning,
        )
    except Exception as exc:
        logger.exception("comparativo index failed")
        flash("No se pudo cargar el comparativo semanal. Revisa la configuración de Headcount.", "error")
        return render_template(
            "comparativo/index.html",
            clientes=[],
            headcount={
                "exists": False,
                "total_activos": 0,
                "fecha_actualizacion": None,
                "clientes_detectados": [],
                "activos_por_cliente": {},
            },
            headcount_warning=str(exc),
        )


@comparativo_bp.post("/actualizar-headcount")
@_login_required_page
def actualizar_headcount_route():
    _ensure_dirs()
    try:
        from flask import current_app

        from modules.headcount.snapshot_service import refresh_headcount_snapshot

        db_path = str(current_app.config["DATABASE"])
        result = refresh_headcount_snapshot(db_path, force=True, source="comparativo_manual")
        if result.get("ok"):
            return jsonify(
                {
                    "ok": True,
                    "message": (
                        f"Headcount actualizado: {result.get('activos_count', 0)} activos, "
                        f"{result.get('total_rows', 0)} registros."
                    ),
                    **result,
                }
            )
        if result.get("skipped"):
            return jsonify({"ok": False, "error": "refresh_already_running", "skipped": True}), 409
        return jsonify({"ok": False, "error": result.get("error") or "refresh_failed"}), 500
    except Exception as exc:
        flash(str(exc), "error")
        return jsonify({"ok": False, "error": str(exc)}), 500


@comparativo_bp.get("/activos-por-cliente")
@_login_required_page
def activos_por_cliente():
    try:
        meta = _headcount_meta()
        return jsonify(
            {
                "ok": True,
                "total_activos": meta.get("total_activos", 0),
                "fecha_actualizacion": meta.get("fecha_actualizacion"),
                "clientes": meta.get("clientes_detectados", []),
                "activos_por_cliente": meta.get("activos_por_cliente", {}),
                "headcount_error": meta.get("headcount_error"),
            }
        )
    except Exception as exc:
        logger.warning("activos_por_cliente failed: %s", exc)
        return jsonify(
            {
                "ok": True,
                "total_activos": 0,
                "fecha_actualizacion": None,
                "clientes": [],
                "activos_por_cliente": {},
                "headcount_error": str(exc),
            }
        )


@comparativo_bp.post("/preview-nomina")
@_login_required_page
def preview_nomina():
    file = request.files.get("nomina_file")
    if not file:
        return jsonify({"error": "No se recibió archivo de nómina."}), 400
    try:
        nombres = parsear_nomina(file)
        return jsonify({"nombres": nombres, "total": len(nombres)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@comparativo_bp.post("/verificar-similitudes")
@_login_required_page
def verificar_similitudes():
    data = request.get_json(silent=True) or {}
    nombres_nomina = data.get("nombres_nomina") or []
    cliente = str(data.get("cliente") or "").strip()
    if not isinstance(nombres_nomina, list):
        return jsonify({"error": "nombres_nomina debe ser una lista."}), 400
    if not cliente:
        return jsonify({"error": "cliente es obligatorio."}), 400
    try:
        activos = _resolver_activos_por_selector(cliente)
        lista_activos = [a.get("nombre_completo", "") for a in activos]
        similitudes = detectar_similitudes(nombres_nomina, lista_activos)
        return jsonify({"similitudes": similitudes, "hay_similitudes": bool(similitudes)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@comparativo_bp.post("/confirmar-alias")
@_login_required_page
def confirmar_alias():
    data = request.get_json(silent=True) or {}
    confirmaciones = data.get("confirmaciones") or []
    if not isinstance(confirmaciones, list):
        return jsonify({"error": "confirmaciones debe ser una lista."}), 400
    guardados = 0
    for item in confirmaciones:
        if not isinstance(item, dict):
            continue
        if bool(item.get("es_mismo")):
            nomina = str(item.get("nomina") or "").strip()
            headcount = str(item.get("headcount") or "").strip()
            if nomina and headcount:
                alias_service.guardar_alias(nomina, headcount)
                guardados += 1
    return jsonify({"ok": True, "aliases_guardados": guardados})


@comparativo_bp.post("/semanal")
@_login_required_page
def comparativo_semanal():
    _ensure_dirs()
    cliente = (request.form.get("cliente") or "").strip()
    periodo_inicio = (request.form.get("periodo_inicio") or "").strip()
    periodo_fin = (request.form.get("periodo_fin") or "").strip()
    nombres_json = request.form.get("nombres_json") or "[]"
    forzar = str(request.form.get("forzar") or "").strip().lower() in {"1", "true", "on", "yes", "si"}
    if not cliente or not periodo_inicio or not periodo_fin:
        return jsonify({"error": "Faltan datos para generar el comparativo semanal."}), 400

    try:
        duplicado, duplicado_path = _buscar_comparativo_duplicado(cliente, periodo_inicio, periodo_fin)
        if duplicado and not forzar:
            return jsonify(
                {
                    "duplicado": True,
                    "comparativo_id_existente": str(duplicado.get("id", "")),
                    "mensaje": "Ya existe un comparativo para este periodo.",
                }
            )
        if duplicado and forzar and duplicado_path:
            try:
                os.remove(duplicado_path)
            except OSError:
                pass

        lista_nomina_raw = json.loads(nombres_json)
        if not isinstance(lista_nomina_raw, list):
            return jsonify({"error": "nombres_json debe contener una lista JSON."}), 400

        activos_cliente = _resolver_activos_por_selector(cliente)
        if not activos_cliente:
            return jsonify(
                {
                    "error": (
                        "No hay registros suficientes en Headcount para este cliente. "
                        "Verifica la configuración o selecciona otro cliente."
                    )
                }
            ), 400
        lista_activos = [item.get("nombre_completo", "") for item in activos_cliente]
        resultado = aplicar_aliases_y_comparar(lista_nomina_raw, lista_activos)
        guardar_nomina_semana(cliente, periodo_inicio, periodo_fin, lista_nomina_raw)
        comparativo = guardar_comparativo_semanal(
            resultado=resultado,
            cliente=cliente,
            periodo_inicio=periodo_inicio,
            periodo_fin=periodo_fin,
            fecha_baja_asumida=periodo_fin,
        )
        return jsonify(
            {
                "duplicado": False,
                "comparativo_id": comparativo["id"],
                "altas": [
                    {"nombre": nombre, "fecha_alta_sugerida": periodo_inicio}
                    for nombre in resultado.get("altas", [])
                ],
                "bajas": [
                    {"nombre": nombre, "fecha_baja_sugerida": periodo_fin}
                    for nombre in resultado.get("bajas", [])
                ],
                "permanencias": resultado.get("permanencias", []),
                "total_nomina": resultado.get("total_nomina", 0),
                "total_activos": resultado.get("total_activos", 0),
                "aliases_aplicados": resultado.get("aliases_aplicados", []),
            }
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@comparativo_bp.post("/agrupaciones")
@_login_required_page
def guardar_agrupacion_route():
    data = request.get_json(silent=True) or {}
    nombre = str(data.get("nombre") or "").strip()
    clientes = data.get("clientes") or []
    if not nombre or not isinstance(clientes, list):
        return jsonify({"error": "Nombre y lista de clientes son obligatorios."}), 400
    alias_service.guardar_agrupacion(nombre, [str(c).strip() for c in clientes if str(c).strip()])
    return jsonify({"ok": True})


@comparativo_bp.get("/agrupaciones")
@_login_required_page
def listar_agrupaciones_route():
    return jsonify(alias_service.obtener_agrupaciones())


@comparativo_bp.delete("/agrupaciones/<nombre>")
@_login_required_page
def eliminar_agrupacion_route(nombre: str):
    alias_service.eliminar_agrupacion(unquote(nombre))
    return jsonify({"ok": True})


@comparativo_bp.get("/historial")
@_login_required_page
def historial_paginado():
    cliente = (request.args.get("cliente") or "").strip() or None
    try:
        page = max(1, int(request.args.get("page", "1")))
    except ValueError:
        page = 1
    try:
        per_page = max(1, int(request.args.get("per_page", "10")))
    except ValueError:
        per_page = 10

    items = obtener_historial(cliente=cliente)
    total = len(items)
    pages = max(1, math.ceil(total / per_page)) if total else 1
    start = (page - 1) * per_page
    end = start + per_page
    return jsonify(
        {
            "items": items[start:end],
            "total": total,
            "pages": pages,
            "current_page": page,
        }
    )


@comparativo_bp.delete("/historial/<comparativo_id>")
@_login_required_page
def eliminar_historial_item(comparativo_id: str):
    path, _ = _find_comparativo_file(comparativo_id)
    if path:
        try:
            os.remove(path)
            return jsonify({"ok": True})
        except OSError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": False, "error": "Comparativo no encontrado."}), 404


@comparativo_bp.get("/historial/<comparativo_id>/detalle")
@_login_required_page
def historial_detalle(comparativo_id: str):
    path, data = _find_comparativo_file(comparativo_id)
    if not path or not isinstance(data, dict):
        return jsonify({"ok": False, "error": "Comparativo no encontrado."}), 404
    return jsonify({"ok": True, "comparativo": data})


@comparativo_bp.post("/historial/<comparativo_id>/actualizar")
@_login_required_page
def historial_actualizar(comparativo_id: str):
    payload = request.get_json(silent=True) or {}
    altas_raw = payload.get("altas")
    bajas_raw = payload.get("bajas")
    if not isinstance(altas_raw, list) or not isinstance(bajas_raw, list):
        return jsonify({"ok": False, "error": "altas y bajas deben ser listas."}), 400

    def _norm_rows(rows: list, key_fecha: str) -> tuple[list[str], list[dict]]:
        nombres: list[str] = []
        detalle: list[dict] = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            nombre = str(item.get("nombre", "")).strip().upper()
            fecha = str(item.get(key_fecha, "")).strip()
            if not nombre:
                continue
            nombres.append(nombre)
            detalle.append({"nombre": nombre, "fecha": fecha})
        return nombres, detalle

    nuevas_altas, altas_detalle = _norm_rows(altas_raw, "fecha_alta")
    nuevas_bajas, bajas_detalle = _norm_rows(bajas_raw, "fecha_baja")

    path, data = _find_comparativo_file(comparativo_id)
    if not path or not isinstance(data, dict):
        return jsonify({"ok": False, "error": "Comparativo no encontrado."}), 404

    data["altas"] = nuevas_altas
    data["bajas"] = nuevas_bajas
    data["altas_detalle"] = altas_detalle
    data["bajas_detalle"] = bajas_detalle
    data["total_nomina"] = len(nuevas_altas) + len(list(data.get("permanencias", [])))
    data["total_activos"] = len(nuevas_bajas) + len(list(data.get("permanencias", [])))

    try:
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@comparativo_bp.get("/reporte-mensual")
@_login_required_page
def reporte_mensual_index():
    return render_template("comparativo/reporte_mensual.html")


@comparativo_bp.get("/reporte-mensual/resumen-clientes")
@_login_required_page
def reporte_mensual_resumen_clientes():
    try:
        resumen = obtener_resumen_nominas_por_cliente()
        historial = obtener_historial_reportes()
        return jsonify({"ok": True, "resumen": resumen, "historial_reportes": historial})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@comparativo_bp.post("/reporte-mensual/generar")
@_login_required_page
def reporte_mensual_generar():
    data = request.get_json(silent=True) or {}
    cliente = str(data.get("cliente") or "").strip()
    mes = data.get("mes")
    anio = data.get("anio")
    if not cliente or mes is None or anio is None:
        return jsonify({"ok": False, "error": "cliente, mes y anio son obligatorios."}), 400
    try:
        reporte = generar_reporte_mensual_v2(cliente, int(mes), int(anio))
        guardar_reporte_mensual(reporte)
        return jsonify({"ok": True, "reporte": reporte})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@comparativo_bp.post("/reporte-mensual/confirmar-similitudes")
@_login_required_page
def reporte_mensual_confirmar_similitudes():
    data = request.get_json(silent=True) or {}
    confirmaciones = data.get("confirmaciones") or []
    cliente = str(data.get("cliente") or "").strip()
    mes = data.get("mes")
    anio = data.get("anio")
    if not cliente or mes is None or anio is None:
        return jsonify({"ok": False, "error": "cliente, mes y anio son obligatorios."}), 400
    if not isinstance(confirmaciones, list):
        return jsonify({"ok": False, "error": "confirmaciones debe ser lista."}), 400
    guardados = 0
    for item in confirmaciones:
        if not isinstance(item, dict):
            continue
        es_mismo = bool(item.get("es_mismo"))
        if not es_mismo:
            continue
        nombre_a = str(item.get("nombre_a") or item.get("nomina") or "").strip()
        nombre_b = str(item.get("nombre_b") or item.get("headcount") or "").strip()
        if nombre_a and nombre_b:
            alias_service.guardar_alias(nombre_a, nombre_b)
            guardados += 1
    try:
        reporte = generar_reporte_mensual_v2(cliente, int(mes), int(anio))
        guardar_reporte_mensual(reporte)
        return jsonify({"ok": True, "aliases_guardados": guardados, "reporte": reporte})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@comparativo_bp.post("/reporte-mensual/actualizar-fecha")
@_login_required_page
def reporte_mensual_actualizar_fecha():
    data = request.get_json(silent=True) or {}
    cliente = str(data.get("cliente") or "").strip()
    mes = data.get("mes")
    anio = data.get("anio")
    nombre = str(data.get("nombre") or "").strip().upper()
    tipo_fecha = str(data.get("tipo_fecha") or "").strip().lower()
    nueva_fecha = str(data.get("nueva_fecha") or "").strip()
    indice_raw = data.get("indice", 0)
    try:
        indice = int(indice_raw)
    except (TypeError, ValueError):
        indice = 0
    if indice < 0:
        indice = 0
    if not cliente or mes is None or anio is None or not nombre or tipo_fecha not in {"alta", "baja"}:
        return jsonify({"ok": False, "error": "Datos incompletos para actualizar fecha."}), 400
    path = _reporte_path(cliente, int(anio), int(mes))
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": "Reporte no encontrado."}), 404
    try:
        with open(path, "r", encoding="utf-8") as fh:
            reporte = json.load(fh)
        if not isinstance(reporte, dict):
            return jsonify({"ok": False, "error": "Reporte inválido."}), 400
        rotativos = reporte.get("rotativos")
        if not isinstance(rotativos, list):
            return jsonify({"ok": False, "error": "Estructura de rotativos inválida."}), 400
        updated = False
        for r in rotativos:
            if not isinstance(r, dict):
                continue
            if str(r.get("nombre", "")).strip().upper() != nombre:
                continue
            key_lista = f"fechas_{tipo_fecha}"
            key_scalar = f"fecha_{tipo_fecha}"
            fechas = r.get(key_lista)
            if not isinstance(fechas, list):
                fechas = []
            fechas = [str(v or "").strip() for v in fechas]
            if not fechas:
                fallback = str(r.get(key_scalar) or "").strip()
                if fallback:
                    fechas.append(fallback)
            while len(fechas) <= indice:
                fechas.append("")
            fechas[indice] = nueva_fecha or ""
            r[key_lista] = fechas
            r[key_scalar] = fechas[0] if fechas and fechas[0] else None

            alertas = r.get("alertas")
            if not isinstance(alertas, list):
                alertas = []
            if nueva_fecha:
                if tipo_fecha == "alta":
                    alertas = [a for a in alertas if "Sin fecha de alta" not in str(a)]
                else:
                    alertas = [a for a in alertas if "Sin fecha de baja" not in str(a)]
            else:
                msg = f"Sin fecha de {tipo_fecha} - requiere captura manual"
                if msg not in [str(a) for a in alertas]:
                    alertas.append(msg)
            r["alertas"] = alertas
            updated = True
            break
        if not updated:
            return jsonify({"ok": False, "error": "Empleado no encontrado en rotativos."}), 404
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(reporte, fh, ensure_ascii=False, indent=2)
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@comparativo_bp.get("/reporte-mensual/exportar")
@_login_required_page
def reporte_mensual_exportar():
    cliente = str(request.args.get("cliente") or "").strip()
    mes_raw = str(request.args.get("mes") or "").strip()
    anio_raw = str(request.args.get("anio") or "").strip()
    if not cliente or not mes_raw or not anio_raw:
        return jsonify({"ok": False, "error": "Parámetros cliente, mes y anio requeridos."}), 400
    reporte = _load_reporte_guardado(cliente, int(anio_raw), int(mes_raw))
    if not reporte:
        return jsonify({"ok": False, "error": "Reporte mensual no encontrado."}), 404

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Personal del mes"
    ws_all.append(["Nombre", "Tipo", "Fecha Alta", "Fecha Baja", "En Headcount", "Alertas"])
    for f in reporte.get("fijos", []):
        if not isinstance(f, dict):
            continue
        ws_all.append(
            [
                f.get("nombre", ""),
                "Fijo",
                "",
                "",
                "SI" if f.get("en_headcount") else "NO",
                f.get("alerta") or "",
            ]
        )
    for r in reporte.get("rotativos", []):
        if not isinstance(r, dict):
            continue
        ws_all.append(
            [
                r.get("nombre", ""),
                "Rotativo",
                r.get("fecha_alta") or "",
                r.get("fecha_baja") or "",
                "SI" if r.get("en_headcount") else "NO",
                " | ".join([str(a) for a in r.get("alertas", [])]),
            ]
        )

    ws_fijos = wb.create_sheet("Fijos")
    ws_fijos.append(["Nombre", "En Headcount", "Alerta"])
    for f in reporte.get("fijos", []):
        if not isinstance(f, dict):
            continue
        ws_fijos.append([f.get("nombre", ""), "SI" if f.get("en_headcount") else "NO", f.get("alerta") or ""])

    ws_rot = wb.create_sheet("Rotativos")
    ws_rot.append(["Nombre", "Fecha Alta", "Fecha Baja", "En Headcount", "Alertas", "Semanas"])
    for r in reporte.get("rotativos", []):
        if not isinstance(r, dict):
            continue
        ws_rot.append(
            [
                r.get("nombre", ""),
                r.get("fecha_alta") or "",
                r.get("fecha_baja") or "",
                "SI" if r.get("en_headcount") else "NO",
                " | ".join([str(a) for a in r.get("alertas", [])]),
                " | ".join([str(s) for s in r.get("semanas_presente", [])]),
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"reporte_mensual_{cliente}_{int(anio_raw):04d}-{int(mes_raw):02d}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@comparativo_bp.delete("/reporte-mensual/<cliente>/<anio>/<mes>")
@_login_required_page
def reporte_mensual_eliminar(cliente: str, anio: str, mes: str):
    try:
        path = _reporte_path(unquote(cliente), int(anio), int(mes))
    except ValueError:
        return jsonify({"ok": False, "error": "Parámetros inválidos."}), 400
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": "Reporte no encontrado."}), 404
    try:
        os.remove(path)
        return jsonify({"ok": True})
    except OSError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@comparativo_bp.get("/reporte-mensual/cargar/<cliente>/<anio>/<mes>")
@_login_required_page
def reporte_mensual_cargar(cliente: str, anio: str, mes: str):
    try:
        reporte = _load_reporte_guardado(unquote(cliente), int(anio), int(mes))
    except ValueError:
        return jsonify({"ok": False, "error": "Parámetros inválidos."}), 400
    if not isinstance(reporte, dict):
        return jsonify({"ok": False, "error": "Reporte no encontrado"}), 404
    return jsonify({"ok": True, "reporte": reporte})


@comparativo_bp.post("/reporte-mensual/actualizar-reporte")
@_login_required_page
def reporte_mensual_actualizar_reporte():
    reporte = request.get_json(silent=True) or {}
    if not isinstance(reporte, dict):
        return jsonify({"ok": False, "error": "Payload inválido."}), 400
    required = ("cliente", "mes", "anio", "fijos", "rotativos")
    if any(k not in reporte for k in required):
        return jsonify({"ok": False, "error": "Faltan campos requeridos en el reporte."}), 400
    if not isinstance(reporte.get("fijos"), list) or not isinstance(reporte.get("rotativos"), list):
        return jsonify({"ok": False, "error": "fijos y rotativos deben ser listas."}), 400
    try:
        guardar_reporte_mensual(reporte)
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@comparativo_bp.get("/mensual")
@_login_required_page
def reporte_mensual():
    _ensure_dirs()
    cliente = (request.args.get("cliente") or "").strip()
    mes_raw = (request.args.get("mes") or "").strip()
    anio_raw = (request.args.get("anio") or "").strip()
    if not cliente or not mes_raw or not anio_raw:
        flash("Selecciona cliente, mes y año para generar el reporte mensual.", "error")
        return redirect(url_for("comparativo.index"))
    try:
        reporte = generar_reporte_mensual(cliente=cliente, mes=int(mes_raw), anio=int(anio_raw))
        return render_template("comparativo/resultado_mensual.html", reporte=reporte)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@comparativo_bp.get("/exportar-semanal/<comparativo_id>")
@_login_required_page
def exportar_semanal(comparativo_id: str):
    _ensure_dirs()
    historial = obtener_historial()
    comp = next((item for item in historial if str(item.get("id")) == str(comparativo_id)), None)
    if not comp:
        flash("Comparativo no encontrado.", "error")
        return redirect(url_for("comparativo.index"))

    wb = Workbook()
    ws_altas = wb.active
    ws_altas.title = "Altas"
    ws_altas.append(["Nombre", "Fecha Alta"])
    altas_json = request.args.get("altas_json")
    bajas_json = request.args.get("bajas_json")

    altas_rows = None
    bajas_rows = None
    if altas_json:
        try:
            parsed = json.loads(altas_json)
            if isinstance(parsed, list):
                altas_rows = parsed
        except json.JSONDecodeError:
            altas_rows = None
    if bajas_json:
        try:
            parsed = json.loads(bajas_json)
            if isinstance(parsed, list):
                bajas_rows = parsed
        except json.JSONDecodeError:
            bajas_rows = None

    if altas_rows is None:
        detalle_guardado = comp.get("altas_detalle")
        if isinstance(detalle_guardado, list) and detalle_guardado:
            altas_rows = detalle_guardado
        else:
            altas_rows = [{"nombre": nombre, "fecha": comp.get("periodo_inicio", "")} for nombre in comp.get("altas", [])]
    for item in altas_rows:
        nombre = str(item.get("nombre", "") if isinstance(item, dict) else item)
        fecha = (
            str(item.get("fecha", "") if isinstance(item, dict) else comp.get("periodo_inicio", ""))
            or comp.get("periodo_inicio", "")
        )
        ws_altas.append([nombre, fecha])

    ws_bajas = wb.create_sheet("Bajas")
    ws_bajas.append(["Nombre", "Fecha Baja"])
    if bajas_rows is None:
        detalle_guardado = comp.get("bajas_detalle")
        if isinstance(detalle_guardado, list) and detalle_guardado:
            bajas_rows = detalle_guardado
        else:
            bajas_rows = [{"nombre": nombre, "fecha": comp.get("fecha_baja_asumida", "")} for nombre in comp.get("bajas", [])]
    for item in bajas_rows:
        nombre = str(item.get("nombre", "") if isinstance(item, dict) else item)
        fecha = (
            str(item.get("fecha", "") if isinstance(item, dict) else comp.get("fecha_baja_asumida", ""))
            or comp.get("fecha_baja_asumida", "")
        )
        ws_bajas.append([nombre, fecha])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"comparativo_semanal_{comparativo_id}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@comparativo_bp.get("/exportar-mensual")
@_login_required_page
def exportar_mensual():
    _ensure_dirs()
    cliente = (request.args.get("cliente") or "").strip()
    mes_raw = (request.args.get("mes") or "").strip()
    anio_raw = (request.args.get("anio") or "").strip()
    if not cliente or not mes_raw or not anio_raw:
        flash("Faltan parámetros para exportar el reporte mensual.", "error")
        return redirect(url_for("comparativo.index"))

    try:
        reporte = generar_reporte_mensual(cliente=cliente, mes=int(mes_raw), anio=int(anio_raw))
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("comparativo.index"))

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["Cliente", reporte.get("cliente", "")])
    ws_resumen.append(["Mes", reporte.get("mes", "")])
    ws_resumen.append(["Año", reporte.get("anio", "")])
    ws_resumen.append(["Total altas", len(reporte.get("altas_mes", []))])
    ws_resumen.append(["Total bajas", len(reporte.get("bajas_mes", []))])
    ws_resumen.append([])
    ws_resumen.append(["Semanas del mes"])
    ws_resumen.append(["Periodo inicio", "Periodo fin", "Altas", "Bajas"])
    for semana in reporte.get("semanas", []):
        ws_resumen.append(
            [
                semana.get("periodo_inicio", ""),
                semana.get("periodo_fin", ""),
                len(semana.get("altas", [])),
                len(semana.get("bajas", [])),
            ]
        )

    ws_altas = wb.create_sheet("Altas mes")
    ws_altas.append(["Nombre", "Fecha Alta"])
    for item in reporte.get("altas_mes", []):
        ws_altas.append([item.get("nombre", ""), item.get("fecha_alta", "")])

    ws_bajas = wb.create_sheet("Bajas mes")
    ws_bajas.append(["Nombre", "Fecha Baja"])
    for item in reporte.get("bajas_mes", []):
        ws_bajas.append([item.get("nombre", ""), item.get("fecha_baja", "")])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"comparativo_mensual_{cliente}_{mes_raw}_{anio_raw}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
