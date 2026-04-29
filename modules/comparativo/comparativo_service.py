from __future__ import annotations

import json
import math
import os
import uuid
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz

from modules.comparativo import alias_service
from modules.comparativo.headcount_service import buscar_trabajador

DATA_DIR = os.environ.get("DATA_DIR", "./data")
COMPARATIVOS_DIR = os.path.join(DATA_DIR, "comparativos")
NOMINAS_DIR = os.path.join(DATA_DIR, "nominas")
REPORTES_MENSUALES_DIR = os.path.join(DATA_DIR, "reportes_mensuales")


def _normalize_spaces(value: str) -> str:
    return " ".join((value or "").split())


def _normalize_name(value: Any) -> str:
    return _normalize_spaces(str(value or "").upper().strip())


def _parse_ddmmyyyy(value: str) -> datetime | None:
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _parse_period_sort_key(value: str) -> datetime:
    parsed = _parse_ddmmyyyy(value)
    if parsed is not None:
        return parsed
    return datetime.min


def _ddmmyyyy(dt: date | datetime | None) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y")
    return dt.strftime("%d/%m/%Y")


def _date_from_str(value: str | None) -> date | None:
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _safe_slug_cliente(cliente: str) -> str:
    return _normalize_spaces(str(cliente or "general")).replace(" ", "_").replace("/", "-")


def _iter_json_dicts(path: str) -> list[dict[str, Any]]:
    if not os.path.exists(path):
        return []
    out: list[dict[str, Any]] = []
    for name in os.listdir(path):
        if not name.lower().endswith(".json"):
            continue
        fp = os.path.join(path, name)
        try:
            with open(fp, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                out.append(data)
        except Exception:
            continue
    return out


def _mes_key(anio: int, mes: int) -> str:
    return f"{anio:04d}-{mes:02d}"


def parsear_nomina(file) -> list[str]:
    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        match_row = None
        match_col = None

        for row in sheet.iter_rows():
            for cell in row:
                text = _normalize_spaces(str(cell.value or "").strip().upper())
                if text == "NOMBRE DE EMPLEADO":
                    match_row = cell.row
                    match_col = cell.column
                    break
            if match_row is not None:
                break

        if match_row is None or match_col is None:
            raise ValueError("No se encontró la celda 'NOMBRE DE EMPLEADO' en la nómina.")

        nombres: list[str] = []
        row_idx = match_row + 1
        while row_idx <= sheet.max_row:
            value = sheet.cell(row=row_idx, column=match_col).value
            if value is None or str(value).strip() == "":
                break
            if isinstance(value, (int, float)):
                row_idx += 1
                continue
            nombre = _normalize_name(value)
            if nombre:
                nombres.append(nombre)
            row_idx += 1
        return nombres
    except Exception as exc:
        raise ValueError(f"No se pudo parsear la nómina: {exc}") from exc


def comparar_listas(lista_nomina: list[str], lista_activos: list[str]) -> dict[str, Any]:
    try:
        nomina_set = {_normalize_name(n) for n in lista_nomina if _normalize_name(n)}
        activos_set = {_normalize_name(n) for n in lista_activos if _normalize_name(n)}
        altas = sorted(nomina_set - activos_set)
        bajas = sorted(activos_set - nomina_set)
        permanencias = sorted(nomina_set & activos_set)
        return {
            "altas": altas,
            "bajas": bajas,
            "permanencias": permanencias,
            "total_nomina": len(nomina_set),
            "total_activos": len(activos_set),
        }
    except Exception as exc:
        raise ValueError(f"No se pudo comparar listas: {exc}") from exc


def guardar_comparativo_semanal(
    resultado: dict[str, Any],
    cliente: str,
    periodo_inicio: str,
    periodo_fin: str,
    fecha_baja_asumida: str,
) -> dict[str, Any]:
    try:
        os.makedirs(COMPARATIVOS_DIR, exist_ok=True)
        try:
            os.makedirs("/app/data/comparativos", exist_ok=True)
        except OSError:
            pass
        comparativo = {
            "id": str(uuid.uuid4()),
            "cliente": str(cliente or "").strip(),
            "periodo_inicio": str(periodo_inicio or "").strip(),
            "periodo_fin": str(periodo_fin or "").strip(),
            "fecha_baja_asumida": str(fecha_baja_asumida or "").strip(),
            "fecha_generacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "altas": list(resultado.get("altas", [])),
            "bajas": list(resultado.get("bajas", [])),
            "permanencias": list(resultado.get("permanencias", [])),
            "total_nomina": int(resultado.get("total_nomina", 0)),
            "total_activos": int(resultado.get("total_activos", 0)),
        }
        safe_cliente = _normalize_spaces(str(cliente or "general")).replace(" ", "_").replace("/", "-")
        safe_inicio = str(periodo_inicio or "").replace("/", "-")
        safe_fin = str(periodo_fin or "").replace("/", "-")
        filename = f"{safe_cliente}_{safe_inicio}_{safe_fin}.json"
        file_path = os.path.join(COMPARATIVOS_DIR, filename)
        with open(file_path, "w", encoding="utf-8") as fh:
            json.dump(comparativo, fh, ensure_ascii=False, indent=2)
        return comparativo
    except Exception as exc:
        raise ValueError(f"No se pudo guardar comparativo semanal: {exc}") from exc


def obtener_historial(cliente: str | None = None) -> list[dict[str, Any]]:
    try:
        os.makedirs(COMPARATIVOS_DIR, exist_ok=True)
        items: list[dict[str, Any]] = []
        for name in os.listdir(COMPARATIVOS_DIR):
            if not name.lower().endswith(".json"):
                continue
            path = os.path.join(COMPARATIVOS_DIR, name)
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if not isinstance(data, dict):
                continue
            if cliente and str(data.get("cliente", "")).strip().casefold() != str(cliente).strip().casefold():
                continue
            items.append(data)
        items.sort(key=lambda x: str(x.get("fecha_generacion", "")), reverse=True)
        return items
    except Exception as exc:
        raise ValueError(f"No se pudo obtener historial de comparativos: {exc}") from exc


def generar_reporte_mensual(cliente: str, mes: int, anio: int) -> dict[str, Any]:
    try:
        semanas: list[dict[str, Any]] = []
        for comp in obtener_historial(cliente):
            inicio = _parse_ddmmyyyy(comp.get("periodo_inicio", ""))
            fin = _parse_ddmmyyyy(comp.get("periodo_fin", ""))
            if (inicio and inicio.month == int(mes) and inicio.year == int(anio)) or (
                fin and fin.month == int(mes) and fin.year == int(anio)
            ):
                semanas.append(comp)

        altas_map: dict[str, str] = {}
        bajas_map: dict[str, str] = {}
        personal_activo_set: set[str] = set()
        altas_detalle: list[dict[str, Any]] = []
        bajas_detalle: list[dict[str, Any]] = []

        for comp in semanas:
            fecha_alta = str(comp.get("periodo_inicio", "")).strip()
            fecha_baja = str(comp.get("fecha_baja_asumida", "")).strip()
            for nombre in comp.get("altas", []):
                n = _normalize_name(nombre)
                if not n:
                    continue
                personal_activo_set.add(n)
                old = _parse_ddmmyyyy(altas_map.get(n, ""))
                current = _parse_ddmmyyyy(fecha_alta)
                if old is None or (current is not None and current < old):
                    altas_map[n] = fecha_alta
            for nombre in comp.get("bajas", []):
                n = _normalize_name(nombre)
                if not n:
                    continue
                personal_activo_set.add(n)
                old = _parse_ddmmyyyy(bajas_map.get(n, ""))
                current = _parse_ddmmyyyy(fecha_baja)
                if old is None or (current is not None and current > old):
                    bajas_map[n] = fecha_baja
            for nombre in comp.get("permanencias", []):
                n = _normalize_name(nombre)
                if n:
                    personal_activo_set.add(n)

        for nombre, fecha in sorted(altas_map.items()):
            altas_detalle.append({"nombre": nombre, "fecha_alta": fecha, "trabajador": buscar_trabajador(nombre)})
        for nombre, fecha in sorted(bajas_map.items()):
            bajas_detalle.append({"nombre": nombre, "fecha_baja": fecha, "trabajador": buscar_trabajador(nombre)})

        return {
            "cliente": cliente,
            "mes": int(mes),
            "anio": int(anio),
            "semanas": semanas,
            "altas_mes": [{"nombre": n, "fecha_alta": altas_map[n]} for n in sorted(altas_map)],
            "bajas_mes": [{"nombre": n, "fecha_baja": bajas_map[n]} for n in sorted(bajas_map)],
            "personal_activo_mes": sorted(personal_activo_set),
            "altas_mes_detalle": altas_detalle,
            "bajas_mes_detalle": bajas_detalle,
        }
    except Exception as exc:
        raise ValueError(f"No se pudo generar reporte mensual: {exc}") from exc


def detectar_similitudes(lista_nomina: list[str], lista_activos: list[str], umbral: int = 88) -> list[dict[str, Any]]:
    try:
        similitudes: list[dict[str, Any]] = []
        activos_norm = [_normalize_name(n) for n in lista_activos if _normalize_name(n)]

        for nombre_raw in lista_nomina:
            nomina_norm = _normalize_name(nombre_raw)
            if not nomina_norm:
                continue

            alias = alias_service.obtener_alias(nomina_norm)
            if alias:
                # Alias conocido: se toma como resuelto y se omite sugerencia.
                continue

            for activo in activos_norm:
                if nomina_norm == activo:
                    continue
                score = float(fuzz.ratio(nomina_norm, activo))
                if score >= float(umbral):
                    similitudes.append(
                        {
                            "nomina": nomina_norm,
                            "headcount": activo,
                            "score": score,
                        }
                    )

        similitudes.sort(key=lambda x: float(x.get("score", 0.0)), reverse=True)
        return similitudes
    except Exception as exc:
        raise ValueError(f"No se pudo detectar similitudes: {exc}") from exc


def aplicar_aliases_y_comparar(lista_nomina_raw: list[str], lista_activos: list[str]) -> dict[str, Any]:
    try:
        lista_resuelta: list[str] = []
        aliases_aplicados: list[dict[str, str]] = []

        for nombre_raw in lista_nomina_raw:
            original = _normalize_name(nombre_raw)
            if not original:
                continue
            alias = alias_service.obtener_alias(original)
            resuelto = _normalize_name(alias) if alias else original
            lista_resuelta.append(resuelto)
            if resuelto != original:
                aliases_aplicados.append({"original": original, "resuelto": resuelto})

        resultado = comparar_listas(lista_resuelta, lista_activos)
        resultado["aliases_aplicados"] = aliases_aplicados
        return resultado
    except Exception as exc:
        raise ValueError(f"No se pudo aplicar aliases y comparar: {exc}") from exc


def guardar_nomina_semana(
    cliente: str,
    periodo_inicio: str,
    periodo_fin: str,
    lista_nombres: list[str],
) -> dict[str, Any]:
    try:
        os.makedirs(NOMINAS_DIR, exist_ok=True)
        nombres_norm = sorted({_normalize_name(n) for n in lista_nombres if _normalize_name(n)})
        payload: dict[str, Any] = {
            "cliente": str(cliente or "").strip(),
            "periodo_inicio": str(periodo_inicio or "").strip(),
            "periodo_fin": str(periodo_fin or "").strip(),
            "fecha_guardado": datetime.now().isoformat(),
            "empleados": nombres_norm,
        }

        agrupaciones = alias_service.obtener_agrupaciones()
        if payload["cliente"] in agrupaciones:
            raw_clients = agrupaciones.get(payload["cliente"], [])
            payload["clientes_agrupados"] = [str(c).strip() for c in raw_clients if str(c).strip()]

        safe_cliente = _normalize_spaces(payload["cliente"] or "general").replace(" ", "_").replace("/", "-")
        safe_inicio = payload["periodo_inicio"].replace("/", "-")
        safe_fin = payload["periodo_fin"].replace("/", "-")
        out_path = os.path.join(NOMINAS_DIR, f"{safe_cliente}_{safe_inicio}_{safe_fin}.json")
        with open(out_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2)
        return payload
    except Exception as exc:
        raise ValueError(f"No se pudo guardar nómina semanal: {exc}") from exc


def obtener_nominas_guardadas(cliente: str | None = None) -> list[dict[str, Any]]:
    try:
        os.makedirs(NOMINAS_DIR, exist_ok=True)
        items: list[dict[str, Any]] = []
        for name in os.listdir(NOMINAS_DIR):
            if not name.lower().endswith(".json"):
                continue
            path = os.path.join(NOMINAS_DIR, name)
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if not isinstance(data, dict):
                continue
            if cliente:
                objetivo = str(cliente).strip().casefold()
                cliente_nomina = str(data.get("cliente", "")).strip().casefold()
                agrupados = [str(c).strip().casefold() for c in data.get("clientes_agrupados", []) if str(c).strip()]
                if objetivo != cliente_nomina and objetivo not in agrupados:
                    continue
            items.append(data)
        items.sort(key=lambda x: _parse_period_sort_key(str(x.get("periodo_inicio", ""))), reverse=True)
        return items
    except Exception as exc:
        raise ValueError(f"No se pudo obtener nóminas guardadas: {exc}") from exc


def obtener_resumen_nominas_por_cliente() -> dict[str, Any]:
    try:
        os.makedirs(COMPARATIVOS_DIR, exist_ok=True)
        comparativos = _iter_json_dicts(COMPARATIVOS_DIR)
        comparativo_map: dict[tuple[str, str, str], str] = {}
        for comp in comparativos:
            key = (
                str(comp.get("cliente", "")).strip(),
                str(comp.get("periodo_inicio", "")).strip(),
                str(comp.get("periodo_fin", "")).strip(),
            )
            comparativo_map[key] = str(comp.get("id", "")).strip()

        resumen: dict[str, Any] = {}
        for comp in comparativos:
            cliente = str(comp.get("cliente", "")).strip()
            if not cliente:
                continue
            periodo_inicio = str(comp.get("periodo_inicio", "")).strip()
            periodo_fin = str(comp.get("periodo_fin", "")).strip()
            try:
                inicio_dt = datetime.strptime(periodo_inicio, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue
            try:
                fin_dt = datetime.strptime(periodo_fin, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue
            cliente_block = resumen.setdefault(cliente, {"periodos": [], "meses": {}})
            cliente_block["periodos"].append(
                {
                    "periodo_inicio": periodo_inicio,
                    "periodo_fin": periodo_fin,
                    "comparativo_id": comparativo_map.get((cliente, periodo_inicio, periodo_fin), ""),
                    "_inicio_dt": inicio_dt,
                    "_fin_dt": fin_dt,
                }
            )

        for cliente, block in resumen.items():
            periodos = block.get("periodos", [])
            periodos.sort(key=lambda p: p.get("_inicio_dt") or date.min, reverse=False)
            meses_map: dict[str, Any] = {}
            for p in periodos:
                ini_dt = p.get("_inicio_dt")
                fin_dt = p.get("_fin_dt")
                if not isinstance(ini_dt, date) or not isinstance(fin_dt, date):
                    continue
                if ini_dt > fin_dt:
                    continue
                cursor_mes = date(ini_dt.year, ini_dt.month, 1)
                fin_mes = date(fin_dt.year, fin_dt.month, 1)
                while cursor_mes <= fin_mes:
                    anio_mes = cursor_mes.year
                    mes_mes = cursor_mes.month
                    primer_dia_mes = date(anio_mes, mes_mes, 1)
                    if mes_mes == 12:
                        primer_dia_mes_sig = date(anio_mes + 1, 1, 1)
                    else:
                        primer_dia_mes_sig = date(anio_mes, mes_mes + 1, 1)
                    ultimo_dia_mes = date.fromordinal(primer_dia_mes_sig.toordinal() - 1)

                    # Criterio de intersección de periodos con mes.
                    if ini_dt <= ultimo_dia_mes and fin_dt >= primer_dia_mes:
                        key = _mes_key(anio_mes, mes_mes)
                        m = meses_map.setdefault(
                            key,
                            {
                                "mes": mes_mes,
                                "anio": anio_mes,
                                "periodos_count": 0,
                                "fecha_min_dt": None,
                                "fecha_max_dt": None,
                            },
                        )
                        m["periodos_count"] += 1
                        inicio_clip = max(ini_dt, primer_dia_mes)
                        fin_clip = min(fin_dt, ultimo_dia_mes)
                        if m["fecha_min_dt"] is None or inicio_clip < m["fecha_min_dt"]:
                            m["fecha_min_dt"] = inicio_clip
                        if m["fecha_max_dt"] is None or fin_clip > m["fecha_max_dt"]:
                            m["fecha_max_dt"] = fin_clip

                    if cursor_mes.month == 12:
                        cursor_mes = date(cursor_mes.year + 1, 1, 1)
                    else:
                        cursor_mes = date(cursor_mes.year, cursor_mes.month + 1, 1)

            normalized_meses: dict[str, Any] = {}
            for key, m in sorted(meses_map.items()):
                fecha_min_dt = m.get("fecha_min_dt")
                fecha_max_dt = m.get("fecha_max_dt")
                comp = calcular_completitud_mes(cliente, int(m["mes"]), int(m["anio"]))
                normalized_meses[key] = {
                    "mes": int(m["mes"]),
                    "anio": int(m["anio"]),
                    "periodos_count": int(comp.get("periodos_count", m["periodos_count"])),
                    "fecha_min": _ddmmyyyy(fecha_min_dt),
                    "fecha_max": _ddmmyyyy(fecha_max_dt),
                    "dias_cubiertos": int(comp.get("dias_cubiertos", 0)),
                    "dias_totales_mes": int(comp.get("dias_totales_mes", 0)),
                    "completo": bool(comp.get("completo")),
                    "semanas_faltantes": comp.get("advertencia") or "",
                    "dias_faltantes": comp.get("dias_faltantes", []),
                }
            block["meses"] = normalized_meses
            block["periodos"] = sorted(
                [
                    {
                        "periodo_inicio": p.get("periodo_inicio", ""),
                        "periodo_fin": p.get("periodo_fin", ""),
                        "comparativo_id": p.get("comparativo_id", ""),
                    }
                    for p in periodos
                ],
                key=lambda p: (_date_from_str(p.get("periodo_inicio", "")) or date.min),
                reverse=True,
            )
        return resumen
    except Exception as exc:
        raise ValueError(f"No se pudo obtener resumen de nóminas por cliente: {exc}") from exc


def calcular_completitud_mes(cliente: str, mes: int, anio: int) -> dict[str, Any]:
    try:
        mes = int(mes)
        anio = int(anio)
        mes_inicio = date(anio, mes, 1)
        if mes == 12:
            siguiente_mes = date(anio + 1, 1, 1)
        else:
            siguiente_mes = date(anio, mes + 1, 1)
        mes_fin = siguiente_mes.fromordinal(siguiente_mes.toordinal() - 1)

        periodos_count = 0
        dias_cubiertos_set: set[date] = set()

        for nomina in obtener_nominas_guardadas(cliente=cliente):
            periodo_inicio_raw = str(nomina.get("periodo_inicio", "")).strip()
            periodo_fin_raw = str(nomina.get("periodo_fin", "")).strip()
            try:
                periodo_inicio_dt = datetime.strptime(periodo_inicio_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue
            try:
                periodo_fin_dt = datetime.strptime(periodo_fin_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue

            if periodo_inicio_dt > mes_fin or periodo_fin_dt < mes_inicio:
                continue

            periodos_count += 1
            inicio_efectivo = max(periodo_inicio_dt, mes_inicio)
            fin_efectivo = min(periodo_fin_dt, mes_fin)
            cursor = inicio_efectivo
            while cursor <= fin_efectivo:
                dias_cubiertos_set.add(cursor)
                cursor = date.fromordinal(cursor.toordinal() + 1)

        dias_requeridos_set: set[date] = set()
        cursor = mes_inicio
        while cursor <= mes_fin:
            dias_requeridos_set.add(cursor)
            cursor = date.fromordinal(cursor.toordinal() + 1)

        dias_faltantes_dt = sorted(dias_requeridos_set - dias_cubiertos_set)
        dias_faltantes = [d.strftime("%d/%m/%Y") for d in dias_faltantes_dt]
        dias_totales_mes = len(dias_requeridos_set)
        dias_cubiertos = len(dias_cubiertos_set)
        completo = len(dias_faltantes) == 0
        advertencia = None if completo else f"Faltan {len(dias_faltantes)} día(s) sin cobertura en el mes."
        return {
            "completo": completo,
            "dias_cubiertos": dias_cubiertos,
            "dias_totales_mes": dias_totales_mes,
            "periodos_count": periodos_count,
            "dias_faltantes": dias_faltantes,
            "advertencia": advertencia,
        }
    except Exception as exc:
        raise ValueError(f"No se pudo calcular completitud del mes: {exc}") from exc


def generar_reporte_mensual_v2(cliente: str, mes: int, anio: int) -> dict[str, Any]:
    try:
        mes = int(mes)
        anio = int(anio)
        mes_inicio = date(anio, mes, 1)
        if mes == 12:
            siguiente_mes = date(anio + 1, 1, 1)
        else:
            siguiente_mes = date(anio, mes + 1, 1)
        mes_fin = date.fromordinal(siguiente_mes.toordinal() - 1)

        agrupaciones = alias_service.obtener_agrupaciones()
        clientes_objetivo = (
            [str(c).strip() for c in agrupaciones.get(cliente, []) if str(c).strip()]
            if cliente in agrupaciones
            else [cliente]
        )
        if not clientes_objetivo:
            clientes_objetivo = [cliente]

        all_nominas = _iter_json_dicts(NOMINAS_DIR)
        semanas: list[dict[str, Any]] = []
        for nomina in all_nominas:
            c = str(nomina.get("cliente", "")).strip()
            if c not in clientes_objetivo:
                continue
            periodo_inicio_raw = str(nomina.get("periodo_inicio", "")).strip()
            periodo_fin_raw = str(nomina.get("periodo_fin", "")).strip()
            try:
                periodo_inicio_dt = datetime.strptime(periodo_inicio_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue
            try:
                periodo_fin_dt = datetime.strptime(periodo_fin_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue

            # Criterio de intersección con el mes del reporte.
            if periodo_inicio_dt > mes_fin or periodo_fin_dt < mes_inicio:
                continue

            empleados_raw = nomina.get("empleados", [])
            empleados_norm = set()
            if isinstance(empleados_raw, list):
                for emp in empleados_raw:
                    nom = _normalize_name(emp)
                    if not nom:
                        continue
                    alias = alias_service.obtener_alias(nom)
                    nom = _normalize_name(alias) if alias else nom
                    if nom:
                        empleados_norm.add(nom)
            semanas.append(
                {
                    "cliente": c,
                    "periodo_inicio": periodo_inicio_raw,
                    "periodo_fin": periodo_fin_raw,
                    "_periodo_inicio_dt": periodo_inicio_dt,
                    "_periodo_fin_dt": periodo_fin_dt,
                    "empleados": empleados_norm,
                }
            )
        semanas.sort(key=lambda s: s.get("_periodo_inicio_dt") or date.min)

        if not semanas:
            base_comp = calcular_completitud_mes(cliente, mes, anio)
            return {
                "cliente": cliente,
                "mes": mes,
                "anio": anio,
                "completo": bool(base_comp.get("completo")),
                "advertencia_completitud": base_comp.get("advertencia"),
                "total_personal_mes": 0,
                "fijos": [],
                "rotativos": [],
                "similitudes_detectadas": [],
                "semanas": [],
            }

        semana_sets = [s["empleados"] for s in semanas]
        todos_del_mes: set[str] = set().union(*semana_sets)

        comparativos_mes: list[dict[str, Any]] = []
        for comparativo in _iter_json_dicts(COMPARATIVOS_DIR):
            c_cliente = str(comparativo.get("cliente", "")).strip()
            if c_cliente not in clientes_objetivo:
                continue
            c_inicio_raw = str(comparativo.get("periodo_inicio", "")).strip()
            c_fin_raw = str(comparativo.get("periodo_fin", "")).strip()
            try:
                c_inicio_dt = datetime.strptime(c_inicio_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue
            try:
                c_fin_dt = datetime.strptime(c_fin_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                continue
            if c_inicio_dt > mes_fin or c_fin_dt < mes_inicio:
                continue
            comparativos_mes.append(comparativo)

        altas_internas: dict[str, list[date]] = {}
        bajas_internas: dict[str, list[date]] = {}
        for comparativo in comparativos_mes:
            alta_raw = str(comparativo.get("periodo_inicio", "")).strip()
            baja_raw = str(comparativo.get("fecha_baja_asumida", "")).strip()
            alta_dt = None
            baja_dt = None
            try:
                alta_dt = datetime.strptime(alta_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                alta_dt = None
            try:
                baja_dt = datetime.strptime(baja_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                baja_dt = None

            for raw in comparativo.get("altas", []):
                nom_base = _normalize_name(raw)
                if not nom_base:
                    continue
                alias = alias_service.obtener_alias(nom_base)
                nombre = _normalize_name(alias) if alias else nom_base
                if alta_dt and mes_inicio <= alta_dt <= mes_fin:
                    altas_internas.setdefault(nombre, []).append(alta_dt)

            for raw in comparativo.get("bajas", []):
                nom_base = _normalize_name(raw)
                if not nom_base:
                    continue
                alias = alias_service.obtener_alias(nom_base)
                nombre = _normalize_name(alias) if alias else nom_base
                if baja_dt and mes_inicio <= baja_dt <= mes_fin:
                    bajas_internas.setdefault(nombre, []).append(baja_dt)

        for nombre in list(altas_internas.keys()):
            altas_internas[nombre] = sorted(set(altas_internas[nombre]))
        for nombre in list(bajas_internas.keys()):
            bajas_internas[nombre] = sorted(set(bajas_internas[nombre]))

        total_semanas = len(semanas)
        fijos: list[dict[str, Any]] = []
        rotativos: list[dict[str, Any]] = []
        periodos_ordenados = [s.get("periodo_inicio", "") for s in semanas]

        for nombre in sorted(todos_del_mes):
            presencia_por_periodo = [nombre in s.get("empleados", set()) for s in semanas]
            semanas_presente = [s.get("periodo_inicio", "") for idx, s in enumerate(semanas) if presencia_por_periodo[idx]]
            en_todas = len(semanas_presente) == total_semanas
            trabajador_hc = buscar_trabajador(nombre)
            en_headcount = trabajador_hc is not None

            if en_todas:
                alerta = None if en_headcount else "No encontrado en Headcount"
                fijos.append(
                    {
                        "nombre": nombre,
                        "en_headcount": en_headcount,
                        "alerta": alerta,
                    }
                )
                continue

            alertas: list[str] = []
            bloques_presencia = 0
            en_bloque = False
            indices_inicio_bloques: list[int] = []
            indices_fin_bloques: list[int] = []
            for idx, presente in enumerate(presencia_por_periodo):
                if presente and not en_bloque:
                    bloques_presencia += 1
                    en_bloque = True
                    indices_inicio_bloques.append(idx)
                if en_bloque and (not presente):
                    en_bloque = False
                    indices_fin_bloques.append(idx - 1)
            if en_bloque:
                indices_fin_bloques.append(len(presencia_por_periodo) - 1)

            tiene_reingreso = bloques_presencia > 1

            altas_mes_disponibles = list(altas_internas.get(nombre, []))
            bajas_mes_disponibles = list(bajas_internas.get(nombre, []))
            fecha_hc_raw = str((trabajador_hc or {}).get("fecha_ingreso", "")).strip() if trabajador_hc else ""
            fecha_hc_dt = None
            try:
                fecha_hc_dt = datetime.strptime(fecha_hc_raw, "%d/%m/%Y").date()
            except (TypeError, ValueError):
                fecha_hc_dt = None
            fecha_hc_mes = fecha_hc_dt if (fecha_hc_dt and mes_inicio <= fecha_hc_dt <= mes_fin) else None
            existe_alta_previa = bool(fecha_hc_dt and fecha_hc_dt < mes_inicio)

            fechas_alta_dt: list[date] = []
            fechas_baja_dt: list[date] = []

            for pos, idx_inicio in enumerate(indices_inicio_bloques):
                semana_inicio_dt = semanas[idx_inicio].get("_periodo_inicio_dt")
                if not isinstance(semana_inicio_dt, date):
                    continue
                if idx_inicio > 0:
                    fecha_alta_bloque = max(semana_inicio_dt, mes_inicio)
                    if mes_inicio <= fecha_alta_bloque <= mes_fin:
                        fechas_alta_dt.append(fecha_alta_bloque)
                    continue

                # Primer bloque: solo mostrar alta si cae dentro del mes.
                fecha_alta_bloque = None
                if pos == 0 and fecha_hc_mes is not None:
                    fecha_alta_bloque = fecha_hc_mes
                elif altas_mes_disponibles:
                    fecha_alta_bloque = altas_mes_disponibles.pop(0)

                if fecha_alta_bloque is not None and mes_inicio <= fecha_alta_bloque <= mes_fin:
                    fechas_alta_dt.append(fecha_alta_bloque)
                elif not existe_alta_previa:
                    alertas.append("Sin fecha de alta - requiere captura manual")

            for idx_fin in indices_fin_bloques:
                # Solo hay baja si desaparece dentro del mes (hay periodo posterior).
                if idx_fin >= len(semanas) - 1:
                    continue
                fecha_baja_bloque = None
                if bajas_mes_disponibles:
                    fecha_baja_bloque = bajas_mes_disponibles.pop(0)

                if fecha_baja_bloque is not None and mes_inicio <= fecha_baja_bloque <= mes_fin:
                    fechas_baja_dt.append(fecha_baja_bloque)
                else:
                    alertas.append("Sin fecha de baja - requiere captura manual")

            # Regla de rango: conservar únicamente fechas dentro del mes.
            fechas_alta_dt = [d for d in fechas_alta_dt if mes_inicio <= d <= mes_fin]
            fechas_baja_dt = [d for d in fechas_baja_dt if mes_inicio <= d <= mes_fin]
            fechas_alta = [_ddmmyyyy(d) for d in fechas_alta_dt]
            fechas_baja = [_ddmmyyyy(d) for d in fechas_baja_dt]
            alertas = sorted(set(alertas))

            rotativos.append(
                {
                    "nombre": nombre,
                    "tiene_reingreso": tiene_reingreso,
                    "fechas_alta": fechas_alta,
                    "fechas_baja": fechas_baja,
                    # Compatibilidad con consumidores previos.
                    "fecha_alta": fechas_alta[0] if fechas_alta else None,
                    "fecha_baja": fechas_baja[0] if fechas_baja else None,
                    "alertas": alertas,
                    "semanas_presente": semanas_presente,
                    "presencia_por_periodo": presencia_por_periodo,
                    "en_headcount": en_headcount,
                }
            )

        similitudes_detectadas: list[dict[str, Any]] = []
        nombres_reporte = sorted(todos_del_mes)
        for i in range(len(nombres_reporte)):
            for j in range(i + 1, len(nombres_reporte)):
                a = nombres_reporte[i]
                b = nombres_reporte[j]
                if a == b:
                    continue
                score = float(fuzz.ratio(a, b))
                if score >= 88:
                    similitudes_detectadas.append({"nombre_a": a, "nombre_b": b, "score": score})
        similitudes_detectadas.sort(key=lambda x: float(x.get("score", 0)), reverse=True)

        comp = calcular_completitud_mes(cliente, mes, anio)
        return {
            "cliente": cliente,
            "mes": mes,
            "anio": anio,
            "completo": bool(comp.get("completo")),
            "advertencia_completitud": comp.get("advertencia"),
            "total_personal_mes": len(todos_del_mes),
            "fijos": fijos,
            "rotativos": rotativos,
            "similitudes_detectadas": similitudes_detectadas,
            "semanas": periodos_ordenados,
        }
    except Exception as exc:
        raise ValueError(f"No se pudo generar reporte mensual v2: {exc}") from exc


def guardar_reporte_mensual(reporte_dict: dict[str, Any]) -> dict[str, Any]:
    try:
        os.makedirs(REPORTES_MENSUALES_DIR, exist_ok=True)
        cliente = _safe_slug_cliente(str(reporte_dict.get("cliente", "")))
        anio = int(reporte_dict.get("anio"))
        mes = int(reporte_dict.get("mes"))
        path = os.path.join(REPORTES_MENSUALES_DIR, f"{cliente}_{anio:04d}-{mes:02d}.json")
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(reporte_dict, fh, ensure_ascii=False, indent=2)
        return reporte_dict
    except Exception as exc:
        raise ValueError(f"No se pudo guardar reporte mensual: {exc}") from exc


def obtener_historial_reportes(cliente: str | None = None) -> list[dict[str, Any]]:
    try:
        os.makedirs(REPORTES_MENSUALES_DIR, exist_ok=True)
        items: list[dict[str, Any]] = []
        for rep in _iter_json_dicts(REPORTES_MENSUALES_DIR):
            cli = str(rep.get("cliente", "")).strip()
            if cliente and cli.casefold() != str(cliente).strip().casefold():
                continue
            fijos = rep.get("fijos") if isinstance(rep.get("fijos"), list) else []
            rot = rep.get("rotativos") if isinstance(rep.get("rotativos"), list) else []
            items.append(
                {
                    "cliente": cli,
                    "mes": int(rep.get("mes", 0)),
                    "anio": int(rep.get("anio", 0)),
                    "total_personal_mes": int(rep.get("total_personal_mes", len(fijos) + len(rot))),
                    "fijos_count": len(fijos),
                    "rotativos_count": len(rot),
                    "completo": bool(rep.get("completo", False)),
                }
            )
        items.sort(key=lambda x: (int(x.get("anio", 0)), int(x.get("mes", 0))), reverse=True)
        return items
    except Exception as exc:
        raise ValueError(f"No se pudo obtener historial de reportes mensuales: {exc}") from exc
