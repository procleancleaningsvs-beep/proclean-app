from __future__ import annotations

import sqlite3
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.facturacion.db import adjuntos_for_facturas, dashboard_stats, list_notas_credito
from modules.facturacion.normalize import parse_alertas_json

MES_NOMBRES = (
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)

_FILL_HEADER = PatternFill("solid", fgColor="1E3A5F")
_FILL_SOFT = PatternFill("solid", fgColor="E8F4FC")
_FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
_FONT_TITLE = Font(bold=True, size=14, color="1E3A5F")
_THIN = Side(style="thin", color="B8D4E8")


def _style_header(cell) -> None:
    cell.fill = _FILL_HEADER
    cell.font = _FONT_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_body(cell, *, alt: bool = False) -> None:
    if alt:
        cell.fill = _FILL_SOFT
    cell.font = Font(size=10, color="334155")
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_facturacion_export_bytes(
    conn: sqlite3.Connection,
    *,
    mes: int,
    anio: int,
) -> bytes:
    stats = dashboard_stats(conn, mes=mes, anio=anio)
    rows = conn.execute(
        """
        SELECT * FROM facturacion_facturas
        WHERE es_factura_activa = 1 AND mes = ? AND anio = ?
        ORDER BY cliente, numero_factura
        """,
        (mes, anio),
    ).fetchall()
    ids = [int(r["id"]) for r in rows]
    adjm = adjuntos_for_facturas(conn, ids)

    wb = openpyxl.Workbook()
    ws_d = wb.active
    ws_d.title = "Dashboard"
    mes_nombre = MES_NOMBRES[mes] if 1 <= mes <= 12 else str(mes)
    ws_d["A1"] = f"Facturación ProClean — {mes_nombre} {anio}"
    ws_d["A1"].font = _FONT_TITLE
    ws_d.merge_cells("A1:F1")

    labels = [
        ("Avance (LISTO)", f"{stats['avance_pct']}%"),
        ("Total facturas", stats["total"]),
        ("Listas", stats["listo"]),
        ("Pendientes", stats["pendientes"]),
        ("Portal", stats["portal"]),
        ("Pendiente NR", stats["pendiente_nr"]),
        ("Sin PO/OC (alerta)", stats["sin_po_oc"]),
        ("Atoradas / atención", stats["atoradas"]),
        ("Pagadas (pago)", stats["pagadas"]),
        ("Refacturaciones (vínculo)", stats["refacturaciones"]),
    ]
    r0 = 3
    for i, (k, v) in enumerate(labels):
        ws_d.cell(r0 + i, 1, k).font = Font(bold=True)
        ws_d.cell(r0 + i, 2, v)
    ws_d.column_dimensions["A"].width = 28
    ws_d.column_dimensions["B"].width = 18

    ws_f = wb.create_sheet("Facturas")
    headers = [
        "ID",
        "Mes",
        "Año",
        "Cliente",
        "Planta",
        "Usuario",
        "Factura",
        "PO/OC",
        "Subtotal",
        "IVA",
        "Total",
        "Fecha factura",
        "Vencimiento",
        "Estatus operativo",
        "Estatus pago",
        "Alertas",
        "Comentarios",
        "PDF",
        "XML",
        "Req. portal",
        "Activa",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws_f.cell(1, col, h)
        _style_header(c)
    for ri, row in enumerate(rows, start=2):
        adj = adjm.get(int(row["id"]), {})
        pdf = "Sí" if adj.get("pdf") else "No"
        xml = "Sí" if adj.get("xml") else "No"
        vals = [
            row["id"],
            row["mes"],
            row["anio"],
            row["cliente"],
            row["planta_servicio"],
            row["usuario_contacto"],
            row["numero_factura"],
            row["po_oc"],
            row["subtotal"],
            row["iva"],
            row["total"],
            row["fecha_factura"],
            row["fecha_vencimiento"],
            row["estatus_operativo"],
            row["estatus_pago"],
            ", ".join(parse_alertas_json(row["alertas_json"])),
            row["comentarios"],
            pdf,
            xml,
            "Sí" if row["requiere_portal"] else "No",
            "Sí" if row["es_factura_activa"] else "No",
        ]
        for col, v in enumerate(vals, start=1):
            c = ws_f.cell(ri, col, v)
            _style_body(c, alt=ri % 2 == 0)
    for col in range(1, len(headers) + 1):
        ws_f.column_dimensions[get_column_letter(col)].width = 14

    ref_rows = conn.execute(
        """
        SELECT o.id, o.numero_factura AS numero_anterior, o.cliente, o.mes, o.anio,
               o.factura_reemplazada_por_id,
               n.numero_factura AS numero_nueva,
               n.refacturacion_motivo, n.refacturacion_fecha
        FROM facturacion_facturas o
        LEFT JOIN facturacion_facturas n ON n.id = o.factura_reemplazada_por_id
        WHERE o.mes = ? AND o.anio = ? AND o.factura_reemplazada_por_id IS NOT NULL
        """,
        (mes, anio),
    ).fetchall()
    if ref_rows:
        ws_r = wb.create_sheet("Refacturaciones")
        rh = ["ID anterior", "Factura anterior", "Cliente", "Mes", "Año", "ID nueva", "Factura nueva", "Motivo", "Fecha"]
        for col, h in enumerate(rh, start=1):
            c = ws_r.cell(1, col, h)
            _style_header(c)
        for ri, row in enumerate(ref_rows, start=2):
            vals = [
                row["id"],
                row["numero_anterior"],
                row["cliente"],
                row["mes"],
                row["anio"],
                row["factura_reemplazada_por_id"],
                row["numero_nueva"],
                row["refacturacion_motivo"],
                row["refacturacion_fecha"],
            ]
            for col, v in enumerate(vals, start=1):
                _style_body(ws_r.cell(ri, col, v), alt=ri % 2 == 0)

    adj_all = conn.execute(
        """
        SELECT a.*, f.numero_factura, f.cliente
        FROM facturacion_adjuntos a
        JOIN facturacion_facturas f ON f.id = a.factura_id
        WHERE f.mes = ? AND f.anio = ?
        """,
        (mes, anio),
    ).fetchall()
    if adj_all:
        ws_a = wb.create_sheet("Adjuntos")
        ah = ["Factura ID", "Cliente", "Número", "Tipo", "Archivo original", "Ruta interna"]
        for col, h in enumerate(ah, start=1):
            _style_header(ws_a.cell(1, col, h))
        for ri, row in enumerate(adj_all, start=2):
            vals = [
                row["factura_id"],
                row["cliente"],
                row["numero_factura"],
                row["tipo"],
                row["original_name"],
                row["file_path"],
            ]
            for col, v in enumerate(vals, start=1):
                _style_body(ws_a.cell(ri, col, v), alt=ri % 2 == 0)

    notas = [n for n in list_notas_credito(conn, limit=500) if (n.get("anio") == anio and (n.get("mes") in (mes, None)))]
    if notas:
        ws_n = wb.create_sheet("NotasCredito")
        nh = ["ID", "Cliente", "Nota", "Factura ID", "Monto", "Fecha", "Comentario"]
        for col, h in enumerate(nh, start=1):
            _style_header(ws_n.cell(1, col, h))
        for ri, row in enumerate(notas, start=2):
            vals = [
                row["id"],
                row["cliente"],
                row["numero_nota"],
                row["factura_id"],
                row["monto"],
                row["fecha"],
                row["comentario"],
            ]
            for col, v in enumerate(vals, start=1):
                _style_body(ws_n.cell(ri, col, v), alt=ri % 2 == 0)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
