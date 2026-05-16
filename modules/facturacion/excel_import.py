from __future__ import annotations

import re
import sqlite3
from collections import Counter
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl

from modules.facturacion.cliente_catalog import CatalogMaps, load_catalog_maps, resolve_cliente_principal
from modules.facturacion.config import CLIENTE_POR_CLASIFICAR, DOMINIOS_CORREO_PUBLICO
from modules.facturacion.db import insert_factura, insert_nota_credito, record_correo_learned_from_import
from modules.facturacion.normalize import (
    alertas_desde_texto_excel,
    coerce_operativo_o_default,
    fix_cliente_name,
    merge_alertas,
    norm_key,
    normalize_estatus_operativo,
    normalize_estatus_pago,
    parse_mes_num,
    split_operativo_y_pago,
)

MES_SHEETS = frozenset(
    {"ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"}
)


def _norm_cell(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and str(x).endswith(".0"):
        try:
            return str(int(x))
        except ValueError:
            pass
    return str(x).strip()


def _fecha_sql(val: Any) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()[:10]
    return s or None


def _float_val(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", ""))
    except ValueError:
        return None


def _infer_cliente(usuario: str, mes_col: str) -> str | None:
    """
    Infiere cliente desde correo, bloques conocidos (CARRIER/GEPP) o texto en MES que no sea mes calendario.
    Catálogo correo/dominio se aplica en resolve_cliente_principal; aquí solo heurísticas legacy.
    """
    u = (usuario or "").strip().lower()
    m_raw = (mes_col or "").strip()
    m = m_raw.upper()
    if m == "CARRIER":
        return "CARRIER"
    if m in ("GEPP", "GEEP"):
        return "GEPP"
    if "@gepp.com" in u or "gepp.com" in u:
        return "GEPP"
    if "@carrier.com" in u:
        return "CARRIER"
    if "@" in u:
        dom = u.split("@", 1)[-1].strip().lower()
        if dom in DOMINIOS_CORREO_PUBLICO:
            return None
        dom_label = dom.split(".")[0]
        if dom_label:
            return fix_cliente_name(dom_label.upper())
    # Texto en columna MES que no sea un mes (p. ej. sección / cliente en layouts atípicos)
    if m_raw and parse_mes_num(m_raw) is None:
        mk = " ".join(m.split())
        if len(mk) >= 3 and mk not in ("TOTAL", "NONE", "CARRIER", "GEPP", "GEEP"):
            if re.search(r"[A-Za-z]", mk):
                return fix_cliente_name(mk)
    return None


def _try_bloque_cliente_desde_mes(mes_txt: str, maps: CatalogMaps) -> str | None:
    """Fila tipo encabezado: columna MES trae el cliente principal y no hay número de factura válido."""
    from modules.facturacion.cliente_catalog import normalize_razon_key

    raw = (mes_txt or "").strip()
    if not raw or parse_mes_num(raw) is not None:
        return None
    cand = fix_cliente_name(raw)
    nkb = normalize_razon_key(cand)
    if not nkb:
        return None
    if nkb in maps.known_principales_norm or nkb in maps.bloque_hints_norm:
        return cand
    return None


def _find_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    best_row = 0
    mapping: dict[str, int] = {}
    for ri, row in enumerate(ws.iter_rows(max_row=40, values_only=True), start=1):
        cells = [_norm_cell(c).upper() for c in row]
        if "FACTURA" in cells:
            mapping = {}
            for j, name in enumerate(cells):
                if not name:
                    continue
                key = re.sub(r"\s+", " ", name).strip()
                # El Excel duplica "FACTURA" / "Factura" en columnas de pipeline; conservar la primera (datos reales).
                if key in mapping:
                    continue
                mapping[key] = j
            best_row = ri
            break
    return best_row, mapping


def _col(mapping: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        k = c.upper().strip()
        if k in mapping:
            return mapping[k]
    for mk, idx in mapping.items():
        for c in candidates:
            if mk.replace(" ", "") == c.replace(" ", "").upper():
                return idx
    return None


def _row_vals(row: tuple[Any, ...], mapping: dict[str, int]) -> dict[str, Any]:
    def g(*names: str) -> Any:
        idx = _col(mapping, *names)
        if idx is None:
            return None
        if idx >= len(row):
            return None
        return row[idx]

    return {
        "mes": g("MES"),
        "asistencia": g("ASISTENCIA DE"),
        "planta": g("PLANTA"),
        "usuario": g("USUARIO"),
        "factura": g("FACTURA"),
        "po": g("PO"),
        "subtotal": g("SUBTOTAL"),
        "iva": g("IVA"),
        "total": g("TOTAL"),
        "fecha_factura": g("FECHA FACTURA", "FECHA FACTURA "),
        "fecha_venc": g("FECHA DE VENCIMIENTO", "FECHA DE VENCIMIENTO "),
        "estatus": g("ESTATUS"),
        "comentarios": g("COMENTARIOS"),
        "fecha_pago": g("FECHA DE PAGO", "FECHA DE PAGO "),
        "razon_social": g(
            "RAZON SOCIAL",
            "RAZÓN SOCIAL",
            "RAZON SOCIAL ",
            "NOMBRE O RAZON SOCIAL",
            "NOMBRE O RAZÓN SOCIAL",
            "CLIENTE RAZON SOCIAL",
        ),
    }


def _skip_numero(num: str) -> bool:
    u = num.strip().upper()
    if not u or u in ("TOTAL", "NONE"):
        return True
    if u == "FALTA":
        return True
    if not re.search(r"\d", u):
        return True
    return False


def _import_notas_credito_sheet(
    conn: sqlite3.Connection,
    ws: Any,
    anio: int,
    user_id: int | None,
    now: str,
) -> int:
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cliente = _norm_cell(row[0])
        num = row[1]
        fecha = row[2]
        com = row[3] if len(row) > 3 else None
        if not num:
            continue
        insert_nota_credito(
            conn,
            {
                "cliente": fix_cliente_name(cliente),
                "numero_nota": _norm_cell(num),
                "factura_id": None,
                "monto": None,
                "comentario": _norm_cell(com) if com else None,
                "fecha": _fecha_sql(fecha) or now[:10],
                "mes": None,
                "anio": anio,
            },
            user_id=user_id,
            now=now,
        )
        n += 1
    return n


def import_facturacion_excel(
    conn: sqlite3.Connection,
    content: bytes,
    *,
    anio_default: int,
    user_id: int | None,
    now: str,
    original_filename: str | None = None,
) -> dict[str, Any]:
    bio = BytesIO(content)
    wb = openpyxl.load_workbook(bio, data_only=True)

    hojas_procesadas: list[str] = []
    filas_leidas = 0
    inserted = 0
    duplicados = 0
    por_clasificar = 0
    sin_mes = 0
    omitidas_sin_numero = 0
    errores: list[str] = []
    clientes_detectados: set[str] = set()
    estatus_norm_counter: Counter[str] = Counter()
    advertencias: list[str] = []
    notas_insertadas = 0

    maps = load_catalog_maps(conn)

    for sheet_name in wb.sheetnames:
        key = str(sheet_name).strip().upper()
        if key == "NOTAS DE CREDITO":
            notas_insertadas += _import_notas_credito_sheet(conn, wb[sheet_name], anio_default, user_id, now)
            continue
        if key in ("DASHBOARD", "COMPLEMENTOS DE PAGO"):
            continue
        if key not in MES_SHEETS and parse_mes_num(key) is None:
            continue

        ws = wb[sheet_name]
        header_row, mapping = _find_header_row(ws)
        if not mapping or "FACTURA" not in [k.upper() for k in mapping.keys()]:
            advertencias.append(f"Hoja «{sheet_name}»: sin fila de encabezado con columna FACTURA; omitida.")
            continue

        hojas_procesadas.append(sheet_name)
        sheet_mes_num = parse_mes_num(key)
        context_bloque: str | None = None

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            rv = _row_vals(tuple(row), mapping)
            num_raw = _norm_cell(rv["factura"])
            mes_txt = _norm_cell(rv["mes"])
            usuario = _norm_cell(rv["usuario"])

            if _skip_numero(num_raw):
                bloque = _try_bloque_cliente_desde_mes(mes_txt, maps)
                if bloque:
                    context_bloque = fix_cliente_name(bloque)
                    continue
                omitidas_sin_numero += 1
                continue

            filas_leidas += 1
            razon_excel = _norm_cell(rv.get("razon_social")) or None
            cli_infer = _infer_cliente(usuario, mes_txt)
            cli, razon_guardada = resolve_cliente_principal(
                maps,
                razon_social_excel=razon_excel,
                cli_infer=cli_infer,
                fix_cliente_name_fn=fix_cliente_name,
                por_clasificar=CLIENTE_POR_CLASIFICAR,
                email_contacto=usuario or None,
                bloque_cliente_excel=context_bloque,
            )
            if cli == CLIENTE_POR_CLASIFICAR:
                por_clasificar += 1

            mes_fact = parse_mes_num(mes_txt) or sheet_mes_num or parse_mes_num(key)
            if mes_fact is None:
                sin_mes += 1
                advertencias.append(f"{sheet_name} factura {num_raw}: sin mes deducible.")
                continue

            asist = parse_mes_num(_norm_cell(rv["asistencia"]))

            est_raw = rv["estatus"]
            est_raw_s = str(est_raw).strip() if est_raw is not None else ""
            op_frag, pago_frag = split_operativo_y_pago(est_raw)
            tiene_fp = _fecha_sql(rv["fecha_pago"]) is not None
            est_pago = normalize_estatus_pago(pago_frag, tiene_fecha_pago=tiene_fp)

            op_norm = normalize_estatus_operativo(est_raw if op_frag is None else op_frag)
            if op_norm is None and est_raw:
                op_norm = normalize_estatus_operativo(est_raw)
            if op_norm is None and pago_frag == "PAGADO":
                operativo = "LISTO"
            else:
                operativo = coerce_operativo_o_default(op_norm or est_raw, fallback="EN COLA")

            if "NO PORTAL" in est_raw_s.upper():
                estatus_norm_counter["NO PORTAL → ENVIADO (requiere_portal=0)"] += 1
            elif est_raw_s and norm_key(est_raw_s) != norm_key(operativo):
                estatus_norm_counter[f"{est_raw_s[:80]} → {operativo}"] += 1

            extra_events: list[dict[str, Any]] = []
            req_portal = 0
            if "NO PORTAL" in est_raw_s.upper():
                req_portal = 0
                operativo = "ENVIADO"
                extra_events.append(
                    {
                        "tipo": "NORMALIZACION_IMPORT",
                        "detalle": {
                            "estatus_excel_original": est_raw_s,
                            "estatus_operativo": operativo,
                            "requiere_portal": False,
                            "regla": "NO_PORTAL",
                        },
                    }
                )
            elif operativo == "PORTAL":
                req_portal = 1

            comentarios = _norm_cell(rv["comentarios"]) or None
            if "NO PORTAL" in est_raw_s.upper():
                tag = "[Import] Excel estatus: NO PORTAL (conservado; operativo=ENVIADO, sin portal)"
                comentarios = f"{comentarios} | {tag}" if comentarios else tag
            if cli == CLIENTE_POR_CLASIFICAR:
                tag2 = "[Import] Cliente asignado como POR CLASIFICAR (revisar y corregir)."
                comentarios = f"{comentarios} | {tag2}" if comentarios else tag2

            extra_alerts = alertas_desde_texto_excel(_norm_cell(rv["comentarios"]), est_raw)

            data: dict[str, Any] = {
                "mes": int(mes_fact),
                "anio": int(anio_default),
                "asistencia_mes": int(asist) if asist else None,
                "asistencia_anio": int(anio_default),
                "cliente": cli,
                "razon_social": razon_guardada,
                "planta_servicio": _norm_cell(rv["planta"]) or None,
                "usuario_contacto": usuario or None,
                "numero_factura": num_raw.upper() if num_raw.isalnum() else num_raw,
                "po_oc": _norm_cell(rv["po"]) or None,
                "requiere_portal": req_portal,
                "subtotal": _float_val(rv["subtotal"]),
                "iva": _float_val(rv["iva"]),
                "total": _float_val(rv["total"]),
                "fecha_factura": _fecha_sql(rv["fecha_factura"]),
                "fecha_vencimiento": _fecha_sql(rv["fecha_venc"]),
                "estatus_operativo": operativo,
                "estatus_pago": est_pago,
                "alertas": merge_alertas(extra_alerts, []),
                "comentarios": comentarios,
                "_extra_eventos": extra_events,
            }
            try:
                insert_factura(conn, data, user_id=user_id, now=now)
                inserted += 1
                clientes_detectados.add(cli)
                record_correo_learned_from_import(conn, email=usuario, cliente_principal=cli, now=now)
                maps = load_catalog_maps(conn)
            except sqlite3.IntegrityError:
                duplicados += 1
            except Exception as exc:  # noqa: BLE001
                errores.append(f"{sheet_name} {num_raw}: {exc}")

    wb.close()

    summary: dict[str, Any] = {
        "archivo": original_filename,
        "hojas_procesadas": hojas_procesadas,
        "filas_leidas": filas_leidas,
        "filas_importadas": inserted,
        "duplicados_omitidos": duplicados,
        "por_clasificar_cliente": por_clasificar,
        "omitidas_sin_numero_factura": omitidas_sin_numero,
        "omitidas_sin_mes": sin_mes,
        "filas_con_error": len(errores),
        "detalle_errores": errores[:200],
        "estatus_normalizados": dict(estatus_norm_counter),
        "clientes_detectados": sorted(clientes_detectados),
        "advertencias": advertencias,
        "notas_credito_filas": notas_insertadas,
    }
    return summary
